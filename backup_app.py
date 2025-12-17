"""
Performance Management System - Main Application
A comprehensive PMS built with Streamlit and Supabase
"""
import sys
sys.setrecursionlimit(3000)

import streamlit as st
import pandas as pd
from datetime import datetime, date, timedelta
import pytz
import plotly.express as px
import plotly.graph_objects as go
import calendar
import io
from plotly.subplots import make_subplots
from datetime import datetime, timedelta
import numpy as np
from io import BytesIO
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from dotenv import load_dotenv
from dateutil.relativedelta import relativedelta
from database import get_supabase_client
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import secrets
import os
from datetime import datetime, timedelta
import streamlit.components.v1 as components
from monthly_reminder import start_reminder_scheduler, test_send_reminder
from monthly_reminder import send_goal_completion_email


def safe_float(value, default=None):  # Changed default from 0 to None
    """Safely convert value to float, handling None and invalid types"""
    if value is None or value == '':
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default
    
def get_feedback_giver_role(target_role):
    """Who is allowed to give official feedback to this role?"""
    mapping = {
        'VP': 'CMD',
        'HR': 'VP',
        'Manager': 'VP',
        'Employee': 'Manager',
        'CMD': None
    }
    return mapping.get(target_role)

def get_role_hierarchy():
    """Define role hierarchy levels"""
    return {
        'CMD': 5,
        'VP': 4,
        'HR': 3,
        'Manager': 2,
        'Employee': 1
    }

def create_notification(notification_data):
    """Create a new notification in the database"""
    try:
        result = supabase.table('notifications').insert(notification_data).execute()
        return result.data[0] if result.data else None
    except Exception as e:
        print(f"Error creating notification: {str(e)}")
        return None

def notify_goal_created(goal_data, creator_user):
    """Notify relevant users when a goal is created"""
    creator_role = creator_user['role']
    
    # Employee creates goal -> Notify Manager (pending approval)
    if creator_role == 'Employee':
        if creator_user.get('manager_id'):
            manager = db.get_user_by_id(creator_user['manager_id'])
            if manager:
                create_notification({
                    'user_id': manager['id'],
                    'action_by': creator_user['id'],
                    'action_by_name': creator_user['name'],
                    'action_type': 'goal_created',
                    'details': f"{creator_user['name']} created a new goal: '{goal_data['goal_title']}' (Pending Approval)",
                    'is_read': False,
                    'created_at': datetime.now(IST).isoformat()
                })
    
    # Manager creates goal -> Notify VP
    elif creator_role == 'Manager':
        all_users = db.get_all_users()
        vp_users = [u for u in all_users if u['role'] == 'VP']
        for vp in vp_users:
            create_notification({
                'user_id': vp['id'],
                'action_by': creator_user['id'],
                'action_by_name': creator_user['name'],
                'action_type': 'goal_created',
                'details': f"Manager {creator_user['name']} created goal: '{goal_data['goal_title']}'",
                'is_read': False,
                'created_at': datetime.now(IST).isoformat()
            })
    
    # HR creates goal -> Notify VP
    elif creator_role == 'HR':
        all_users = db.get_all_users()
        vp_users = [u for u in all_users if u['role'] == 'VP']
        for vp in vp_users:
            create_notification({
                'user_id': vp['id'],
                'action_by': creator_user['id'],
                'action_by_name': creator_user['name'],
                'action_type': 'goal_created',
                'details': f"HR {creator_user['name']} created goal: '{goal_data['goal_title']}'",
                'is_read': False,
                'created_at': datetime.now(IST).isoformat()
            })
    
    # VP creates goal -> Notify CMD
    elif creator_role == 'VP':
        all_users = db.get_all_users()
        cmd_users = [u for u in all_users if u['role'] == 'CMD']
        for cmd in cmd_users:
            create_notification({
                'user_id': cmd['id'],
                'action_by': creator_user['id'],
                'action_by_name': creator_user['name'],
                'action_type': 'goal_created',
                'details': f"VP {creator_user['name']} created goal: '{goal_data['goal_title']}'",
                'is_read': False,
                'created_at': datetime.now(IST).isoformat()
            })

def notify_goal_approved(goal, approver_user, employee_user):
    """Notify employee when their goal is approved"""
    create_notification({
        'user_id': employee_user['id'],
        'action_by': approver_user['id'],
        'action_by_name': approver_user['name'],
        'action_type': 'goal_approved',
        'details': f"Your goal '{goal['goal_title']}' has been approved by {approver_user['name']}",
        'is_read': False,
        'created_at': datetime.now(IST).isoformat()
    })
    
    # Also notify HR that an approved goal exists
    all_users = db.get_all_users()
    hr_users = [u for u in all_users if u['role'] == 'HR']
    for hr in hr_users:
        create_notification({
            'user_id': hr['id'],
            'action_by': employee_user['id'],
            'action_by_name': employee_user['name'],
            'action_type': 'goal_created',
            'details': f"{employee_user['name']}'s goal '{goal['goal_title']}' was approved by {approver_user['name']}",
            'is_read': False,
            'created_at': datetime.now(IST).isoformat()
        })

def notify_weekly_achievement_updated(goal, updater_user, week_num):
    """Notify relevant users when weekly achievement is updated"""
    updater_role = updater_user['role']
    goal_owner = db.get_user_by_id(goal['user_id'])
    
    # Employee updates -> Notify Manager
    if updater_role == 'Employee':
        if goal_owner.get('manager_id'):
            manager = db.get_user_by_id(goal_owner['manager_id'])
            if manager:
                create_notification({
                    'user_id': manager['id'],
                    'action_by': updater_user['id'],
                    'action_by_name': updater_user['name'],
                    'action_type': 'weekly_achievement_updated',
                    'details': f"{updater_user['name']} updated Week {week_num} achievement for '{goal['goal_title']}'",
                    'is_read': False,
                    'created_at': datetime.now(IST).isoformat()
                })
        
        # Also notify HR
        all_users = db.get_all_users()
        hr_users = [u for u in all_users if u['role'] == 'HR']
        for hr in hr_users:
            create_notification({
                'user_id': hr['id'],
                'action_by': updater_user['id'],
                'action_by_name': updater_user['name'],
                'action_type': 'weekly_achievement_updated',
                'details': f"{updater_user['name']} updated Week {week_num} achievement for '{goal['goal_title']}'",
                'is_read': False,
                'created_at': datetime.now(IST).isoformat()
            })
    
    # Manager/HR updates -> Notify VP
    elif updater_role in ['Manager', 'HR']:
        all_users = db.get_all_users()
        vp_users = [u for u in all_users if u['role'] == 'VP']
        for vp in vp_users:
            create_notification({
                'user_id': vp['id'],
                'action_by': updater_user['id'],
                'action_by_name': updater_user['name'],
                'action_type': 'weekly_achievement_updated',
                'details': f"{updater_role} {updater_user['name']} updated Week {week_num} achievement for '{goal['goal_title']}'",
                'is_read': False,
                'created_at': datetime.now(IST).isoformat()
            })
    
    # VP updates -> Notify CMD
    elif updater_role == 'VP':
        all_users = db.get_all_users()
        cmd_users = [u for u in all_users if u['role'] == 'CMD']
        for cmd in cmd_users:
            create_notification({
                'user_id': cmd['id'],
                'action_by': updater_user['id'],
                'action_by_name': updater_user['name'],
                'action_type': 'weekly_achievement_updated',
                'details': f"VP {updater_user['name']} updated Week {week_num} achievement for '{goal['goal_title']}'",
                'is_read': False,
                'created_at': datetime.now(IST).isoformat()
            })

def notify_goal_completed(goal, completer_user):
    """Notify relevant users when a goal is completed"""
    completer_role = completer_user['role']
    goal_owner = db.get_user_by_id(goal['user_id'])
    
    # Employee completes -> Notify Manager
    if completer_role == 'Employee':
        if goal_owner.get('manager_id'):
            manager = db.get_user_by_id(goal_owner['manager_id'])
            if manager:
                create_notification({
                    'user_id': manager['id'],
                    'action_by': completer_user['id'],
                    'action_by_name': completer_user['name'],
                    'action_type': 'goal_completed',
                    'details': f"✅ {completer_user['name']} completed goal: '{goal['goal_title']}'",
                    'is_read': False,
                    'created_at': datetime.now(IST).isoformat()
                })
        
        # Notify HR
        all_users = db.get_all_users()
        hr_users = [u for u in all_users if u['role'] == 'HR']
        for hr in hr_users:
            create_notification({
                'user_id': hr['id'],
                'action_by': completer_user['id'],
                'action_by_name': completer_user['name'],
                'action_type': 'goal_completed',
                'details': f"✅ {completer_user['name']} completed goal: '{goal['goal_title']}'",
                'is_read': False,
                'created_at': datetime.now(IST).isoformat()
            })
    
    # Manager/HR completes -> Notify VP
    elif completer_role in ['Manager', 'HR']:
        all_users = db.get_all_users()
        vp_users = [u for u in all_users if u['role'] == 'VP']
        for vp in vp_users:
            create_notification({
                'user_id': vp['id'],
                'action_by': completer_user['id'],
                'action_by_name': completer_user['name'],
                'action_type': 'goal_completed',
                'details': f"✅ {completer_role} {completer_user['name']} completed goal: '{goal['goal_title']}'",
                'is_read': False,
                'created_at': datetime.now(IST).isoformat()
            })
    
    # VP completes -> Notify CMD
    elif completer_role == 'VP':
        all_users = db.get_all_users()
        cmd_users = [u for u in all_users if u['role'] == 'CMD']
        for cmd in cmd_users:
            create_notification({
                'user_id': cmd['id'],
                'action_by': completer_user['id'],
                'action_by_name': completer_user['name'],
                'action_type': 'goal_completed',
                'details': f"✅ VP {completer_user['name']} completed goal: '{goal['goal_title']}'",
                'is_read': False,
                'created_at': datetime.now(IST).isoformat()
            })

def notify_goal_not_completed(goal, goal_owner):
    """Notify when goal is not completed by end date"""
    owner_role = goal_owner['role']
    
    # Notify the goal owner themselves
    create_notification({
        'user_id': goal_owner['id'],
        'action_by': goal_owner['id'],
        'action_by_name': 'System',
        'action_type': 'goal_not_completed',
        'details': f"⚠️ Your goal '{goal['goal_title']}' was not completed by the deadline",
        'is_read': False,
        'created_at': datetime.now(IST).isoformat()
    })
    
    # Employee -> Notify Manager
    if owner_role == 'Employee':
        if goal_owner.get('manager_id'):
            manager = db.get_user_by_id(goal_owner['manager_id'])
            if manager:
                create_notification({
                    'user_id': manager['id'],
                    'action_by': goal_owner['id'],
                    'action_by_name': goal_owner['name'],
                    'action_type': 'goal_not_completed',
                    'details': f"⚠️ {goal_owner['name']}'s goal '{goal['goal_title']}' was not completed by deadline",
                    'is_read': False,
                    'created_at': datetime.now(IST).isoformat()
                })
        
        # Notify HR
        all_users = db.get_all_users()
        hr_users = [u for u in all_users if u['role'] == 'HR']
        for hr in hr_users:
            create_notification({
                'user_id': hr['id'],
                'action_by': goal_owner['id'],
                'action_by_name': goal_owner['name'],
                'action_type': 'goal_not_completed',
                'details': f"⚠️ {goal_owner['name']}'s goal '{goal['goal_title']}' was not completed by deadline",
                'is_read': False,
                'created_at': datetime.now(IST).isoformat()
            })
    
    # Manager/HR -> Notify VP
    elif owner_role in ['Manager', 'HR']:
        all_users = db.get_all_users()
        vp_users = [u for u in all_users if u['role'] == 'VP']
        for vp in vp_users:
            create_notification({
                'user_id': vp['id'],
                'action_by': goal_owner['id'],
                'action_by_name': goal_owner['name'],
                'action_type': 'goal_not_completed',
                'details': f"⚠️ {owner_role} {goal_owner['name']}'s goal '{goal['goal_title']}' was not completed by deadline",
                'is_read': False,
                'created_at': datetime.now(IST).isoformat()
            })
    
    # VP -> Notify CMD
    elif owner_role == 'VP':
        all_users = db.get_all_users()
        cmd_users = [u for u in all_users if u['role'] == 'CMD']
        for cmd in cmd_users:
            create_notification({
                'user_id': cmd['id'],
                'action_by': goal_owner['id'],
                'action_by_name': goal_owner['name'],
                'action_type': 'goal_not_completed',
                'details': f"⚠️ VP {goal_owner['name']}'s goal '{goal['goal_title']}' was not completed by deadline",
                'is_read': False,
                'created_at': datetime.now(IST).isoformat()
            })

def notify_feedback_given(goal, feedback_giver, goal_owner):
    """Notify when feedback is given"""
    giver_role = feedback_giver['role']
    owner_role = goal_owner['role']
    
    # Always notify the goal owner
    create_notification({
        'user_id': goal_owner['id'],
        'action_by': feedback_giver['id'],
        'action_by_name': feedback_giver['name'],
        'action_type': 'feedback_received',
        'details': f"{feedback_giver['name']} gave feedback on your goal: '{goal['goal_title']}'",
        'is_read': False,
        'created_at': datetime.now(IST).isoformat()
    })
    
    # Manager gives feedback to Employee -> Notify HR
    if giver_role == 'Manager' and owner_role == 'Employee':
        all_users = db.get_all_users()
        hr_users = [u for u in all_users if u['role'] == 'HR']
        for hr in hr_users:
            create_notification({
                'user_id': hr['id'],
                'action_by': feedback_giver['id'],
                'action_by_name': feedback_giver['name'],
                'action_type': 'feedback_given',
                'details': f"Manager {feedback_giver['name']} gave feedback to {goal_owner['name']} on '{goal['goal_title']}'",
                'is_read': False,
                'created_at': datetime.now(IST).isoformat()
            })
    
    # VP gives feedback to Manager/HR -> Notify them
    elif giver_role == 'VP' and owner_role in ['Manager', 'HR']:
        create_notification({
            'user_id': goal_owner['id'],
            'action_by': feedback_giver['id'],
            'action_by_name': feedback_giver['name'],
            'action_type': 'feedback_received',
            'details': f"VP {feedback_giver['name']} gave feedback on your goal: '{goal['goal_title']}'",
            'is_read': False,
            'created_at': datetime.now(IST).isoformat()
        })
    
    # CMD gives feedback to VP -> Notify VP
    elif giver_role == 'CMD' and owner_role == 'VP':
        create_notification({
            'user_id': goal_owner['id'],
            'action_by': feedback_giver['id'],
            'action_by_name': feedback_giver['name'],
            'action_type': 'feedback_received',
            'details': f"CMD {feedback_giver['name']} gave feedback on your goal: '{goal['goal_title']}'",
            'is_read': False,
            'created_at': datetime.now(IST).isoformat()
        })

def notify_feedback_reply(feedback, replier_user, original_feedback_giver):
    """Notify when someone replies to feedback"""
    replier_role = replier_user['role']
    
    # Notify the original feedback giver
    create_notification({
        'user_id': original_feedback_giver['id'],
        'action_by': replier_user['id'],
        'action_by_name': replier_user['name'],
        'action_type': 'feedback_reply',
        'details': f"{replier_user['name']} replied to your feedback",
        'is_read': False,
        'created_at': datetime.now(IST).isoformat()
    })
    
    # Employee replies to Manager feedback -> Manager already notified above
    # VP replies to CMD feedback -> CMD already notified above
    # CMD replies to VP feedback -> VP already notified above

def notify_goal_edited(goal, editor_user, goal_owner):
    """Notify when a goal is edited"""
    editor_role = editor_user['role']
    owner_role = goal_owner['role']
    
    # If editing someone else's goal
    if editor_user['id'] != goal_owner['id']:
        # Notify the goal owner
        create_notification({
            'user_id': goal_owner['id'],
            'action_by': editor_user['id'],
            'action_by_name': editor_user['name'],
            'action_type': 'goal_edited',
            'details': f"{editor_user['name']} edited your goal: '{goal['goal_title']}'",
            'is_read': False,
            'created_at': datetime.now(IST).isoformat()
        })
        
        # Notify HR when CMD/VP/Manager edits goals
        if editor_role in ['CMD', 'VP', 'Manager']:
            all_users = db.get_all_users()
            hr_users = [u for u in all_users if u['role'] == 'HR']
            for hr in hr_users:
                create_notification({
                    'user_id': hr['id'],
                    'action_by': editor_user['id'],
                    'action_by_name': editor_user['name'],
                    'action_type': 'goal_edited',
                    'details': f"{editor_role} {editor_user['name']} edited {goal_owner['name']}'s goal: '{goal['goal_title']}'",
                    'is_read': False,
                    'created_at': datetime.now(IST).isoformat()
                })
    
    # If Manager edits own goal -> Notify VP
    elif editor_role == 'Manager':
        all_users = db.get_all_users()
        vp_users = [u for u in all_users if u['role'] == 'VP']
        for vp in vp_users:
            create_notification({
                'user_id': vp['id'],
                'action_by': editor_user['id'],
                'action_by_name': editor_user['name'],
                'action_type': 'goal_edited',
                'details': f"Manager {editor_user['name']} edited their goal: '{goal['goal_title']}'",
                'is_read': False,
                'created_at': datetime.now(IST).isoformat()
            })
    
    # If VP edits own goal -> Notify CMD
    elif editor_role == 'VP':
        all_users = db.get_all_users()
        cmd_users = [u for u in all_users if u['role'] == 'CMD']
        for cmd in cmd_users:
            create_notification({
                'user_id': cmd['id'],
                'action_by': editor_user['id'],
                'action_by_name': editor_user['name'],
                'action_type': 'goal_edited',
                'details': f"VP {editor_user['name']} edited their goal: '{goal['goal_title']}'",
                'is_read': False,
                'created_at': datetime.now(IST).isoformat()
            })

def notify_goal_deleted(goal, deleter_user, goal_owner):
    """Notify when a goal is deleted"""
    deleter_role = deleter_user['role']
    owner_role = goal_owner['role']
    
    # If deleting someone else's goal
    if deleter_user['id'] != goal_owner['id']:
        # Notify the goal owner
        create_notification({
            'user_id': goal_owner['id'],
            'action_by': deleter_user['id'],
            'action_by_name': deleter_user['name'],
            'action_type': 'goal_deleted',
            'details': f"⚠️ {deleter_user['name']} deleted your goal: '{goal['goal_title']}'",
            'is_read': False,
            'created_at': datetime.now(IST).isoformat()
        })
        
        # Notify HR when CMD/VP/Manager deletes goals
        if deleter_role in ['CMD', 'VP', 'Manager']:
            all_users = db.get_all_users()
            hr_users = [u for u in all_users if u['role'] == 'HR']
            for hr in hr_users:
                create_notification({
                    'user_id': hr['id'],
                    'action_by': deleter_user['id'],
                    'action_by_name': deleter_user['name'],
                    'action_type': 'goal_deleted',
                    'details': f"⚠️ {deleter_role} {deleter_user['name']} deleted {goal_owner['name']}'s goal: '{goal['goal_title']}'",
                    'is_read': False,
                    'created_at': datetime.now(IST).isoformat()
                })
    
    # If Manager deletes own goal -> Notify VP
    elif deleter_role == 'Manager':
        all_users = db.get_all_users()
        vp_users = [u for u in all_users if u['role'] == 'VP']
        for vp in vp_users:
            create_notification({
                'user_id': vp['id'],
                'action_by': deleter_user['id'],
                'action_by_name': deleter_user['name'],
                'action_type': 'goal_deleted',
                'details': f"⚠️ Manager {deleter_user['name']} deleted their goal: '{goal['goal_title']}'",
                'is_read': False,
                'created_at': datetime.now(IST).isoformat()
            })
    
    # If VP deletes own goal -> Notify CMD
    elif deleter_role == 'VP':
        all_users = db.get_all_users()
        cmd_users = [u for u in all_users if u['role'] == 'CMD']
        for cmd in cmd_users:
            create_notification({
                'user_id': cmd['id'],
                'action_by': deleter_user['id'],
                'action_by_name': deleter_user['name'],
                'action_type': 'goal_deleted',
                'details': f"⚠️ VP {deleter_user['name']} deleted their goal: '{goal['goal_title']}'",
                'is_read': False,
                'created_at': datetime.now(IST).isoformat()
            })

def notify_goal_due_soon(goal, goal_owner, days_remaining):
    """Notify when goal is due in X days"""
    create_notification({
        'user_id': goal_owner['id'],
        'action_by': goal_owner['id'],
        'action_by_name': 'System',
        'action_type': 'goal_due_soon',
        'details': f"⏰ Your goal '{goal['goal_title']}' is due in {days_remaining} day(s)",
        'is_read': False,
        'created_at': datetime.now(IST).isoformat()
    })

def notify_goal_not_updated(goal, goal_owner, week_num):
    """Notify when weekly achievement is not updated (shows -)"""
    owner_role = goal_owner['role']
    
    # Notify the goal owner
    create_notification({
        'user_id': goal_owner['id'],
        'action_by': goal_owner['id'],
        'action_by_name': 'System',
        'action_type': 'goal_not_updated',
        'details': f"⚠️ Week {week_num} achievement not updated for goal: '{goal['goal_title']}'",
        'is_read': False,
        'created_at': datetime.now(IST).isoformat()
    })
    
    # Employee -> Notify Manager
    if owner_role == 'Employee':
        if goal_owner.get('manager_id'):
            manager = db.get_user_by_id(goal_owner['manager_id'])
            if manager:
                create_notification({
                    'user_id': manager['id'],
                    'action_by': goal_owner['id'],
                    'action_by_name': goal_owner['name'],
                    'action_type': 'goal_not_updated',
                    'details': f"⚠️ {goal_owner['name']} has not updated Week {week_num} for '{goal['goal_title']}'",
                    'is_read': False,
                    'created_at': datetime.now(IST).isoformat()
                })

def check_and_notify_due_dates_and_missing_updates():
    """Daily check for goals due soon and missing weekly updates"""
    today = date.today()
    all_users = db.get_all_users()
    
    for user in all_users:
        goals = db.get_user_all_goals(user['id'])
        
        for goal in goals:
            if goal.get('status') == 'Active':
                # Check due dates
                end_date_str = goal.get('end_date')
                if end_date_str:
                    try:
                        end_date = datetime.strptime(str(end_date_str), '%Y-%m-%d').date()
                        days_remaining = (end_date - today).days
                        
                        # Notify if due in 5, 4, 3, 2, or 1 days
                        if 1 <= days_remaining <= 5:
                            notify_goal_due_soon(goal, user, days_remaining)
                        
                        # Check if goal is past due and not completed
                        if days_remaining < 0:
                            monthly_achievement = goal.get('monthly_achievement', 0)
                            monthly_target = goal.get('monthly_target', 1)
                            
                            if monthly_achievement is None:
                                monthly_achievement = 0
                            
                            progress = (monthly_achievement / monthly_target * 100) if monthly_target > 0 else 0
                            
                            if progress < 100:
                                notify_goal_not_completed(goal, user)
                    except:
                        pass
                
                # Check for missing weekly updates (-)
                current_year = today.year
                current_month = today.month
                
                if goal['year'] == current_year and goal.get('month') == current_month:
                    # Check each week
                    for week_num in range(1, 5):
                        week_achievement = goal.get(f'week{week_num}_achievement')
                        
                        # If achievement is None (showing as '-'), notify
                        if week_achievement is None:
                            # Only notify if we're past that week's date
                            week_start, week_end = get_week_dates(current_year, current_month, week_num)
                            
                            if today > week_end:
                                notify_goal_not_updated(goal, user, week_num)
             # Check for all roles (Manager, HR, VP, CMD) about their own goals
        if user['role'] in ['Manager', 'HR', 'VP', 'CMD']:
            goals = db.get_user_all_goals(user['id'])
            
            for goal in goals:
                if goal.get('status') == 'Active':
                    end_date_str = goal.get('end_date')
                    if end_date_str:
                        try:
                            end_date = datetime.strptime(str(end_date_str), '%Y-%m-%d').date()
                            days_remaining = (end_date - today).days
                            
                            # Notify if due in 5, 4, 3, 2, or 1 days
                            if 1 <= days_remaining <= 5:
                                notify_goal_due_soon(goal, user, days_remaining)
                            
                            # Check if goal is past due and not completed
                            if days_remaining < 0:
                                monthly_achievement = goal.get('monthly_achievement', 0)
                                monthly_target = goal.get('monthly_target', 1)
                                
                                if monthly_achievement is None:
                                    monthly_achievement = 0
                                
                                progress = (monthly_achievement / monthly_target * 100) if monthly_target > 0 else 0
                                
                                if progress < 100:
                                    notify_goal_not_completed(goal, user)
                        except:
                            pass
                    
                    # Check for missing weekly updates
                    current_year = today.year
                    current_month = today.month
                    
                    if goal['year'] == current_year and goal.get('month') == current_month:
                        for week_num in range(1, 5):
                            week_achievement = goal.get(f'week{week_num}_achievement')
                            
                            if week_achievement is None:
                                week_start, week_end = get_week_dates(current_year, current_month, week_num)
                                
                                if today > week_end:
                                    notify_goal_not_updated(goal, user, week_num)

def get_user_notifications(user_id, limit=50):
    """Get notifications for a specific user"""
    try:
        result = supabase.table('notifications').select('*').eq(
            'user_id', user_id
        ).order('created_at', desc=True).limit(limit).execute()
        return result.data if result.data else []
    except Exception as e:
        print(f"Error getting notifications: {str(e)}")
        return []

def mark_notification_read(notification_id):
    """Mark a notification as read"""
    try:
        result = supabase.table('notifications').update({
            'is_read': True,
            'read_at': datetime.now(IST).isoformat()
        }).eq('id', notification_id).execute()
        return True
    except Exception as e:
        print(f"Error marking notification as read: {str(e)}")
        return False


def can_modify_user(current_user_role, target_user_role):
    """Check if current user can modify target user based on hierarchy"""
    hierarchy = get_role_hierarchy()
    current_level = hierarchy.get(current_user_role, 0)
    target_level = hierarchy.get(target_user_role, 0)
    
    # Can only modify users at lower hierarchy levels
    return current_level > target_level


def get_modifiable_roles(user_role):
    """Get list of roles that a user can modify"""
    hierarchy = get_role_hierarchy()
    user_level = hierarchy.get(user_role, 0)
    
    modifiable = []
    for role, level in hierarchy.items():
        if level < user_level:
            modifiable.append(role)
    
    return modifiable

def get_viewable_roles(user_role):
    """Get list of roles that a user can view"""
    # All roles can view users at same or lower level
    hierarchy = get_role_hierarchy()
    user_level = hierarchy.get(user_role, 0)
    
    viewable = []
    for role, level in hierarchy.items():
        if level <= user_level:
            viewable.append(role)
    
    return viewable

def check_password_strength(password):
    """Check password strength and return score, color, and feedback"""
    if not password:
        return 0, "#94a3b8", "No password entered", []
    
    score = 0
    feedback = []
    
    # Length check
    if len(password) >= 8:
        score += 25
    elif len(password) >= 6:
        score += 15
        feedback.append("🔸 Use at least 8 characters")
    else:
        feedback.append("❌ Too short (minimum 6 characters)")
    
    # Uppercase check
    if any(c.isupper() for c in password):
        score += 20
    else:
        feedback.append("🔸 Add uppercase letters (A-Z)")
    
    # Lowercase check
    if any(c.islower() for c in password):
        score += 20
    else:
        feedback.append("🔸 Add lowercase letters (a-z)")
    
    # Number check
    if any(c.isdigit() for c in password):
        score += 20
    else:
        feedback.append("🔸 Add numbers (0-9)")
    
    # Special character check
    special_chars = "!@#$%^&*()_+-=[]{}|;:,.<>?"
    if any(c in special_chars for c in password):
        score += 15
    else:
        feedback.append("🔸 Add special characters (!@#$%)")
    
    # Determine strength level and color
    if score >= 85:
        strength = "Very Strong 💪"
        color = "#10b981"  # Green
    elif score >= 70:
        strength = "Strong 👍"
        color = "#22c55e"  # Light green
    elif score >= 50:
        strength = "Medium ⚡"
        color = "#f59e0b"  # Orange
    elif score >= 30:
        strength = "Weak ⚠️"
        color = "#fb923c"  # Light orange
    else:
        strength = "Very Weak ❌"
        color = "#ef4444"  # Red
    
    return score, color, strength, feedback

def send_achievement_approval_email(manager_email, employee_name, goal_data, achievements):
    """Send achievement approval request email to manager"""
    try:
        smtp_server = os.getenv('SMTP_SERVER', 'smtp.gmail.com')
        smtp_port = int(os.getenv('SMTP_PORT', '587'))
        smtp_email = os.getenv('SMTP_EMAIL')
        smtp_password = os.getenv('SMTP_PASSWORD')
        
        if not smtp_email or not smtp_password:
            print("Email configuration not found")
            return False
        
        subject = f"📊 Achievement Approval Request from {employee_name}"
        
        # Format achievements display
        achievements_html = ""
        for week, value in achievements.items():
            if week != 'monthly':
                display_value = value if value is not None else '-'
                achievements_html += f"""
                <tr>
                    <td style="padding: 8px; border: 1px solid #ddd;">{week.title()}</td>
                    <td style="padding: 8px; border: 1px solid #ddd; text-align: center;">{display_value}</td>
                </tr>
                """
        
        monthly_achievement = achievements.get('monthly', 0)
        if monthly_achievement is None:
            monthly_achievement = 0
        
        body = f"""
        <html>
            <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
                <div style="max-width: 600px; margin: 0 auto; padding: 20px; border: 1px solid #ddd; border-radius: 10px;">
                    <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                                padding: 20px; border-radius: 10px 10px 0 0; text-align: center;">
                        <h1 style="color: white; margin: 0; font-size: 28px;">📊 Achievement Approval Required</h1>
                    </div>
                    
                    <div style="padding: 30px 20px;">
                        <p style="font-size: 16px; margin-bottom: 20px;">
                            <strong>{employee_name}</strong> has updated their goal achievements and needs your approval:
                        </p>
                        
                        <div style="background: #f3f4f6; padding: 20px; border-radius: 8px; margin: 20px 0;">
                            <p style="margin: 0 0 10px 0;"><strong>Goal:</strong> {goal_data['goal_title']}</p>
                            <p style="margin: 0 0 10px 0;"><strong>Department:</strong> {goal_data.get('department', 'N/A')}</p>
                            <p style="margin: 0;"><strong>Monthly Target:</strong> {goal_data.get('monthly_target', 0)}</p>
                        </div>
                        
                        <h3 style="color: #1e40af; margin-top: 20px;">Reported Achievements:</h3>
                        <table style="width: 100%; border-collapse: collapse; margin: 15px 0;">
                            <thead>
                                <tr style="background: #3b82f6; color: white;">
                                    <th style="padding: 10px; border: 1px solid #ddd; text-align: left;">Week</th>
                                    <th style="padding: 10px; border: 1px solid #ddd; text-align: center;">Achievement</th>
                                </tr>
                            </thead>
                            <tbody>
                                {achievements_html}
                                <tr style="background: #dbeafe; font-weight: bold;">
                                    <td style="padding: 10px; border: 1px solid #ddd;">Monthly Total</td>
                                    <td style="padding: 10px; border: 1px solid #ddd; text-align: center;">{monthly_achievement}</td>
                                </tr>
                            </tbody>
                        </table>
                        
                        <p style="margin-top: 30px; padding-top: 20px; border-top: 1px solid #ddd; color: #64748b; font-size: 14px;">
                            Please log in to the Performance Management System to approve or reject these achievements.
                        </p>
                    </div>
                    
                    <div style="text-align: center; padding: 20px; background: #f9fafb; border-radius: 0 0 10px 10px;">
                        <p style="margin: 0; font-size: 12px; color: #64748b;">
                            This is an automated email from Performance Management System.<br>
                            Please do not reply to this email.
                        </p>
                    </div>
                </div>
            </body>
        </html>
        """
        
        message = MIMEMultipart("alternative")
        message["Subject"] = subject
        message["From"] = smtp_email
        message["To"] = manager_email
        
        html_part = MIMEText(body, "html")
        message.attach(html_part)
        
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(smtp_email, smtp_password)
            server.sendmail(smtp_email, manager_email, message.as_string())
        
        print(f"✅ Achievement approval email sent to {manager_email}")
        return True
    except Exception as e:
        print(f"❌ Failed to send achievement approval email: {str(e)}")
        return False
    
def send_goal_completion_email(manager_email, employee_name, goal_title, completed=True):
    """Send email to manager when employee completes or misses goal"""
    try:
        smtp_server = os.getenv('SMTP_SERVER', 'smtp.gmail.com')
        smtp_port = int(os.getenv('SMTP_PORT', '587'))
        smtp_email = os.getenv('SMTP_EMAIL')
        smtp_password = os.getenv('SMTP_PASSWORD')
        
        if not smtp_email or not smtp_password:
            print("Email configuration not found")
            return False
        
        if completed:
            subject = f"✅ Goal Completed: {employee_name}"
            status_color = "#10b981"
            status_icon = "✅"
            status_text = "COMPLETED"
            message_text = f"{employee_name} has successfully completed the goal"
        else:
            subject = f"⚠️ Goal Deadline Missed: {employee_name}"
            status_color = "#ef4444"
            status_icon = "⚠️"
            status_text = "DEADLINE MISSED"
            message_text = f"{employee_name} did not complete the goal by the deadline"
        
        body = f"""
        <html>
            <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
                <div style="max-width: 600px; margin: 0 auto; padding: 20px; border: 1px solid #ddd; border-radius: 10px;">
                    <div style="background: {status_color}; padding: 20px; border-radius: 10px 10px 0 0; text-align: center;">
                        <h1 style="color: white; margin: 0; font-size: 28px;">{status_icon} Goal {status_text}</h1>
                    </div>
                    
                    <div style="padding: 30px 20px;">
                        <p style="font-size: 16px; margin-bottom: 20px;">{message_text}:</p>
                        
                        <div style="background: #f3f4f6; padding: 20px; border-radius: 8px; margin: 20px 0;">
                            <p style="margin: 0 0 10px 0;"><strong>Employee:</strong> {employee_name}</p>
                            <p style="margin: 0;"><strong>Goal:</strong> {goal_title}</p>
                        </div>
                        
                        <p style="margin-top: 30px; padding-top: 20px; border-top: 1px solid #ddd; color: #64748b; font-size: 14px;">
                            Please review the goal details in the Performance Management System.
                        </p>
                    </div>
                    
                    <div style="text-align: center; padding: 20px; background: #f9fafb; border-radius: 0 0 10px 10px;">
                        <p style="margin: 0; font-size: 12px; color: #64748b;">
                            This is an automated email from Performance Management System.<br>
                            Please do not reply to this email.
                        </p>
                    </div>
                </div>
            </body>
        </html>
        """
        
        message = MIMEMultipart("alternative")
        message["Subject"] = subject
        message["From"] = smtp_email
        message["To"] = manager_email
        
        html_part = MIMEText(body, "html")
        message.attach(html_part)
        
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(smtp_email, smtp_password)
            server.sendmail(smtp_email, manager_email, message.as_string())
        
        print(f"Email sent to {manager_email}")
        return True
    except Exception as e:
        print(f"Failed to send email: {str(e)}")
        return False

def send_goal_approval_email(manager_email, employee_name, goal_data, goal_id):
    """Send approval request email to manager with approve/reject links"""
    try:
        smtp_server = os.getenv('SMTP_SERVER', 'smtp.gmail.com')
        smtp_port = int(os.getenv('SMTP_PORT', '587'))
        smtp_email = os.getenv('SMTP_EMAIL')
        smtp_password = os.getenv('SMTP_PASSWORD')
        
        if not smtp_email or not smtp_password:
            print("Email configuration not found")
            return False
        
        # Generate approval tokens
        import secrets
        approve_token = secrets.token_urlsafe(32)
        reject_token = secrets.token_urlsafe(32)
        
        # Store tokens in database (you'll need to create this table)
        db.store_approval_token(goal_id, approve_token, 'approve')
        db.store_approval_token(goal_id, reject_token, 'reject')
        
        # Create approval links (replace with your actual domain)
        app_url = os.getenv('APP_URL', 'http://localhost:8501')
        approve_link = f"{app_url}/?action=approve&token={approve_token}"
        reject_link = f"{app_url}/?action=reject&token={reject_token}"
        
        subject = f"🔔 Goal Approval Request from {employee_name}"
        
        body = f"""
        <html>
            <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
                <div style="max-width: 600px; margin: 0 auto; padding: 20px; border: 1px solid #ddd; border-radius: 10px;">
                    <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 20px; border-radius: 10px 10px 0 0; text-align: center;">
                        <h1 style="color: white; margin: 0; font-size: 28px;">🔔 Goal Approval Required</h1>
                    </div>
                    
                    <div style="padding: 30px 20px;">
                        <p style="font-size: 16px; margin-bottom: 20px;">
                            <strong>{employee_name}</strong> has created a new goal and needs your approval:
                        </p>
                        
                        <div style="background: #f3f4f6; padding: 20px; border-radius: 8px; margin: 20px 0;">
                            <p style="margin: 0 0 10px 0;"><strong>Goal Title:</strong> {goal_data['goal_title']}</p>
                            <p style="margin: 0 0 10px 0;"><strong>Department:</strong> {goal_data.get('department', 'N/A')}</p>
                            <p style="margin: 0 0 10px 0;"><strong>KPI:</strong> {goal_data.get('kpi', 'N/A')}</p>
                            <p style="margin: 0 0 10px 0;"><strong>Monthly Target:</strong> {goal_data.get('monthly_target', 0)}</p>
                            <p style="margin: 0 0 10px 0;"><strong>Period:</strong> {goal_data.get('year')}-Q{goal_data.get('quarter')}-M{goal_data.get('month')}</p>
                            <p style="margin: 0;"><strong>Description:</strong> {goal_data.get('goal_description', 'No description')}</p>
                        </div>
                        
                        <div style="text-align: center; margin: 30px 0;">
                            <a href="{approve_link}" style="display: inline-block; background: #10b981; color: white; padding: 15px 40px; text-decoration: none; border-radius: 8px; font-weight: bold; margin: 10px;">
                                ✅ Approve Goal
                            </a>
                            <a href="{reject_link}" style="display: inline-block; background: #ef4444; color: white; padding: 15px 40px; text-decoration: none; border-radius: 8px; font-weight: bold; margin: 10px;">
                                ❌ Reject Goal
                            </a>
                        </div>
                        
                        <p style="margin-top: 30px; padding-top: 20px; border-top: 1px solid #ddd; color: #64748b; font-size: 14px;">
                            You can also approve/reject this goal by logging into the Performance Management System.
                        </p>
                    </div>
                    
                    <div style="text-align: center; padding: 20px; background: #f9fafb; border-radius: 0 0 10px 10px;">
                        <p style="margin: 0; font-size: 12px; color: #64748b;">
                            This is an automated email from Performance Management System.<br>
                            Please do not reply to this email.
                        </p>
                    </div>
                </div>
            </body>
        </html>
        """
        
        message = MIMEMultipart("alternative")
        message["Subject"] = subject
        message["From"] = smtp_email
        message["To"] = manager_email
        
        html_part = MIMEText(body, "html")
        message.attach(html_part)
        
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(smtp_email, smtp_password)
            server.sendmail(smtp_email, manager_email, message.as_string())
        
        print(f"✅ Approval email sent to {manager_email}")
        return True
    except Exception as e:
        print(f"❌ Failed to send approval email: {str(e)}")
        return False
       
def calculate_performance_metrics(goals):
    """Calculate comprehensive performance metrics"""
    if not goals:
        return None
    
    total_goals = len(goals)
    completed = len([g for g in goals if g.get('status') == 'Completed'])
    active = len([g for g in goals if g.get('status') == 'Active'])
    on_hold = len([g for g in goals if g.get('status') == 'On Hold'])
    cancelled = len([g for g in goals if g.get('status') == 'Cancelled'])
    
    # Calculate average progress
    total_progress = 0
    goal_count = 0
    for goal in goals:
        progress = calculate_progress(
            goal.get('monthly_achievement', 0),
            goal.get('monthly_target', 1)
        )
        total_progress += progress
        goal_count += 1
    
    avg_progress = total_progress / goal_count if goal_count > 0 else 0
    
    # Calculate completion rate
    completion_rate = (completed / total_goals * 100) if total_goals > 0 else 0
    
    # Calculate on-time completion rate
    on_time = 0
    overdue = 0
    today = date.today()
    
    for goal in goals:
        if goal.get('status') == 'Completed':
            end_date = datetime.strptime(str(goal.get('end_date')), '%Y-%m-%d').date()
            if end_date >= today:
                on_time += 1
            else:
                overdue += 1
    
    on_time_rate = (on_time / completed * 100) if completed > 0 else 0
    
    return {
        'total_goals': total_goals,
        'completed': completed,
        'active': active,
        'on_hold': on_hold,
        'cancelled': cancelled,
        'avg_progress': avg_progress,
        'completion_rate': completion_rate,
        'on_time': on_time,
        'overdue': overdue,
        'on_time_rate': on_time_rate
    }


def get_trend_data(user_id, months=6):
    """Get performance trend data for last N months"""
    trends = []
    today = date.today()
    
    for i in range(months):
        month_date = today - relativedelta(months=i)
        year = month_date.year
        month = month_date.month
        quarter = ((month - 1) // 3) + 1
        
        # Get goals for this month
        try:
            goals = db.get_month_goals(user_id, year, quarter, month)
            
            if goals and len(goals) > 0:
                metrics = calculate_performance_metrics(goals)
                trends.append({
                    'month': month_date.strftime('%b %Y'),
                    'completion_rate': metrics['completion_rate'],
                    'avg_progress': metrics['avg_progress'],
                    'total_goals': metrics['total_goals'],
                    'completed': metrics['completed']
                })
        except Exception as e:
            # Skip months with errors
            print(f"Error getting data for {month_date}: {str(e)}")
            continue
    
    return list(reversed(trends))


# ============================================
# ADVANCED CHARTS
# ============================================

def create_performance_gauge(value, title="Performance"):
    """Create a gauge chart for performance"""
    fig = go.Figure(go.Indicator(
        mode="gauge+number+delta",
        value=value,
        domain={'x': [0, 1], 'y': [0, 1]},
        title={'text': title, 'font': {'size': 24}},
        delta={'reference': 75, 'increasing': {'color': "green"}},
        gauge={
            'axis': {'range': [None, 100], 'tickwidth': 1, 'tickcolor': "darkblue"},
            'bar': {'color': "darkblue"},
            'bgcolor': "white",
            'borderwidth': 2,
            'bordercolor': "gray",
            'steps': [
                {'range': [0, 30], 'color': '#fee2e2'},
                {'range': [30, 60], 'color': '#fed7aa'},
                {'range': [60, 80], 'color': '#fef3c7'},
                {'range': [80, 100], 'color': '#d1fae5'}
            ],
            'threshold': {
                'line': {'color': "red", 'width': 4},
                'thickness': 0.75,
                'value': 90
            }
        }
    ))
    
    fig.update_layout(
        height=300,
        margin=dict(l=20, r=20, t=60, b=20),
        paper_bgcolor="white",
        font={'color': "darkblue", 'family': "Arial"}
    )
    
    return fig


def create_trend_chart(trend_data):
    """Create performance trend line chart"""
    if not trend_data:
        return None
    
    df = pd.DataFrame(trend_data)
    
    fig = make_subplots(
        rows=2, cols=1,
        subplot_titles=('Goal Completion Trend', 'Average Progress Trend'),
        vertical_spacing=0.15
    )
    
    # Completion rate trend
    fig.add_trace(
        go.Scatter(
            x=df['month'],
            y=df['completion_rate'],
            mode='lines+markers',
            name='Completion Rate',
            line=dict(color='#3b82f6', width=3),
            marker=dict(size=8)
        ),
        row=1, col=1
    )
    
    # Average progress trend
    fig.add_trace(
        go.Scatter(
            x=df['month'],
            y=df['avg_progress'],
            mode='lines+markers',
            name='Avg Progress',
            line=dict(color='#10b981', width=3),
            marker=dict(size=8)
        ),
        row=2, col=1
    )
    
    fig.update_xaxes(title_text="Month", row=2, col=1)
    fig.update_yaxes(title_text="Completion %", row=1, col=1)
    fig.update_yaxes(title_text="Progress %", row=2, col=1)
    
    fig.update_layout(
        height=600,
        showlegend=True,
        title_text="Performance Trends (Last 6 Months)",
        title_font_size=20
    )
    
    return fig


def create_status_distribution_chart(goals):
    """Create pie chart for goal status distribution"""
    status_counts = {}
    for goal in goals:
        status = goal.get('status', 'Active')
        status_counts[status] = status_counts.get(status, 0) + 1
    
    colors_map = {
        'Active': '#3b82f6',
        'Completed': '#10b981',
        'On Hold': '#f59e0b',
        'Cancelled': '#ef4444'
    }
    
    fig = go.Figure(data=[go.Pie(
        labels=list(status_counts.keys()),
        values=list(status_counts.values()),
        hole=0.4,
        marker_colors=[colors_map.get(k, '#64748b') for k in status_counts.keys()]
    )])
    
    fig.update_layout(
        title_text="Goal Status Distribution",
        title_font_size=20,
        height=400,
        showlegend=True
    )
    
    return fig


def create_department_performance_chart(goals):
    """Create bar chart for performance by department"""
    department_data = {}
    
    for goal in goals:
        department = goal.get('department', 'Unassigned')
        if department not in department_data:
            department_data[department] = {'total': 0, 'completed': 0, 'progress': []}
        
        department_data[department]['total'] += 1
        if goal.get('status') == 'Completed':
            department_data[department]['completed'] += 1
        
        progress = calculate_progress(
            goal.get('monthly_achievement', 0),
            goal.get('monthly_target', 1)
        )
        department_data[department]['progress'].append(progress)
    
    verticals = list(department_data.keys())
    completion_rates = [
        (department_data[v]['completed'] / department_data[v]['total'] * 100) 
        if department_data[v]['total'] > 0 else 0 
        for v in verticals
    ]
    avg_progress = [
        sum(department_data[v]['progress']) / len(department_data[v]['progress']) 
        if department_data[v]['progress'] else 0 
        for v in verticals
    ]
    
    fig = go.Figure(data=[
        go.Bar(name='Completion Rate', x=verticals, y=completion_rates, marker_color='#3b82f6'),
        go.Bar(name='Avg Progress', x=verticals, y=avg_progress, marker_color='#10b981')
    ])
    
    fig.update_layout(
        barmode='group',
        title_text="Performance by Vertical",
        title_font_size=20,
        xaxis_title="Vertical",
        yaxis_title="Percentage (%)",
        height=400
    )
    
    return fig


def create_heatmap_calendar(goals, year, month):
    """Create calendar heatmap for daily goal achievements"""
    # Get days in month
    days_in_month = calendar.monthrange(year, month)[1]
    
    # Initialize achievement data
    daily_achievements = {}
    for day in range(1, days_in_month + 1):
        daily_achievements[day] = 0
    
    # Calculate daily achievements
    for goal in goals:
        for week in range(1, 5):
            week_achievement = goal.get(f'week{week}_achievement', 0)
            if week_achievement > 0:
                # Distribute across week days
                week_start, week_end = get_week_dates(year, month, week)
                days_in_week = (week_end - week_start).days + 1
                daily_avg = week_achievement / days_in_week
                
                for day in range(week_start.day, min(week_end.day + 1, days_in_month + 1)):
                    daily_achievements[day] += daily_avg
    
    # Create matrix for heatmap
    weeks = 5
    days = 7
    matrix = np.zeros((weeks, days))
    day_labels = []
    
    first_day = date(year, month, 1)
    first_weekday = first_day.weekday()
    
    day_counter = 1
    for week in range(weeks):
        for day in range(days):
            if week == 0 and day < first_weekday:
                matrix[week][day] = np.nan
            elif day_counter > days_in_month:
                matrix[week][day] = np.nan
            else:
                matrix[week][day] = daily_achievements.get(day_counter, 0)
                day_counter += 1
    
    fig = go.Figure(data=go.Heatmap(
        z=matrix,
        x=['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun'],
        y=[f'Week {i+1}' for i in range(weeks)],
        colorscale='Blues',
        showscale=True
    ))
    
    fig.update_layout(
        title=f"Daily Achievement Heatmap - {get_month_name(month)} {year}",
        title_font_size=20,
        height=400
    )
    
    return fig


# ============================================
# PDF REPORT GENERATION
# ============================================

def generate_performance_report_pdf(user, goals, period="Monthly"):
    """Generate comprehensive PDF performance report"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    story = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1e40af'),
        alignment=TA_CENTER,
        spaceAfter=30
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#3b82f6'),
        spaceAfter=12
    )
    
    # Title
    story.append(Paragraph(f"Performance Report - {period}", title_style))
    story.append(Spacer(1, 0.2*inch))
    
    # User Information
    story.append(Paragraph("Employee Information", heading_style))
    user_data = [
        ['Name:', user['name']],
        ['Email:', user['email']],
        ['Role:', user['role']],
        ['Department:', user.get('department', 'N/A')],
        ['Report Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
    ]
    
    user_table = Table(user_data, colWidths=[2*inch, 4*inch])
    user_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e0e7ff')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey)
    ]))
    
    story.append(user_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Performance Metrics
    metrics = calculate_performance_metrics(goals)
    
    if metrics:
        story.append(Paragraph("Performance Summary", heading_style))
        
        metrics_data = [
            ['Metric', 'Value'],
            ['Total Goals', str(metrics['total_goals'])],
            ['Completed Goals', str(metrics['completed'])],
            ['Active Goals', str(metrics['active'])],
            ['Completion Rate', f"{metrics['completion_rate']:.1f}%"],
            ['Average Progress', f"{metrics['avg_progress']:.1f}%"],
            ['On-Time Completion Rate', f"{metrics['on_time_rate']:.1f}%"]
        ]
        
        metrics_table = Table(metrics_data, colWidths=[3*inch, 2*inch])
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(metrics_table)
        story.append(Spacer(1, 0.3*inch))
    
    # Goals Details
    story.append(Paragraph("Goals Details", heading_style))
    
    goals_data = [['Goal Title', 'Vertical', 'Target', 'Achievement', 'Progress', 'Status']]
    
    for goal in goals:
        progress = calculate_progress(
            goal.get('monthly_achievement', 0),
            goal.get('monthly_target', 1)
        )
        
        goals_data.append([
            goal['goal_title'][:30],
            goal.get('vertical', 'N/A'),
            str(goal.get('monthly_target', 0)),
            str(goal.get('monthly_achievement', 0)),
            f"{progress:.1f}%",
            goal.get('status', 'Active')
        ])
    
    goals_table = Table(goals_data, colWidths=[2*inch, 1.2*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch])
    goals_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    
    story.append(goals_table)
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    return buffer



# ============================================
# ANALYTICS PAGE
# ============================================

def display_analytics_page():
    """Display comprehensive analytics page"""
    user = st.session_state.user
    if not user:
        st.warning("⚠️ Session expired. Please login again.")
        st.rerun()
    role = user['role']
    
    st.title("📊 Advanced Analytics & Reports")
    
    # Filters
    col_filter1, col_filter2, col_filter3 = st.columns(3)
    
    with col_filter1:
        if role == 'HR':
            all_users = db.get_all_users()
            selected_user = st.selectbox(
                "Select User",
                ["My Analytics"] + [f"{u['name']} ({u['email']})" for u in all_users]
            )
            if selected_user == "My Analytics":
                analysis_user = user
            else:
                user_email = selected_user.split('(')[1].strip(')')
                analysis_user = next(u for u in all_users if u['email'] == user_email)
        elif role == 'Manager':
            team_members = db.get_team_members(user['id'])
            selected_user = st.selectbox(
                "Select Team Member",
                ["My Analytics"] + [f"{m['name']} ({m['email']})" for m in team_members]
            )
            if selected_user == "My Analytics":
                analysis_user = user
            else:
                user_email = selected_user.split('(')[1].strip(')')
                analysis_user = next(m for m in team_members if m['email'] == user_email)
        else:
            analysis_user = user
    
    with col_filter2:
        analysis_period = st.selectbox(
            "Analysis Period",
            ["Current Month", "Current Quarter", "Current Year", "Last 6 Months", "All Time"]
        )
    
    with col_filter3:
        view_type = st.selectbox(
            "View Type",
            ["Overview", "Trends", "Comparisons", "Detailed"]
        )
    
    # Get goals based on period
    all_goals = db.get_user_all_goals(analysis_user['id'])
    
    if analysis_period == "Current Month":
        today = date.today()
        goals = [g for g in all_goals if g['year'] == today.year and g.get('month') == today.month]
    elif analysis_period == "Current Quarter":
        today = date.today()
        quarter = ((today.month - 1) // 3) + 1
        goals = [g for g in all_goals if g['year'] == today.year and g.get('quarter') == quarter]
    elif analysis_period == "Current Year":
        today = date.today()
        goals = [g for g in all_goals if g['year'] == today.year]
    else:
        goals = all_goals
    
    if not goals:
        st.info(f"No goals found for {analysis_user['name']} in selected period")
        return
    
    st.markdown("---")
    
    # Calculate metrics
    metrics = calculate_performance_metrics(goals)
    
    # Overview Tab
    if view_type == "Overview":
        st.subheader(f"📈 Performance Overview - {analysis_user['name']}")
        
        # KPI Cards
        col1, col2, col3, col4, col5 = st.columns(5)
        
        with col1:
            render_metric_card("Total Goals", str(metrics['total_goals']), color="#3b82f6")
        with col2:
            render_metric_card("Completed", str(metrics['completed']), color="#10b981")
        with col3:
            render_metric_card("Active", str(metrics['active']), color="#f59e0b")
        with col4:
            render_metric_card("Completion Rate", f"{int(metrics['completion_rate'])}%", color="#8b5cf6")
        with col5:
            render_metric_card("Avg Progress", str(metrics['avg_progress']), color="#ec4899")
        
        st.markdown("---")
        
        # Charts Row 1
        col_chart1, col_chart2 = st.columns(2)
        
        with col_chart1:
            # Performance Gauge
            gauge_fig = create_performance_gauge(metrics['avg_progress'], "Overall Performance")
            st.plotly_chart(gauge_fig, use_container_width=True)
        
        with col_chart2:
            # Status Distribution
            status_fig = create_status_distribution_chart(goals)
            st.plotly_chart(status_fig, use_container_width=True)
        
        st.markdown("---")
        
        # Charts Row 2
        col_chart3, col_chart4 = st.columns(2)
        
        with col_chart3:
            # Vertical Performance
            vertical_fig = create_department_performance_chart(goals)
            st.plotly_chart(vertical_fig, use_container_width=True)
        
        with col_chart4:
            # On-time vs Overdue
            on_time_data = pd.DataFrame({
                'Category': ['On-Time', 'Overdue'],
                'Count': [metrics['on_time'], metrics['overdue']]
            })
            fig_ontime = px.bar(on_time_data, x='Category', y='Count',
                               title="On-Time vs Overdue Completion",
                               color='Category',
                               color_discrete_map={'On-Time': '#10b981', 'Overdue': '#ef4444'})
            fig_ontime.update_layout(showlegend=False, height=400)
            st.plotly_chart(fig_ontime, use_container_width=True)
    
    # Trends Tab
    elif view_type == "Trends":
        st.subheader(f"📈 Performance Trends - {analysis_user['name']}")
        
        trend_data = get_trend_data(analysis_user['id'], months=6)
        
        if trend_data and len(trend_data) > 0:
            # Main trend chart
            trend_fig = create_trend_chart(trend_data)
            st.plotly_chart(trend_fig, use_container_width=True)
            
            st.markdown("---")
            
            # Monthly Achievement Trend
            st.subheader("Monthly Goal Achievement Trend")
            
            monthly_data = []
            for month_info in trend_data:
                monthly_data.append({
                    'Month': month_info['month'],
                    'Goals': month_info['total_goals'],
                    'Completed': month_info['completed']
                })
            
            df_monthly = pd.DataFrame(monthly_data)
            
            if not df_monthly.empty:
                fig_monthly = px.line(df_monthly, x='Month', y=['Goals', 'Completed'],
                                     title="Goals Created vs Completed",
                                     markers=True)
                fig_monthly.update_layout(height=400)
                st.plotly_chart(fig_monthly, use_container_width=True)
            else:
                st.info("No monthly data available for visualization")
        else:
            st.info("Not enough data for trend analysis. Please ensure you have goals spanning multiple months.")
    
    # Comparisons Tab
    elif view_type == "Comparisons":
        st.subheader(f"🔄 Performance Comparisons - {analysis_user['name']}")
        
        if role in ['HR', 'Manager']:
            # Get comparison users based on the selected user's role
            if role == 'HR':
                # HR: Compare with same role users
                selected_role = analysis_user['role']
                all_users = db.get_all_users()
                
                if selected_role == 'Manager':
                    compare_users = [u for u in all_users if u['role'] == 'Manager']
                    comparison_title = "Compare with Other Managers"
                elif selected_role == 'Employee':
                    compare_users = [u for u in all_users if u['role'] == 'Employee']
                    comparison_title = "Compare with Other Employees"
                else:  # HR viewing another HR
                    compare_users = [u for u in all_users if u['role'] == 'HR']
                    comparison_title = "Compare with Other HR Members"
                    
            else:  # Manager
                # Manager compares team members only (not themselves)
                if analysis_user['id'] == user['id']:
                    # Manager viewing their own analytics - compare with team only
                    compare_users = db.get_team_members(user['id'])
                    comparison_title = "Compare Team Members"
                else:
                    # Manager viewing a team member - compare with other team members
                    team_members = db.get_team_members(user['id'])
                    compare_users = [m for m in team_members]
                    comparison_title = "Compare with Team Members"
            
            if compare_users:
                # Show appropriate info message
                if role == 'Manager' and analysis_user['id'] == user['id']:
                    st.info(f"**{comparison_title}** ({len(compare_users)} team members)\n\n💡 *Note: You are viewing your team's performance comparison (manager is not included in rankings)*")
                else:
                    st.info(f"**{comparison_title}** ({len(compare_users)} users)")
                
                comparison_data = []
                
                # Add all users (including current)
                for comp_user in compare_users:
                    comp_goals = db.get_user_all_goals(comp_user['id'])
                    if comp_goals:
                        comp_metrics = calculate_performance_metrics(comp_goals)
                        is_current = (comp_user['id'] == analysis_user['id'])
                        comparison_data.append({
                            'Name': f"⭐ {comp_user['name']}" if is_current else comp_user['name'],
                            'Display_Name': comp_user['name'],
                            'Total Goals': comp_metrics['total_goals'],
                            'Completion Rate': comp_metrics['completion_rate'],
                            'Avg Progress': comp_metrics['avg_progress'],
                            'Is_Current': is_current,
                            'User_ID': comp_user['id']
                        })
                
                if comparison_data and len(comparison_data) > 0:
                    df_compare = pd.DataFrame(comparison_data)
                    
                    # Highlight current user
                    st.success(f"📍 **{analysis_user['name']}** is highlighted with ⭐ in the comparisons below")
                    
                    # Comparison bar chart
                    fig_compare = px.bar(
                        df_compare, 
                        x='Name', 
                        y=['Completion Rate', 'Avg Progress'],
                        title=f"{comparison_title} - Performance Metrics",
                        barmode='group',
                        height=500,
                        color_discrete_map={
                            'Completion Rate': '#3b82f6',
                            'Avg Progress': '#10b981'
                        }
                    )
                    fig_compare.update_layout(
                        xaxis_tickangle=-45,
                        xaxis_title="",
                        yaxis_title="Percentage (%)"
                    )
                    st.plotly_chart(fig_compare, use_container_width=True)
                    
                    st.markdown("---")
                    
                    # Comparison table with ranking
                    st.subheader("📊 Detailed Comparison Table")
                    
                    # Sort by completion rate for ranking
                    df_sorted = df_compare.sort_values('Completion Rate', ascending=False).reset_index(drop=True)
                    df_sorted.insert(0, 'Rank', range(1, len(df_sorted) + 1))
                    
                    # Create display dataframe
                    df_display = df_sorted[['Rank', 'Name', 'Total Goals', 'Completion Rate', 'Avg Progress']].copy()
                    df_display['Completion Rate'] = df_display['Completion Rate'].apply(lambda x: f"{x:.1f}%")
                    df_display['Avg Progress'] = df_display['Avg Progress'].apply(lambda x: f"{x:.1f}%")
                    
                    st.dataframe(df_display, use_container_width=True, height=400)
                    
                    # Performance insights
                    st.markdown("---")
                    st.subheader("💡 Performance Insights")
                    
                    # Calculate metrics correctly
                    avg_completion = df_compare['Completion Rate'].mean()
                    avg_progress = df_compare['Avg Progress'].mean()
                    
                    # Get current user's metrics
                    current_user_data = df_compare[df_compare['Is_Current'] == True]
                    
                    if not current_user_data.empty:
                        user_completion = current_user_data['Completion Rate'].iloc[0]
                        user_progress = current_user_data['Avg Progress'].iloc[0]
                        
                        # Get current user's rank
                        current_rank_row = df_sorted[df_sorted['Is_Current'] == True]
                        if not current_rank_row.empty:
                            current_rank = current_rank_row.iloc[0]['Rank']
                            
                            # Display rank
                            col_rank, col_space = st.columns([1, 3])
                            with col_rank:
                                st.metric("Current User's Rank", f"#{current_rank} out of {len(df_sorted)}")
                        
                        col_insight1, col_insight2 = st.columns(2)
                        
                        with col_insight1:
                            st.markdown("**Completion Rate Analysis:**")
                            diff = user_completion - avg_completion
                            if diff > 0:
                                st.success(f"✅ **Above Average**\n\n{analysis_user['name']}'s completion rate: **{user_completion:.1f}%**\n\nGroup average: **{avg_completion:.1f}%**\n\nDifference: **+{diff:.1f}%** higher")
                            elif diff < 0:
                                st.warning(f"⚠️ **Below Average**\n\n{analysis_user['name']}'s completion rate: **{user_completion:.1f}%**\n\nGroup average: **{avg_completion:.1f}%**\n\nDifference: **{abs(diff):.1f}%** lower")
                            else:
                                st.info(f"📊 **Average Performance**\n\n{analysis_user['name']}'s completion rate: **{user_completion:.1f}%**\n\nMatches group average: **{avg_completion:.1f}%**")
                        
                        with col_insight2:
                            st.markdown("**Progress Analysis:**")
                            prog_diff = user_progress - avg_progress
                            if prog_diff > 0:
                                st.success(f"✅ **Above Average**\n\n{analysis_user['name']}'s avg progress: **{user_progress:.1f}%**\n\nGroup average: **{avg_progress:.1f}%**\n\nDifference: **+{prog_diff:.1f}%** higher")
                            elif prog_diff < 0:
                                st.warning(f"⚠️ **Below Average**\n\n{analysis_user['name']}'s avg progress: **{user_progress:.1f}%**\n\nGroup average: **{avg_progress:.1f}%**\n\nDifference: **{abs(prog_diff):.1f}%** lower")
                            else:
                                st.info(f"📊 **Average Performance**\n\n{analysis_user['name']}'s avg progress: **{user_progress:.1f}%**\n\nMatches group average: **{avg_progress:.1f}%**")
                        
                        # Top performer
                        st.markdown("---")
                        top_performer = df_sorted.iloc[0]
                        if top_performer['Is_Current']:
                            st.success(f"🏆 **Congratulations!** {analysis_user['name']} has the highest completion rate ({top_performer['Completion Rate']:.1f}%) in the group!")
                        else:
                            st.info(f"🏆 **Top Performer:** {top_performer['Display_Name']} leads with **{top_performer['Completion Rate']:.1f}%** completion rate")
                    
                else:
                    st.info("No comparison data available - users need to have goals to be included in comparison")
            else:
                st.info(f"No {comparison_title.lower()} available for comparison")
        else:
            st.warning("⚠️ Comparison view is only available for HR and Managers")
    
    # Detailed Tab
    elif view_type == "Detailed":
        st.subheader(f"📋 Detailed Goal Analysis - {analysis_user['name']}")
        
        # Goals breakdown by status
        st.markdown("### Goals by Status")
        
        status_groups = {}
        for goal in goals:
            status = goal.get('status', 'Active')
            if status not in status_groups:
                status_groups[status] = []
            status_groups[status].append(goal)
        
        for status, status_goals in status_groups.items():
            with st.expander(f"{status} Goals ({len(status_goals)})"):
                for goal in status_goals:
                    progress = calculate_progress(
                        goal.get('monthly_achievement', 0),
                        goal.get('monthly_target', 1)
                    )
                    
                    col1, col2, col3 = st.columns([3, 1, 1])
                    with col1:
                        st.markdown(f"**{goal['goal_title']}**")
                        st.caption(f"Vertical: {goal.get('vertical', 'N/A')} | KPI: {goal.get('kpi', 'N/A')}")
                    with col2:
                        st.metric("Target", goal.get('monthly_target', 0))
                    with col3:
                        st.metric("Progress", f"{progress:.1f}%")
                    
                    st.markdown("---")
        
        st.markdown("---")
        
        # Weekly breakdown
        st.markdown("### Weekly Performance Breakdown")
        
        weekly_data = {
            'Week 1': {'target': 0, 'achievement': 0},
            'Week 2': {'target': 0, 'achievement': 0},
            'Week 3': {'target': 0, 'achievement': 0},
            'Week 4': {'target': 0, 'achievement': 0}
        }
        
        for goal in goals:
            for week in range(1, 5):
                weekly_data[f'Week {week}']['target'] += goal.get(f'week{week}_target', 0)
                weekly_data[f'Week {week}']['achievement'] += goal.get(f'week{week}_achievement', 0)
        
        weekly_df = pd.DataFrame({
            'Week': list(weekly_data.keys()),
            'Target': [weekly_data[w]['target'] for w in weekly_data.keys()],
            'Achievement': [weekly_data[w]['achievement'] for w in weekly_data.keys()]
        })
        
        fig_weekly = px.bar(weekly_df, x='Week', y=['Target', 'Achievement'],
                           title="Weekly Targets vs Achievements",
                           barmode='group',
                           height=400)
        st.plotly_chart(fig_weekly, use_container_width=True)
    
    # Export Report
    st.markdown("---")
    st.subheader("📥 Export Report")
    
    col_export1, col_export2, col_export3 = st.columns(3)
    
    with col_export1:
        if st.button("📄 Export to PDF", use_container_width=True):
            with st.spinner("Generating PDF report..."):
                pdf_buffer = generate_performance_report_pdf(analysis_user, goals, analysis_period)
                st.download_button(
                    label="📥 Download PDF Report",
                    data=pdf_buffer,
                    file_name=f"Performance_Report_{analysis_user['name']}_{datetime.now().strftime('%Y%m%d')}.pdf",
                    mime="application/pdf",
                    use_container_width=True
                )
    
    with col_export2:
        if st.button("📊 Export to Excel", use_container_width=True):
            # Create Excel with metrics
            excel_buffer = export_goals_to_excel(analysis_user['id'], 
                                                 goals[0]['year'] if goals else date.today().year,
                                                 goals[0].get('quarter', 1) if goals else 1,
                                                 goals[0].get('month', 1) if goals else 1)
            if excel_buffer:
                st.download_button(
                    label="📥 Download Excel Report",
                    data=excel_buffer,
                    file_name=f"Goals_Report_{analysis_user['name']}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
def get_completable_goals(user_id):
    """Get all goals that can be auto-completed"""
    all_goals = db.get_user_all_goals(user_id)
    
    completable = []
    for goal in all_goals:
        if goal.get('status') == 'Active':
            progress = calculate_progress(
                goal.get('monthly_achievement', 0),
                goal.get('monthly_target', 1)
            )
            if progress >= 100:
                completable.append(goal)
    
    return completable


def auto_complete_goal(goal_id, completed_by=None, remarks=None):
    """Auto-complete a goal that has reached 100%"""
    updates = {
        'status': 'Completed',
        'completed_at': datetime.now(IST).isoformat()
    }
    
    if completed_by:
        updates['completed_by'] = completed_by
    
    if remarks:
        updates['completion_remarks'] = remarks
    
    return db.update_goal(goal_id, updates)
def display_auto_complete_banner():
    """Show banner with auto-complete suggestions"""
    user = st.session_state.user
    
    completable_goals = get_completable_goals(user['id'])
    
    if completable_goals:
        with st.container():
            st.markdown("""
            <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                        padding: 20px; border-radius: 10px; margin-bottom: 20px;">
                <h3 style="color: white; margin: 0;">🎉 Goals Ready for Completion!</h3>
            </div>
            """, unsafe_allow_html=True)
            
            st.info(f"You have **{len(completable_goals)} goal(s)** at 100% progress")
            
            with st.expander(f"📋 View {len(completable_goals)} Completable Goals", expanded=True):
                for goal in completable_goals:
                    col1, col2, col3 = st.columns([3, 1, 1])
                    
                    with col1:
                        st.markdown(f"**{goal['goal_title']}**")
                        st.caption(f"Target: {goal.get('monthly_target', 0)}")
                    
                    with col2:
                        st.success("100% ✅")
                    
                    with col3:
                        if st.button("✅ Complete", key=f"complete_{goal['goal_id']}", use_container_width=True):
                            if auto_complete_goal(goal['goal_id'], completed_by=user['id']):
                                st.success(f"Completed!")
                                st.rerun()
                
                st.markdown("---")
                
                col_bulk1, col_bulk2 = st.columns(2)
                
                with col_bulk1:
                    if st.button("✅ Complete All", type="primary", use_container_width=True):
                        for goal in completable_goals:
                            auto_complete_goal(goal['goal_id'], completed_by=user['id'])
                        st.success(f"✅ Completed {len(completable_goals)} goals!")
                        st.balloons()
                        st.rerun()
                
                with col_bulk2:
                    if st.button("⏭️ Remind Later", use_container_width=True):
                        st.rerun()

def calculate_performance_score(stats):
    """Calculate weighted performance score for ranking"""
    # Weighted scoring system
    completion_weight = 0.4  # 40%
    progress_weight = 0.3    # 30%
    total_goals_weight = 0.2  # 20%
    on_time_weight = 0.1     # 10%
    
    completion_score = stats.get('completion_rate', 0) * completion_weight
    progress_score = stats.get('avg_progress', 0) * progress_weight
    goals_score = min(stats.get('total_goals', 0) * 5, 100) * total_goals_weight  # Cap at 100
    on_time_score = stats.get('on_time_rate', 0) * on_time_weight
    
    total_score = completion_score + progress_score + goals_score + on_time_score
    
    return round(total_score, 2)


def get_current_team_rankings(manager_id, year=None, month=None):
    """Get current team rankings with scores"""
    if year is None:
        year = date.today().year
    if month is None:
        month = date.today().month
    
    team_members = db.get_team_members(manager_id)
    
    rankings = []
    for member in team_members:
        # Get goals for specific month
        quarter = ((month - 1) // 3) + 1
        goals = db.get_month_goals(member['id'], year, quarter, month)
        
        if goals:
            stats = calculate_performance_metrics(goals)
            score = calculate_performance_score(stats)
            
            rankings.append({
                'employee_id': member['id'],
                'name': member['name'],
                'email': member['email'],
                'designation': member.get('designation', 'N/A'),
                'department': member.get('department', 'N/A'),
                'total_goals': stats['total_goals'],
                'completed_goals': stats['completed'],
                'completion_rate': stats['completion_rate'],
                'avg_progress': stats['avg_progress'],
                'on_time_rate': stats.get('on_time_rate', 0),
                'score': score
            })
    
    # Sort by score (highest first)
    rankings.sort(key=lambda x: x['score'], reverse=True)
    
    # Assign ranks
    for idx, ranking in enumerate(rankings):
        ranking['rank'] = idx + 1
    
    return rankings


def save_monthly_rankings(manager_id, year, month):
    """Save monthly rankings to database"""
    rankings = get_current_team_rankings(manager_id, year, month)
    
    if not rankings:
        return False
    
    saved_count = 0
    for ranking in rankings:
        try:
            # Check if ranking already exists
            existing = supabase.table('team_rankings').select('id').eq(
                'manager_id', manager_id
            ).eq('employee_id', ranking['employee_id']).eq(
                'year', year
            ).eq('month', month).execute()
            
            ranking_data = {
                'manager_id': manager_id,
                'employee_id': ranking['employee_id'],
                'year': year,
                'month': month,
                'rank': ranking['rank'],
                'total_goals': ranking['total_goals'],
                'completed_goals': ranking['completed_goals'],
                'completion_rate': ranking['completion_rate'],
                'avg_progress': ranking['avg_progress'],
                'score': ranking['score']
            }
            
            if existing.data:
                # Update existing
                supabase.table('team_rankings').update(ranking_data).eq(
                    'id', existing.data[0]['id']
                ).execute()
            else:
                # Insert new
                supabase.table('team_rankings').insert(ranking_data).execute()
            
            saved_count += 1
        except Exception as e:
            print(f"Error saving ranking: {str(e)}")
            continue
    
    return saved_count > 0


def get_historical_rankings(manager_id, employee_id, months=6):
    """Get historical rankings for an employee"""
    try:
        result = supabase.table('team_rankings').select('*').eq(
            'manager_id', manager_id
        ).eq('employee_id', employee_id).order(
            'year', desc=True
        ).order('month', desc=True).limit(months).execute()
        
        return result.data if result.data else []
    except Exception as e:
        print(f"Error getting historical rankings: {str(e)}")
        return []


def get_average_ranking(manager_id, employee_id, months=6):
    """Calculate average ranking for an employee over last N months"""
    history = get_historical_rankings(manager_id, employee_id, months)
    
    if not history:
        return None
    
    total_rank = sum(r['rank'] for r in history)
    avg_rank = total_rank / len(history)
    
    return {
        'avg_rank': round(avg_rank, 1),
        'best_rank': min(r['rank'] for r in history),
        'worst_rank': max(r['rank'] for r in history),
        'months_tracked': len(history)
    }

def render_password_strength_meter(password, key_suffix=""):
    """Render password strength meter with visual feedback"""
    score, color, strength, feedback = check_password_strength(password)
    
    # Strength meter HTML
    st.markdown(f"""
    <div style="margin: 15px 0;">
        <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 8px;">
            <span style="font-size: 13px; font-weight: 600; color: #64748b;">Password Strength:</span>
            <span style="font-size: 14px; font-weight: bold; color: {color};">{strength}</span>
        </div>
        <div style="width: 100%; height: 8px; background: #e2e8f0; border-radius: 10px; overflow: hidden;">
            <div style="width: {score}%; height: 100%; background: {color}; transition: width 0.3s ease;"></div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # Show feedback if password is not very strong
    if score < 85 and feedback:
        with st.expander("💡 Tips to strengthen your password", expanded=False):
            for tip in feedback:
                st.markdown(f"- {tip}")
            
            st.markdown("""
            **Best Practices:**
            - Use a mix of uppercase and lowercase letters
            - Include numbers and special characters
            - Avoid common words or patterns
            - Use at least 8-12 characters
            - Don't reuse passwords from other accounts
            """)
def send_password_reset_email(email, reset_token):
    """Send password reset email"""
    try:
        smtp_server = os.getenv('SMTP_SERVER', 'smtp.gmail.com')
        smtp_port = int(os.getenv('SMTP_PORT', '587'))
        smtp_email = os.getenv('SMTP_EMAIL')
        smtp_password = os.getenv('SMTP_PASSWORD')
        
        if not smtp_email or not smtp_password:
            st.error("Email configuration not found. Please contact administrator.")
            return False
        
        # Create reset link (you'll need to handle this in your app)
        reset_link = f"http://your-app-url.com/reset-password?token={reset_token}"
        
        # Email content
        subject = "Password Reset Request - Performance Management System"
        body = f"""
        <html>
            <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
                <div style="max-width: 600px; margin: 0 auto; padding: 20px; border: 1px solid #ddd; border-radius: 10px;">
                    <h2 style="color: #3b82f6; text-align: center;">🎯 Password Reset Request</h2>
                    <p>Hello,</p>
                    <p>You requested to reset your password for the Performance Management System.</p>
                    <p><strong>Your Reset Token:</strong></p>
                    <div style="background: #f3f4f6; padding: 15px; border-radius: 5px; text-align: center; font-size: 24px; font-weight: bold; letter-spacing: 2px; color: #1e40af;">
                        {reset_token}
                    </div>
                    <p style="margin-top: 20px;">Please use this token to reset your password. This token will expire in 1 hour.</p>
                    <p style="color: #ef4444;"><strong>⚠️ Important:</strong> If you did not request this password reset, please ignore this email or contact your administrator.</p>
                    <hr style="margin: 30px 0; border: none; border-top: 1px solid #ddd;">
                    <p style="font-size: 12px; color: #64748b; text-align: center;">
                        This is an automated email from Performance Management System.<br>
                        Please do not reply to this email.
                    </p>
                </div>
            </body>
        </html>
        """
        
        # Create message
        message = MIMEMultipart("alternative")
        message["Subject"] = subject
        message["From"] = smtp_email
        message["To"] = email
        
        html_part = MIMEText(body, "html")
        message.attach(html_part)
        
        # Send email
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(smtp_email, smtp_password)
            server.sendmail(smtp_email, email, message.as_string())
        
        return True
    except Exception as e:
        st.error(f"Failed to send email: {str(e)}")
        return False


# ✅ PAGE CONFIG MUST BE FIRST STREAMLIT COMMAND
st.set_page_config(
    page_title="Performance Management System",
    page_icon="🎯",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Load environment variables
load_dotenv()

# Initialize Supabase client ONCE
supabase = get_supabase_client()


# Import our modules
from database import Database
from helper import (
    apply_theme, init_session_state, get_quarter_months, get_month_name,
    get_quarter_name, calculate_progress, format_goal_table_data,
    render_user_avatar, render_card, render_metric_card, render_progress_bar,
    render_feedback_card, validate_goal_data
)



# Initialize session state and database
init_session_state()
db = Database()
apply_theme()

# IST Timezone
IST = pytz.timezone('Asia/Kolkata')

# Initialize session state and database
init_session_state()
db = Database()
apply_theme()

# IST Timezone
IST = pytz.timezone('Asia/Kolkata')

if 'reminder_scheduler_started' not in st.session_state:
    start_reminder_scheduler(db)
    st.session_state.reminder_scheduler_started = True

# Start daily notification checker
if 'notification_checker_started' not in st.session_state:
    import threading
    import time
    
    def run_daily_checker():
        while True:
            check_and_notify_due_dates_and_missing_updates()
            # Run once per day (86400 seconds)
            time.sleep(86400)
    
    checker_thread = threading.Thread(target=run_daily_checker, daemon=True)
    checker_thread.start()
    st.session_state.notification_checker_started = True

# ============================================
# SESSION PERSISTENCE
# ============================================
def save_session_to_storage():
    """Save session state to browser storage"""
    if st.session_state.user:
        try:
            # Store in Streamlit's query params (URL-based persistence)
            st.query_params['user_id'] = str(st.session_state.user['id'])
            st.query_params['page'] = st.session_state.page
            
            # ✅ Save navigation state
            if st.session_state.get('selected_year'):
                st.query_params['year'] = str(st.session_state.selected_year)
            if st.session_state.get('selected_quarter'):
                st.query_params['quarter'] = str(st.session_state.selected_quarter)
            if st.session_state.get('selected_month'):
                st.query_params['month'] = str(st.session_state.selected_month)
            
            # ✅ Save active tab states
            if st.session_state.get('active_month_tab'):
                st.query_params['month_tab'] = str(st.session_state.active_month_tab)
            if st.session_state.get('active_hr_tab'):
                st.query_params['hr_tab'] = str(st.session_state.active_hr_tab)
            
            # ✅ Save employee viewing state
            if st.session_state.get('viewing_employee'):
                st.query_params['viewing_emp_id'] = str(st.session_state.viewing_employee['id'])
            
            if st.session_state.get('viewing_employee_year'):
                st.query_params['viewing_emp_year'] = 'true'
            
        except Exception as e:
            print(f"Session save error: {str(e)}")


def restore_session_from_storage():
    """Restore session state from browser storage"""
    try:
        # Check URL params for user_id
        user_id = st.query_params.get('user_id')
        if user_id:
            # Restore user from database
            user = db.get_user_by_id(user_id)
            if user:
                st.session_state.user = user
                st.session_state.page = st.query_params.get('page', 'dashboard')
                
                # ✅ Restore navigation state
                year_param = st.query_params.get('year')
                if year_param:
                    st.session_state.selected_year = int(year_param)
                
                quarter_param = st.query_params.get('quarter')
                if quarter_param:
                    st.session_state.selected_quarter = int(quarter_param)
                
                month_param = st.query_params.get('month')
                if month_param:
                    st.session_state.selected_month = int(month_param)
                
                # ✅ Restore tab states
                month_tab = st.query_params.get('month_tab')
                if month_tab:
                    st.session_state.active_month_tab = int(month_tab)
                
                hr_tab = st.query_params.get('hr_tab')
                if hr_tab:
                    st.session_state.active_hr_tab = int(hr_tab)
                
                # ✅ Restore employee viewing state
                viewing_emp_id = st.query_params.get('viewing_emp_id')
                if viewing_emp_id:
                    viewing_employee = db.get_user_by_id(viewing_emp_id)
                    if viewing_employee:
                        st.session_state.viewing_employee = viewing_employee
                
                viewing_emp_year = st.query_params.get('viewing_emp_year')
                if viewing_emp_year == 'true':
                    st.session_state.viewing_employee_year = True
                
                return True
    except Exception as e:
        print(f"Session restore error: {str(e)}")
    
    return False


# ✅ RESTORE SESSION ON PAGE LOAD (if not already logged in)
if not st.session_state.user:
    restore_session_from_storage()

# ✅ SAVE SESSION ON CHANGES (if user is logged in)
if st.session_state.user:
    save_session_to_storage()



# ============================================
# WEEK DATE CALCULATION HELPERS
# ============================================
def get_week_dates(year, month, week_num):
    """Get start and end dates for a specific week in a month"""
    # Get first day of month
    first_day = date(year, month, 1)
    # Get last day of month
    last_day = date(year, month, calendar.monthrange(year, month)[1])
    
    # Calculate week ranges (approximate 7-day periods)
    if week_num == 1:
        start_date = first_day
        end_date = min(first_day + timedelta(days=6), last_day)
    elif week_num == 2:
        start_date = first_day + timedelta(days=7)
        end_date = min(first_day + timedelta(days=13), last_day)
    elif week_num == 3:
        start_date = first_day + timedelta(days=14)
        end_date = min(first_day + timedelta(days=20), last_day)
    else:  # week 4
        start_date = first_day + timedelta(days=21)
        end_date = last_day
    
    return start_date, end_date


def get_week_for_date(year, month, target_date):
    """Determine which week a date falls into"""
    first_day = date(year, month, 1)
    day_diff = (target_date - first_day).days
    
    if day_diff < 7:
        return 1
    elif day_diff < 14:
        return 2
    elif day_diff < 21:
        return 3
    else:
        return 4

def create_goal_with_date_based_placement(goal_data):
    """Create goal and place it in the correct month based on start/end dates"""
    start_date = datetime.strptime(goal_data['start_date'], '%Y-%m-%d').date()
    end_date = datetime.strptime(goal_data['end_date'], '%Y-%m-%d').date()
    
    # Determine the month based on start date
    goal_year = start_date.year
    goal_month = start_date.month
    goal_quarter = ((goal_month - 1) // 3) + 1
    
    # Update goal data with correct year, quarter, month
    goal_data['year'] = goal_year
    goal_data['quarter'] = goal_quarter
    goal_data['month'] = goal_month
    
    # Create the goal
    result = db.create_goal(goal_data)
    
    return result, goal_year, goal_quarter, goal_month
# ============================================
# LOGIN PAGE
# ============================================
import base64

def get_base64_image(image_path):
    with open(image_path, "rb") as img_file:
        encoded = base64.b64encode(img_file.read()).decode()
    return encoded
# ============================================
# LOGIN PAGE
# ============================================
def login_page():
    """Display login page with forgot password option"""
    base64_img = get_base64_image("infopacee.jpg")

    st.markdown(f"""
    <div style="display:flex; justify-content:center; align-items:center; width:100%; margin-top:10px;">
        <img src="data:image/jpg;base64,{base64_img}" style="width:120px; height:auto;">
    </div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <style>
        .stApp {
            background-color: white !important;
        }
    </style>
    """, unsafe_allow_html=True)
    """Display login page with forgot password option"""
    st.markdown("<h1 style='text-align: center;'>🎯 Performance Management System</h1>", unsafe_allow_html=True)
    st.markdown("<p style='text-align: center; color: #64748b;'>Sign in to continue</p>", unsafe_allow_html=True)
    
    # Check if showing forgot password or reset password
    if 'show_forgot_password' not in st.session_state:
        st.session_state.show_forgot_password = False
    
    if 'show_reset_password' not in st.session_state:
        st.session_state.show_reset_password = False
    
    col1, col2, col3 = st.columns([1, 2, 1])
    
    with col2:
        # ===== NORMAL LOGIN FORM =====
        if not st.session_state.show_forgot_password and not st.session_state.show_reset_password:
            with st.form("login_form"):
                email = st.text_input("📧 Email", placeholder="your@email.com")
                password = st.text_input("🔒 Password", type="password", placeholder="••••••••")
                submit = st.form_submit_button("Sign In", use_container_width=True)
                
                if submit:
                    if email and password:
                        user = db.authenticate_user(email, password)
                        if user:
                            st.session_state.user = user
                            st.session_state.page = 'dashboard'  # ✅ Set default page
                            save_session_to_storage()  # ✅ Save session
                            st.success("✅ Login successful!")
                            st.rerun()
                        else:
                            st.error("❌ Invalid credentials")
                    else:
                        st.warning("⚠️ Please enter both email and password")
            
            # Forgot Password Link
            col_fp1, col_fp2, col_fp3 = st.columns([1, 2, 1])
            with col_fp2:
                if st.button("🔑 Forgot Password?", use_container_width=True):
                    st.session_state.show_forgot_password = True
                    st.rerun()
            
            
        
        # ===== FORGOT PASSWORD FORM =====
        elif st.session_state.show_forgot_password:
            st.markdown("### 🔑 Forgot Password")
            st.info("Enter your email address and we'll send you a reset token.")
            
            with st.form("forgot_password_form"):
                reset_email = st.text_input("📧 Email Address", placeholder="your@email.com")
                submit_reset = st.form_submit_button("Send Reset Token", use_container_width=True)
                
                if submit_reset:
                    if reset_email:
                        # Generate token
                        token = db.create_password_reset_token(reset_email)
                        
                        if token:
                            # Send email
                            if send_password_reset_email(reset_email, token):
                                st.success("✅ Password reset token sent to your email!")
                                st.info(f"**Your Reset Token:** `{token}`\n\nPlease check your email for the reset token. It will expire in 1 hour.")
                                st.session_state.show_forgot_password = False
                                st.session_state.show_reset_password = True
                                st.rerun()
                            else:
                                st.warning(f"⚠️ Could not send email, but here's your reset token: **{token}**")
                                st.session_state.show_forgot_password = False
                                st.session_state.show_reset_password = True
                                st.rerun()
                        else:
                            st.error("❌ Email not found in system")
                    else:
                        st.warning("⚠️ Please enter your email address")
            
            # Back to Login
            if st.button("← Back to Login", use_container_width=True):
                st.session_state.show_forgot_password = False
                st.rerun()
        
        # ===== RESET PASSWORD FORM =====
        elif st.session_state.show_reset_password:
            st.markdown("### 🔒 Reset Password")
            st.info("Enter the reset token you received via email and your new password.")
            
            with st.form("reset_password_form"):
                reset_token = st.text_input("🎫 Reset Token", placeholder="Enter 8-character token")
                new_password = st.text_input("🔒 New Password", type="password", placeholder="••••••••", key="login_new_pass")
                
                # Password strength meter
                if new_password:
                    render_password_strength_meter(new_password, "login_reset")
                
                confirm_password = st.text_input("🔒 Confirm Password", type="password", placeholder="••••••••", key="login_confirm_pass")
                submit_new_password = st.form_submit_button("Reset Password", use_container_width=True)
                
                if submit_new_password:
                    if reset_token and new_password and confirm_password:
                        if new_password == confirm_password:
                            if len(new_password) >= 6:
                                # Check password strength
                                score, _, strength, _ = check_password_strength(new_password)
                                
                                if score < 30:
                                    st.warning(f"⚠️ Your password is {strength}. Consider making it stronger for better security.")
                                
                                # Reset password
                                if db.reset_password_with_token(reset_token, new_password):
                                    st.success("✅ Password reset successful! You can now login with your new password.")
                                    st.balloons()
                                    st.session_state.show_reset_password = False
                                    st.query_params.clear()
                                    st.rerun()
                                else:
                                    st.error("❌ Invalid or expired token")
                            else:
                                st.error("❌ Password must be at least 6 characters")
                        else:
                            st.error("❌ Passwords don't match")
                    else:
                        st.warning("⚠️ Please fill all fields")
            
            # Back to Login
            if st.button("← Back to Login", use_container_width=True):
                st.session_state.show_reset_password = False
                st.rerun()

# ============================================
# ENHANCED DASHBOARD PAGE
# ============================================
def display_dashboard():
    """Display enhanced professional dashboard"""
    user = st.session_state.user
    role = user['role']
    
    if not user:
        st.warning("⚠️ Session expired. Please login again.")
        st.rerun()

    # ✅ INITIALIZE ALL VARIABLES AT THE START
    avg_progress = 0
    total_goals = 0
    completed_goals = 0
    active_goals = 0
    overdue_goals = 0
    user_goals = []

    # Welcome Header with gradient
    role_greetings = {
        'CMD': '👑 Chief Managing Director',
        'VP': '🎖️ Vice President',
        'HR': '🏢 Human Resources',
        'Manager': '👔 Manager',
        'Employee': '👤 Employee'
    }
    
    st.markdown(f"""
    <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                padding: 30px; border-radius: 15px; margin-bottom: 30px; box-shadow: 0 4px 6px rgba(0,0,0,0.1);">
        <h1 style="color: white; margin: 0; font-size: 32px;">Welcome back, {user['name']}! 👋</h1>
        <p style="color: rgba(255,255,255,0.9); margin: 10px 0 0 0; font-size: 16px;">
            {role_greetings.get(role, user.get('designation', 'Employee'))} • {datetime.now().strftime('%A, %B %d, %Y')}
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    # Rest of dashboard remains similar but with role-based views...
    # Add CMD/VP specific organization-wide metrics
    
    if role in ['CMD', 'VP','HR']:
    
        all_users = db.get_all_users()
        total_cmd = len([u for u in all_users if u['role'] == 'CMD'])
        total_vp = len([u for u in all_users if u['role'] == 'VP'])
        total_hr = len([u for u in all_users if u['role'] == 'HR'])
        total_managers = len([u for u in all_users if u['role'] == 'Manager'])
        total_employees = len([u for u in all_users if u['role'] == 'Employee'])
        
        st.markdown("###  Your Performance Overview")

        # Get all user goals
        user_goals = db.get_user_all_goals(user['id'])
        
        # ✅ Calculate metrics FIRST (before using them in HTML)
        total_goals = len(user_goals)
        completed_goals = len([g for g in user_goals if g.get('status') == 'Completed'])
        active_goals = len([g for g in user_goals if g.get('status') == 'Active'])
        
        # Calculate average progress
        total_progress = 0
        goals_with_progress = 0
        
        for goal in user_goals:
            monthly_achievement = goal.get('monthly_achievement')
            if monthly_achievement is not None:
                monthly_target = goal.get('monthly_target', 1)
                if monthly_target > 0:
                    progress = (monthly_achievement / monthly_target * 100)
                    total_progress += progress
                    goals_with_progress += 1
        
        avg_progress = (total_progress / goals_with_progress) if goals_with_progress > 0 else 0
        completion_rate = (completed_goals / total_goals * 100) if total_goals > 0 else 0
        
        # Count overdue goals
        overdue_goals = 0
        today = date.today()
        for goal in user_goals:
            if goal.get('status') == 'Active':
                end_date_str = goal.get('end_date')
                if end_date_str:
                    try:
                        end_date = datetime.strptime(str(end_date_str), '%Y-%m-%d').date()
                        if today > end_date:
                            overdue_goals += 1
                    except:
                        pass
        
        # NOW display the metrics (after all calculations are done)
        col_perf1, col_perf2, col_perf3, col_perf4, col_perf5 = st.columns(5)
        
        with col_perf1:
            st.markdown(f"""
            <div style="background: #FFFFFF; width: 100%; height: 160px;
                    padding: 20px; border-radius: 10px; text-align: center; box-shadow:0 2px 4px rgba(0,0,0,0.08),0 0 18px rgba(59,130,246,0.25);display: flex; flex-direction: column; justify-content: center;">
                <div style="font-size: 36px; margin-bottom: 10px;"></div>
                <div style="font-size: 32px; font-weight: 700; color: #3B82F6; margin-bottom: 6px;">{total_goals}</div>
                <div style="font-size: 13px; font-weight: 700; text-transform: uppercase; color: #3B82F6;
                    padding: 4px 10px; border-radius: 6px; display: inline-block;">Total Goals</div>
            </div>
            """, unsafe_allow_html=True)
        
        with col_perf2:
            st.markdown(f"""
            <div style="background:#FFF; width:100%; height:160px; padding:20px; border-radius:10px;
                text-align:center; display:flex; flex-direction:column; justify-content:center;
                box-shadow:0 2px 4px rgba(0,0,0,0.08), 0 0 18px rgba(16,185,129,0.22);">
                <div style="font-size:32px; font-weight:700; color:#10B981; margin-bottom:6px;">
                    {completed_goals}
                </div>
                <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:#10B981;
                     padding:4px 10px; border-radius:6px; display:inline-block;">
                    Completed</div>
            </div>
            """, unsafe_allow_html=True)

        
        with col_perf3:
            st.markdown(f"""
            <div style="background:#FFF; width:100%; height:160px; padding:20px; border-radius:10px;
                text-align:center; display:flex; flex-direction:column; justify-content:center;
                box-shadow:0 2px 4px rgba(0,0,0,0.08), 0 0 18px rgba(245,87,108,0.23);">
                <div style="font-size:32px; font-weight:700; color:#F5576C; margin-bottom:6px;">
                    {active_goals}
                </div>
                <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:#F5576C;
                    padding:4px 10px; border-radius:6px; display:inline-block;">
                    Active
                </div>
            </div>
            """, unsafe_allow_html=True)
        
        with col_perf4:
            st.markdown(f"""
            <div style=" background:#FFF; width:100%; height:160px; padding:20px; border-radius:10px;
                text-align:center; display:flex; flex-direction:column; justify-content:center;
                box-shadow:0 2px 4px rgba(0,0,0,0.08), 0 0 18px rgba(0,201,255,0.23);">
                <div style="font-size:32px; font-weight:700; color:#00C9FF; margin-bottom:6px;">
                    {avg_progress:.1f}%
                </div>
                <div style="
                    font-size:13px; font-weight:700; text-transform:uppercase; color:#00C9FF;
                     padding:4px 10px; border-radius:6px; display:inline-block;">
                    Avg Progress
                </div>
            </div>
            """, unsafe_allow_html=True)
        
        with col_perf5:
            overdue_color = "#EF4444" if overdue_goals > 0 else "#10B981"
            label_bg = "#FEF2F2" if overdue_goals > 0 else "#ECFDF5"
            glow = "rgba(239,68,68,0.23)" if overdue_goals > 0 else "rgba(16,185,129,0.23)"

            st.markdown(f"""
            <div style="background:#FFF; width:100%; height:160px; padding:20px; border-radius:10px;
                text-align:center; display:flex; flex-direction:column; justify-content:center;
                box-shadow:0 2px 4px rgba(0,0,0,0.08), 0 0 18px {glow};">
                <div style="font-size:32px; font-weight:700; color:{overdue_color}; margin-bottom:6px;">
                    {overdue_goals}
                </div>
                <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:{overdue_color};
                     padding:4px 10px; border-radius:6px; display:inline-block;">Overdue
                </div>
            </div>
            """, unsafe_allow_html=True)

        
        st.markdown("---")
        # Department-wise breakdown
        col_left, col_right = st.columns([1, 1])
        
        with col_left:
            st.markdown("#### 📊 Department Distribution")
            dept_count = {}
            for u in all_users:
                dept = u.get('department', 'Unassigned')
                dept_count[dept] = dept_count.get(dept, 0) + 1
            
            if dept_count:
                fig_dept = px.pie(
                    values=list(dept_count.values()),
                    names=list(dept_count.keys()),
                    title="Employees by Department",
                    hole=0.4,
                    color_discrete_sequence=px.colors.qualitative.Set3
                )
                fig_dept.update_traces(textposition='inside', textinfo='percent+label')
                fig_dept.update_layout(height=350, showlegend=True)
                st.plotly_chart(fig_dept, use_container_width=True)
        
        with col_right:
            st.markdown("#### 🎯 Goals Performance")
            total_goals_org = sum([len(db.get_user_all_goals(u['id'])) for u in all_users])
            completed_goals_org = sum([db.get_user_goal_stats(u['id']).get('completed_goals', 0) for u in all_users])
            active_goals_org = sum([db.get_user_goal_stats(u['id']).get('active_goals', 0) for u in all_users])
            
            fig_goals = go.Figure(data=[
                go.Bar(name='Total', x=['Goals'], y=[total_goals_org], marker_color='#3b82f6'),
                go.Bar(name='Completed', x=['Goals'], y=[completed_goals_org], marker_color='#10b981'),
                go.Bar(name='Active', x=['Goals'], y=[active_goals_org], marker_color='#f59e0b')
            ])
            fig_goals.update_layout(
                title="Organization Goals Overview",
                barmode='group',
                height=350,
                showlegend=True
            )
            st.plotly_chart(fig_goals, use_container_width=True)
        
        st.markdown("---")
        
        # ✅ FIXED: Top Performers - Role-based display
        st.markdown("#### 🏆 Top Performers (This Month)")
        rankings = []
        
        # Determine which employees to rank based on role
        if role == 'Manager':
            # Manager sees only their team members
            team_members = db.get_team_members(user['id'])
            employees_to_rank = [m for m in team_members if m['role'] == 'Employee']
            top_count = 3  # Show top 3 for managers
            performer_scope = "Team"
        else:
            # CMD, VP, HR see all employees in organization
            employees_to_rank = [u for u in all_users if u['role'] == 'Employee']
            top_count = 5  # Show top 5 for CMD/VP/HR
            performer_scope = "Organization"
        
        for emp in employees_to_rank:
            emp_stats = db.get_user_goal_stats(emp['id'])
            if emp_stats.get('total_goals', 0) > 0:
                rankings.append({
                    'Name': emp['name'],
                    'Department': emp.get('department', 'N/A'),
                    'Designation': emp.get('designation', 'N/A'),
                    'Total Goals': emp_stats.get('total_goals', 0),
                    'Completed': emp_stats.get('completed_goals', 0),
                    'Progress %': f"{emp_stats.get('avg_progress', 0):.1f}%",
                    'Progress_Val': emp_stats.get('avg_progress', 0)
                })
        
        if rankings:
            df_rank = pd.DataFrame(rankings)
            df_rank = df_rank.sort_values('Progress_Val', ascending=False).head(top_count)
            df_rank = df_rank.drop('Progress_Val', axis=1)
            df_rank.insert(0, 'Rank', range(1, len(df_rank) + 1))
            
            # Show context-aware caption
            st.caption(f"📊 Showing Top {min(len(df_rank), top_count)} of {len(rankings)} {performer_scope} Performers")
            st.dataframe(df_rank, use_container_width=True, height=300)
        else:
            st.info(f"No performance data available for {performer_scope.lower()}")
            
    elif role == 'Manager':
        # Personal Performance Metrics
        st.markdown("### 📊 Your Performance Overview")
        
        # Get all user goals
        user_goals = db.get_user_all_goals(user['id'])
        
        # ✅ Calculate metrics (Managers see all their goals - no approval needed)
        total_goals = len(user_goals)
        completed_goals = len([g for g in user_goals if g.get('status') == 'Completed'])
        active_goals = len([g for g in user_goals if g.get('status') == 'Active'])
        
        # Calculate average progress
        total_progress = 0
        goals_with_progress = 0
        
        for goal in user_goals:
            monthly_achievement = goal.get('monthly_achievement')
            if monthly_achievement is not None:
                monthly_target = goal.get('monthly_target', 1)
                if monthly_target > 0:
                    progress = (monthly_achievement / monthly_target * 100)
                    total_progress += progress
                    goals_with_progress += 1
        
        avg_progress = (total_progress / goals_with_progress) if goals_with_progress > 0 else 0
        
        # Count overdue goals
        overdue_goals = 0
        today = date.today()
        for goal in user_goals:
            if goal.get('status') == 'Active':
                end_date_str = goal.get('end_date')
                if end_date_str:
                    try:
                        end_date = datetime.strptime(str(end_date_str), '%Y-%m-%d').date()
                        if today > end_date:
                            overdue_goals += 1
                    except:
                        pass
        
        col_perf1, col_perf2, col_perf3, col_perf4, col_perf5 = st.columns(5)
        
        # ===========================
# CLEAN PASTEL METRIC CARDS
# ===========================

        with col_perf1:
            st.markdown(f"""
            <div style="background:#FFF; width:100%; height:160px; padding:20px; border-radius:10px;
                text-align:center; display:flex; flex-direction:column; justify-content:center;
                box-shadow:0 2px 4px rgba(0,0,0,0.08),0 0 18px rgba(59,130,246,0.25);">
                <div style="font-size:32px; font-weight:700; color:#3B82F6; margin-bottom:6px;">
                    {total_goals}
                </div>
                <div style="font-size:13px; font-weight:700; text-transform:uppercase;
                    color:#3B82F6;padding:4px 10px; border-radius:6px; display:inline-block;">Total Goals
                </div>
            </div>
            """, unsafe_allow_html=True)

        with col_perf2:
            st.markdown(f"""
            <div style="background:#FFF; width:100%; height:160px; padding:20px; border-radius:10px;
                text-align:center; display:flex; flex-direction:column; justify-content:center;
                box-shadow:0 2px 4px rgba(0,0,0,0.08),0 0 18px rgba(16,185,129,0.22);">
                <div style="font-size:32px; font-weight:700; color:#10B981; margin-bottom:6px;">
                    {completed_goals}
                </div>
                <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:#10B981; padding:4px 10px; border-radius:6px; display:inline-block;"> Completed
                </div>
            </div>
            """, unsafe_allow_html=True)

        with col_perf3:
            st.markdown(f"""
            <div style="background:#FFF; width:100%; height:160px; padding:20px; border-radius:10px;
                text-align:center; display:flex; flex-direction:column; justify-content:center;
                box-shadow:0 2px 4px rgba(0,0,0,0.08),0 0 18px rgba(245,87,108,0.23);">
                <div style="font-size:32px; font-weight:700; color:#F5576C; margin-bottom:6px;">
                    {active_goals}
                </div>
                <div style="font-size:13px; font-weight:700; text-transform:uppercase;
                    color:#F5576C;padding:4px 10px; border-radius:6px; display:inline-block;">Active
                </div>
            </div>
            """, unsafe_allow_html=True)


        with col_perf4:
            st.markdown(f"""
            <div style="background:#FFF; width:100%; height:160px; padding:20px; border-radius:10px;
                text-align:center; display:flex; flex-direction:column; justify-content:center;
                box-shadow:0 2px 4px rgba(0,0,0,0.08),0 0 18px rgba(0,201,255,0.23);">
                <div style="font-size:32px; font-weight:700; color:#00C9FF; margin-bottom:6px;">
                    {avg_progress:.1f}%
                </div>
                <div style="font-size:13px; font-weight:700; text-transform:uppercase;color:#00C9FF;
                    padding:4px 10px; border-radius:6px; display:inline-block;">Avg Progress
                </div>
            </div>
            """, unsafe_allow_html=True)

        with col_perf5:
            overdue_color = "#EF4444" if overdue_goals > 0 else "#10B981"
            label_bg = "#FEF2F2" if overdue_goals > 0 else "#ECFDF5"
            glow = "rgba(239,68,68,0.23)" if overdue_goals > 0 else "rgba(16,185,129,0.23)"

            st.markdown(f"""
            <div style="background:#FFF; width:100%; height:160px; padding:20px; border-radius:10px;
                text-align:center; display:flex; flex-direction:column; justify-content:center;
                box-shadow:0 2px 4px rgba(0,0,0,0.08),0 0 18px {glow};">
                <div style="font-size:32px; font-weight:700; color:{overdue_color}; margin-bottom:6px;">
                    {overdue_goals}
                </div>
                <div style="font-size:13px; font-weight:700; text-transform:uppercase;
                    color:{overdue_color};padding:4px 10px; border-radius:6px; display:inline-block;">Overdue
                </div>
            </div>
            """, unsafe_allow_html=True)        
        st.markdown("---")
        
        # Team Overview
        team_members = db.get_team_members(user['id'])
        
        st.markdown("###  Team Overview")
        
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.markdown(f"""
            <div style="background:#FFF; width:100%; height:160px; padding:20px; border-radius:10px;
                text-align:center; display:flex; flex-direction:column; justify-content:center;
                box-shadow:0 2px 4px rgba(0,0,0,0.08), 0 0 18px rgba(118,75,162,0.23);">
                <div style="font-size:28px; font-weight:700; color:#764BA2; margin-bottom:6px;">
                    {len(team_members)}
                </div>
                <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:#764BA2;
                    padding:4px 10px; border-radius:6px; display:inline-block;">Team Members
                </div>
            </div>
            """, unsafe_allow_html=True)

        with col2:
            team_goals = sum([len(db.get_user_all_goals(m['id'])) for m in team_members])
            st.markdown(f"""
                <div style="background:#FFF; width:100%; height:160px; padding:20px; border-radius:10px;
                    text-align:center; display:flex; flex-direction:column; justify-content:center;
                    box-shadow:0 2px 4px rgba(0,0,0,0.08), 0 0 18px rgba(56,249,215,0.23);">
                    <div style="font-size:28px; font-weight:700; color:#10B981; margin-bottom:6px;">
                        {team_goals}
                    </div>
                    <div style="font-size:13px; font-weight:700; text-transform:uppercase;
                        color:#10B981; padding:4px 10px; border-radius:6px; display:inline-block;">Team Goals
                    </div>
                </div>
                """, unsafe_allow_html=True)

        
        with col3:
            team_completed = sum([db.get_user_goal_stats(m['id']).get('completed_goals', 0) for m in team_members])
            st.markdown(f"""
                <div style="background:#FFF; width:100%; height:160px; padding:20px; border-radius:10px;
                    text-align:center; display:flex; flex-direction:column; justify-content:center;
                    box-shadow:0 2px 4px rgba(0,0,0,0.08), 0 0 18px rgba(245,87,108,0.23);">
                    <div style="font-size:28px; font-weight:700; color:#F5576C; margin-bottom:6px;">
                        {team_completed}
                    </div>
                    <div style="font-size:13px; font-weight:700; text-transform:uppercase;color:#F5576C;padding:4px 10px; border-radius:6px; display:inline-block;">
                    Completed
                    </div>
                </div>
                """, unsafe_allow_html=True)
            
        with col4:
            st.markdown(f"""
                <div style="background:#FFF; width:100%; height:160px; padding:20px; border-radius:10px;
                text-align:center; display:flex; flex-direction:column; justify-content:center;
                box-shadow:0 2px 4px rgba(0,0,0,0.08), 0 0 18px rgba(245,158,11,0.23);">
                <div style="font-size:28px; font-weight:700; color:#F59E0B; margin-bottom:6px;">
                    {total_goals}
                </div>
                <div style="font-size:13px; font-weight:700; text-transform:uppercase;color:#F59E0B;
                    padding:4px 10px; border-radius:6px; display:inline-block;">My Goals
                </div>
            </div>
            """, unsafe_allow_html=True)

        
        st.markdown("---")
        
        # Top Performers from Team
        st.markdown("### 🏆 Top 3 Team Performers (This Month)")
        team_rankings = []
        for member in team_members:
            member_stats = db.get_user_goal_stats(member['id'])
            if member_stats.get('total_goals', 0) > 0:
                team_rankings.append({
                    'Name': member['name'],
                    'Department': member.get('department', 'N/A'),
                    'Designation': member.get('designation', 'N/A'),
                    'Total Goals': member_stats.get('total_goals', 0),
                    'Completed': member_stats.get('completed_goals', 0),
                    'Progress %': f"{member_stats.get('avg_progress', 0):.1f}%",
                    'Progress_Val': member_stats.get('avg_progress', 0)
                })
        
        if team_rankings:
            df_team_rank = pd.DataFrame(team_rankings)
            df_team_rank = df_team_rank.sort_values('Progress_Val', ascending=False).head(3)  # Show top 3
            df_team_rank = df_team_rank.drop('Progress_Val', axis=1)
            df_team_rank.insert(0, 'Rank', range(1, len(df_team_rank) + 1))
            df_team_rank = df_team_rank.reset_index(drop=True)
            st.dataframe(df_team_rank, use_container_width=True, height=300, hide_index=True)
        else:
            st.info("No team performance data available yet")
  

    else:  # Employee
    # Personal Performance Metrics for Employee
        st.markdown("### 📊 Your Performance Overview")
        
                # Get all user goals
        all_user_goals = db.get_user_all_goals(user['id'])

        # ✅ CRITICAL FIX: Filter to only APPROVED goals for employees
        if user['role'] == 'Employee':
            user_goals = [g for g in all_user_goals if g.get('approval_status') == 'approved']
            
            # Show pending count
            pending_count = len([g for g in all_user_goals if g.get('approval_status') == 'pending'])
            if pending_count > 0:
                st.info(f"ℹ️ You have {pending_count} goal(s) pending manager approval (not shown in metrics)")
        else:
            # Managers/HR/VP/CMD see all their goals
            user_goals = all_user_goals

        # ✅ Calculate metrics from filtered goals
        total_goals = len(user_goals)
        completed_goals = len([g for g in user_goals if g.get('status') == 'Completed'])
        active_goals = len([g for g in user_goals if g.get('status') == 'Active'])
        
        # Count overdue goals
        overdue_goals = 0
        today = date.today()
        for goal in user_goals:
            if goal.get('status') == 'Active':
                end_date_str = goal.get('end_date')
                if end_date_str:
                    try:
                        end_date = datetime.strptime(str(end_date_str), '%Y-%m-%d').date()
                        if today > end_date:
                            overdue_goals += 1
                    except:
                        pass
        
        col_perf1, col_perf2, col_perf3, col_perf4, col_perf5 = st.columns(5)
        
            # =============================
        # CLEAN WHITE + GLOW KPI CARDS
        # =============================

        with col_perf1:
            st.markdown(f"""
            <div style="background:#FFF; width:100%; height:160px; padding:20px;
                border-radius:10px; text-align:center;
                display:flex; flex-direction:column; justify-content:center;
                box-shadow:0 2px 4px rgba(0,0,0,0.08),
                            0 0 18px rgba(59,130,246,0.25);">
                <div style="font-size:32px; font-weight:700; color:#3B82F6; margin-bottom:6px;">{total_goals}
                </div>
                <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:#3B82F6;
                    padding:4px 10px; border-radius:6px; display:inline-block;">Total Goals
                </div>
            </div>
            """, unsafe_allow_html=True)

        with col_perf2:
            st.markdown(f"""
            <div style="background:#FFF; width:100%; height:160px; padding:20px; border-radius:10px; text-align:center;
                display:flex; flex-direction:column; justify-content:center;
                box-shadow:0 2px 4px rgba(0,0,0,0.08), 0 0 18px rgba(16,185,129,0.22);">
                <div style="font-size:32px; font-weight:700; color:#10B981; margin-bottom:6px;">
                    {completed_goals}
                </div>
                <div style=" font-size:13px; font-weight:700; text-transform:uppercase;
                    color:#10B981; padding:4px 10px; border-radius:6px; display:inline-block;">Completed
                </div>
            </div>
            """, unsafe_allow_html=True)

        with col_perf3:
            st.markdown(f"""
            <div style=" background:#FFF; width:100%; height:160px; padding:20px;
                border-radius:10px; text-align:center;
                display:flex; flex-direction:column; justify-content:center;
                box-shadow:0 2px 4px rgba(0,0,0,0.08), 0 0 18px rgba(245,87,108,0.23);">
                <div style="font-size:32px; font-weight:700; color:#F5576C; margin-bottom:6px;">
                    {active_goals}
                </div>
                <div style=" font-size:13px; font-weight:700; text-transform:uppercase;
                    color:#F5576C; padding:4px 10px; border-radius:6px; display:inline-block;">Active
                </div>
            </div>
            """, unsafe_allow_html=True)

        with col_perf4:
            st.markdown(f"""
            <div style=" background:#FFF; width:100%; height:160px; padding:20px; border-radius:10px; text-align:center;
                display:flex; flex-direction:column; justify-content:center;
                box-shadow:0 2px 4px rgba(0,0,0,0.08), 0 0 18px rgba(0,201,255,0.23);">
                <div style="font-size:32px; font-weight:700; color:#00C9FF; margin-bottom:6px;">
                    {avg_progress:.1f}%
                </div>
                <div style=" font-size:13px; font-weight:700; text-transform:uppercase;
                    color:#00C9FF; padding:4px 10px; border-radius:6px; display:inline-block;"> Avg Progress
                </div>
            </div>
            """, unsafe_allow_html=True)

        with col_perf5:
            overdue_color = "#EF4444" if overdue_goals > 0 else "#10B981"
            label_bg = "#FEF2F2" if overdue_goals > 0 else "#ECFDF5"
            glow = "rgba(239,68,68,0.23)" if overdue_goals > 0 else "rgba(16,185,129,0.23)"

            st.markdown(f"""
            <div style=" background:#FFF; width:100%; height:160px; padding:20px;
                border-radius:10px; text-align:center;
                display:flex; flex-direction:column; justify-content:center;
                box-shadow:0 2px 4px rgba(0,0,0,0.08), 0 0 18px rgba(16,185,129,0.23);">
                <div style="font-size:32px; font-weight:700; color:{overdue_color}; margin-bottom:6px;"> {overdue_goals} </div> 
                <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:{overdue_color}; padding:4px 10px; border-radius:6px; display:inline-block;"> Overdue </div>
            </div>
            """, unsafe_allow_html=True)

   # Notifications Section
    st.markdown("---")
    
    col_notif_header, col_notif_actions, col_notif_clear = st.columns([3, 1, 1])
    with col_notif_header:
        st.markdown("### 🔔 Recent Activity & Reminders")
    
    # ADD THIS NEW SUMMARY SECTION
    notifications = get_enhanced_notifications(user)
    unread_count = len([n for n in notifications if not n.get('is_read', False)])

    # Count by type
    notif_counts = {
        'critical': 0,  # goal_not_completed, overdue, goal_not_updated
        'important': 0,  # deadline, feedback_received
        'normal': 0  # everything else
    }

    for notif in notifications:
        if not notif.get('is_read', False):
            notif_type = notif.get('type', '')
            if notif_type in ['goal_not_completed', 'overdue', 'goal_not_updated']:
                notif_counts['critical'] += 1
            elif notif_type in ['deadline', 'feedback_received', 'goal_approved']:
                notif_counts['important'] += 1
            else:
                notif_counts['normal'] += 1

    # Display summary badges
    # Display summary badges
    col_badge1, col_badge2, col_badge3, col_badge4 = st.columns(4)

        # =============================
    # CLEAN WHITE + GLOW BADGE CARDS
    # =============================

    with col_badge1:  # Critical
        st.markdown(f"""
        <div style="background:#EF4444; width:100%; padding:14px; border-radius:10px;
            text-align:center; display:flex; flex-direction:column; justify-content:center;
            box-shadow:0 2px 4px rgba(0,0,0,0.06), 0 0 14px rgba(239,68,68,0.25);">
            <div style="font-size:22px; font-weight:700; color:#FEE2E2;">
                {notif_counts['critical']}
            </div> 
            <div style=" margin-top:4px; font-size:11px; font-weight:700; text-transform:uppercase;
                 color:#FEE2E2; padding:3px 8px; border-radius:5px; display:inline-block;">Critical
            </div>
        </div>
        """, unsafe_allow_html=True)

    with col_badge2:  # Important
        st.markdown(f"""
        <div style=" background:#F59E0B; width:100%; padding:14px; border-radius:10px;
            text-align:center; display:flex; flex-direction:column; justify-content:center;
            box-shadow:0 2px 4px rgba(0,0,0,0.06), 0 0 14px rgba(245,158,11,0.25);">
            <div style="font-size:22px; font-weight:700; color:#FFF7DA;">
                {notif_counts['important']}
            </div>
            <div style=" margin-top:4px; font-size:11px; font-weight:700; text-transform:uppercase;
                color:#FFF7DA; padding:3px 8px; border-radius:5px; display:inline-block;"> Important
            </div>
        </div>
        """, unsafe_allow_html=True)

    with col_badge3:  # Updates (Normal)
        st.markdown(f"""
        <div style=" background:#3B82F6; width:100%; padding:14px; border-radius:10px;
            text-align:center; display:flex; flex-direction:column; justify-content:center;
            box-shadow:0 2px 4px rgba(0,0,0,0.06), 0 0 14px rgba(59,130,246,0.25); ">
            <div style="font-size:22px; font-weight:700; color:#DBEAFE;">
                {notif_counts['normal']}
            </div>
            <div style=" margin-top:4px; font-size:11px; font-weight:700; text-transform:uppercase; color:#DBEAFE;
                padding:3px 8px; border-radius:5px; display:inline-block;"> Updates
            </div>
        </div>
        """, unsafe_allow_html=True)

    with col_badge4:  # Total Unread
        st.markdown(f"""
        <div style=" background: #64748B; width:100%; padding:14px; border-radius:10px;
            text-align:center; display:flex; flex-direction:column; justify-content:center;
            box-shadow:0 2px 4px rgba(0,0,0,0.06), 0 0 14px rgba(100,116,139,0.25); ">
            <div style="font-size:22px; font-weight:700; color:#E2E8F0;">
                {unread_count}
            </div> 
            <div style=" margin-top:4px; font-size:11px; font-weight:700; text-transform:uppercase; color:#E2E8F0; padding:3px 8px; border-radius:5px; display:inline-block;">Total Unread
            </div>
        </div>
        """, unsafe_allow_html=True)

    with col_notif_actions:
        if st.button("View All", use_container_width=True):
            st.session_state.show_all_notifications = True
    with col_notif_clear:
        if st.button("✓ Mark All Read", use_container_width=True):
            try:
                # Mark all unread notifications as read
                supabase.table('notifications').update({
                    'is_read': True,
                    'read_at': datetime.now(IST).isoformat()
                }).eq('user_id', user['id']).eq('is_read', False).execute()
                
                st.success("✅ All notifications marked as read!")
                st.rerun()
            except Exception as e:
                st.error(f"Error: {str(e)}")
   
        
    notifications = get_enhanced_notifications(user)
    unread_count = len([n for n in notifications if not n.get('is_read', False)])

    st.caption(f"**{unread_count} unread notifications**")
    
    # ✅ FILTER: Only show UNREAD notifications
    unread_notifications = [n for n in notifications if not n.get('is_read', False)]

    if unread_notifications:
        # Show top 5 or all based on state
        display_count = len(unread_notifications) if st.session_state.get('show_all_notifications') else 5
        
        for notif in unread_notifications[:display_count]:
            notif_type = notif['type']
            
            if notif_type == 'goal_created':
                icon = "📝"
                color = "#3b82f6"
            elif notif_type == 'goal_approved':
                icon = "✅"
                color = "#10b981"
            elif notif_type == 'goal_edited':
                icon = "✏️"
                color = "#f59e0b"
            elif notif_type == 'goal_deleted':
                icon = "🗑️"
                color = "#ef4444"
            elif notif_type == 'achievement':
                icon = "🎉"
                color = "#10b981"
            elif notif_type == 'goal_not_completed':
                icon = "❌"
                color = "#ef4444"
            elif notif_type == 'update':
                icon = "📊"
                color = "#3b82f6"
            elif notif_type == 'goal_not_updated':
                icon = "⚠️"
                color = "#f97316"
            elif notif_type == 'feedback':
                icon = "💬"
                color = "#8b5cf6"
            elif notif_type == 'feedback_given':
                icon = "✍️"
                color = "#6366f1"
            elif notif_type == 'feedback_reply':
                icon = "↩️"
                color = "#8b5cf6"
            elif notif_type == 'deadline':
                icon = "⏰"
                color = "#f59e0b"
            elif notif_type == 'overdue':
                icon = "🚨"
                color = "#ef4444"
            elif notif_type == 'assignment':
                icon = "📬"
                color = "#3b82f6"
            else:
                icon = "ℹ️"
                color = "#64748b"
            
            # Create columns for inline tick button
            col_notif, col_tick = st.columns([20, 1])
            
            with col_notif:
                st.markdown(f'''
                <div style="background: #ffffff; padding: 12px 16px; border-radius: 8px; border-left: 4px solid {color}; 
                            box-shadow: 0 1px 3px rgba(0,0,0,0.1);">
                    <div style="display: flex; justify-content: space-between; align-items: start;">
                        <div style="flex: 1;">
                            <div style="font-weight: 600; color: #1e293b; margin-bottom: 4px;">
                                {icon} {notif['title']}
                            </div>
                            <div style="color: #475569; font-size: 14px;">
                                {notif['message']}
                            </div>
                            <div style="font-size: 11px; color: #94a3b8; margin-top: 6px;">
                                {notif.get('time', '')}
                            </div>
                        </div>
                    </div>
                </div>
                ''', unsafe_allow_html=True)
            
            # Tick mark button on the right
            with col_tick:
                if notif.get('notification_id'):
                    if st.button("✓", key=f"mark_read_{notif.get('notification_id')}", help="Mark as read"):
                        if mark_notification_read(notif.get('notification_id')):
                            st.rerun()
        
        # Show more button only if there are more unread notifications
        if len(unread_notifications) > 5 and not st.session_state.get('show_all_notifications'):
            if st.button("📋 Show All Notifications", use_container_width=True):
                st.session_state.show_all_notifications = True
                st.rerun()
        elif st.session_state.get('show_all_notifications'):
            if st.button("📋 Show Less", use_container_width=True):
                st.session_state.show_all_notifications = False
                st.rerun()
    else:
        st.info("✨ All caught up! No new notifications.")

def get_enhanced_notifications(user):
    """Get enhanced notifications from database only"""
    notifications = []
    role = user['role']
    today = date.today()
    # ✅ FIX: Use timezone-aware IST datetime
    now = datetime.now(IST)
    
    # ✅ Fetch database notifications
    db_notifications = get_user_notifications(user['id'], limit=100)
    
    for notif in db_notifications:
        try:
            created_at = notif.get('created_at')
            notif_datetime = None
            
            if created_at:
                try:
                    if isinstance(created_at, str):
                        # Parse ISO format timestamp
                        if 'T' in created_at:
                            # Handle timezone info in string
                            if '+' in created_at or created_at.endswith('Z'):
                                # Parse with timezone and convert to IST
                                notif_datetime = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                                notif_datetime = notif_datetime.astimezone(IST)
                            else:
                                # Parse as UTC and convert to IST
                                notif_datetime = datetime.strptime(created_at[:19], '%Y-%m-%dT%H:%M:%S')
                                notif_datetime = pytz.utc.localize(notif_datetime).astimezone(IST)
                        else:
                            # Date only - use IST midnight
                            notif_datetime = datetime.strptime(created_at[:10], '%Y-%m-%d')
                            notif_datetime = IST.localize(notif_datetime)
                    elif isinstance(created_at, datetime):
                        # Already datetime object
                        if created_at.tzinfo is None:
                            # Assume UTC if no timezone
                            notif_datetime = pytz.utc.localize(created_at).astimezone(IST)
                        else:
                            # Convert to IST
                            notif_datetime = created_at.astimezone(IST)
                    else:
                        notif_datetime = now
                except Exception as e:
                    print(f"Error parsing date '{created_at}': {str(e)}")
                    notif_datetime = now
            else:
                notif_datetime = now
            
            # ✅ Ensure both datetimes are timezone-aware
            if notif_datetime.tzinfo is None:
                notif_datetime = IST.localize(notif_datetime)
            
            # Calculate time difference
            time_diff = now - notif_datetime
            days_since = time_diff.days
            
            # ✅ Only show notifications with valid (non-negative) time
            if days_since <= 90 and days_since >= 0:
                hours_since = time_diff.total_seconds() / 3600
                
                # Format time ago
                if hours_since < 0:
                    time_ago = "Just now"
                elif hours_since < 1:
                    minutes = max(0, int(hours_since * 60))
                    time_ago = f"{minutes} minute{'s' if minutes != 1 else ''} ago"
                elif hours_since < 24:
                    hours = int(hours_since)
                    time_ago = f"{hours} hour{'s' if hours != 1 else ''} ago"
                elif days_since == 1:
                    time_ago = "Yesterday"
                else:
                    time_ago = f"{days_since} day{'s' if days_since != 1 else ''} ago"
                
                action_type = notif.get('action_type', 'update')
                
                # Map action types to notification types and icons
                type_map = {
                    'goal_created': ('goal_created', '📝'),
                    'goal_approved': ('goal_approved', '✅'),
                    'goal_edited': ('goal_edited', '✏️'),
                    'goal_deleted': ('goal_deleted', '🗑️'),
                    'goal_completed': ('achievement', '🎉'),
                    'goal_not_completed': ('goal_not_completed', '❌'),
                    'goal_assigned': ('assignment', '🎯'),
                    'weekly_achievement_updated': ('update', '📊'),
                    'goal_not_updated': ('goal_not_updated', '⚠️'),
                    'feedback_received': ('feedback', '💬'),
                    'feedback_given': ('feedback_given', '✍️'),
                    'feedback_reply': ('feedback_reply', '↩️'),
                    'goal_due_soon': ('deadline', '⏰'),
                    'goal_overdue': ('overdue', '🚨')
                }
                
                notif_type, icon = type_map.get(action_type, ('update', 'ℹ️'))
                
                notifications.append({
                    'type': notif_type,
                    'title': notif.get('action_type', 'Update').replace('_', ' ').title(),
                    'message': notif.get('details', 'No details available'),
                    'time': time_ago,
                    'timestamp': notif_datetime,
                    'priority': 0 if not notif.get('is_read') else 5,
                    'is_read': notif.get('is_read', False),
                    'notification_id': notif.get('id')
                })
                
        except Exception as e:
            print(f"Error parsing notification: {str(e)}")
            continue
    
    # Sort by priority and timestamp
    priority_weights = {
        'goal_not_completed': 1,
        'overdue': 1,
        'goal_not_updated': 2,
        'deadline': 3,
        'feedback_received': 4,
        'goal_approved': 5,
        'achievement': 5,
        'goal_created': 6,
        'update': 7,
        'feedback_reply': 8,
        'goal_edited': 9,
        'goal_deleted': 10
    }

    notifications.sort(key=lambda x: (
        priority_weights.get(x.get('type', ''), 999),
        -(x.get('timestamp', datetime.min.replace(tzinfo=IST)).timestamp())
    ))
    
    # Remove duplicates based on notification_id
    seen = set()
    unique_notifications = []
    for n in notifications:
        nid = n.get('notification_id')
        if nid and nid not in seen:
            seen.add(nid)
            unique_notifications.append(n)
        elif not nid:
            unique_notifications.append(n)

    return unique_notifications[:50]

def check_and_notify_missed_deadlines():
    """Check for goals past their deadline and notify managers"""
    try:
        today = date.today()
        all_users = db.get_all_users()
        
        for user in all_users:
            if user['role'] in ['Employee', 'Manager', 'HR', 'VP']:
                goals = db.get_user_all_goals(user['id'])
                
                for goal in goals:
                    if goal.get('status') == 'Active':
                        end_date_str = goal.get('end_date')
                        
                        if end_date_str:
                            try:
                                end_date = datetime.strptime(str(end_date_str), '%Y-%m-%d').date()
                                
                                # Check if deadline was yesterday (to send notification once)
                                if end_date == today - timedelta(days=1):
                                    # Check if goal was completed
                                    monthly_achievement = goal.get('monthly_achievement', 0)
                                    monthly_target = goal.get('monthly_target', 1)
                                    
                                    if monthly_achievement is None:
                                        monthly_achievement = 0
                                    
                                    progress = (monthly_achievement / monthly_target * 100) if monthly_target > 0 else 0
                                    
                                    # If not completed (less than 100%), notify manager
                                    if progress < 100:
                                        if user.get('manager_id'):
                                            manager = db.get_user_by_id(user['manager_id'])
                                            if manager and manager.get('email'):
                                                send_goal_completion_email(
                                                    manager['email'],
                                                    user['name'],
                                                    goal['goal_title'],
                                                    completed=False
                                                )
                                                
                                                # Also create in-app notification
                                                notification_data = {
                                                    'user_id': manager['id'],
                                                    'action_by': user['id'],
                                                    'action_by_name': user['name'],
                                                    'action_type': 'goal_deadline_missed',
                                                    'details': f"{user['name']}'s goal '{goal['goal_title']}' deadline was missed ({progress:.1f}% completed)",
                                                    'is_read': False,
                                                    'created_at': datetime.now(IST).isoformat()
                                                }
                                                create_notification(notification_data)
                            except Exception as e:
                                print(f"Error checking deadline for goal {goal['goal_id']}: {str(e)}")
                                continue
    except Exception as e:
        print(f"Error in check_and_notify_missed_deadlines: {str(e)}")

# ============================================
# VIEW ALL GOALS PAGE (NEW)
# ============================================
def display_view_all_goals():
    """View all goals of a user with edit/delete options"""
    user = st.session_state.user
    role = user['role']
    
    if not user:
        st.warning("⚠️ Session expired. Please login again.")
        st.rerun()

    # Back button
    col1, col2 = st.columns([1, 5])
    with col1:
        if st.button("← Back to My Goals"):
            st.session_state.page = 'my_goals'
            st.rerun()
    with col2:
        st.title("📋 View All Goals")
    
    # Select user (HR can see all, Manager can see team, Employee sees own)
    if role == 'HR':
        all_users = db.get_all_users()
        selected_user = st.selectbox(
            "Select User",
            [f"{u['name']} ({u['email']})" for u in all_users]
        )
        user_email = selected_user.split('(')[1].strip(')')
        selected_user_obj = next(u for u in all_users if u['email'] == user_email)
        view_user_id = selected_user_obj['id']
    elif role == 'Manager':
        team_members = db.get_team_members(user['id'])
        if team_members:
            selected_user = st.selectbox(
                "Select Team Member",
                [user['name']] + [f"{m['name']} ({m['email']})" for m in team_members]
            )
            if selected_user == user['name']:
                view_user_id = user['id']
            else:
                user_email = selected_user.split('(')[1].strip(')')
                selected_user_obj = next(m for m in team_members if m['email'] == user_email)
                view_user_id = selected_user_obj['id']
        else:
            view_user_id = user['id']
    else:
        view_user_id = user['id']
    
    # Get all goals for selected user
    all_goals = db.get_user_all_goals(view_user_id)
    
    if not all_goals:
        st.info("No goals found for this user")
        return
    
    # Filter options
    col1, col2, col3 = st.columns(3)
    with col1:
        filter_year = st.selectbox("Filter by Year", ["All"] + sorted(list(set([g['year'] for g in all_goals])), reverse=True))
    with col2:
        filter_status = st.selectbox("Filter by Status", ["All", "Active", "Completed", "On Hold", "Cancelled"])
    with col3:
        search_term = st.text_input("🔍 Search Goal Title")
    
    # Apply filters
    filtered_goals = all_goals
    if filter_year != "All":
        filtered_goals = [g for g in filtered_goals if g['year'] == filter_year]
    if filter_status != "All":
        filtered_goals = [g for g in filtered_goals if g.get('status') == filter_status]
    if search_term:
        filtered_goals = [g for g in filtered_goals if search_term.lower() in g['goal_title'].lower()]
    
    st.markdown(f"**Showing {len(filtered_goals)} of {len(all_goals)} goals**")
    
    # Display goals in expandable cards
    for goal in filtered_goals:
        progress = calculate_progress(goal.get('monthly_achievement', 0), goal.get('monthly_target', 1))
        
        # Determine status color
        if progress >= 100:
            status_color = "#10b981"
            status_text = "Completed"
        elif progress >= 60:
            status_color = "#f59e0b"
            status_text = "On Track"
        else:
            status_color = "#ef4444"
            status_text = "At Risk"
        
        with st.expander(f"🎯 {goal['goal_title']} - {goal['year']}/Q{goal.get('quarter', 'N/A')}/M{goal.get('month', 'N/A')}"):
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                st.markdown(f"**Department:** {goal.get('department', 'N/A')}")
                st.markdown(f"**KPI:** {goal.get('kpi', 'N/A')}")
            
            with col2:
                st.markdown(f"**Start:** {goal['start_date']}")
                st.markdown(f"**End:** {goal['end_date']}")
            
            with col3:
                st.markdown(f"**Target:** {goal.get('monthly_target', 0)}")
                st.markdown(f"**Achievement:** {goal.get('monthly_achievement', 0)}")
            
            with col4:
                st.markdown(f"**Status:** <span style='color: {status_color}; font-weight: bold;'>{goal.get('status', 'Active')}</span>", unsafe_allow_html=True)
                st.markdown(f"**Progress:** <span style='color: {status_color}; font-weight: bold;'>{progress:.1f}%</span>", unsafe_allow_html=True)
            
            st.markdown(f"**Description:** {goal.get('goal_description', 'No description')}")
            
            # Edit and Delete buttons
            col_edit, col_delete, col_space = st.columns([1, 1, 3])
            
            with col_edit:
                if st.button("✏️ Edit", key=f"edit_goal_{goal['goal_id']}", use_container_width=True):
                    st.session_state.editing_goal = goal
                    st.rerun()
            
            with col_delete:
                if st.button("🗑️ Delete", key=f"delete_goal_{goal['goal_id']}", use_container_width=True):
                    goal_owner = db.get_user_by_id(goal['user_id'])
                    if db.delete_goal(goal['goal_id']):
                        if goal_owner:
                            notify_goal_deleted(goal, user, goal_owner)
                        st.success("Goal deleted!")
                        st.rerun()
    
    # Edit Goal Modal
    if 'editing_goal' in st.session_state:
        st.markdown("---")
        st.subheader("✏️ Edit Goal")
        
        edit_goal = st.session_state.editing_goal
        
        with st.form("edit_all_goals_form"):
            col1, col2 = st.columns(2)
            
            with col1:
                new_department = st.text_input("Department*", value=edit_goal.get('department', ''))
                new_title = st.text_input("Goal Title*", value=edit_goal['goal_title'])
                new_kpi = st.text_input("KPI*", value=edit_goal.get('kpi', ''))
                new_monthly_target = st.number_input("Monthly Target", min_value=0.0, value=float(edit_goal.get('monthly_target', 0)))
            
            with col2:
                new_status = st.selectbox("Status", ['Active', 'Completed', 'On Hold', 'Cancelled'],
                                         index=['Active', 'Completed', 'On Hold', 'Cancelled'].index(edit_goal.get('status', 'Active')))
                new_monthly_achievement = st.number_input("Monthly Achievement", min_value=0.0, value=float(edit_goal.get('monthly_achievement', 0)))
                new_description = st.text_area("Description", value=edit_goal.get('goal_description', ''))
            
            col_save, col_cancel = st.columns(2)
            
            with col_save:
                if st.form_submit_button("💾 Save Changes", use_container_width=True):
                    updates = {
                        'department': new_department,
                        'goal_title': new_title,
                        'kpi': new_kpi,
                        'monthly_target': new_monthly_target,
                        'monthly_achievement': new_monthly_achievement,
                        'goal_description': new_description,
                        'status': new_status
                    }
                    if db.update_goal(edit_goal['goal_id'], updates):
                        goal_owner = db.get_user_by_id(edit_goal['user_id'])
                        if goal_owner:
                            notify_goal_edited(edit_goal, user, goal_owner)
                        st.success("✅ Goal updated!")
                        del st.session_state.editing_goal
                        st.rerun()
            
            with col_cancel:
                if st.form_submit_button("❌ Cancel", use_container_width=True):
                    del st.session_state.editing_goal
                    st.rerun()



# ============================================
# HR INFO PAGE (FIXED)
# ============================================
def display_hr_info():
    """Display all HR information with delete user option"""
    user = st.session_state.user
    
    if not user:
        st.warning("⚠️ Session expired. Please login again.")
        st.rerun()

    # Allow CMD, VP, and HR to access
    if user['role'] not in ['CMD', 'VP', 'HR']:
        st.warning("⚠️ You don't have permission to access this page")
        return
    
    st.title("🏢 HR Information Dashboard")
    
    all_users = db.get_all_users()
    
    # HR Stats
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        total_users = len(all_users)
        render_metric_card("Total Users", str(total_users), color="#3b82f6")
    with col2:
        total_goals = sum([len(db.get_user_all_goals(u['id'])) for u in all_users])
        render_metric_card("Total Goals", str(total_goals), color="#8b5cf6")
    with col3:
        total_feedback = len(db.get_all_feedback())
        render_metric_card("Total Feedback", str(total_feedback), color="#10b981")
    with col4:
        active_goals = len(db.get_all_active_goals())
        render_metric_card("Active Goals", str(active_goals), color="#f59e0b")
    
    st.markdown("---")
    
    # Tabs for different views
    tab1, tab2, tab3, tab4 = st.tabs(["👥 All Users", "📊 Department Stats", "🎯 Goal Summary", "💬 Feedback Summary"])
    
    with tab1:
        st.subheader("All Users in System")
        
        users_data = []
        for u in all_users:
            stats = db.get_user_goal_stats(u['id'])
            manager_name = "N/A"
            if u.get('manager_id'):
                manager = db.get_user_by_id(u['manager_id'])
                if manager:
                    manager_name = manager['name']
            
            users_data.append({
                'Name': u['name'],
                'Email': u['email'],
                'Role': u['role'],
                'Department': u.get('department', 'N/A'),
                'Manager': manager_name,
                'Joining Date': u.get('joining_date', 'N/A'),
                'End Date': u.get('end_date', 'N/A'),
                'Total Goals': stats.get('total_goals', 0),
                'Completed': stats.get('completed_goals', 0),
                'Progress %': f"{stats.get('avg_progress', 0):.1f}%",
                'User_ID': u['id']  # ✅ Changed from 'User ID' to 'User_ID'
            })
        
        df_users = pd.DataFrame(users_data)
        
        # Display without User_ID column
        display_df = df_users.drop('User_ID', axis=1)  # ✅ Fixed: Use 'User_ID' instead of 'User ID'
        st.dataframe(display_df, use_container_width=True, height=500)
        
        # Delete User option - only show deletable users based on hierarchy
        st.markdown("---")
        st.subheader("⚠️ Delete User")
        
        
        
        # Get deletable users based on current user's role
        deletable_users = [u for u in all_users 
                          if u['id'] != user['id'] 
                          and can_modify_user(user['role'], u['role'])]
        
        if deletable_users:
            col_del1, col_del2 = st.columns([2, 1])
            with col_del1:
                delete_user_select = st.selectbox(
                    "Select User to Delete",
                    [f"{u['name']} ({u['role']}) - {u['email']}" for u in deletable_users]
                )
            
            with col_del2:
                if st.button("🗑️ Delete Selected User", type="primary", use_container_width=True):
                    user_email = delete_user_select.split(' - ')[1]
                    delete_user_obj = next(u for u in deletable_users if u['email'] == user_email)
                    st.session_state['user_to_delete'] = delete_user_obj
            
            # Confirmation prompt
            if 'user_to_delete' in st.session_state:
                delete_user_obj = st.session_state['user_to_delete']
                
                st.warning(f"Are you sure you want to delete **{delete_user_obj['name']}** ({delete_user_obj['role']})?")
                st.error("⚠️ This will also delete all their goals and feedback!")
                
                confirm_col1, confirm_col2 = st.columns(2)
                with confirm_col1:
                    if st.button("Yes, Delete", key="confirm_delete_user"):
                        result = db.delete_user(delete_user_obj['id'])
                        if result:
                            st.success(f"✅ User {delete_user_obj['name']} deleted successfully!")
                            del st.session_state['user_to_delete']
                            st.rerun()
                        else:
                            st.error("❌ Failed to delete user")
                
                with confirm_col2:
                    if st.button("Cancel", key="cancel_delete_user"):
                        del st.session_state['user_to_delete']
                        st.info("❎ Deletion cancelled.")
        else:
            st.info("No users available for deletion based on your permissions")

    
    with tab2:
        st.subheader("Department-wise Statistics")
        
        dept_stats = {}
        for u in all_users:
            dept = u.get('department', 'Unassigned')
            if dept not in dept_stats:
                dept_stats[dept] = {'users': 0, 'goals': 0, 'completed': 0}
            dept_stats[dept]['users'] += 1
            stats = db.get_user_goal_stats(u['id'])
            dept_stats[dept]['goals'] += stats.get('total_goals', 0)
            dept_stats[dept]['completed'] += stats.get('completed_goals', 0)
        
        dept_data = []
        for dept, stats in dept_stats.items():
            dept_data.append({
                'Department': dept,
                'Total Users': stats['users'],
                'Total Goals': stats['goals'],
                'Completed Goals': stats['completed'],
                'Completion Rate': f"{(stats['completed'] / stats['goals'] * 100) if stats['goals'] > 0 else 0:.1f}%"
            })
        
        df_dept = pd.DataFrame(dept_data)
        st.dataframe(df_dept, use_container_width=True)
    
    with tab3:
        st.subheader("Goal Summary")
        
        all_goals = []
        for u in all_users:
            all_goals.extend(db.get_user_all_goals(u['id']))
        
        # Status breakdown
        status_count = {}
        for goal in all_goals:
            status = goal.get('status', 'Active')
            status_count[status] = status_count.get(status, 0) + 1
        
        col_status1, col_status2, col_status3, col_status4 = st.columns(4)
        with col_status1:
            st.metric("Active", status_count.get('Active', 0))
        with col_status2:
            st.metric("Completed", status_count.get('Completed', 0))
        with col_status3:
            st.metric("On Hold", status_count.get('On Hold', 0))
        with col_status4:
            st.metric("Cancelled", status_count.get('Cancelled', 0))
        
        # Year-wise breakdown
        st.markdown("---")
        st.subheader("Year-wise Goal Distribution")
        
        year_stats = {}
        for goal in all_goals:
            year = goal['year']
            if year not in year_stats:
                year_stats[year] = 0
            year_stats[year] += 1
        
        if year_stats:
            # Create dataframe for proper plotting
            df_year = pd.DataFrame({
                'Year': list(year_stats.keys()),
                'Goals': list(year_stats.values())
            })
            
            fig = px.bar(
                df_year,
                x='Year',
                y='Goals',
                title='Goals by Year',
                color='Goals',
                color_continuous_scale='Blues'
            )
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("No goals data available")
    
    with tab4:
        st.subheader("Feedback Summary")
        
        # User filter
        all_users = db.get_all_users()
        user_filter_options = ["All Users"] + [f"{u['name']} ({u['email']})" for u in all_users]
        selected_user_filter = st.selectbox("Filter by User", user_filter_options)
        
        # Get feedback based on filter
        if selected_user_filter == "All Users":
            all_feedback = db.get_all_feedback()
            filter_display = "All Users"
        else:
            user_email = selected_user_filter.split('(')[1].strip(')')
            filtered_user = next(u for u in all_users if u['email'] == user_email)
            all_feedback = db.get_user_all_feedback(filtered_user['id'])
            filter_display = filtered_user['name']
        
        st.markdown(f"**Showing feedback for: {filter_display}**")
        st.markdown("---")
        
        
        
        # Average ratings
        
        st.subheader("Average Ratings")
        
        if all_feedback:
            total_rating = sum([fb.get('rating', 0) for fb in all_feedback])
            avg_rating = total_rating / len(all_feedback) if all_feedback else 0
            
            st.metric(f"Average Rating for {filter_display}", f"{avg_rating:.2f} ⭐")
            
            # Rating distribution
            st.markdown("---")
            st.subheader("Rating Distribution")
            
            rating_count = {}
            for fb in all_feedback:
                rating = fb.get('rating', 0)
                rating_count[rating] = rating_count.get(rating, 0) + 1
            
            if rating_count:
                df_rating = pd.DataFrame({
                    'Rating': list(rating_count.keys()),
                    'Count': list(rating_count.values())
                })
                
                fig2 = px.bar(
                    df_rating,
                    x='Rating',
                    y='Count',
                    title=f'Feedback Rating Distribution - {filter_display}',
                    color='Rating',
                    color_continuous_scale='RdYlGn'
                )
                st.plotly_chart(fig2, use_container_width=True)
            else:
                st.info("No rating data available")
        else:
            st.info(f"No feedback data available for {filter_display}")
# ============================================
# EMPLOYEES PAGE
# ============================================
def display_employees_page():
    """Display employees with hierarchical drill-down"""
    user = st.session_state.user
    role = user['role']
    
    if not user:
        st.warning("⚠️ Session expired. Please login again.")
        st.rerun()

    # Get all users
    all_users = db.get_all_users()
    
    # Determine what to show based on role
    if role == 'CMD':
        # CMD sees VP, HR, and Manager cards at top level
        employees = [u for u in all_users if u['role'] in ['VP', 'HR', 'Manager']]
        view_title = "Organization Overview"
    elif role == 'VP':
        # VP sees HR and Manager cards
        employees = [u for u in all_users if u['role'] in ['HR', 'Manager']]
        view_title = "HR & Managers"
    elif role == 'HR':
        # HR sees only Manager cards
        employees = [u for u in all_users if u['role'] == 'Manager']
        view_title = "Managers"
    elif role == 'Manager':
        # Manager sees their team employees only
        employees = db.get_team_members(user['id'])
        view_title = "My Team"
    else:
        st.warning("⚠️ You don't have permission to view this page")
        return
    
    if not employees:
        st.info("No employees found")
        return
    
    st.title(f"👥 {view_title}")
    
    # Assign Goal Section (for users who can modify)
    if role in ['CMD', 'VP', 'HR', 'Manager']:
        with st.expander("➕ Assign Goal to Employee"):
            display_quick_assign_goal_form(user, employees)
    
    # Search and Filter
    col1, col2 = st.columns([2, 1])
    with col1:
        search = st.text_input("🔍 Search by name or email", "")
    with col2:
        filter_dept = st.selectbox("Filter by Department", ["All"] + list(set([e.get('department', 'N/A') for e in employees])))
    
    # Filter employees
    filtered_employees = employees
    if search:
        filtered_employees = [e for e in filtered_employees if search.lower() in e['name'].lower() or search.lower() in e['email'].lower()]
    if filter_dept != "All":
        filtered_employees = [e for e in filtered_employees if e.get('department') == filter_dept]
    
    # Display employee cards
    st.markdown("---")
    cols = st.columns(3)
    for idx, emp in enumerate(filtered_employees):
        with cols[idx % 3]:
            stats = db.get_user_goal_stats(emp['id'])
            
            # Role-based badge colors
            role_colors = {
                'CMD': '#8B0000',
                'VP': '#FF4500',
                'HR': '#4facfe',
                'Manager': '#f093fb',
                'Employee': '#dbeafe'
            }
            role_color = role_colors.get(emp['role'], '#dbeafe')
            
            st.markdown(f"""
            <div class='hierarchy-card'>
                <div style='text-align: center;'>
                    <div style='width: 60px; height: 60px; background: linear-gradient(135deg, #2DCCFF, #9BBCE0);
                                border-radius: 50%; margin: 0 auto 10px; display: flex; align-items: center; 
                                justify-content: center; color: white; font-size: 24px; font-weight: bold;'>
                        {emp['name'][0].upper()}
                    </div>
                    <h3 style='margin: 5px 0;'>{emp['name']}</h3>
                    <p style='color: #64748b; font-size: 14px;'>{emp.get('designation', 'Employee')}</p>
                    <p style='color: #64748b; font-size: 12px;'>{emp.get('department', 'N/A')}</p>
                    <div style='margin-top: 10px;'>
                        <span style='background: #dbeafe; color: #1e40af; padding: 3px 10px; 
                                     border-radius: 10px; font-size: 11px; font-weight: bold;'>
                            {emp['role']}
                        </span>
                    </div>
                    <div style='margin-top: 15px; font-size: 13px;'>
                        <p>Goals: {stats.get('total_goals', 0)} | Progress: {stats.get('avg_progress', 0):.1f}%</p>
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)
            
            # Determine button actions based on role and employee type
            if emp['role'] == 'VP':
                # VP card - show their goals and team button
                col_goals, col_team = st.columns(2)
                with col_goals:
                    if st.button("📊", key=f"view_vp_goals_{emp['id']}_{idx}", use_container_width=True, help="View Goals"):
                        st.session_state.viewing_employee = emp
                        st.session_state.page = 'employee_goals'
                        st.rerun()
                with col_team:
                    if st.button("👥", key=f"view_vp_team_{emp['id']}_{idx}", use_container_width=True, help="View Team"):
                        st.session_state.viewing_vp_team = emp
                        st.session_state.page = 'vp_team_view'
                        st.rerun()
            
            elif emp['role'] == 'HR':
                # HR card - show their goals and team button
                col_goals, col_team = st.columns(2)
                with col_goals:
                    if st.button("📊", key=f"view_hr_goals_{emp['id']}_{idx}", use_container_width=True, help="View Goals"):
                        st.session_state.viewing_employee = emp
                        st.session_state.page = 'employee_goals'
                        st.rerun()
                with col_team:
                    if st.button("👥", key=f"view_hr_team_{emp['id']}_{idx}", use_container_width=True, help="View Team"):
                        st.session_state.viewing_hr_team = emp
                        st.session_state.page = 'hr_team_view'
                        st.rerun()
            
            elif emp['role'] == 'Manager':
                # Manager card - show their goals and team button
                col_goals, col_team = st.columns(2)
                with col_goals:
                    if st.button("📊", key=f"view_mgr_goals_{emp['id']}_{idx}", use_container_width=True, help="View Goals"):
                        st.session_state.viewing_employee = emp
                        st.session_state.page = 'employee_goals'
                        st.rerun()
                with col_team:
                    if st.button("👥", key=f"view_mgr_team_{emp['id']}_{idx}", use_container_width=True, help="View Team"):
                        st.session_state.viewing_manager_team = emp
                        st.session_state.page = 'manager_team_view'
                        st.rerun()
            
            else:  # Employee
                # Employee card - only goals button
                if st.button("👁️ View Goals", key=f"view_emp_{emp['id']}_{idx}", use_container_width=True):
                    st.session_state.viewing_employee = emp
                    st.session_state.page = 'employee_goals'
                    st.rerun()

    
    # Edit Employee Modal - Show at top with expander
    if 'editing_employee' in st.session_state:
        st.markdown("---")
        
        
        edit_emp = st.session_state.editing_employee
        
        # Prominent header
        st.markdown(f"""
        <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                    padding: 20px; border-radius: 10px; margin-bottom: 20px;">
            <h2 style="color: white; margin: 0;">✏️ Editing: {edit_emp['name']}</h2>
        </div>
        """, unsafe_allow_html=True)
        
        with st.form("edit_employee_form"):
            col1, col2 = st.columns(2)
            
            with col1:
                new_name = st.text_input("Full Name*", value=edit_emp['name'])
                new_email = st.text_input("Email*", value=edit_emp['email'])
                new_designation = st.text_input("Designation", value=edit_emp.get('designation', ''))
            
            with col2:
                new_role = st.selectbox(
                    "Role*", 
                    ["Employee", "Manager", "HR"],
                    index=["Employee", "Manager", "HR"].index(edit_emp['role'])
                )
                new_department = st.text_input("Department", value=edit_emp.get('department', ''))
                
                # Manager assignment
                managers = [u for u in db.get_all_users() if u['role'] == 'Manager' and u['id'] != edit_emp['id']]
                manager_options = ["None"] + [f"{m['name']} ({m['email']})" for m in managers]
                
                current_manager_idx = 0
                if edit_emp.get('manager_id'):
                    current_manager = db.get_user_by_id(edit_emp['manager_id'])
                    if current_manager:
                        current_manager_str = f"{current_manager['name']} ({current_manager['email']})"
                        if current_manager_str in manager_options:
                            current_manager_idx = manager_options.index(current_manager_str)
                
                selected_manager = st.selectbox("Assign to Manager", manager_options, index=current_manager_idx)
            
            col_save, col_cancel = st.columns(2)
            
            with col_save:
                if st.form_submit_button("💾 Save Changes", use_container_width=True):
                    if new_name and new_email and new_role:
                        manager_id = None
                        if selected_manager != "None":
                            manager_email = selected_manager.split('(')[1].strip(')')
                            manager_id = next((m['id'] for m in managers if m['email'] == manager_email), None)
                        
                        updates = {
                            'name': new_name,
                            'email': new_email.lower().strip(),
                            'designation': new_designation,
                            'role': new_role,
                            'department': new_department,
                            'manager_id': manager_id
                        }
                        
                        if db.update_user(edit_emp['id'], updates):
                            st.success(f"✅ Employee {new_name} updated successfully!")
                            # No notification needed for employee creation per new requirements
                            del st.session_state.editing_employee
                            st.rerun()
                        else:
                            st.error("❌ Failed to update employee")
                    else:
                        st.error("❌ Please fill all required fields")
            
            with col_cancel:
                if st.form_submit_button("❌ Cancel", use_container_width=True):
                    del st.session_state.editing_employee
                    st.rerun()
    
    # Delete Employee Modal - Show at top with expander
    if 'deleting_employee' in st.session_state:
        st.markdown("---")
        st.markdown("---")
        
        del_emp = st.session_state.deleting_employee
        
        # Prominent warning header
        st.markdown(f"""
        <div style="background: linear-gradient(135deg, #ef4444 0%, #dc2626 100%); 
                    padding: 20px; border-radius: 10px; margin-bottom: 20px;">
            <h2 style="color: white; margin: 0;">⚠️ Delete Employee: {del_emp['name']}</h2>
        </div>
        """, unsafe_allow_html=True)
        
        st.warning(f"**Are you sure you want to delete {del_emp['name']}?**")
        st.error("⚠️ This will also delete all their goals and feedback! This action cannot be undone.")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.info(f"**Name:** {del_emp['name']}")
        with col2:
            st.info(f"**Email:** {del_emp['email']}")
        with col3:
            st.info(f"**Role:** {del_emp['role']}")
        
        # Get stats
        emp_goals = db.get_user_all_goals(del_emp['id'])
        st.warning(f"📊 This will delete **{len(emp_goals)} goals** associated with this employee")
        
        confirm = st.checkbox("I understand this action cannot be undone", key="confirm_emp_delete")
        
        col_del1, col_del2, col_del3 = st.columns([1, 1, 1])
        
        with col_del2:
            if st.button(
                "🗑️ Delete Employee", 
                disabled=not confirm,
                use_container_width=True,
                type="primary",
                key="execute_emp_delete"
            ):
                if db.delete_user(del_emp['id']):
                    st.success(f"✅ Employee {del_emp['name']} deleted successfully!")
                    # No notification needed for employee creation per new requirements
                    del st.session_state.deleting_employee
                    st.balloons()
                    st.rerun()
                else:
                    st.error("❌ Failed to delete employee")
        
        with col_del3:
            if st.button("❌ Cancel", use_container_width=True, key="cancel_emp_delete"):
                del st.session_state.deleting_employee
                st.rerun()

# ============================================
# VP TEAM VIEW
# ============================================
def display_vp_team_view():
    """Show employees under a VP"""
    if not st.session_state.get('viewing_vp_team'):
        st.warning("⚠️ VP data lost. Returning to employees page...")
        st.session_state.page = 'employees'
        st.rerun()
    
    vp = st.session_state.viewing_vp_team
    user = st.session_state.user
    
    col1, col2 = st.columns([1, 5])
    with col1:
        if st.button("← Back"):
            st.session_state.pop('viewing_vp_team', None)
            st.session_state.page = 'employees'
            st.rerun()
    with col2:
        st.title(f"👥 {vp['name']}'s Team")
    
    # Get all employees under this VP (all employees managed by managers under this VP)
    all_users = db.get_all_users()
    
    # Get all managers under this VP
    vp_managers = [u for u in all_users if u['role'] == 'Manager' and u.get('manager_id') == vp['id']]
    
    # Get all employees under those managers
    vp_employees = []
    for manager in vp_managers:
        team_members = db.get_team_members(manager['id'])
        vp_employees.extend(team_members)
    
    if not vp_employees:
        st.info(f"No employees found under {vp['name']}")
        return
    
    st.markdown(f"**Total Team Members: {len(vp_employees)}**")
    st.markdown("---")
    
    # Display employee cards
    cols = st.columns(3)
    for idx, emp in enumerate(vp_employees):
        with cols[idx % 3]:
            stats = db.get_user_goal_stats(emp['id'])
            
            st.markdown(f"""
            <div class='hierarchy-card'>
                <div style='text-align: center;'>
                    <div style='width: 60px; height: 60px; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                                border-radius: 50%; margin: 0 auto 10px; display: flex; align-items: center; 
                                justify-content: center; color: white; font-size: 24px; font-weight: bold;'>
                        {emp['name'][0].upper()}
                    </div>
                    <h3 style='margin: 5px 0;'>{emp['name']}</h3>
                    <p style='color: #64748b; font-size: 14px;'>{emp.get('designation', 'Employee')}</p>
                    <p style='color: #64748b; font-size: 12px;'>{emp.get('department', 'N/A')}</p>
                    <div style='margin-top: 10px;'>
                        <span style='background: #dbeafe; color: #1e40af; padding: 3px 10px; 
                                     border-radius: 10px; font-size: 11px; font-weight: bold;'>
                            Employee
                        </span>
                    </div>
                    <div style='margin-top: 15px; font-size: 13px;'>
                        <p>Goals: {stats.get('total_goals', 0)} | Progress: {stats.get('avg_progress', 0):.1f}%</p>
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)
            
            if st.button("👁️ View Goals", key=f"view_vp_emp_{emp['id']}_{idx}", use_container_width=True):
                st.session_state.viewing_employee = emp
                st.session_state.page = 'employee_goals'
                st.rerun()


# ============================================
# HR TEAM VIEW
# ============================================
def display_hr_team_view():
    """Show employees under an HR"""
    if not st.session_state.get('viewing_hr_team'):
        st.warning("⚠️ HR data lost. Returning to employees page...")
        st.session_state.page = 'employees'
        st.rerun()
    
    hr = st.session_state.viewing_hr_team
    user = st.session_state.user
    
    col1, col2 = st.columns([1, 5])
    with col1:
        if st.button("← Back"):
            st.session_state.pop('viewing_hr_team', None)
            st.session_state.page = 'employees'
            st.rerun()
    with col2:
        st.title(f"👥 {hr['name']}'s Team")
    
    # Get all employees under this HR (employees managed by this HR directly)
    hr_employees = db.get_team_members(hr['id'])
    
    if not hr_employees:
        st.info(f"No employees found under {hr['name']}")
        return
    
    st.markdown(f"**Total Team Members: {len(hr_employees)}**")
    st.markdown("---")
    
    # Display employee cards
    cols = st.columns(3)
    for idx, emp in enumerate(hr_employees):
        with cols[idx % 3]:
            stats = db.get_user_goal_stats(emp['id'])
            
            st.markdown(f"""
            <div class='hierarchy-card'>
                <div style='text-align: center;'>
                    <div style='width: 60px; height: 60px; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                                border-radius: 50%; margin: 0 auto 10px; display: flex; align-items: center; 
                                justify-content: center; color: white; font-size: 24px; font-weight: bold;'>
                        {emp['name'][0].upper()}
                    </div>
                    <h3 style='margin: 5px 0;'>{emp['name']}</h3>
                    <p style='color: #64748b; font-size: 14px;'>{emp.get('designation', 'Employee')}</p>
                    <p style='color: #64748b; font-size: 12px;'>{emp.get('department', 'N/A')}</p>
                    <div style='margin-top: 10px;'>
                        <span style='background: #dbeafe; color: #1e40af; padding: 3px 10px; 
                                     border-radius: 10px; font-size: 11px; font-weight: bold;'>
                            Employee
                        </span>
                    </div>
                    <div style='margin-top: 15px; font-size: 13px;'>
                        <p>Goals: {stats.get('total_goals', 0)} | Progress: {stats.get('avg_progress', 0):.1f}%</p>
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)
            
            if st.button("👁️ View Goals", key=f"view_hr_emp_{emp['id']}_{idx}", use_container_width=True):
                st.session_state.viewing_employee = emp
                st.session_state.page = 'employee_goals'
                st.rerun()


# ============================================
# MANAGER TEAM VIEW
# ============================================
def display_manager_team_view():
    """Show employees under a Manager"""
    if not st.session_state.get('viewing_manager_team'):
        st.warning("⚠️ Manager data lost. Returning to employees page...")
        st.session_state.page = 'employees'
        st.rerun()
    
    manager = st.session_state.viewing_manager_team
    user = st.session_state.user
    
    col1, col2 = st.columns([1, 5])
    with col1:
        if st.button("← Back"):
            st.session_state.pop('viewing_manager_team', None)
            st.session_state.page = 'employees'
            st.rerun()
    with col2:
        st.title(f"👥 {manager['name']}'s Team")
    
    # Get team members
    team_members = db.get_team_members(manager['id'])
    
    if not team_members:
        st.info(f"No team members found under {manager['name']}")
        return
    
    st.markdown(f"**Total Team Members: {len(team_members)}**")
    st.markdown("---")
    
    # Display employee cards
    cols = st.columns(3)
    for idx, emp in enumerate(team_members):
        with cols[idx % 3]:
            stats = db.get_user_goal_stats(emp['id'])
            
            st.markdown(f"""
            <div class='hierarchy-card'>
                <div style='text-align: center;'>
                    <div style='width: 60px; height: 60px; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                                border-radius: 50%; margin: 0 auto 10px; display: flex; align-items: center; 
                                justify-content: center; color: white; font-size: 24px; font-weight: bold;'>
                        {emp['name'][0].upper()}
                    </div>
                    <h3 style='margin: 5px 0;'>{emp['name']}</h3>
                    <p style='color: #64748b; font-size: 14px;'>{emp.get('designation', 'Employee')}</p>
                    <p style='color: #64748b; font-size: 12px;'>{emp.get('department', 'N/A')}</p>
                    <div style='margin-top: 10px;'>
                        <span style='background: #dbeafe; color: #1e40af; padding: 3px 10px; 
                                     border-radius: 10px; font-size: 11px; font-weight: bold;'>
                            Employee
                        </span>
                    </div>
                    <div style='margin-top: 15px; font-size: 13px;'>
                        <p>Goals: {stats.get('total_goals', 0)} | Progress: {stats.get('avg_progress', 0):.1f}%</p>
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)
            
            if st.button("👁️ View Goals", key=f"view_mgr_emp_{emp['id']}_{idx}", use_container_width=True):
                st.session_state.viewing_employee = emp
                st.session_state.page = 'employee_goals'
                st.rerun()

def display_quick_assign_goal_form(user, employees):
    """Quick assign goal form in employees page"""
    # Filter based on role hierarchy
    if user['role'] == 'HR':
        employees = [e for e in employees if e.get('manager_id') == user['id']]
        if not employees:
            st.info("You can only assign goals to employees in your team.")
            return
    elif user['role'] == 'VP':
        # VP can only assign to HR and Manager
        employees = [e for e in employees if e['role'] in ['HR', 'Manager']]
        if not employees:
            st.info("You can only assign goals to HR and Manager roles.")
            return
    elif user['role'] == 'CMD':
        # CMD can only assign to VP
        employees = [e for e in employees if e['role'] == 'VP']
        if not employees:
            st.info("You can only assign goals to VP roles.")
            return
    
    with st.form("quick_assign_goal"):
        st.subheader("Quick Assign Goal")
        
        col1, col2 = st.columns(2)
        
        with col1:
            selected_emp = st.selectbox(
                "Assign To*",
                [f"{e['name']} ({e['email']})" for e in employees]
            )
            
            emp_email = selected_emp.split('(')[1].strip(')')
            emp_id = next(e['id'] for e in employees if e['email'] == emp_email)
            
            year = st.number_input("Year", min_value=2020, max_value=2100, value=datetime.now().year)
            quarter = st.selectbox("Quarter", [1, 2, 3, 4])
            month = st.selectbox("Month", list(range(1, 13)), index=datetime.now().month - 1)
        
        with col2:
            department = st.text_input("Department*")
            title = st.text_input("Goal Title*")
            kpi = st.text_input("KPI*")
        
        # Auto-fill dates based on month
            month_start = date(year, month, 1)
            month_end = date(year, month, calendar.monthrange(year, month)[1])
            
            col_date1, col_date2 = st.columns(2)
            with col_date1:
                start_date = st.date_input("Start Date", value=month_start)
            with col_date2:
                end_date = st.date_input("End Date", value=month_end)
            
            description = st.text_area("Description")
            monthly_target = st.number_input("Monthly Target*", min_value=0.0)
            
            st.markdown("**Weekly Targets**")
            auto_divide = st.checkbox("Auto-divide monthly target equally into 4 weeks", value=True, key="quick_auto_divide")
            
            if auto_divide and monthly_target > 0:
                week_target = monthly_target / 4
            else:
                week_target = 0.0
            
            col3, col4, col5, col6 = st.columns(4)
            with col3:
                w1_t = st.number_input("Week 1 Target", min_value=0.0, value=week_target, key="quick_w1")
            with col4:
                w2_t = st.number_input("Week 2 Target", min_value=0.0, value=week_target, key="quick_w2")
            with col5:
                w3_t = st.number_input("Week 3 Target", min_value=0.0, value=week_target, key="quick_w3")
            with col6:
                w4_t = st.number_input("Week 4 Target", min_value=0.0, value=week_target, key="quick_w4")
        # ========== ADD THIS ENTIRE SECTION ==========
        st.markdown("**Weekly Ratings (Optional)**")
        col_r1, col_r2, col_r3, col_r4 = st.columns(4)
        with col_r1:
            w1_rating = st.selectbox("Week 1 Rating", [0, 1, 2, 3, 4], key="quick_w1_rating")
        with col_r2:
            w2_rating = st.selectbox("Week 2 Rating", [0, 1, 2, 3, 4], key="quick_w2_rating")
        with col_r3:
            w3_rating = st.selectbox("Week 3 Rating", [0, 1, 2, 3, 4], key="quick_w3_rating")
        with col_r4:
            w4_rating = st.selectbox("Week 4 Rating", [0, 1, 2, 3, 4], key="quick_w4_rating")
        # ========== END OF ADDED SECTION ==========
        if st.form_submit_button("✅ Assign Goal", use_container_width=True):
            if department and title and kpi:
                goal_data = {
                    'user_id': emp_id,
                    'year': year,
                    'quarter': quarter,
                    'month': month,
                    'department': department,
                    'goal_title': title,
                    'goal_description': description,
                    'kpi': kpi,
                    'monthly_target': monthly_target,
                    'week1_target': w1_t,
                    'week2_target': w2_t,
                    'week3_target': w3_t,
                    'week4_target': w4_t,
                    'week1_remarks': w1_rating,  # ✅ ADD
                    'week2_remarks': w2_rating,  # ✅ ADD
                    'week3_remarks': w3_rating,  # ✅ ADD
                    'week4_remarks': w4_rating,
                    'start_date': str(start_date),
                    'end_date': str(end_date),
                    'created_by': user['id']
                }
                
                if db.create_goal(goal_data):
                    # Get assigned employee
                    assigned_emp = db.get_user_by_id(emp_id)
                    if assigned_emp:
                        # Create goal data for notification
                        assigned_goal_data = {
                            'goal_title': title,
                            'user_id': emp_id
                        }
                        notify_goal_created(assigned_goal_data, user)
                    st.rerun()
            else:
                st.error("❌ Please fill all required fields")



# ============================================
# EMPLOYEE GOALS VIEW
# ============================================
def display_employee_goals():
    """Display goals for a specific employee"""
    if not st.session_state.get('viewing_employee'):
        st.warning("⚠️ Employee data lost. Returning to employees page...")
        st.session_state.page = 'employees'
        st.rerun()
    emp = st.session_state.viewing_employee
    user = st.session_state.user
    
    col1, col2 = st.columns([1, 5])
    with col1:
        if st.button("← Back to Employees"):
            st.session_state.page = 'employees'
            st.rerun()
    with col2:
        st.title(f"📊 {emp['name']}'s Goals")
    
    # Show year selection for this employee
    years = db.get_years(emp['id'])
    
    if not years:
        st.info(f"No goals found for {emp['name']}")
        return
    
    st.subheader("📆 Select Year")
    
    # Display years in rows with goal counts
    sorted_years = sorted(years.items(), reverse=True)
    for year, summary in sorted_years:
        # Get goal count for this year
        year_goals = [g for g in db.get_user_all_goals(emp['id']) if g['year'] == year]
        goal_count = len(year_goals)
        
        st.markdown(f"""
        <div class='hierarchy-card' style='cursor: pointer;'>
            <h2 style='margin:0;'>📅 {year} <span style='color: #64748b; font-size: 16px;'>({goal_count} goals)</span></h2>
            <p style='color: #64748b; margin-top: 8px;'>{summary[:80] if summary else 'Click to view quarters'}</p>
        </div>
        """, unsafe_allow_html=True)
        
        if st.button(f"View {year}", key=f"emp_year_{year}", use_container_width=True):
            st.session_state.selected_year = year
            st.session_state.viewing_employee_year = True
            st.session_state.page = 'employee_quarters'
            st.rerun()
        
        st.markdown("<br>", unsafe_allow_html=True)


# ============================================
# MY GOALS PAGE (UPDATED)
# ============================================
def display_my_goals():
    """Display user's own goals with month search"""
    user = st.session_state.user
    if not user:
        st.warning("⚠️ Session expired. Please login again.")
        st.session_state.page = 'login'
        st.rerun()
    role = user['role']
    # Header with three-dot menu
    col_title, col_menu = st.columns([10, 1])
    with col_title:
        st.title(f"📅 My Goals - {user['name']}")
        st.caption(f"{user.get('designation', 'Employee')} • {user['role']}")
    
    with col_menu:
        # Three-dot menu using popover
        with st.popover("⋮", use_container_width=True):
            if st.button("📋 View All Goals", use_container_width=True, key="view_all_goals_menu"):
                st.session_state.page = 'view_all_goals'
                st.rerun()
    
    # Month Quick Search
    
    st.subheader("🔍 Quick Search by Month")
    
    col_search1, col_search2 = st.columns([2, 1])
    with col_search1:
        search_month = st.selectbox(
            "Select Month to View Across All Years",
            ["None"] + [get_month_name(i) for i in range(1, 13)]
        )
    
    if search_month != "None":
        month_num = [get_month_name(i) for i in range(1, 13)].index(search_month) + 1
        
        st.subheader(f"📅 {search_month} Goals Across All Years")
        
        all_goals = db.get_user_all_goals(user['id'])
        month_goals = [g for g in all_goals if g.get('month') == month_num]
        
        if month_goals:
            # Group by year
            year_groups = {}
            for goal in month_goals:
                year = goal['year']
                if year not in year_groups:
                    year_groups[year] = []
                year_groups[year].append(goal)
            
            for year in sorted(year_groups.keys(), reverse=True):
                with st.expander(f"📅 {search_month} {year} ({len(year_groups[year])} goals)"):
                    for goal in year_groups[year]:
                        progress = calculate_progress(
                            goal.get('monthly_achievement', 0),
                            goal.get('monthly_target', 1)
                        )
                        
                        col1, col2, col3 = st.columns([2, 1, 1])
                        with col1:
                            st.markdown(f"**{goal['goal_title']}**")
                            st.caption(f"Vertical: {goal.get('vertical', 'N/A')} | KPI: {goal.get('kpi', 'N/A')}")
                        with col2:
                            st.metric("Target", goal.get('monthly_target', 0))
                        with col3:
                            st.metric("Progress", f"{progress:.1f}%")
                        
                        if st.button(f"View Goal", key=f"view_month_goal_{goal['goal_id']}"):
                            st.session_state.selected_year = year
                            st.session_state.selected_quarter = goal.get('quarter')
                            st.session_state.selected_month = month_num
                            st.session_state.page = 'month_goals'
                            st.rerun()
                        
                        st.markdown("---")
        else:
            st.info(f"No goals found for {search_month}")
        
        st.markdown("---")
    
    # Regular year display
    # Regular year display
    years = db.get_years(user['id'])
    current_year = datetime.now().year
    if current_year not in years:
        years[current_year] = ""
    st.markdown("---")
    st.subheader("📆 Browse by Year")
    
    
    # Add Create New Year button - Only for HR and Manager
    if role in ['HR', 'Manager']:
        col_header1, col_header2 = st.columns([3, 1])
        with col_header2:
            if st.button("➕ Create New Year", use_container_width=True, key="create_new_year_btn"):
                st.session_state.creating_new_year = True
    
    # Show create year form if button clicked
    if st.session_state.get('creating_new_year'):
        with st.expander("➕ Create New Year", expanded=True):
            with st.form("create_new_year_form"):
                new_year = st.number_input("Year*", min_value=2020, max_value=2100, value=datetime.now().year + 1)
                new_year_summary = st.text_area("Year Summary (Optional)", placeholder="Enter goals/plans for this year...")
                
                col_create1, col_create2 = st.columns(2)
                with col_create1:
                    if st.form_submit_button("✅ Create Year", use_container_width=True):
                        if new_year:
                            # Check if year already exists
                            if new_year in years:
                                st.error(f"❌ Year {new_year} already exists!")
                            else:
                                # Create year entry
                                if db.update_year_summary(user['id'], new_year, new_year_summary):
                                    st.success(f"✅ Year {new_year} created successfully!")
                                    st.session_state.creating_new_year = False
                                    st.rerun()
                                else:
                                    st.error("❌ Failed to create year")
                        else:
                            st.error("❌ Please enter a valid year")
                
                with col_create2:
                    if st.form_submit_button("❌ Cancel", use_container_width=True):
                        st.session_state.creating_new_year = False
                        st.rerun()
        st.markdown("---")
    
    # Display years in rows with goal counts
    sorted_years = sorted(years.items(), reverse=True)
    for year, summary in sorted_years:
        # Get goal count for this year
        year_goals = [g for g in db.get_user_all_goals(user['id']) if g['year'] == year]
        goal_count = len(year_goals)
        
        st.markdown(f"""
        <div class='hierarchy-card' style='cursor: pointer;'>
            <h2 style='margin:0;'>📅 {year} <span style='color: #64748b; font-size: 16px;'>({goal_count} goals)</span></h2>
            <p style='color: #64748b; margin-top: 8px;'>{summary[:80] if summary else 'Click to view quarters'}</p>
        </div>
        """, unsafe_allow_html=True)
        
        # Show centered view button for Employee, CMD, VP
        if role in ['Employee', 'CMD', 'VP']:
            if st.button(f"View {year}", key=f"year_view_{year}", use_container_width=True):
                st.session_state.selected_year = year
                st.session_state.page = 'quarters'
                st.rerun()
        
        else:
            # HR and Manager get edit/delete options
            col_btn1, col_btn2, col_btn3 = st.columns([2, 1, 1])
            
            with col_btn1:
                if st.button(f"View {year}", key=f"year_view_{year}", use_container_width=True):
                    st.session_state.selected_year = year
                    st.session_state.page = 'quarters'
                    st.rerun()
            
            with col_btn2:
                if st.button("✏️ Edit", key=f"year_edit_{year}", use_container_width=True):
                    st.session_state.editing_year = year
                    st.session_state.editing_year_summary = summary
            
            with col_btn3:
                if st.button("🗑️ Delete", key=f"year_del_{year}", use_container_width=True):
                    if db.delete_year(user['id'], year):
                        st.success(f"Year {year} deleted!")
                        st.rerun()
        
        st.markdown("<br>", unsafe_allow_html=True)
    
    if user['role'] == 'Employee':
        all_user_goals = db.get_user_all_goals(user['id'])
        rejected_goals = [g for g in all_user_goals if g.get('approval_status') == 'rejected']
        
        if rejected_goals:
            st.markdown("---")
            st.error(f"⚠️ You have {len(rejected_goals)} rejected goal(s)")
            
            with st.expander("View Rejected Goals"):
                for goal in rejected_goals:
                    st.markdown(f"**{goal['goal_title']}**")
                    st.caption(f"Period: {goal['year']}-Q{goal['quarter']}-M{goal['month']}")
                    st.warning(f"**Reason:** {goal.get('rejection_reason', 'No reason provided')}")
                    
                    if st.button(f"Revise & Resubmit", key=f"revise_{goal['goal_id']}"):
                        # Allow editing and resubmit
                        st.session_state.revising_goal = goal
                        st.rerun()
                    
                    st.markdown("---")
    # Rest of the existing code for edit year and add new year
    # ...

# ============================================
# QUARTER SELECTION PAGE (UPDATED WITH GOAL COUNT)
# ============================================
def display_quarter_selection():
    """Display quarter selection page with goal counts"""
    user = st.session_state.user
    year = st.session_state.selected_year
    
    if not year:
        st.warning("⚠️ Navigation state lost. Returning to My Goals...")
        st.session_state.page = 'my_goals'
        st.rerun()
    
    # Check if viewing employee goals
    if st.session_state.get('viewing_employee_year'):
        # ✅ Safety check for employee
        if not st.session_state.get('viewing_employee'):
            st.warning("⚠️ Employee data lost. Returning to employees page...")
            st.session_state.page = 'employees'
            st.rerun()
        emp = st.session_state.viewing_employee
        
        # Check if viewing a Manager's goals
        if emp['role'] == 'Manager':
            # Show header with team dropdown
            col1, col2, col_menu = st.columns([1, 7, 1])
            with col1:
                if st.button("← Back"):
                    st.session_state.viewing_employee_year = False
                    st.session_state.page = 'employee_goals'
                    st.rerun()
            with col2:
                st.title(f"📊 {emp['name']}'s Year {year} - Quarters")
            with col_menu:
                # Three-dot menu for team members
                with st.popover("⋮", use_container_width=True):
                    st.markdown("**View Team Member:**")
                    
                    # Get team members under this manager
                    team_members = db.get_team_members(emp['id'])
                    
                    if team_members:
                        for member in team_members:
                            if st.button(
                                f"👤 {member['name']}", 
                                key=f"switch_quarter_{member['id']}", 
                                use_container_width=True
                            ):
                                # Store current manager for back navigation
                                st.session_state.previous_manager = emp
                                # Switch to viewing this team member
                                st.session_state.viewing_employee = member
                                st.rerun()
                    else:
                        st.caption("No team members")
        else:
            # Regular header
            col1, col2 = st.columns([1, 5])
            with col1:
                # Check if we came from a manager's view
                if st.session_state.get('previous_manager'):
                    if st.button("← Back to Manager"):
                        # Go back to the manager's goal sheet
                        st.session_state.viewing_employee = st.session_state.previous_manager
                        st.session_state.pop('previous_manager', None)
                        st.rerun()
                else:
                    if st.button("← Back"):
                        st.session_state.viewing_employee_year = False
                        st.session_state.page = 'employee_goals'
                        st.rerun()
            with col2:
                st.title(f"📊 {emp['name']}'s Year {year} - Quarters")
        
        user_id = emp['id']
    else:
        col1, col2 = st.columns([1, 5])
        with col1:
            if st.button("← Back to Years"):
                st.session_state.page = 'my_goals'
                st.rerun()
        with col2:
            st.title(f"📊 Year {year} - Quarters")
        user_id = user['id']

    
    quarters = db.get_quarters(user_id, year)
    for q in [1, 2, 3, 4]:
        if q not in quarters:
            quarters[q] = ""
    
    cols = st.columns(2)
    for idx, (quarter, summary) in enumerate(sorted(quarters.items())):
        with cols[idx % 2]:
            # Get goal count for this quarter
            quarter_goals = [g for g in db.get_user_all_goals(user_id) if g['year'] == year and g.get('quarter') == quarter]
            goal_count = len(quarter_goals)
            
            st.markdown(f"""
            <div class='hierarchy-card'>
                <h2>📈 Quarter {quarter} <span style='color: #64748b; font-size: 14px;'>({goal_count} goals)</span></h2>
                <p style='color: #64748b;'>{get_quarter_name(quarter)}</p>
                <p style='margin-top: 8px;'>{summary[:80] if summary else 'Click to view months'}</p>
            </div>
            """, unsafe_allow_html=True)
            
            if st.button(f"Open Q{quarter}", key=f"q_{quarter}", use_container_width=True):
                st.session_state.selected_quarter = quarter
                st.session_state.page = 'employee_months' if st.session_state.get('viewing_employee_year') else 'months'
                st.rerun()
            
            # Only show edit for own goals
            if not st.session_state.get('viewing_employee_year'):
                with st.expander(f"✏️ Edit Q{quarter} Summary"):
                    with st.form(f"edit_q{quarter}"):
                        new_summary = st.text_area("Summary", value=summary, key=f"qsum_{quarter}")
                        if st.form_submit_button("Save"):
                            if db.update_quarter_summary(user_id, year, quarter, new_summary):
                                st.success("✅ Saved!")
                                st.rerun()


# ============================================
# MONTH SELECTION PAGE (UPDATED WITH GOAL COUNT)
# ============================================
def display_month_selection():
    """Display month selection page with goal counts"""
    user = st.session_state.user
    year = st.session_state.selected_year
    quarter = st.session_state.selected_quarter
    
    if not year or not quarter:
        st.warning("⚠️ Navigation state lost. Returning to My Goals...")
        st.session_state.page = 'my_goals'
        st.rerun()
    
    # Check if viewing employee goals
    if st.session_state.get('viewing_employee_year'):
        if not st.session_state.get('viewing_employee'):
            st.warning("⚠️ Employee data lost. Returning to employees page...")
            st.session_state.page = 'employees'
            st.rerun()
        emp = st.session_state.viewing_employee
        
        # Check if viewing a Manager's goals
        if emp['role'] == 'Manager':
            # Show header with team dropdown
            col1, col2, col_menu = st.columns([1, 7, 1])
            with col1:
                if st.button("← Back to Quarters"):
                    st.session_state.page = 'employee_quarters'
                    st.rerun()
            with col2:
                st.title(f"📅 {emp['name']}'s Year {year} - Q{quarter} - Months")
            with col_menu:
                # Three-dot menu for team members
                with st.popover("⋮", use_container_width=True):
                    st.markdown("**View Team Member:**")
                    
                    # Get team members under this manager
                    team_members = db.get_team_members(emp['id'])
                    
                    if team_members:
                        for member in team_members:
                            if st.button(
                                f"👤 {member['name']}", 
                                key=f"switch_month_{member['id']}", 
                                use_container_width=True
                            ):
                                # Store current manager for back navigation
                                st.session_state.previous_manager = emp
                                # Switch to viewing this team member
                                st.session_state.viewing_employee = member
                                st.rerun()
                    else:
                        st.caption("No team members")
        else:
            # Regular header
            col1, col2 = st.columns([1, 5])
            with col1:
                # Check if we came from a manager's view
                if st.session_state.get('previous_manager'):
                    if st.button("← Back to Manager"):
                        # Go back to the manager's goal sheet
                        st.session_state.viewing_employee = st.session_state.previous_manager
                        st.session_state.pop('previous_manager', None)
                        st.rerun()
                else:
                    if st.button("← Back to Quarters"):
                        st.session_state.page = 'employee_quarters'
                        st.rerun()
            with col2:
                st.title(f"📅 {emp['name']}'s Year {year} - Q{quarter} - Months")
        
        user_id = emp['id']
    else:
        col1, col2 = st.columns([1, 5])
        with col1:
            if st.button("← Back to Quarters"):
                st.session_state.page = 'quarters'
                st.rerun()
        with col2:
            st.title(f"📅 Year {year} - Q{quarter} - Months")
        user_id = user['id']
    
    quarter_month_nums = get_quarter_months(quarter)
    months = db.get_months(user_id, year, quarter)
    
    cols = st.columns(3)
    for idx, month_num in enumerate(quarter_month_nums):
        with cols[idx]:
            month_name = get_month_name(month_num)
            summary = months.get(month_num, "")
            month_goals = db.get_month_goals(user_id, year, quarter, month_num)
            goal_count = len(month_goals)

            st.markdown(f"""
            <div class='month-card'>
                <div class='month-card-content'>
                    <h2>📆 {month_name}</h2>
                    <p>{goal_count} goals</p>
                </div>
            </div>
            """, unsafe_allow_html=True)

            if st.button(
                f"Open {month_name}",
                key=f"m_{user_id}_{year}_{quarter}_{month_num}", 
                use_container_width=True):
                st.session_state.selected_month = month_num
                st.session_state.page = 'employee_month_goals' if st.session_state.get('viewing_employee_year') else 'month_goals'
                st.rerun()
            
            # Only show edit for own goals
            if not st.session_state.get('viewing_employee_year'):
                with st.expander(f"✏️ Edit {month_name} Summary"):
                    with st.form(f"edit_m{month_num}"):
                        new_summary = st.text_area("Summary", value=summary, key=f"msum_{month_num}")
                        if st.form_submit_button("Save"):
                            if db.update_month_summary(user_id, year, quarter, month_num, new_summary):
                                st.success("✅ Saved!")
                                st.rerun()


# ============================================
# MONTH GOALS PAGE (WITH ASSIGN IN MONTHLY VIEW)
# ============================================
def display_month_goals():
    """Display month goals with week tabs"""
    user = st.session_state.user
    year = st.session_state.selected_year
    quarter = st.session_state.selected_quarter
    month = st.session_state.selected_month

    if not year or not quarter or not month:
        st.warning("⚠️ Navigation state lost. Returning to dashboard...")
        st.session_state.page = 'dashboard'
        st.rerun()

    month_name = get_month_name(month)
    
    # ✅ CRITICAL: Determine viewing permissions based on role hierarchy
    is_read_only = False
    viewing_employee = st.session_state.get('viewing_employee')
    
    if st.session_state.get('viewing_employee_year') and viewing_employee:
        viewing_employee_role = viewing_employee['role']
        
        # HR can only edit their own team members
        if user['role'] == 'HR':
            if viewing_employee_role == 'CMD':
                is_read_only = True
            else:
                is_read_only = False 
                
        # Manager can only edit their team members (already handled by employees page access)
        elif user['role'] == 'Manager':
            if viewing_employee.get('manager_id') != user['id']:
                is_read_only = True
                
        # VP can edit HR and Manager only
        elif user['role'] == 'VP':
            if viewing_employee_role not in ['HR', 'Manager']:
                is_read_only = True
                
        # CMD can edit VP only
        elif user['role'] == 'CMD':
            if viewing_employee_role != 'VP':
                is_read_only = True
    
    # ===== HEADER WITH TEAM DROPDOWN FOR MANAGER =====
    if st.session_state.get('viewing_employee_year'):
        # Safety check for employee
        if not viewing_employee:
            st.warning("⚠️ Employee data lost. Returning to employees page...")
            st.session_state.page = 'employees'
            st.rerun()
        
        # Check if viewing a Manager's goals
        if viewing_employee['role'] == 'Manager':
            # Show header with team dropdown
            col1, col2, col_menu = st.columns([1, 7, 1])
            with col1:
                if st.button("← Back to Months"):
                    st.session_state.page = 'employee_months'
                    st.rerun()
            with col2:
                st.title(f"📊 {viewing_employee['name']}'s {month_name} {year} Goals")
            with col_menu:
                # Three-dot menu for team members
                with st.popover("⋮", use_container_width=True):
                    st.markdown("**View Team Member:**")
                    
                    # Get team members under this manager
                    team_members = db.get_team_members(viewing_employee['id'])
                    
                    if team_members:
                        for member in team_members:
                            if st.button(
                                f"👤 {member['name']}", 
                                key=f"switch_to_{member['id']}", 
                                use_container_width=True
                            ):
                                # Store current manager for back navigation
                                st.session_state.previous_manager = viewing_employee
                                # Switch to viewing this team member
                                st.session_state.viewing_employee = member
                                st.rerun()
                    else:
                        st.caption("No team members")
        else:
            # Regular header for non-managers
            col1, col2 = st.columns([1, 5])
            with col1:
                # Check if we came from a manager's view
                if st.session_state.get('previous_manager'):
                    if st.button("← Back to Manager"):
                        # Go back to the manager's goal sheet
                        st.session_state.viewing_employee = st.session_state.previous_manager
                        st.session_state.pop('previous_manager', None)
                        st.rerun()
                else:
                    if st.button("← Back to Months"):
                        st.session_state.page = 'employee_months'
                        st.rerun()
            with col2:
                st.title(f"📊 {viewing_employee['name']}'s {month_name} {year} Goals")
        
        display_user = viewing_employee
    else:
        # User viewing their own goals
        col1, col2 = st.columns([1, 5])
        with col1:
            if st.button("← Back to Months"):
                st.session_state.page = 'months'
                st.rerun()
        with col2:
            st.title(f"📊 {month_name} {year} Goal Sheet")
        display_user = user
        is_read_only = False  # Always False for own goals
    
    # Rest of the function remains the same...
    # Create tabs
    tab_all, tab_w1, tab_w2, tab_w3, tab_w4 = st.tabs([
        "📋 Monthly View",
        "📅 Week 1",
        "📅 Week 2",
        "📅 Week 3",
        "📅 Week 4"
    ])

    
    # Monthly view - PASS THE FLAG
    with tab_all:
        display_monthly_view(display_user, year, quarter, month,is_read_only)
    
    # Week views
    for week_num, tab in enumerate([tab_w1, tab_w2, tab_w3, tab_w4], 1):
        with tab:
            display_week_view(display_user, year, quarter, month, week_num)

def export_goals_to_excel(user_id, year, quarter, month):
    """Export goals to Excel with proper formatting including Monthly Achievement"""
    
    goals = db.get_month_goals(user_id, year, quarter, month)
    
    if not goals:
        st.warning("No goals to export")
        return None
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"{get_month_name(month)} {year}"
    
    # Define styles
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    subheader_fill = PatternFill(start_color="93C5FD", end_color="93C5FD", fill_type="solid")
    subheader_font = Font(bold=True, color="FFFFFF", size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "Department", "Goal Title", "KPI", "Monthly Target", "Start Date", "End Date"
    ]
    
    # Weekly headers
    weekly_headers = ["Week 1", "Week 2", "Week 3", "Week 4"]
    
    # Row 1: Main headers
    current_col = 1
    for header in headers:
        cell = ws.cell(row=1, column=current_col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
        current_col += 1
    
    # Add "Weekly Target" merged header
    ws.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col + 3)
    cell = ws.cell(row=1, column=current_col)
    cell.value = "Weekly Target"
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border
    current_col += 4
    
    # Add "Weekly Achievement" merged header
    ws.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col + 3)
    cell = ws.cell(row=1, column=current_col)
    cell.value = "Weekly Achievement"
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border
    current_col += 4
    
    # Add "Monthly Achievement" header (single column)
    cell = ws.cell(row=1, column=current_col, value="Monthly Achievement")
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border
    ws.merge_cells(start_row=1, start_column=current_col, end_row=2, end_column=current_col)
    current_col += 1
    
    # Weekly ratings (instead of remarks)
    for week in range(1, 5):
        rating_value = goal.get(f'week{week}_rating', 0)
        cell = ws.cell(row=row_num, column=col_num, value=rating_value)
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        # Color based on rating
        if rating_value == 1:
            cell.fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")  # Light red
        elif rating_value == 2:
            cell.fill = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")  # Light amber
        elif rating_value == 3:
            cell.fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")  # Light yellow
        elif rating_value == 4:
            cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
        
        col_num += 1
    
    # Row 2: Week subheaders
    current_col = 7  # Start after main headers
    
    # Week numbers under "Weekly Target"
    for week in weekly_headers:
        cell = ws.cell(row=2, column=current_col)
        cell.value = week
        cell.fill = subheader_fill
        cell.font = subheader_font
        cell.alignment = center_align
        cell.border = border
        current_col += 1
    
    # Week numbers under "Weekly Achievement"
    for week in weekly_headers:
        cell = ws.cell(row=2, column=current_col)
        cell.value = week
        cell.fill = subheader_fill
        cell.font = subheader_font
        cell.alignment = center_align
        cell.border = border
        current_col += 1
    
    # Skip Monthly Achievement column (already merged)
    current_col += 1
    
    # Week numbers under "Remarks"
    for week in weekly_headers:
        cell = ws.cell(row=2, column=current_col)
        cell.value = week
        cell.fill = subheader_fill
        cell.font = subheader_font
        cell.alignment = center_align
        cell.border = border
        current_col += 1
    
    # Fill main header cells in row 2
    for col in range(1, 7):
        cell = ws.cell(row=2, column=col)
        cell.fill = header_fill
        cell.border = border
    
    # Data rows
    row_num = 3
    for goal in goals:
        col_num = 1
        
        # Main goal info
        ws.cell(row=row_num, column=col_num, value=goal.get('department', '')).border = border
        col_num += 1
        ws.cell(row=row_num, column=col_num, value=goal['goal_title']).border = border
        col_num += 1
        ws.cell(row=row_num, column=col_num, value=goal.get('kpi', '')).border = border
        col_num += 1
        ws.cell(row=row_num, column=col_num, value=goal.get('monthly_target', 0)).border = border
        col_num += 1
        ws.cell(row=row_num, column=col_num, value=goal.get('start_date', '')).border = border
        col_num += 1
        ws.cell(row=row_num, column=col_num, value=goal.get('end_date', '')).border = border
        col_num += 1
        
        # Weekly targets
        for week in range(1, 5):
            ws.cell(row=row_num, column=col_num, value=goal.get(f'week{week}_target', 0)).border = border
            col_num += 1
        
        # Weekly achievements
        # Weekly achievements
        for week in range(1, 5):
            achievement_val = goal.get(f'week{week}_achievement')
            cell_value = achievement_val if achievement_val is not None else '-'
            ws.cell(row=row_num, column=col_num, value=cell_value).border = border
            col_num += 1

        # Monthly achievement
        monthly_ach = goal.get('monthly_achievement')
        cell = ws.cell(row=row_num, column=col_num, value=monthly_ach if monthly_ach is not None else '-')
        cell.border = border
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        col_num += 1
                
        
        
        # Weekly remarks
        for week in range(1, 5):
            cell = ws.cell(row=row_num, column=col_num, value=goal.get(f'week{week}_remarks', ''))
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            col_num += 1
        
        row_num += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15  # Vertical
    ws.column_dimensions['B'].width = 30  # Goal Title
    ws.column_dimensions['C'].width = 15  # KPI
    ws.column_dimensions['D'].width = 15  # Monthly Target
    ws.column_dimensions['E'].width = 12  # Start Date
    ws.column_dimensions['F'].width = 12  # End Date
    
    # Weekly columns (targets and achievements)
    for col in range(7, 15):  # Weeks 1-4 for targets and achievements
        ws.column_dimensions[chr(64 + col)].width = 12
    
    # Monthly Achievement column
    ws.column_dimensions['O'].width = 15
    
    # Remarks columns (wider)
    for col in range(16, 20):
        ws.column_dimensions[chr(64 + col)].width = 12
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output
# ============================================
# MONTHLY VIEW (WITH ASSIGN GOAL)
# ============================================
def display_monthly_view(user, year, quarter, month, is_read_only=False):
    """Display monthly goals view with Excel-like format and assign goal option"""
    # Don't show any title - just the content
    if st.session_state.get('show_create_goal_form'):
        if not st.session_state.get('viewing_employee_year') or not is_read_only:
            with st.expander("➕ Create New Monthly Goal", expanded=True):
                display_add_goal_form_inline(user, year, quarter, month)
            st.markdown("---")
    
    # Assign Goal Section (for HR and Manager) - NOW IN MONTHLY VIEW
    if user['role'] in ['HR', 'Manager'] and not st.session_state.get('viewing_employee_year'):
        with st.expander("➕ Assign Goal to Employee"):
            display_assign_goal_form_monthly(user, year, quarter, month)
    
    # Get goals - only show approved goals for employees
    all_goals = db.get_month_goals(user['id'], year, quarter, month)

    if user['role'] == 'Employee':
        # Show only approved goals
        goals = [g for g in all_goals if g.get('approval_status') == 'approved']
        
        # Show pending count
        pending_count = len([g for g in all_goals if g.get('approval_status') == 'pending'])
        if pending_count > 0:
            st.info(f"ℹ️ You have {pending_count} goal(s) pending manager approval")
    else:
        # Managers/HR/VP/CMD see all goals
        goals = all_goals
    
    # Create & Export Buttons - Only show for own goals or if not read-only
    if not st.session_state.get('viewing_employee_year') or not is_read_only:
        col1, col2, col3 = st.columns([3, 1, 1])

        with col2:
            if st.button("➕ Create New Goal", key="create_goal_month", use_container_width=True):
                st.session_state.show_create_goal_form = True

        with col3:
            if st.button("📥 Export", key="export_excel_month", use_container_width=True):
                excel_file = export_goals_to_excel(user['id'], year, quarter, month)
                if excel_file:
                    st.download_button(
                        label="📄 Download",
                        data=excel_file,
                        file_name=f"Goals_{get_month_name(month)}_{year}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
        
    else:
        # Read-only view - only show export button
        col_space, col_export = st.columns([5, 1])
        with col_export:
            if st.button("📥 Export", key="export_excel_month_readonly", use_container_width=True):
                excel_file = export_goals_to_excel(user['id'], year, quarter, month)
                if excel_file:
                    st.download_button(
                        label="📄 Download",
                        data=excel_file,
                        file_name=f"Goals_{get_month_name(month)}_{year}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
        
    
    if goals:
        # Display goals in Excel-like format
        st.markdown("### 📋 Monthly Goal Sheet")
        
        # Create the formatted table data
        table_data = []
        for goal in goals:
            # ✅ FIX: Properly handle None values and convert to '-'
            def format_achievement(value):
                if value is None:
                    return '-'
                return value
            
            row = {
                'Department': goal.get('department', ''),
                'Goal Title': goal['goal_title'],
                'KPI': goal.get('kpi', ''),
                'Monthly Target': goal.get('monthly_target', 0),
                'Start Date': goal.get('start_date', ''),
                'End Date': goal.get('end_date', ''),
                # Weekly Targets
                'Week 1 Target': goal.get('week1_target', 0),
                'Week 2 Target': goal.get('week2_target', 0),
                'Week 3 Target': goal.get('week3_target', 0),
                'Week 4 Target': goal.get('week4_target', 0),
                # ✅ Weekly Achievements - properly formatted
                'Week 1 Achievement': format_achievement(goal.get('week1_achievement')),
                'Week 2 Achievement': format_achievement(goal.get('week2_achievement')),
                'Week 3 Achievement': format_achievement(goal.get('week3_achievement')),
                'Week 4 Achievement': format_achievement(goal.get('week4_achievement')),
                # ✅ Monthly Achievement - properly formatted
                'Monthly Achievement': format_achievement(goal.get('monthly_achievement')),
                # Weekly Remarks
                'Week 1 Remarks': goal.get('week1_remarks', ''),
                'Week 2 Remarks': goal.get('week2_remarks', ''),
                'Week 3 Remarks': goal.get('week3_remarks', ''),
                'Week 4 Remarks': goal.get('week4_remarks', '')
            }
            table_data.append(row)

        df = pd.DataFrame(table_data)
        
        # Display with better styling using custom HTML/CSS
        st.markdown("""
        <style>
        .excel-table {
            width: 100%;
            overflow-x: auto;
            overflow-y: auto;
            max-height: 600px;
            display: block;             /* ✅ Required for overflow to work */
        }

        /* Table Base */
        .excel-table table {
            border-collapse: collapse;
            width: 100%;
            min-width: 2000px;  /* Increased to ensure all columns fit */
            table-layout: fixed;          /* ✅ Force horizontal scroll for wide tables */
            font-size: 13px;
            background-color: #ffffff;
            color: #222222;
            border: 1px solid #ddd;
        }

        /* Unified Header Styling */
        .excel-table th,
        .excel-table .section-header {
            background-color: #3B82F6;
            color: #ffffff;
            padding: 10px;
            text-align: center;
            border: 1px solid #cfcfcf;
            font-weight: 600;
            position: sticky;
            top: 0;
            z-index: 10;
        }

        /* Table Cells */
        .excel-table td {
            border: 1px solid #e0e0e0;
            padding: 8px;
            text-align: center;
            background-color: #ffffff;
            vertical-align: middle;
        }

        /* Target columns */
        .excel-table .target-col {
            background-color: #F3F8FF;
        }

        /* Achievement columns */
        .excel-table .achievement-col {
            background-color: #F6FFF8;
        }

        /* Monthly Achievement */
        .excel-table .monthly-achievement-col {
            background-color: #E9FCEB;
            font-weight: bold;
        }

        /* Remarks columns */
        .excel-table .remarks-col {
            background-color: #FFFDF4;
            color: #444;
            word-wrap: break-word;
            max-width: 180px;
        }

        /* ✅ CUSTOM COLUMN WIDTHS - FIXED FOR ALL ROLES */

        /* Department column */
        .excel-table td:nth-child(1),
        .excel-table th:nth-child(1) {
            min-width: 120px;
            max-width: 120px;
            word-wrap: break-word;
            white-space: normal;
            text-align: center;
            line-height: 1.4;
            font-size: 12px;
        }

        /* Goal Title - FIXED WIDTH */
        .excel-table td:nth-child(2),
        .excel-table th:nth-child(2) {
            min-width: 200px;
            max-width: 200px;
            word-wrap: break-word;
            white-space: normal;
            text-align: left;
            line-height: 1.4;
            padding: 8px;
            font-size: 12px;
        }

        /* KPI - FIXED WIDTH */
        .excel-table th:nth-child(3),
        .excel-table td:nth-child(3) {
            min-width: 150px;
            max-width: 150px;
            word-wrap: break-word;
            white-space: normal;
            text-align: center;
            line-height: 1.4;
            font-size: 12px;
        }

        /* Monthly Target column */
        .excel-table td:nth-child(4),
        .excel-table th:nth-child(4) {
            min-width: 100px;
            max-width: 100px;
            text-align: center;
            font-size: 12px;
        }

        /* Start Date - FIXED WIDTH */
        .excel-table td:nth-child(5),
        .excel-table th:nth-child(5) {
            min-width: 100px;
            max-width: 100px;
            text-align: center;
            white-space: nowrap;
            font-size: 11px;
            padding: 4px 6px;
        }

        /* End Date - FIXED WIDTH */
        .excel-table td:nth-child(6),
        .excel-table th:nth-child(6) {
            min-width: 100px;
            max-width: 100px;
            text-align: center;
            white-space: nowrap;
            font-size: 11px;
            padding: 4px 6px;
        }

        /* Weekly Target columns (7-10) */
        .excel-table td:nth-child(7),
        .excel-table td:nth-child(8),
        .excel-table td:nth-child(9),
        .excel-table td:nth-child(10),
        .excel-table th:nth-child(7),
        .excel-table th:nth-child(8),
        .excel-table th:nth-child(9),
        .excel-table th:nth-child(10) {
            min-width: 90px;
            max-width: 90px;
            text-align: center;
            font-size: 12px;
        }

        /* Weekly Achievement columns (11-14) */
        .excel-table td:nth-child(11),
        .excel-table td:nth-child(12),
        .excel-table td:nth-child(13),
        .excel-table td:nth-child(14),
        .excel-table th:nth-child(11),
        .excel-table th:nth-child(12),
        .excel-table th:nth-child(13),
        .excel-table th:nth-child(14) {
            min-width: 90px;
            max-width: 90px;
            text-align: center;
            font-size: 12px;
        }

        /* Monthly Achievement column (15) */
        .excel-table td:nth-child(15),
        .excel-table th:nth-child(15) {
            min-width: 110px;
            max-width: 110px;
            text-align: center;
            font-weight: bold;
            font-size: 12px;
        }

        /* Rating columns (16-19) */
        .excel-table td:nth-child(16),
        .excel-table td:nth-child(17),
        .excel-table td:nth-child(18),
        .excel-table td:nth-child(19),
        .excel-table th:nth-child(16),
        .excel-table th:nth-child(17),
        .excel-table th:nth-child(18),
        .excel-table th:nth-child(19) {
            min-width: 80px;
            max-width: 80px;
            text-align: center;
            font-size: 12px;
        }

        /* Hover effect */
        .excel-table tr:hover td {
            background-color: #F9FAFB;
        }

        /* Bold key columns */
        .excel-table th,
        .excel-table td:first-child,
        .excel-table td:nth-child(2) {
            font-weight: 500;
        }

        /* Clean borders */
        .excel-table table,
        .excel-table th,
        .excel-table td {
            border: 1px solid #e6e6e6;
        }

        /* Goal title button styling */
        .goal-title-btn {
            background: none;
            border: none;
            color: #3B82F6;
            text-decoration: underline;
            font-weight: bold;
            cursor: pointer;
            padding: 0;
            font-size: inherit;
        }

        .goal-title-btn:hover {
            color: #1E40AF;
        }

        /* Modal styling */
        .goal-modal {
            display: none;
            position: fixed;
            z-index: 10000;
            left: 0;
            top: 0;
            width: 100%;
            height: 100%;
            background-color: rgba(0,0,0,0.5);
            overflow: auto;
        }

        .modal-content {
            background-color: #fefefe;
            margin: 5% auto;
            padding: 20px;
            border-radius: 10px;
            width: 80%;
            max-width: 600px;
            box-shadow: 0 4px 20px rgba(0,0,0,0.3);
            position: relative;
            animation: slideIn 0.3s ease-out;
        }

        @keyframes slideIn {
            from { transform: translateY(-50px); opacity: 0; }
            to { transform: translateY(0); opacity: 1; }
        }

        .close-btn {
            position: absolute;
            right: 15px;
            top: 15px;
            background: #ef4444;
            color: white;
            border: none;
            border-radius: 5px;
            padding: 8px 15px;
            cursor: pointer;
            font-weight: bold;
            font-size: 18px;
        }

        .close-btn:hover {
            background: #dc2626;
        }

        /* Responsive styling */
        @media (max-width: 768px) {
            .excel-table table {
                font-size: 11px;
            }
            .excel-table th,
            .excel-table td {
                padding: 6px;
            }
        }
        </style>
        """, unsafe_allow_html=True)

        # Store the complete HTML in a variable for components
        complete_html = f'''
        <!DOCTYPE html>
        <html>
        <head>
        <style>
        .excel-table {{
            width: 100%;
            overflow-x: auto;
        }}

        .excel-table table {{
            border-collapse: collapse;
            width: 100%;
            min-width: 1800px;
            font-size: 13px;
            background-color: #ffffff;
            color: #222222;
            border: 1px solid #ddd;
        }}

        .excel-table th {{
            background-color: #3B82F6;
            color: #ffffff;
            padding: 10px;
            text-align: center;
            border: 1px solid #cfcfcf;
            font-weight: 600;
        }}

        .excel-table td {{
            border: 1px solid #e0e0e0;
            padding: 8px;
            text-align: center;
            background-color: #ffffff;
            vertical-align: middle;
        }}

        .target-col {{ background-color: #F3F8FF; }}
        .achievement-col {{ background-color: #F6FFF8; }}
        .monthly-achievement-col {{ background-color: #E9FCEB; font-weight: bold; }}
        .remarks-col {{ background-color: #FFFDF4; max-width: 180px; word-wrap: break-word; }}

        .goal-title-btn {{
            background: none;
            border: none;
            color: #3B82F6;
            text-decoration: underline;
            font-weight: bold;
            cursor: pointer;
            padding: 0;
            font-size: inherit;
        }}

        .goal-title-btn:hover {{
            color: #1E40AF;
        }}

        .modal {{
            display: none;
            position: fixed;
            z-index: 10000;
            left: 0;
            top: 0;
            width: 100%;
            height: 100%;
            background-color: rgba(0,0,0,0.5);
            overflow: auto;
        }}

        .modal-content {{
            background-color: #fefefe;
            margin: 5% auto;
            padding: 20px;
            border-radius: 10px;
            width: 80%;
            max-width: 600px;
            box-shadow: 0 4px 20px rgba(0,0,0,0.3);
            position: relative;
            animation: slideIn 0.3s ease-out;
        }}

        @keyframes slideIn {{
            from {{ transform: translateY(-50px); opacity: 0; }}
            to {{ transform: translateY(0); opacity: 1; }}
        }}

        .close-btn {{
            position: absolute;
            right: 15px;
            top: 15px;
            background: #ef4444;
            color: white;
            border: none;
            border-radius: 5px;
            padding: 8px 15px;
            cursor: pointer;
            font-weight: bold;
            font-size: 18px;
        }}

        .close-btn:hover {{
            background: #dc2626;
        }}
        </style>
        </head>
        <body>

        <div class="excel-table">
        <table>
        <thead>
        <tr>
            <th rowspan="2">Department</th>
            <th rowspan="2">Goal Title</th>
            <th rowspan="2">KPI</th>
            <th rowspan="2">Monthly Target</th>
            <th rowspan="2">Start Date</th>
            <th rowspan="2">End Date</th>
            <th colspan="4">Weekly Target</th>
            <th colspan="4">Weekly Achievement</th>
            <th rowspan="2">Monthly Achievement</th>
            <th colspan="4">Rating</th>
        </tr>
        <tr>
            <th>Week 1</th><th>Week 2</th><th>Week 3</th><th>Week 4</th>
            <th>Week 1</th><th>Week 2</th><th>Week 3</th><th>Week 4</th>
            <th>Week 1</th><th>Week 2</th><th>Week 3</th><th>Week 4</th>
        </tr>
        </thead>
        <tbody>
        '''

        # Add data rows
        for goal_idx, (goal, (_, row)) in enumerate(zip(goals, df.iterrows())):
            goal_desc = goal.get('goal_description', 'No description available').replace("'", "\\'").replace('"', '\\"').replace('\n', '<br>')
            
            complete_html += f'''
        <tr>
            <td>{row["Department"]}</td>
            <td>
                <button class="goal-title-btn" onclick="openModal({goal_idx})">
                    {row["Goal Title"]}
                </button>
            </td>
            <td>{row["KPI"]}</td>
            <td>{row["Monthly Target"]}</td>
            <td>{row["Start Date"]}</td>
            <td>{row["End Date"]}</td>
        '''
            
            # Weekly targets
            for week in range(1, 5):
                complete_html += f'<td class="target-col">{row[f"Week {week} Target"]}</td>'
        
            # ✅ Weekly achievements - check for '-' string
            for week in range(1, 5):
                achievement = row[f"Week {week} Achievement"]
                # Check if it's the '-' string
                if str(achievement) == '-':
                    complete_html += f'<td class="achievement-col" style="color: #94a3b8; font-style: italic;">-</td>'
                else:
                    target = row[f"Week {week} Target"]
                    try:
                        achievement_num = float(achievement)
                        progress = (achievement_num / target * 100) if target > 0 else 0
                        color = '#4CAF50' if progress >= 100 else '#FFC107' if progress >= 60 else '#F44336'
                        complete_html += f'<td class="achievement-col" style="color: {color}; font-weight: bold;">{achievement}</td>'
                    except (ValueError, TypeError):
                        # If conversion fails, show as '-'
                        complete_html += f'<td class="achievement-col" style="color: #94a3b8; font-style: italic;">-</td>'

            # ✅ Monthly achievement - check for '-' string
            monthly_achievement = row["Monthly Achievement"]
            if str(monthly_achievement) == '-':
                complete_html += f'<td class="monthly-achievement-col" style="color: #94a3b8; font-weight: bold; font-style: italic;">-</td>'
            else:
                monthly_target = row["Monthly Target"]
                try:
                    monthly_ach_num = float(monthly_achievement)
                    monthly_progress = (monthly_ach_num / monthly_target * 100) if monthly_target > 0 else 0
                    monthly_color = '#4CAF50' if monthly_progress >= 100 else '#FFC107' if monthly_progress >= 60 else '#F44336'
                    complete_html += f'<td class="monthly-achievement-col" style="color: {monthly_color}; font-weight: bold;">{monthly_achievement}</td>'
                except (ValueError, TypeError):
                    complete_html += f'<td class="monthly-achievement-col" style="color: #94a3b8; font-weight: bold; font-style: italic;">-</td>'
            
            # Weekly ratings (unchanged)
            for week in range(1, 5):
                rating_value = goal.get(f'week{week}_rating', 0)
                
                if rating_value == 1:
                    bg_color = '#FFCCCB'
                elif rating_value == 2:
                    bg_color = '#FFD580'
                elif rating_value == 3:
                    bg_color = '#FFFFE0'
                elif rating_value == 4:
                    bg_color = '#90EE90'
                else:
                    bg_color = '#ffffff'
                
                complete_html += f'<td class="remarks-col" style="background-color: {bg_color}; font-weight: bold;">{rating_value if rating_value else "-"}</td>'
            
            complete_html += '</tr>'
            
            # Add modal
            complete_html += f'''
        <div id="modal{goal_idx}" class="modal">
            <div class="modal-content">
                <button class="close-btn" onclick="closeModal({goal_idx})">✕</button>
                <h3 style="margin: 0 0 15px 0; color: #1E3A8A; padding-right: 40px;">Goal Description</h3>
                <div style="color: #333; line-height: 1.6; max-height: 400px; overflow-y: auto; 
                    padding: 15px; background: #f9fafb; border-radius: 5px;">
                    {goal_desc}
                </div>
            </div>
        </div>
        '''

        complete_html += '''
        </tbody>
        </table>
        </div>

        <script>
        function openModal(index) {
            document.getElementById('modal' + index).style.display = 'block';
        }

        function closeModal(index) {
            document.getElementById('modal' + index).style.display = 'none';
        }

        window.onclick = function(event) {
            const modals = document.querySelectorAll('.modal');
            modals.forEach(modal => {
                if (event.target == modal) {
                    modal.style.display = 'none';
                }
            });
        }

        document.addEventListener('keydown', function(event) {
            if (event.key === 'Escape') {
                const modals = document.querySelectorAll('.modal');
                modals.forEach(modal => {
                    modal.style.display = 'none';
                });
            }
        });
        </script>

        </body>
        </html>
        '''

        # Render using components
        import streamlit.components.v1 as components
        # Calculate dynamic height based on number of goals (tight fit)
        dynamic_height = 150 + (len(goals) * 50)  # Tight calculation: header + rows
        components.html(complete_html, height=dynamic_height, scrolling=False)
        
       # Calculate summary metrics
        # ✅ FIX: Handle None values properly using safe_float helper
        # Calculate summary metrics (find this section around line 3700)
        # ✅ FIX: Handle '-' string in calculations
        total_target = sum(safe_float(g.get('monthly_target'), 0) for g in goals)

        # Calculate achievement - skip '-' values
        total_achievement = 0
        for g in goals:
            ach = g.get('monthly_achievement')
            if ach is not None and str(ach) != '-':
                total_achievement += safe_float(ach, 0)

        avg_progress = calculate_progress(total_achievement, total_target)
        # Display metrics
        st.markdown("---")
        col1, col2, col3 = st.columns(3)
        with col1:
            render_metric_card("Total Target", f"{total_target:.2f}")
        with col2:
            render_metric_card("Total Achievement", f"{total_achievement:.2f}")
        with col3:
            render_metric_card("Progress", f"{avg_progress:.1f}%")
        
        st.markdown("---")
        
        
       # Action tabs: Update, Edit, Delete
        # ✅ NEW: Everyone can update achievements for their own goals
        # Only viewing_employee_year with read_only restrictions apply to edit/delete
        if st.session_state.get('viewing_employee_year') and is_read_only:
            # Read-only view - no tabs
            action_tab1, action_tab2, action_tab3 = None, None, None
        elif st.session_state.get('viewing_employee_year'):
            # Viewing employee (with permissions) - all tabs
            action_tab1, action_tab2, action_tab3 = st.tabs([
                "📝 Update Achievements",
                "✏️ Edit Goal",
                "🗑️ Delete Goal"
            ])
        elif user['role'] in ['HR', 'Manager']:
            # HR/Manager viewing own goals - all tabs
            action_tab1, action_tab2, action_tab3 = st.tabs([
                "📝 Update Achievements",
                "✏️ Edit Goal",
                "🗑️ Delete Goal"
            ])
        else:
            # ✅ CHANGE: Regular employees can now update achievements, but not edit/delete
            action_tab1 = st.tabs(["📝 Update Achievements"])[0]
            action_tab2, action_tab3 = None, None

        # ===== UPDATE ACHIEVEMENTS TAB =====
        
        if action_tab1:
                st.subheader("Update Weekly Achievements")
                
                selected_goal_title = st.selectbox(
                    "Select Goal to Update", 
                    [g['goal_title'] for g in goals],
                    key="update_goal_select"
                )
                selected_goal = next(g for g in goals if g['goal_title'] == selected_goal_title)
                
                # Show current achievement status
                current_monthly = selected_goal.get('monthly_achievement')
                if current_monthly is None:
                    st.info("💡 This goal has no achievements recorded yet. Enter values below to start tracking.")
                
                # ===== SIMPLE VERSION =====
                col1, col2 = st.columns(2)
                
                with col1:
                    current_w1 = selected_goal.get('week1_achievement')
                    w1_input = st.text_input(
                        "Week 1 Achievement (type number or 'none')", 
                        value=str(current_w1) if current_w1 is not None else "none",
                        key="w1_update",
                        help="Enter a number or type 'none' to clear"
                    )
                    
                    current_w2 = selected_goal.get('week2_achievement')
                    w2_input = st.text_input(
                        "Week 2 Achievement (type number or 'none')", 
                        value=str(current_w2) if current_w2 is not None else "none",
                        key="w2_update",
                        help="Enter a number or type 'none' to clear"
                    )
                
                with col2:
                    current_w3 = selected_goal.get('week3_achievement')
                    w3_input = st.text_input(
                        "Week 3 Achievement (type number or 'none')", 
                        value=str(current_w3) if current_w3 is not None else "none",
                        key="w3_update",
                        help="Enter a number or type 'none' to clear"
                    )
                    
                    current_w4 = selected_goal.get('week4_achievement')
                    w4_input = st.text_input(
                        "Week 4 Achievement (type number or 'none')", 
                        value=str(current_w4) if current_w4 is not None else "none",
                        key="w4_update",
                        help="Enter a number or type 'none' to clear"
                    )
                
                # Parse inputs
                def parse_achievement(value_str):
                    """Parse achievement input - returns None or float"""
                    value_str = value_str.strip().lower()
                    if value_str in ['none', 'null', '', '-']:
                        return None
                    try:
                        return float(value_str)
                    except ValueError:
                        return None
                
                w1 = parse_achievement(w1_input)
                w2 = parse_achievement(w2_input)
                w3 = parse_achievement(w3_input)
                w4 = parse_achievement(w4_input)
                
                # Calculate total
                w1_calc = 0 if w1 is None else w1
                w2_calc = 0 if w2 is None else w2
                w3_calc = 0 if w3 is None else w3
                w4_calc = 0 if w4 is None else w4
                total_monthly = w1_calc + w2_calc + w3_calc + w4_calc
                
                st.markdown("---")
                st.markdown(f"**📊 Monthly Achievement (Auto-calculated):** `{total_monthly:.2f}`")
                st.progress(min(total_monthly / selected_goal.get('monthly_target', 1), 1.0))
                
                st.markdown("---")
                st.markdown("**Weekly Ratings (1-4)**")
                col_r1, col_r2 = st.columns(2)
                with col_r1:
                    w1_rating_value = selected_goal.get('week1_rating') or 0
                    w1_rating = st.selectbox("Week 1 Rating", [0, 1, 2, 3, 4], 
                                            index=int(w1_rating_value) if w1_rating_value in [0, 1, 2, 3, 4] else 0, 
                                            key="w1_rating_update",
                                            help="1=Poor (Red), 2=Fair (Amber), 3=Good (Yellow), 4=Excellent (Green)")
                    w2_rating_value = selected_goal.get('week2_rating') or 0
                    w2_rating = st.selectbox("Week 2 Rating", [0, 1, 2, 3, 4], 
                                            index=int(w2_rating_value) if w2_rating_value in [0, 1, 2, 3, 4] else 0, 
                                            key="w2_rating_update",
                                            help="1=Poor (Red), 2=Fair (Amber), 3=Good (Yellow), 4=Excellent (Green)")
                with col_r2:
                    w3_rating_value = selected_goal.get('week3_rating') or 0
                    w3_rating = st.selectbox("Week 3 Rating", [0, 1, 2, 3, 4], 
                                            index=int(w3_rating_value) if w3_rating_value in [0, 1, 2, 3, 4] else 0, 
                                            key="w3_rating_update",
                                            help="1=Poor (Red), 2=Fair (Amber), 3=Good (Yellow), 4=Excellent (Green)")
                    w4_rating_value = selected_goal.get('week4_rating') or 0
                    w4_rating = st.selectbox("Week 4 Rating", [0, 1, 2, 3, 4], 
                                            index=int(w4_rating_value) if w4_rating_value in [0, 1, 2, 3, 4] else 0, 
                                            key="w4_rating_update",
                                            help="1=Poor (Red), 2=Fair (Amber), 3=Good (Yellow), 4=Excellent (Green)")
                
                if st.button("💾 Save Achievements", use_container_width=True, key="save_achievements"):
                    if user['role'] == 'Employee':
                        updates = {
                            'week1_achievement_pending': w1,
                            'week2_achievement_pending': w2,
                            'week3_achievement_pending': w3,
                            'week4_achievement_pending': w4,
                            'monthly_achievement_pending': total_monthly if total_monthly > 0 else None,
                            'week1_rating_pending': w1_rating,
                            'week2_rating_pending': w2_rating,
                            'week3_rating_pending': w3_rating,
                            'week4_rating_pending': w4_rating,
                            'achievement_approval_status': 'pending',
                            'achievement_updated_at': datetime.now(IST).isoformat()
                        }
                        
                        if db.update_goal(selected_goal['goal_id'], updates):
                            st.success("✅ Achievements submitted for manager approval!")
                            st.info("💡 Your manager will review these achievements before they appear in the goal sheet")
                            
                            # Send email to manager
                            if user.get('manager_id'):
                                manager = db.get_user_by_id(user['manager_id'])
                                if manager and manager.get('email'):
                                    send_achievement_approval_email(
                                        manager['email'],
                                        user['name'],
                                        selected_goal,
                                        {
                                            'week1': w1,
                                            'week2': w2,
                                            'week3': w3,
                                            'week4': w4,
                                            'monthly': total_monthly
                                        }
                                    )
                                    st.success("📧 Approval request sent to your manager")
                    else:
                        updates = {
                            'week1_achievement': w1,
                            'week2_achievement': w2,
                            'week3_achievement': w3,
                            'week4_achievement': w4,
                            'monthly_achievement': total_monthly if total_monthly > 0 else None,
                            'week1_rating': w1_rating,
                            'week2_rating': w2_rating,
                            'week3_rating': w3_rating,
                            'week4_rating': w4_rating
                        }
                        
                        if db.update_goal(selected_goal['goal_id'], updates):
                            st.success("✅ Achievements and ratings saved to monthly goal sheet!")
                            st.info("💡 Ratings are now available in the respective week views")
                            
                            # Determine which week was updated (find the last non-None week)
                            updated_week = 0
                            if w4 is not None:
                                updated_week = 4
                            elif w3 is not None:
                                updated_week = 3
                            elif w2 is not None:
                                updated_week = 2
                            elif w1 is not None:
                                updated_week = 1

                            if updated_week > 0:
                                notify_weekly_achievement_updated(selected_goal, user, updated_week)

                            
                            
                            st.rerun()
                        else:
                            st.error("❌ Failed to save achievements")
                        
        # ===== EDIT GOAL TAB =====
        if not is_read_only and action_tab2:
            if action_tab2:
                with action_tab2:
                    st.subheader("Edit Goal Details")
                    
                    edit_goal_title = st.selectbox(
                        "Select Goal to Edit", 
                        [g['goal_title'] for g in goals],
                        key="edit_goal_select"
                    )
                    edit_goal = next(g for g in goals if g['goal_title'] == edit_goal_title)
                    
                    with st.form("edit_goal_form"):
                        st.markdown("**Basic Information**")
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            new_department = st.text_input("Department*", value=edit_goal.get('department', ''))
                            new_title = st.text_input("Goal Title*", value=edit_goal['goal_title'])
                            new_kpi = st.text_input("KPI*", value=edit_goal.get('kpi', ''))
                        
                        with col2:
                            from datetime import datetime as dt
                            start_date_str = edit_goal.get('start_date', str(date.today()))
                            end_date_str = edit_goal.get('end_date', str(date.today()))
                            
                            if isinstance(start_date_str, str):
                                start_date_val = dt.strptime(start_date_str, '%Y-%m-%d').date()
                            else:
                                start_date_val = start_date_str
                            
                            if isinstance(end_date_str, str):
                                end_date_val = dt.strptime(end_date_str, '%Y-%m-%d').date()
                            else:
                                end_date_val = end_date_str
                            
                            new_start_date = st.date_input("Start Date", value=start_date_val)
                            new_end_date = st.date_input("End Date", value=end_date_val)
                            new_status = st.selectbox(
                                "Status", 
                                ['Active', 'Completed', 'On Hold', 'Cancelled'],
                                index=['Active', 'Completed', 'On Hold', 'Cancelled'].index(edit_goal.get('status', 'Active'))
                            )
                        
                        new_description = st.text_area("Description", value=edit_goal.get('goal_description', ''))
                        
                        st.markdown("**Targets**")
                        col3, col4 = st.columns(2)
                        
                        with col3:
                            new_monthly_target = st.number_input(
                                "Monthly Target*", 
                                min_value=0.0, 
                                value=float(edit_goal.get('monthly_target', 0))
                            )
                        
                        st.markdown("**Weekly Targets**")
                        col5, col6, col7, col8 = st.columns(4)
                        
                        with col5:
                            new_w1_target = st.number_input(
                                "Week 1 Target", 
                                min_value=0.0, 
                                value=float(edit_goal.get('week1_target', 0)),
                                key="edit_w1_target"
                            )
                        with col6:
                            new_w2_target = st.number_input(
                                "Week 2 Target", 
                                min_value=0.0, 
                                value=float(edit_goal.get('week2_target', 0)),
                                key="edit_w2_target"
                            )
                        with col7:
                            new_w3_target = st.number_input(
                                "Week 3 Target", 
                                min_value=0.0, 
                                value=float(edit_goal.get('week3_target', 0)),
                                key="edit_w3_target"
                            )
                        with col8:
                            new_w4_target = st.number_input(
                                "Week 4 Target", 
                                min_value=0.0, 
                                value=float(edit_goal.get('week4_target', 0)),
                                key="edit_w4_target"
                            )
                        
                        st.markdown("**Weekly Remarks**")
                        col_r5, col_r6, col_r7, col_r8 = st.columns(4)

                        with col_r5:
                            new_w1_remarks = st.text_area(
                                "Week 1 Remarks",
                                value=edit_goal.get('week1_remarks', ''),
                                key="edit_w1_remarks",
                                height=80
                            )
                        with col_r6:
                            new_w2_remarks = st.text_area(
                                "Week 2 Remarks",
                                value=edit_goal.get('week2_remarks', ''),
                                key="edit_w2_remarks",
                                height=80
                            )
                        with col_r7:
                            new_w3_remarks = st.text_area(
                                "Week 3 Remarks",
                                value=edit_goal.get('week3_remarks', ''),
                                key="edit_w3_remarks",
                                height=80
                            )
                        with col_r8:
                            new_w4_remarks = st.text_area(
                                "Week 4 Remarks",
                                value=edit_goal.get('week4_remarks', ''),
                                key="edit_w4_remarks",
                                height=80
                            )
                        
                        submitted = st.form_submit_button("💾 Save Changes", use_container_width=True)
                        
                        if submitted:
                            if new_department and new_title and new_kpi:
                                updates = {
                                    'department': new_department,
                                    'goal_title': new_title,
                                    'goal_description': new_description,
                                    'kpi': new_kpi,
                                    'monthly_target': new_monthly_target,
                                    'week1_target': new_w1_target,
                                    'week2_target': new_w2_target,
                                    'week3_target': new_w3_target,
                                    'week4_target': new_w4_target,
                                    'week1_remarks': new_w1_remarks,
                                    'week2_remarks': new_w2_remarks,
                                    'week3_remarks': new_w3_remarks,
                                    'week4_remarks': new_w4_remarks,
                                    'start_date': str(new_start_date),
                                    'end_date': str(new_end_date),
                                    'status': new_status
                                }
                                
                                if db.update_goal(edit_goal['goal_id'], updates):
                                    st.success("✅ Goal updated successfully!")
                                    goal_owner = db.get_user_by_id(edit_goal['user_id'])
                                    if goal_owner:
                                        notify_goal_edited(edit_goal, user, goal_owner)
                                    st.rerun()
                                else:
                                    st.error("❌ Failed to update goal")
                            else:
                                st.error("❌ Please fill all required fields (Department, Title, KPI)")
        
        # ===== DELETE GOAL TAB =====
        if not is_read_only and action_tab3:
            if action_tab3:
                with action_tab3:
                    st.subheader("⚠️ Delete Goal")
                    st.warning("**Warning:** Deleting a goal will also delete all associated feedback. This action cannot be undone!")
                    
                    delete_goal_title = st.selectbox(
                        "Select Goal to Delete", 
                        [g['goal_title'] for g in goals],
                        key="delete_goal_select"
                    )
                    delete_goal = next(g for g in goals if g['goal_title'] == delete_goal_title)
                    
                    st.markdown("**Goal Details:**")
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.info(f"**Vertical:** {delete_goal.get('vertical', 'N/A')}")
                    with col2:
                        st.info(f"**KPI:** {delete_goal.get('kpi', 'N/A')}")
                    with col3:
                        st.info(f"**Target:** {delete_goal.get('monthly_target', 0)}")
                    
                    st.markdown(f"**Description:** {delete_goal.get('goal_description', 'No description')}")
                    
                    st.markdown("---")
                    confirm_delete = st.checkbox("I understand this action cannot be undone", key="confirm_delete")
                    
                    col_del1, col_del2, col_del3 = st.columns([1, 1, 1])
                    
                    with col_del2:
                        if st.button(
                            "🗑️ Delete Goal", 
                            disabled=not confirm_delete,
                            use_container_width=True,
                            type="primary"
                        ):
                            if db.delete_goal(delete_goal['goal_id']):
                                st.success("✅ Goal deleted successfully!")
                                goal_owner = db.get_user_by_id(delete_goal['user_id'])
                                if goal_owner:
                                    notify_goal_deleted(delete_goal, user, goal_owner)
                                st.rerun()
                            else:
                                st.error("❌ Failed to delete goal")
        
            
                        # Add new goal (only for own goals)
                    if not st.session_state.get('viewing_employee_year')and not goals:
                        st.markdown("---")
                        display_add_goal_form(user, year, quarter, month)
        
        # Feedback section
    if goals:
        st.markdown("---")
        display_feedback_section(goals, 'month')

    
def display_assign_goal_form_monthly(user, year, quarter, month):
    """Form to assign goals in monthly view"""
    with st.form("assign_goal_monthly"):
        st.markdown(f"**Assigning for:** {get_month_name(month)} {year}, Q{quarter}")
        
        # Get employees based on role
        if user['role'] == 'HR':
            # HR can only assign to their own team members
            employees = db.get_team_members(user['id'])
        elif user['role'] == 'VP':
            # VP can only assign to HR and Manager
            all_users = db.get_all_users()
            employees = [u for u in all_users if u['role'] in ['HR', 'Manager']]
        elif user['role'] == 'CMD':
            # CMD can only assign to VP
            all_users = db.get_all_users()
            employees = [u for u in all_users if u['role'] == 'VP']
        else:  # Manager
            employees = db.get_team_members(user['id'])
        
        if not employees:
            st.info("No employees available")
            return
        
        selected_emp = st.selectbox(
            "Assign To*",
            [f"{e['name']} ({e['email']})" for e in employees]
        )
        
        emp_email = selected_emp.split('(')[1].strip(')')
        emp_id = next(e['id'] for e in employees if e['email'] == emp_email)
        
        col1, col2 = st.columns(2)
        with col1:
            department = st.text_input("Department*")
            title = st.text_input("Goal Title*")
            kpi = st.text_input("KPI*")
        
        with col2:
            description = st.text_area("Description")
            monthly_target = st.number_input("Monthly Target*", min_value=0.0)
        
        st.markdown("**Weekly Targets**")
        auto_divide = st.checkbox("Auto-divide monthly target equally into 4 weeks", value=True, key="monthly_assign_auto")
        
        if auto_divide and monthly_target > 0:
            weekly_target = monthly_target / 4
        else:
            weekly_target = 0.0
        
        col3, col4, col5, col6 = st.columns(4)
        with col3:
            w1_t = st.number_input("Week 1", min_value=0.0, value=weekly_target, key="monthly_assign_w1")
        with col4:
            w2_t = st.number_input("Week 2", min_value=0.0, value=weekly_target, key="monthly_assign_w2")
        with col5:
            w3_t = st.number_input("Week 3", min_value=0.0, value=weekly_target, key="monthly_assign_w3")
        with col6:
            w4_t = st.number_input("Week 4", min_value=0.0, value=weekly_target, key="monthly_assign_w4")
        
        # ========== ADD THIS ENTIRE SECTION ==========
        st.markdown("**Weekly Ratings (Optional)**")
        col_r1, col_r2, col_r3, col_r4 = st.columns(4)
        with col_r1:
            w1_rating = st.selectbox("Week 1 Rating", [0, 1, 2, 3, 4], key="quick_w1_rating")
        with col_r2:
            w2_rating = st.selectbox("Week 2 Rating", [0, 1, 2, 3, 4], key="quick_w2_rating")
        with col_r3:
            w3_rating = st.selectbox("Week 3 Rating", [0, 1, 2, 3, 4], key="quick_w3_rating")
        with col_r4:
            w4_rating = st.selectbox("Week 4 Rating", [0, 1, 2, 3, 4], key="quick_w4_rating")
        # ========== END OF ADDED SECTION ==========

        if st.form_submit_button("✅ Assign Goal", use_container_width=True):
            if department and title and kpi:
                goal_data = {
                    'user_id': emp_id,
                    'year': year,
                    'quarter': quarter,
                    'month': month,
                    'department': department,
                    'goal_title': title,
                    'goal_description': description,
                    'kpi': kpi,
                    'monthly_target': monthly_target,
                    'week1_target': w1_t,
                    'week2_target': w2_t,
                    'week3_target': w3_t,
                    'week4_target': w4_t,
                    'week1_remarks': w1_rating,  # ✅ ADD
                    'week2_remarks': w2_rating,  # ✅ ADD
                    'week3_remarks': w3_rating,  # ✅ ADD
                    'week4_remarks': w4_rating,
                    'start_date': f'{year}-{month:02d}-01',
                    'end_date': f'{year}-{month:02d}-28',
                    'created_by': user['id']
                }
                
                if db.create_goal(goal_data):
                    st.success(f"✅ Goal assigned to {selected_emp}")
    
                    # Notify about goal creation
                    assigned_emp = db.get_user_by_id(emp_id)
                    if assigned_emp:
                        notify_goal_created(goal_data, user)
                    st.rerun()
            else:
                st.error("❌ Please fill all required fields")


# ============================================
# ENHANCED FEEDBACK HISTORY (WITH USER INFO)
# ============================================
def display_feedback_history():
    """Display feedback history with reply and new feedback options"""
    user = st.session_state.user

    if not user:
        st.warning("⚠️ Session expired. Please login again.")
        st.rerun()

    role = user['role']
    
    st.title("💬 Feedback History")
    
    # Add New Feedback Button
    if st.button("➕ Add New Feedback", use_container_width=True):
        st.session_state.adding_new_feedback = True
    
    # New Feedback Modal
    if st.session_state.get('adding_new_feedback'):
        st.markdown("---")
        st.subheader("➕ Add New Feedback")
        
        with st.form("add_new_feedback_form"):
            # Get all goals based on role
            if role == 'HR':
                all_users = db.get_all_users()
                all_goals = []
                for u in all_users:
                    user_goals = db.get_user_all_goals(u['id'])
                    for g in user_goals:
                        g['user_name'] = u['name']
                    all_goals.extend(user_goals)
            elif role == 'Manager':
                team_members = db.get_team_members(user['id'])
                all_goals = []
                for m in team_members:
                    member_goals = db.get_user_all_goals(m['id'])
                    for g in member_goals:
                        g['user_name'] = m['name']
                    all_goals.extend(member_goals)
                # Add own goals
                own_goals = db.get_user_all_goals(user['id'])
                for g in own_goals:
                    g['user_name'] = user['name']
                all_goals.extend(own_goals)
            else:
                all_goals = db.get_user_all_goals(user['id'])
                for g in all_goals:
                    g['user_name'] = user['name']
            
            if all_goals:
                selected_goal_str = st.selectbox(
                    "Select Goal*",
                    [f"{g.get('user_name', 'Unknown')} - {g['goal_title']} ({g['year']}/Q{g.get('quarter', 'N/A')}/M{g.get('month', 'N/A')})" for g in all_goals]
                )
                
                selected_goal = all_goals[[f"{g.get('user_name', 'Unknown')} - {g['goal_title']} ({g['year']}/Q{g.get('quarter', 'N/A')}/M{g.get('month', 'N/A')})" for g in all_goals].index(selected_goal_str)]
                
                # Determine feedback type
                if role == 'HR':
                    fb_types = ["HR", "Manager"] if selected_goal['user_id'] != user['id'] else ["Self Appraisal", "HR"]
                elif role == 'Manager':
                    fb_types = ["Manager"] if selected_goal['user_id'] != user['id'] else ["Self Appraisal", "Manager"]
                else:
                    fb_types = ["Self Appraisal"]
                
                fb_type = st.selectbox("Feedback Type*", fb_types)
                rating = st.slider("Rating", 1, 5, 3)
                comment = st.text_area("Comment*")
                
                col_submit, col_cancel = st.columns(2)
                
                with col_submit:
                    if st.form_submit_button("Submit Feedback", use_container_width=True):
                        if comment.strip():
                            feedback_data = {
                                'goal_id': selected_goal['goal_id'],
                                'user_id': selected_goal['user_id'],
                                'feedback_by': user['id'],
                                'feedback_type': fb_type,
                                'rating': rating,
                                'comment': comment.strip(),
                                'level': 'general'
                            }
                            if db.create_feedback(feedback_data):
                                st.success("✅ Feedback submitted!")
                                del st.session_state.adding_new_feedback
                                st.rerun()
                        else:
                            st.error("❌ Please enter a comment")
                
                with col_cancel:
                    if st.form_submit_button("Cancel", use_container_width=True):
                        del st.session_state.adding_new_feedback
                        st.rerun()
            else:
                st.info("No goals available for feedback")
                if st.form_submit_button("Cancel", use_container_width=True):
                    del st.session_state.adding_new_feedback
                    st.rerun()
        
        st.markdown("---")
    
    # Get feedback based on role
    if role == 'HR':
        all_feedbacks = db.get_all_feedback()
        st.subheader("All Feedback (System-wide)")
    elif role == 'Manager':
        my_feedback = db.get_user_all_feedback(user['id'])
        team_feedback = []
        for member in db.get_team_members(user['id']):
            team_feedback.extend(db.get_user_all_feedback(member['id']))
        all_feedbacks = my_feedback + team_feedback
        st.subheader("My Feedback & Team Feedback")
    else:
        all_feedbacks = db.get_user_all_feedback(user['id'])
        st.subheader("My Feedback")
    
    all_feedbacks.sort(key=lambda x: x.get('created_at', ''), reverse=True)

    if not all_feedbacks:
        st.info("No feedback history found")
        return
    
    # Filters
    col1, col2, col3 = st.columns(3)
    with col1:
        filter_type = st.selectbox("Filter by Type", ["All", "Self Appraisal", "Manager", "HR"])
    with col2:
        filter_rating = st.selectbox("Filter by Rating", ["All", "5", "4", "3", "2", "1"])
    with col3:
        search_goal = st.text_input("Search Goal")
    
    # Filter feedbacks
    filtered = all_feedbacks
    if filter_type != "All":
        filtered = [f for f in filtered if f.get('feedback_type') == filter_type]
    if filter_rating != "All":
        filtered = [f for f in filtered if f.get('rating') == int(filter_rating)]
    if search_goal:
        filtered = [f for f in filtered if search_goal.lower() in f.get('goal_title', '').lower()]
    
    # Display feedback
    st.markdown(f"**Showing {len(filtered)} of {len(all_feedbacks)} feedback entries**")
    
    for feedback in filtered:
        # Get user who received feedback
        feedback_user = db.get_user_by_id(feedback.get('user_id'))
        feedback_user_name = feedback_user['name'] if feedback_user else 'Unknown'
        
        is_new = False
        created_at_str = feedback.get('created_at', '')
        if created_at_str:
            try:
                utc_time = datetime.strptime(created_at_str[:19], '%Y-%m-%dT%H:%M:%S')
                utc_time = pytz.utc.localize(utc_time)
                ist_time = utc_time.astimezone(IST)
                hours_ago = (datetime.now(IST) - ist_time).total_seconds() / 3600
                is_new = hours_ago <= 24
            except:
                pass

        with st.container():
            # Header with user info
            st.markdown(f"### 💬 Feedback for: **{feedback_user_name}**")
            
            col_fb1, col_fb2 = st.columns([3, 1])
            
            with col_fb1:
                st.markdown(f"**Goal:** {feedback.get('goal_title', 'N/A')}")
                st.markdown(f"**Type:** {feedback.get('feedback_type')} | **Rating:** {'⭐' * feedback.get('rating', 0)}")
                st.markdown(f"**Comment:** {feedback.get('comment')}")
                
                # Reply Section
                replies = db.get_feedback_replies(feedback.get('feedback_id'))
                if replies:
                    st.markdown("**Replies:**")
                    for reply in replies:
                        st.info(f"↳ **{reply.get('reply_by_name', 'Unknown')}:** {reply.get('reply_text')} ({reply.get('created_at', '')[:16]})")
                
                # Add Reply Button
                if st.button("💬 Reply", key=f"reply_btn_{feedback.get('feedback_id')}"):
                    st.session_state[f"replying_to_{feedback.get('feedback_id')}"] = True
                
                # Reply Form
                if st.session_state.get(f"replying_to_{feedback.get('feedback_id')}"):
                    with st.form(f"reply_form_{feedback.get('feedback_id')}"):
                        reply_text = st.text_area("Your Reply*", key=f"reply_text_{feedback.get('feedback_id')}")
                        
                        col_reply1, col_reply2 = st.columns(2)
                        with col_reply1:
                            if st.form_submit_button("Send Reply"):
                                if reply_text.strip():
                                    reply_data = {
                                        'feedback_id': feedback.get('feedback_id'),
                                        'reply_by': user['id'],
                                        'reply_text': reply_text.strip()
                                    }
                                    if db.create_feedback_reply(reply_data):
                                        # Notify the original feedback giver
                                        feedback_giver = db.get_user_by_id(feedback.get('feedback_by'))
                                        if feedback_giver:
                                            notify_feedback_reply(feedback, user, feedback_giver)
                                        
                                        st.success("✅ Reply added!")
                                        del st.session_state[f"replying_to_{feedback.get('feedback_id')}"]
                                        st.rerun()
                                else:
                                    st.error("❌ Please enter a reply")
                        
                        with col_reply2:
                            if st.form_submit_button("Cancel"):
                                del st.session_state[f"replying_to_{feedback.get('feedback_id')}"]
                                st.rerun()
            
            with col_fb2:
                st.markdown(f"**Given By:** {feedback.get('feedback_by_name', 'Unknown')}")
                
                # Convert UTC to IST
                created_at_str = feedback.get('created_at', '')
                if created_at_str:
                    try:
                        utc_time = datetime.strptime(created_at_str[:19], '%Y-%m-%dT%H:%M:%S')
                        utc_time = pytz.utc.localize(utc_time)
                        ist_time = utc_time.astimezone(IST)
                        st.markdown(f"**Date:** {ist_time.strftime('%Y-%m-%d')}")
                        st.markdown(f"**Time:** {ist_time.strftime('%I:%M %p IST')}")
                    except:
                        st.markdown(f"**Date:** {feedback.get('date', 'N/A')}")
                else:
                    st.markdown(f"**Date:** {feedback.get('date', 'N/A')}")
                
                # Delete button for HR
                if role == 'HR':
                    if st.button("🗑️ Delete", key=f"del_fb_{feedback.get('feedback_id')}", use_container_width=True):
                        if db.delete_feedback(feedback.get('feedback_id')):
                            st.success("✅ Feedback deleted!")
                            st.rerun()
            
            st.markdown("---")




# ============================================
# WEEK VIEW (Keep existing)
# ============================================
def display_week_view(user, year, quarter, month, week_num):
    """Display week-specific view with remarks from monthly goals"""
    if not year or not quarter or not month or not week_num:
        st.warning("⚠️ Invalid week view parameters. Returning to dashboard...")
        st.session_state.page = 'dashboard'
        st.rerun()
    
    month_name = get_month_name(month)
    st.subheader(f"📅 Week {week_num} - {month_name} {year}")
    
    # Get monthly goals to show breakdown
    monthly_goals = db.get_month_goals(user['id'], year, quarter, month)
    
    if monthly_goals:
        st.markdown("**Weekly Breakdown from Monthly Goals**")
        
        table_data = []
        for goal in monthly_goals:
            week_target = goal.get(f'week{week_num}_target', 0)
            week_achievement_raw = goal.get(f'week{week_num}_achievement')
            week_remarks_raw = goal.get(f'week{week_num}_remarks')
            
            # ✅ FIX: Separate display values from calculation values
            week_achievement_display = '-' if week_achievement_raw is None else str(week_achievement_raw)
            week_achievement_numeric = 0 if week_achievement_raw is None else float(week_achievement_raw)
            week_remarks = '-' if not week_remarks_raw else str(week_remarks_raw)
            
            # ✅ FIX: Use numeric value for progress calculation
            progress = calculate_progress(week_achievement_numeric, week_target)
            
            table_data.append({
                'Department': str(goal.get('department', '')),
                'Title': str(goal['goal_title']),
                'KPI': str(goal.get('kpi', '')),
                'Target': float(week_target),
                'Achievement': week_achievement_display,  # Display string
                'Progress %': f"{progress:.1f}%",
                'Remarks': week_remarks
            })
        
        df = pd.DataFrame(table_data)
        
        # ✅ Ensure proper data types for Arrow serialization
        df['Department'] = df['Department'].astype(str)
        df['Title'] = df['Title'].astype(str)
        df['KPI'] = df['KPI'].astype(str)
        df['Target'] = pd.to_numeric(df['Target'], errors='coerce').fillna(0)
        df['Achievement'] = df['Achievement'].astype(str)
        df['Progress %'] = df['Progress %'].astype(str)
        df['Remarks'] = df['Remarks'].astype(str)
        
        st.dataframe(df, use_container_width=True)
        
        # Show detailed goal cards with remarks
        st.markdown("---")
        st.markdown("**📊 Goal Details**")
        
        for goal in monthly_goals:
            week_target = goal.get(f'week{week_num}_target', 0)
            week_achievement_raw = goal.get(f'week{week_num}_achievement')
            week_remarks = goal.get(f'week{week_num}_remarks', '')
            
            # ✅ FIX: Use 0 instead of None for calculations
            week_achievement_numeric = 0 if week_achievement_raw is None else float(week_achievement_raw)
            progress = calculate_progress(week_achievement_numeric, week_target)
            
            # Display value
            week_achievement_display = '-' if week_achievement_raw is None else week_achievement_raw
            
            with st.expander(f"📌 {goal['goal_title']}"):
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Target", week_target)
                with col2:
                    st.metric("Achievement", week_achievement_display)
                with col3:
                    st.metric("Progress", f"{progress:.1f}%")
                
                render_progress_bar(progress, goal['goal_title'])
                
                st.markdown(f"**Department:** {goal.get('department', 'N/A')}")
                st.markdown(f"**KPI:** {goal.get('kpi', 'N/A')}")
                st.markdown(f"**Description:** {goal.get('goal_description', 'No description')}")
                
                # Show remarks
                st.markdown("---")
                if week_remarks:
                    st.markdown(f"**📝 Week {week_num} Remarks:**")
                    st.info(week_remarks)
                else:
                    st.caption(f"No remarks added for Week {week_num} yet")

    
    # Week-specific goals with management
    st.markdown("---")
    week_goals = db.get_week_goals(user['id'], year, quarter, month, week_num)
    
    if week_goals:
        st.markdown("**Week-Specific Goals**")
        
        # Tabs for managing week goals
        tab1, tab2, tab3 = st.tabs(["📋 View Goals", "✏️ Edit Goal", "🗑️ Delete Goal"])
        
        with tab1:
            for goal in week_goals:
                with st.expander(f"📌 {goal['goal_title']}"):
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Target", goal.get('weekly_target', 0))
                    with col2:
                        st.metric("Achievement", goal.get('weekly_achievement', 0))
                    with col3:
                        progress = calculate_progress(
                            goal.get('weekly_achievement', 0),
                            goal.get('weekly_target', 0)
                        )
                        st.metric("Progress", f"{progress:.1f}%")
                    
                    st.markdown(f"**Department:** {goal.get('department', 'N/A')}")
                    st.markdown(f"**KPI:** {goal.get('kpi', 'N/A')}")
                    st.markdown(f"**Description:** {goal.get('goal_description', 'No description')}")
                    
                    week_rating = goal.get(f'week{week_num}_rating', 0)

                    # Show rating with color
                    st.markdown("---")
                    if week_rating:
                        rating_colors = {1: '🔴', 2: '🟠', 3: '🟡', 4: '🟢'}
                        rating_text = {1: 'Poor', 2: 'Fair', 3: 'Good', 4: 'Excellent'}
                        st.markdown(f"**📊 Week {week_num} Rating:** {rating_colors.get(week_rating, '')} {rating_text.get(week_rating, 'N/A')} ({week_rating}/4)")
                    else:
                        st.caption(f"No rating added for Week {week_num} yet")

        with tab2:
            st.subheader("Edit Week-Specific Goal")
            
            edit_week_goal_title = st.selectbox(
                "Select Goal to Edit",
                [g['goal_title'] for g in week_goals],
                key=f"edit_week_goal_{week_num}"
            )
            edit_week_goal = next(g for g in week_goals if g['goal_title'] == edit_week_goal_title)
            
            with st.form(f"edit_week_goal_form_{week_num}"):
                col1, col2 = st.columns(2)
                with col1:
                    new_department = st.text_input("Department", value=edit_week_goal.get('department', ''))
                    new_title = st.text_input("Goal Title*", value=edit_week_goal['goal_title'])
                with col2:
                    new_kpi = st.text_input("KPI", value=edit_week_goal.get('kpi', ''))
                    new_status = st.selectbox(
                        "Status",
                        ['Active', 'Completed', 'On Hold', 'Cancelled'],
                        index=['Active', 'Completed', 'On Hold', 'Cancelled'].index(edit_week_goal.get('status', 'Active'))
                    )
                
                new_description = st.text_area("Description", value=edit_week_goal.get('goal_description', ''))
                
                col3, col4 = st.columns(2)
                with col3:
                    new_target = st.number_input("Weekly Target", min_value=0.0, value=float(edit_week_goal.get('weekly_target', 0)))
                with col4:
                    new_achievement = st.number_input("Weekly Achievement", min_value=0.0, value=float(edit_week_goal.get('weekly_achievement', 0)))
                
                # Add remarks field for week-specific goals
                new_remarks = st.text_area(f"Week {week_num} Remarks", value=edit_week_goal.get(f'week{week_num}_remarks', ''), height=100)
                
                if st.form_submit_button("💾 Save Changes", use_container_width=True):
                    if new_title:
                        updates = {
                            'department': new_department,
                            'goal_title': new_title,
                            'goal_description': new_description,
                            'kpi': new_kpi,
                            'weekly_target': new_target,
                            'weekly_achievement': new_achievement,
                            'status': new_status,
                            f'week{week_num}_remarks': new_remarks
                        }
                        
                        if db.update_goal(edit_week_goal['goal_id'], updates):
                            goal_owner = db.get_user_by_id(edit_week_goal['user_id'])
                            if goal_owner:
                                notify_goal_edited(edit_week_goal, user, goal_owner)
                            st.success("✅ Week goal updated!")
                            st.rerun()
                    else:
                        st.error("❌ Goal title is required")
        
        with tab3:
            st.subheader("⚠️ Delete Week-Specific Goal")
            st.warning("This action cannot be undone!")
            
            delete_week_goal_title = st.selectbox(
                "Select Goal to Delete",
                [g['goal_title'] for g in week_goals],
                key=f"delete_week_goal_{week_num}"
            )
            delete_week_goal = next(g for g in week_goals if g['goal_title'] == delete_week_goal_title)
            
            st.info(f"**Goal:** {delete_week_goal['goal_title']}")
            st.info(f"**Target:** {delete_week_goal.get('weekly_target', 0)}")
            
            confirm = st.checkbox("I understand this cannot be undone", key=f"confirm_week_del_{week_num}")
            
            if st.button("🗑️ Delete Goal", disabled=not confirm, use_container_width=True):
                goal_owner = db.get_user_by_id(delete_week_goal['user_id'])
                if db.delete_goal(delete_week_goal['goal_id']):
                    if goal_owner:
                        notify_goal_deleted(delete_week_goal, user, goal_owner)
                    st.success("✅ Week goal deleted!")
                    st.rerun()
    
    # Add week-specific goal (only for own goals)
    if not st.session_state.get('viewing_employee_year'):
        with st.expander("➕ Add Week-Specific Goal"):
            with st.form(f"add_week{week_num}_goal"):
                department = st.text_input("Department")
                title = st.text_input("Goal Title*")
                kpi = st.text_input("KPI")
                target = st.number_input("Weekly Target", min_value=0.0)
                description = st.text_area("Description")
                remarks = st.text_area(f"Week {week_num} Remarks", height=100)
                
                if st.form_submit_button("Create Goal"):
                    if title:
                        goal_data = {
                            'user_id': user['id'],
                            'year': year,
                            'quarter': quarter,
                            'month': month,
                            'week': week_num,
                            'department': department,
                            'goal_title': title,
                            'goal_description': description,
                            'kpi': kpi,
                            'weekly_target': target,
                            f'week{week_num}_remarks': remarks,
                            'start_date': str(date.today()),
                            'end_date': str(date.today())
                        }
                        if db.create_goal(goal_data):
                            st.success("✅ Goal created!")
                            st.rerun()
                    else:
                        st.error("❌ Please enter a goal title")
    
    # Week feedback
    st.markdown("---")
    all_week_goals = monthly_goals + week_goals
    if all_week_goals:
        display_feedback_section(all_week_goals, f'week{week_num}')

def display_add_goal_form(user, year, quarter, month):
    """Display form to add new goal"""
    with st.expander("➕ Add New Monthly Goal"):
        with st.form("add_goal"):
            col1, col2 = st.columns(2)
            
            with col1:
                department = st.text_input("Department*")
                title = st.text_input("Goal Title*")
                kpi = st.text_input("KPI*")
                monthly_target = st.number_input("Monthly Target*", min_value=0.0)
            
            with col2:
                start_date = st.date_input("Start Date", value=date.today())
                end_date = st.date_input("End Date", value=date.today())
                description = st.text_area("Description")
            
            st.markdown("**Weekly Targets**")
            auto_divide = st.checkbox("Auto-divide monthly target equally into 4 weeks", value=True)
            
            if auto_divide and monthly_target > 0:
                weekly_target = monthly_target / 4
            else:
                weekly_target = 0.0
            
            col3, col4, col5, col6 = st.columns(4)
            with col3:
                w1_t = st.number_input("Week 1 Target", min_value=0.0, value=weekly_target)
            with col4:
                w2_t = st.number_input("Week 2 Target", min_value=0.0, value=weekly_target)
            with col5:
                w3_t = st.number_input("Week 3 Target", min_value=0.0, value=weekly_target)
            with col6:
                w4_t = st.number_input("Week 4 Target", min_value=0.0, value=weekly_target)

            # ADD THIS NEW SECTION
            st.markdown("**Weekly Ratings (Optional)**")
            col_r1, col_r2, col_r3, col_r4 = st.columns(4)
            with col_r1:
                w1_rating = st.selectbox("Week 1 Rating", [0, 1, 2, 3, 4], key="quick_w1_rating")
            with col_r2:
                w2_rating = st.selectbox("Week 2 Rating", [0, 1, 2, 3, 4], key="quick_w2_rating")
            with col_r3:
                w3_rating = st.selectbox("Week 3 Rating", [0, 1, 2, 3, 4], key="quick_w3_rating")
            with col_r4:
                w4_rating = st.selectbox("Week 4 Rating", [0, 1, 2, 3, 4], key="quick_w4_rating")
                        
            if st.form_submit_button("Create Goal", use_container_width=True):
                if department and title and kpi:
                    goal_data = {
                        'user_id': user['id'],
                        'year': year,
                        'quarter': quarter,
                        'month': month,
                        'department': department,
                        'goal_title': title,
                        'goal_description': description,
                        'kpi': kpi,
                        'monthly_target': monthly_target,
                        'week1_target': w1_t,
                        'week2_target': w2_t,
                        'week3_target': w3_t,
                        'week4_target': w4_t,
                        'week1_remarks': w1_rating,  # ADD
                        'week2_remarks': w2_rating,  # ADD
                        'week3_remarks': w3_rating,  # ADD
                        'week4_remarks': w4_rating,
                        'start_date': str(start_date),
                        'end_date': str(end_date)
                    }
                    
                    is_valid, error_msg = validate_goal_data(goal_data)
                    if is_valid:
                        result, goal_year, goal_quarter, goal_month = create_goal_with_date_based_placement(goal_data)
                        if result:
                            st.success(f"✅ Goal created successfully in {get_month_name(goal_month)} {goal_year}!")
                            notify_goal_created(goal_data, user)
                            st.rerun()
                    else:
                        st.error(f"❌ {error_msg}")
                else:
                    st.error("❌ Please fill all required fields (Department, Title, KPI)")

def display_add_goal_form_inline(user, year, quarter, month):
    """Inline form to add new goal at the top"""
    with st.form("add_goal_inline_top"):
        col1, col2 = st.columns(2)
        
        with col1:
            department = st.text_input("Department*")
            title = st.text_input("Goal Title*")
            kpi = st.text_input("KPI*")
            monthly_target = st.number_input("Monthly Target*", min_value=0.0)
        
        with col2:
            start_date = st.date_input("Start Date", value=date(year, month, 1))
            end_date = st.date_input("End Date", value=date(year, month, calendar.monthrange(year, month)[1]))
            description = st.text_area("Description", height=100)
        
        st.markdown("**Weekly Targets**")
        auto_divide = st.checkbox("Auto-divide monthly target equally into 4 weeks", value=True, key="inline_auto_divide")
        
        if auto_divide and monthly_target > 0:
            weekly_target = monthly_target / 4
        else:
            weekly_target = 0.0
        
        col3, col4, col5, col6 = st.columns(4)
        with col3:
            w1_t = st.number_input("Week 1", min_value=0.0, value=weekly_target, key="inline_top_w1")
        with col4:
            w2_t = st.number_input("Week 2", min_value=0.0, value=weekly_target, key="inline_top_w2")
        with col5:
            w3_t = st.number_input("Week 3", min_value=0.0, value=weekly_target, key="inline_top_w3")
        with col6:
            w4_t = st.number_input("Week 4", min_value=0.0, value=weekly_target, key="inline_top_w4")
        
        st.markdown("**Weekly Ratings (Optional)**")
        col_r1, col_r2, col_r3, col_r4 = st.columns(4)
        with col_r1:
            w1_rating = st.selectbox("Week 1 Rating", [0, 1, 2, 3, 4], key="quick_w1_rating")
        with col_r2:
            w2_rating = st.selectbox("Week 2 Rating", [0, 1, 2, 3, 4], key="quick_w2_rating")
        with col_r3:
            w3_rating = st.selectbox("Week 3 Rating", [0, 1, 2, 3, 4], key="quick_w3_rating")
        with col_r4:
            w4_rating = st.selectbox("Week 4 Rating", [0, 1, 2, 3, 4], key="quick_w4_rating")
        
        col_submit, col_cancel = st.columns(2)
        
        with col_submit:
            if st.form_submit_button("✅ Create Goal", use_container_width=True):
                if department and title and kpi:
                    goal_data = {
                        'user_id': user['id'],
                        'year': year,
                        'quarter': quarter,
                        'month': month,
                        'department': department,
                        'goal_title': title,
                        'goal_description': description,
                        'kpi': kpi,
                        'monthly_target': monthly_target,
                        'week1_target': w1_t,
                        'week2_target': w2_t,
                        'week3_target': w3_t,
                        'week4_target': w4_t,
                        'week1_remarks': w1_rating,
                        'week2_remarks': w2_rating,
                        'week3_remarks': w3_rating,
                        'week4_remarks': w4_rating,
                        'start_date': str(start_date),
                        'end_date': str(end_date),
                        'created_by': user['id']  # Track who created the goal
                    }
                    
                    if db.create_goal(goal_data):
                        if user['role'] == 'Employee' and user.get('manager_id'):
                            manager = db.get_user_by_id(user['manager_id'])
                            if manager and manager.get('email'):
                                send_goal_approval_email(
                                    manager['email'],
                                    user['name'],
                                    goal_data,
                                    goal_data.get('goal_id')  # You'll need to get this from create_goal return
                                )
                                st.success(f"✅ Goal created! Approval request sent to your manager.")
                            else:
                                st.warning(f"✅ Goal created! But manager email not found.")
                        else:
                            st.success(f"✅ Goal created successfully!")
                        
                        notify_goal_created(goal_data, user)
                        st.session_state.show_create_goal_form = False
                        st.rerun()
                    else:
                        st.error("❌ Failed to create goal")
                else:
                    st.error("❌ Please fill all required fields (Department, Title, KPI)")
        
        with col_cancel:
            if st.form_submit_button("❌ Cancel", use_container_width=True):
                st.session_state.show_create_goal_form = False
                st.rerun()

def display_feedback_section(goals, level):
    """Display feedbacks with replies - Official feedback only"""
    if not goals:
        return

    user = st.session_state.user
    target_user_id = goals[0]['user_id']
    target_user = db.get_user_by_id(target_user_id)
    if not target_user:
        st.error("User not found")
        return

    target_role = target_user['role']
    st.subheader("💬 Feedback & Replies")

    # Goal selector
    selected_goal_title = st.selectbox(
        "Select Goal",
        [g['goal_title'] for g in goals],
        key=f"fb_goal_{level}_{target_user_id}"
    )
    selected_goal = next(g for g in goals if g['goal_title'] == selected_goal_title)

    # ──────────────────────────────────────────────
    # 1. Who is allowed to give feedback to this person?
    # ──────────────────────────────────────────────
    feedback_giver_role = get_feedback_giver_role(target_role)   # returns 'CMD', 'VP', 'Manager' or None

    if feedback_giver_role and user['role'] == feedback_giver_role:
        feedback_label = {
            'CMD':     'CMD Feedback',
            'VP':      'VP Feedback',
            'Manager': 'Manager Feedback'
        }[feedback_giver_role]
    else:
        feedback_label = None

    # ──────────────────────────────────────────────
    # 2. Show existing feedback WITH REPLIES
    # ──────────────────────────────────────────────
    feedbacks = db.get_goal_feedback(selected_goal['goal_id'])

    # Ensure feedback_by_name is populated
    for fb in feedbacks:
        if not fb.get('feedback_by_name'):
            feedback_by_id = fb.get('feedback_by')
            if feedback_by_id:
                giver = db.get_user_by_id(feedback_by_id)
                fb['feedback_by_name'] = giver['name'] if giver else 'Unknown'
            else:
                fb['feedback_by_name'] = 'Unknown'

    # Show only official feedbacks
    official_feedbacks = [f for f in feedbacks if f['feedback_type'] in (
        'CMD Feedback', 'VP Feedback', 'Manager Feedback'
    )]

    if official_feedbacks:
        st.markdown("---")
        
        for fb in official_feedbacks:
            label_to_show = fb['feedback_type']
            
            # Get the actual feedback giver's name
            feedback_by_id = fb.get('feedback_by')
            feedback_giver = db.get_user_by_id(feedback_by_id) if feedback_by_id else None
            feedback_giver_name = fb.get('feedback_by_name', 'Unknown')
            
            # Format the created_at date properly
            created_date = fb.get('created_at', '')
            if created_date:
                try:
                    if 'T' in created_date:
                        created_date = created_date.split('T')[0]
                    else:
                        created_date = created_date[:10]
                except:
                    created_date = 'Unknown date'
            
            # Main feedback container
            with st.container():
                st.markdown(f"""
                <div style="
                    background-color: #dbeafe;
                    border-left: 6px solid #3b82f6;
                    padding: 16px;
                    border-radius: 8px;
                    margin: 12px 0;
                    box-shadow: 0 2px 6px rgba(0,0,0,0.1);
                ">
                    <p style="margin:0; font-weight:600; color:#1e40af;">
                        {label_to_show}
                    </p>
                    <p style="margin:8px 0 4px 0; color:#1f2937;">
                        <strong>Rating:</strong> {'★' * fb.get('rating',0)} ({fb.get('rating',0)}/5)
                    </p>
                    <p style="margin:8px 0; color:#374151;">{fb.get('comment','')}</p>
                    <p style="margin:12px 0 0 0; font-size:13px; color:#6b7280;">
                        — {feedback_giver_name} • {created_date}
                    </p>
                </div>
                """, unsafe_allow_html=True)
                
                # ──────────────────────────────────────────────
                # SHOW REPLIES SECTION
                # ──────────────────────────────────────────────
                replies = db.get_feedback_replies(fb.get('feedback_id'))
                
                if replies:
                    st.markdown("**💬 Replies:**")
                    for reply in replies:
                        reply_by_name = reply.get('reply_by_name', 'Unknown')
                        reply_text = reply.get('reply_text', '')
                        reply_date = reply.get('created_at', '')
                        
                        # Format reply date
                        if reply_date:
                            try:
                                if 'T' in reply_date:
                                    reply_date = reply_date.split('T')[0]
                                else:
                                    reply_date = reply_date[:10]
                            except:
                                reply_date = ''
                        
                        st.markdown(f"""
                        <div style="
                            margin-left: 30px;
                            background-color: #f0f9ff;
                            border-left: 3px solid #60a5fa;
                            padding: 12px;
                            border-radius: 5px;
                            margin-top: 8px;
                        ">
                            <p style="margin:0; color:#1f2937; font-size: 14px;">↳ {reply_text}</p>
                            <p style="margin:8px 0 0 0; font-size:12px; color:#6b7280;">
                                — {reply_by_name} • {reply_date}
                            </p>
                        </div>
                        """, unsafe_allow_html=True)
                
                # ──────────────────────────────────────────────
                # REPLY BUTTON & FORM
                # ──────────────────────────────────────────────
                # Anyone can reply (employee can reply to manager's feedback, manager can reply back)
                reply_key = f"reply_btn_{fb.get('feedback_id')}_{level}"
                
                if st.button("💬 Reply to this feedback", key=reply_key):
                    st.session_state[f"replying_to_{fb.get('feedback_id')}_{level}"] = True
                
                # Reply Form
                if st.session_state.get(f"replying_to_{fb.get('feedback_id')}_{level}"):
                    with st.form(f"reply_form_{fb.get('feedback_id')}_{level}"):
                        reply_text = st.text_area(
                            "Your Reply*", 
                            key=f"reply_text_{fb.get('feedback_id')}_{level}",
                            height=100
                        )
                        
                        col_reply1, col_reply2 = st.columns(2)
                        with col_reply1:
                            if st.form_submit_button("📤 Send Reply", use_container_width=True):
                                if reply_text.strip():
                                    reply_data = {
                                        'feedback_id': fb.get('feedback_id'),
                                        'reply_by': user['id'],
                                        'reply_text': reply_text.strip()
                                    }
                                    if db.create_feedback_reply(reply_data):
                                        # Notify the original feedback giver
                                        feedback_giver = db.get_user_by_id(fb.get('feedback_by'))
                                        if feedback_giver:
                                            notify_feedback_reply(fb, user, feedback_giver)
                                        
                                        st.success("✅ Reply added!")
                                        del st.session_state[f"replying_to_{fb.get('feedback_id')}_{level}"]
                                        st.rerun()
                                else:
                                    st.error("❌ Please enter a reply")
                        
                        with col_reply2:
                            if st.form_submit_button("❌ Cancel", use_container_width=True):
                                del st.session_state[f"replying_to_{fb.get('feedback_id')}_{level}"]
                                st.rerun()
                
                st.markdown("---")
    else:
        st.info("💡 No feedback yet for this goal")

    # ──────────────────────────────────────────────
    # 3. Add new feedback form (only if allowed)
    # ──────────────────────────────────────────────
    if feedback_label:   # ← this means the logged-in user IS allowed
        st.markdown("---")
        with st.expander("➕ Add New Feedback", expanded=False):
            with st.form(key=f"fb_form_{selected_goal['goal_id']}_{level}"):
                rating = st.slider("Rating", 1, 5, 3, key=f"rating_{selected_goal['goal_id']}_{level}")
                comment = st.text_area("Comment *", height=120, key=f"comment_{selected_goal['goal_id']}_{level}")

                if st.form_submit_button("📤 Submit Feedback", use_container_width=True):
                    if not comment.strip():
                        st.error("❌ Comment is required")
                    else:
                        feedback_data = {
                            'goal_id': selected_goal['goal_id'],
                            'user_id': target_user_id,
                            'feedback_by': user['id'],
                            'feedback_type': feedback_label,          # ← exactly "CMD Feedback" / "VP Feedback" / "Manager Feedback"
                            'rating': rating,
                            'comment': comment.strip(),
                            'level': level
                        }
                        if db.create_feedback(feedback_data):
                            # Notify goal owner
                            goal_owner = db.get_user_by_id(target_user_id)
                            if goal_owner:
                                notify_feedback_given(selected_goal, user, goal_owner)
                            
                            st.success("✅ Feedback submitted successfully!")
                            st.rerun()
                        else:
                            st.error("❌ Failed to save feedback")
    else:
        if user['id'] != target_user_id:
            st.caption("ℹ️ You do not have permission to give feedback to this user.")

def display_profile():
    """Display and edit user profile"""
    user = st.session_state.user
    
    if not user:
        st.warning("⚠️ Session expired. Please login again.")
        st.rerun()

    st.title("👤 My Profile")
    
    tab1, tab2 = st.tabs(["📝 Edit Profile", "🔒 Change Password"])
    
    with tab1:
        with st.form("edit_profile"):
            st.subheader("Personal Information")
            
            col1, col2 = st.columns(2)
            with col1:
                new_name = st.text_input("Full Name*", value=user['name'])
                new_email = st.text_input("Email*", value=user['email'])
            
            with col2:
                new_designation = st.text_input("Designation", value=user.get('designation', ''))
                new_department = st.text_input("Department", value=user.get('department', ''))
            
            if st.form_submit_button("💾 Save Changes", use_container_width=True):
                if new_name and new_email:
                    updates = {
                        'name': new_name,
                        'email': new_email.lower().strip(),
                        'designation': new_designation,
                        'department': new_department
                    }
                    
                    if db.update_user(user['id'], updates):
                        st.session_state.user.update(updates)
                        st.success("✅ Profile updated successfully!")
                        st.rerun()
                    else:
                        st.error("❌ Failed to update profile")
                else:
                    st.error("❌ Name and Email are required")
    
    with tab2:
        # Initialize session state for forgot password in profile
        if 'profile_forgot_password' not in st.session_state:
            st.session_state.profile_forgot_password = False
        if 'profile_reset_token_sent' not in st.session_state:
            st.session_state.profile_reset_token_sent = False
        
        # Track password input for strength meter (OUTSIDE FORM)
        if 'temp_new_password' not in st.session_state:
            st.session_state.temp_new_password = ""
        
        # ===== NORMAL CHANGE PASSWORD =====
        if not st.session_state.profile_forgot_password and not st.session_state.profile_reset_token_sent:
            st.subheader("Change Password")
            
            old_password = st.text_input("Current Password*", type="password", key="old_pass_input")
            new_password = st.text_input("New Password*", type="password", key="new_pass_input")
            
            # Update temp password for strength meter
            if new_password != st.session_state.temp_new_password:
                st.session_state.temp_new_password = new_password
            
            # Password strength meter (OUTSIDE FORM)
            if new_password:
                render_password_strength_meter(new_password, "profile_change")
            
            confirm_password = st.text_input("Confirm New Password*", type="password", key="confirm_pass_input")
            
            col_submit, col_forgot = st.columns(2)
            
            with col_submit:
                if st.button("🔒 Change Password", use_container_width=True, key="change_pass_btn"):
                    if old_password == user['password']:
                        if new_password == confirm_password:
                            if len(new_password) >= 6:
                                # Check password strength
                                score, _, strength, _ = check_password_strength(new_password)
                                
                                if db.update_user(user['id'], {'password': new_password}):
                                    st.session_state.user['password'] = new_password
                                    st.session_state.temp_new_password = ""
                                    st.success("✅ Password changed successfully!")
                                    if score >= 70:
                                        st.info("🔒 Great! Your password is strong.")
                                    elif score < 30:
                                        st.warning(f"⚠️ Your password is {strength}. Consider making it stronger.")
                                    st.rerun()
                                else:
                                    st.error("❌ Failed to change password")
                            else:
                                st.error("❌ Password must be at least 6 characters")
                        else:
                            st.error("❌ Passwords don't match")
                    else:
                        st.error("❌ Current password is incorrect")
            
            with col_forgot:
                if st.button("🔑 Forgot Password?", use_container_width=True, key="forgot_pass_btn"):
                    st.session_state.profile_forgot_password = True
                    st.rerun()
        
        # ===== FORGOT PASSWORD - REQUEST TOKEN =====
        elif st.session_state.profile_forgot_password and not st.session_state.profile_reset_token_sent:
            st.subheader("🔑 Reset Password via Email")
            st.info("We'll send a reset token to your registered email address.")
            
            st.markdown(f"**Your Email:** {user['email']}")
            st.caption("A reset token will be sent to this email address")
            
            col_send, col_cancel = st.columns(2)
            
            with col_send:
                if st.button("📧 Send Reset Token", use_container_width=True, key="send_token_btn"):
                    # Generate token
                    token = db.create_password_reset_token(user['email'])
                    
                    if token:
                        # Try to send email
                        email_sent = send_password_reset_email(user['email'], token)
                        
                        if email_sent:
                            st.success("✅ Reset token sent to your email!")
                            st.info(f"**Backup Token:** `{token}`\n\n(In case you don't receive the email)")
                        else:
                            st.warning("⚠️ Could not send email. Please use this token:")
                            st.code(token, language=None)
                        
                        st.session_state.profile_reset_token_sent = True
                        st.rerun()
                    else:
                        st.error("❌ Failed to generate reset token")
            
            with col_cancel:
                if st.button("❌ Cancel", use_container_width=True, key="cancel_forgot_btn"):
                    st.session_state.profile_forgot_password = False
                    st.rerun()
        
        # ===== RESET PASSWORD WITH TOKEN =====
        elif st.session_state.profile_reset_token_sent:
            st.subheader("🔒 Enter Reset Token")
            st.info("Check your email for the reset token and enter it below with your new password.")
            
            reset_token = st.text_input("🎫 Reset Token*", placeholder="Enter 8-character token", key="reset_token_input")
            new_password_reset = st.text_input("🔒 New Password*", type="password", placeholder="••••••••", key="new_pass_reset_input")
            
            # Password strength meter (OUTSIDE FORM)
            if new_password_reset:
                render_password_strength_meter(new_password_reset, "profile_reset")
            
            confirm_password_reset = st.text_input("🔒 Confirm New Password*", type="password", placeholder="••••••••", key="confirm_pass_reset_input")
            
            col_reset, col_cancel = st.columns(2)
            
            with col_reset:
                if st.button("✅ Reset Password", use_container_width=True, key="reset_pass_btn"):
                    if reset_token and new_password_reset and confirm_password_reset:
                        if new_password_reset == confirm_password_reset:
                            if len(new_password_reset) >= 6:
                                # Check password strength
                                score, _, strength, _ = check_password_strength(new_password_reset)
                                
                                # Reset password using token
                                if db.reset_password_with_token(reset_token, new_password_reset):
                                    st.success("✅ Password reset successful!")
                                    if score >= 70:
                                        st.info("🔒 Great! Your password is strong.")
                                    elif score < 30:
                                        st.warning(f"⚠️ Your password is {strength}. Consider making it stronger.")
                                    st.balloons()
                                    
                                    # Update session
                                    st.session_state.user['password'] = new_password_reset
                                    st.session_state.profile_forgot_password = False
                                    st.session_state.profile_reset_token_sent = False
                                    st.rerun()
                                else:
                                    st.error("❌ Invalid or expired token. Please request a new one.")
                            else:
                                st.error("❌ Password must be at least 6 characters")
                        else:
                            st.error("❌ Passwords don't match")
                    else:
                        st.warning("⚠️ Please fill all fields")
            
            with col_cancel:
                if st.button("❌ Cancel", use_container_width=True, key="cancel_reset_btn"):
                    st.session_state.profile_forgot_password = False
                    st.session_state.profile_reset_token_sent = False
                    st.rerun()
            
            # Option to resend token
            st.markdown("---")
            if st.button("📧 Resend Reset Token", use_container_width=True, key="resend_token_btn"):
                token = db.create_password_reset_token(user['email'])
                if token:
                    email_sent = send_password_reset_email(user['email'], token)
                    if email_sent:
                        st.success("✅ New token sent to your email!")
                    st.info(f"**Backup Token:** `{token}`")
                else:
                    st.error("❌ Failed to generate new token")
                    
def display_permissions():
    """Manage user permissions (HR only)"""
    user = st.session_state.user
    
    if not user:
        st.warning("⚠️ Session expired. Please login again.")
        st.rerun()

    if user['role'] != 'HR':
        st.warning("⚠️ Only HR can access this page")
        return
    
    st.title("🔐 Permissions Management")
    
    st.info("**Note:** This section allows HR to manage granular permissions for users")
    
    all_users = db.get_all_users()
    
    # Permission categories
    permissions = {
        'view_all_goals': 'View All Goals',
        'edit_all_goals': 'Edit All Goals',
        'delete_all_goals': 'Delete All Goals',
        'view_all_feedback': 'View All Feedback',
        'edit_all_feedback': 'Edit All Feedback',
        'delete_all_feedback': 'Delete All Feedback',
        'create_users': 'Create Users',
        'edit_users': 'Edit Users',
        'delete_users': 'Delete Users',
        'manage_teams': 'Manage Teams',
        'view_analytics': 'View Analytics',
        'export_data': 'Export Data'
    }
    
    selected_user = st.selectbox(
        "Select User",
        [f"{u['name']} ({u['role']}) - {u['email']}" for u in all_users if u['id'] != user['id']]
    )
    
    if selected_user:
        user_email = selected_user.split(' - ')[1]
        selected_user_obj = next(u for u in all_users if u['email'] == user_email)
        
        st.subheader(f"Permissions for {selected_user_obj['name']}")
        
        # Get current permissions
        current_perms = db.get_user_permissions(selected_user_obj['id'])
        
        with st.form("update_permissions"):
            st.markdown("**Grant Permissions:**")
            
            selected_perms = []
            cols = st.columns(2)
            for idx, (perm_key, perm_label) in enumerate(permissions.items()):
                with cols[idx % 2]:
                    if st.checkbox(perm_label, value=perm_key in current_perms, key=perm_key):
                        selected_perms.append(perm_key)
            
            if st.form_submit_button("💾 Update Permissions", use_container_width=True):
                if db.update_user_permissions(selected_user_obj['id'], selected_perms):
                    st.success(f"✅ Permissions updated for {selected_user_obj['name']}")
                    st.rerun()
                else:
                    st.error("❌ Failed to update permissions")
        
        # Show current permissions
        st.markdown("---")
        st.subheader("Current Permissions")
        if current_perms:
            for perm in current_perms:
                st.success(f"✓ {permissions.get(perm, perm)}")
        else:
            st.info("No special permissions granted")


def display_employee_management():
    """Employee management - create users, teams"""
    user = st.session_state.user
    
    if not user:
        st.warning("⚠️ Session expired. Please login again.")
        st.rerun()

    # Check if user has permission to create employees
    if user['role'] not in ['CMD', 'VP', 'HR']:
        st.warning("⚠️ You don't have permission to access this page")
        return
    
    st.title("⚙️ Employee Management")
    
    tab1, tab2, tab3, tab4 = st.tabs(["👤 Create Employee", "👥 All Employees", "👥 Manage Teams", "📋 View All Teams"])
    
    # ===== CREATE EMPLOYEE TAB =====
    # ===== CREATE EMPLOYEE TAB =====
    with tab1:
        st.subheader("Create New Employee")
        
        with st.form("create_employee"):
            col1, col2 = st.columns(2)
            
            with col1:
                new_name = st.text_input("Full Name*")
                new_email = st.text_input("Email*")
                new_password = st.text_input("Password*", type="password")
                new_designation = st.text_input("Designation")
            
            with col2:
                # Show only roles that current user can create
                modifiable_roles = get_modifiable_roles(user['role'])
                new_role = st.selectbox("Role*", modifiable_roles)
                new_department = st.text_input("Department")
                new_joining_date = st.date_input("Joining Date*", value=date.today())
                
                managers = [u for u in db.get_all_users() if u['role'] == 'Manager']
                manager_options = ["None"] + [f"{m['name']} ({m['email']})" for m in managers]
                selected_manager = st.selectbox("Assign to Manager", manager_options)
            
            if st.form_submit_button("Create Employee", use_container_width=True):
                if new_name and new_email and new_password and new_role:
                    manager_id = None
                    if selected_manager != "None":
                        manager_email = selected_manager.split('(')[1].strip(')')
                        manager_id = next((m['id'] for m in managers if m['email'] == manager_email), None)
                    
                    employee_data = {
                        'name': new_name,
                        'email': new_email.lower().strip(),
                        'password': new_password,
                        'designation': new_designation,
                        'role': new_role,
                        'department': new_department,
                        'manager_id': manager_id,
                        'joining_date': str(new_joining_date)
                    }
                    
                    if db.create_user(employee_data):
                        st.success(f"✅ Employee {new_name} created successfully!")
                        # No notification needed for employee creation per new requirements
                    else:
                        st.error("❌ Failed to create employee")
                else:
                    st.error("❌ Please fill all required fields")
    
    # ===== ALL EMPLOYEES TAB (NEW) =====
    with tab2:
        st.subheader("All Employees in Organization")
        
        # Search and filter
        col_search1, col_search2, col_search3 = st.columns(3)
        with col_search1:
            search_emp = st.text_input("🔍 Search by name or email", key="all_emp_search")
        with col_search2:
            filter_role = st.selectbox("Filter by Role", ["All"] + ["CMD", "VP", "HR", "Manager", "Employee"], key="all_emp_role")
        with col_search3:
            filter_dept = st.selectbox("Filter by Department", ["All"] + list(set([u.get('department', 'N/A') for u in db.get_all_users()])), key="all_emp_dept")
        
        all_users_list = db.get_all_users()
        
        # Apply filters
        filtered_users = all_users_list
        if search_emp:
            filtered_users = [u for u in filtered_users if search_emp.lower() in u['name'].lower() or search_emp.lower() in u['email'].lower()]
        if filter_role != "All":
            filtered_users = [u for u in filtered_users if u['role'] == filter_role]
        if filter_dept != "All":
            filtered_users = [u for u in filtered_users if u.get('department', 'N/A') == filter_dept]
        
        st.markdown(f"**Showing {len(filtered_users)} of {len(all_users_list)} employees**")
        st.markdown("---")
        
        # ============================================
        # ✅ NEW: SIMPLE LIST VIEW INSTEAD OF CARDS
        # ============================================
        
        # Create employee list with action buttons
        for idx, emp in enumerate(filtered_users):
            # Role badge color
            role_colors = {
                'CMD': '#8B0000',
                'VP': '#FF4500',
                'HR': '#4facfe',
                'Manager': '#f093fb',
                'Employee': '#dbeafe'
            }
            role_color = role_colors.get(emp['role'], '#dbeafe')
            
            # Create horizontal layout
            col1, col2, col3, col4, col5 = st.columns([3, 2, 2, 1, 1])
            
            with col1:
                st.markdown(f"**{emp['name']}**")
            
            with col2:
                st.markdown(f"{emp.get('designation', 'N/A')}")
            
            with col3:
                st.markdown(f"*{emp.get('department', 'N/A')}*")
            
            with col4:
                st.markdown(f"""<span style='background: {role_color}; color: white; padding: 4px 12px; 
                            border-radius: 12px; font-size: 11px; font-weight: bold; display: inline-block;'>
                    {emp['role']}</span>""", unsafe_allow_html=True)
            
            with col5:
                if st.button("👁️", key=f"view_all_emp_{emp['id']}_{idx}", help=f"View {emp['name']}'s details"):
                    st.session_state.viewing_employee_details = emp
                    st.rerun()
            
            st.markdown("---")
        
        # Employee Details Modal (keep existing code)
        if 'viewing_employee_details' in st.session_state:
            st.markdown("---")
            
            view_emp = st.session_state.viewing_employee_details
            
            st.markdown(f"""
            <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                        padding: 20px; border-radius: 10px; margin-bottom: 20px;">
                <h2 style="color: white; margin: 0;">👤 {view_emp['name']}'s Details</h2>
            </div>
            """, unsafe_allow_html=True)
            
            # Display current details
            col_detail1, col_detail2 = st.columns(2)
            
            with col_detail1:
                st.markdown("**Personal Information**")
                st.info(f"**Name:** {view_emp['name']}")
                st.info(f"**Email:** {view_emp['email']}")
                st.info(f"**Role:** {view_emp['role']}")
                st.info(f"**Designation:** {view_emp.get('designation', 'N/A')}")
            
            with col_detail2:
                st.markdown("**Organization Details**")
                st.info(f"**Department:** {view_emp.get('department', 'N/A')}")
                st.info(f"**Joining Date:** {view_emp.get('joining_date', 'N/A')}")
                st.info(f"**Exit Date:** {view_emp.get('exit_date', 'Not Set')}")
                
                # Manager info
                if view_emp.get('manager_id'):
                    manager = db.get_user_by_id(view_emp['manager_id'])
                    st.info(f"**Manager:** {manager['name'] if manager else 'N/A'}")
                else:
                    st.info("**Manager:** None")
            
            st.markdown("---")
            
            # Edit Employee Details Form
            with st.expander("✏️ Edit Employee Details", expanded=False):
                with st.form(f"edit_emp_details_{view_emp['id']}"):
                    col_edit1, col_edit2 = st.columns(2)
                    
                    with col_edit1:
                        edit_name = st.text_input("Full Name*", value=view_emp['name'])
                        edit_email = st.text_input("Email*", value=view_emp['email'])
                        edit_designation = st.text_input("Designation", value=view_emp.get('designation', ''))
                        edit_department = st.text_input("Department", value=view_emp.get('department', ''))
                    
                    with col_edit2:
                        edit_role = st.selectbox("Role*", ["CMD", "VP", "HR", "Manager", "Employee"], 
                                                index=["CMD", "VP", "HR", "Manager", "Employee"].index(view_emp['role']))
                        
                        # Joining date
                        joining_date_val = view_emp.get('joining_date')
                        if joining_date_val:
                            if isinstance(joining_date_val, str):
                                joining_date_val = datetime.strptime(joining_date_val, '%Y-%m-%d').date()
                        else:
                            joining_date_val = date.today()
                        edit_joining_date = st.date_input("Joining Date", value=joining_date_val)

                        # Exit date - Allow removal
                        exit_date_val = view_emp.get('exit_date')
                        remove_exit = st.checkbox("Remove Exit Date", value=(exit_date_val is None))

                        if not remove_exit:
                            if exit_date_val:
                                if isinstance(exit_date_val, str):
                                    exit_date_val = datetime.strptime(exit_date_val, '%Y-%m-%d').date()
                                edit_exit_date = st.date_input("Exit Date", value=exit_date_val)
                            else:
                                edit_exit_date = st.date_input("Exit Date", value=date.today())
                        else:
                            edit_exit_date = None
                        
                        # Manager assignment
                        managers = [u for u in db.get_all_users() if u['role'] == 'Manager' and u['id'] != view_emp['id']]
                        manager_options = ["None"] + [f"{m['name']} ({m['email']})" for m in managers]
                        
                        current_manager_idx = 0
                        if view_emp.get('manager_id'):
                            current_manager = db.get_user_by_id(view_emp['manager_id'])
                            if current_manager:
                                current_manager_str = f"{current_manager['name']} ({current_manager['email']})"
                                if current_manager_str in manager_options:
                                    current_manager_idx = manager_options.index(current_manager_str)
                        
                        edit_manager = st.selectbox("Assign to Manager", manager_options, index=current_manager_idx)
                    
                    col_save, col_cancel = st.columns(2)
                    
                    with col_save:
                        if st.form_submit_button("💾 Save Changes", use_container_width=True):
                            if edit_name and edit_email and edit_role:
                                manager_id = None
                                if edit_manager != "None":
                                    manager_email = edit_manager.split('(')[1].strip(')')
                                    manager_id = next((m['id'] for m in managers if m['email'] == manager_email), None)
                                
                                updates = {
                                    'name': edit_name,
                                    'email': edit_email.lower().strip(),
                                    'designation': edit_designation,
                                    'role': edit_role,
                                    'department': edit_department,
                                    'manager_id': manager_id,
                                    'joining_date': str(edit_joining_date),
                                    'exit_date': str(edit_exit_date) if edit_exit_date and not remove_exit else None
                                }
                                
                                if db.update_user(view_emp['id'], updates):
                                    st.success(f"✅ Employee {edit_name} updated successfully!")
                                    del st.session_state.viewing_employee_details
                                    st.rerun()
                                else:
                                    st.error("❌ Failed to update employee")
                            else:
                                st.error("❌ Please fill all required fields")
                    
                    with col_cancel:
                        if st.form_submit_button("❌ Cancel", use_container_width=True):
                            # Just close the form, keep modal open
                            pass
            
            # Close Details Button
            if st.button("← Back to All Employees", use_container_width=True):
                del st.session_state.viewing_employee_details
                st.rerun()
    
    
    
    # ===== MANAGE TEAMS TAB =====
    with tab3:
        st.subheader("Assign Employees to Managers")
        
        managers = [u for u in db.get_all_users() if u['role'] == 'Manager']
        employees = [u for u in db.get_all_users() if u['role'] == 'Employee']
        
        if not managers:
            st.info("No managers found. Create a manager first.")
        elif not employees:
            st.info("No employees found. Create employees first.")
        else:
            selected_manager_name = st.selectbox(
                "Select Manager",
                [f"{m['name']} ({m['email']})" for m in managers]
            )
            
            manager_email = selected_manager_name.split('(')[1].strip(')')
            selected_manager = next(m for m in managers if m['email'] == manager_email)
            
            current_team = db.get_team_members(selected_manager['id'])
            st.write(f"**Current Team Size:** {len(current_team)}")
            
            unassigned_employees = [e for e in employees if not e.get('manager_id')]
            
            if unassigned_employees:
                selected_employees = st.multiselect(
                    "Select Employees to Add",
                    [f"{e['name']} ({e['email']})" for e in unassigned_employees]
                )
                
                if st.button("Assign to Team", use_container_width=True):
                    for emp_str in selected_employees:
                        emp_email = emp_str.split('(')[1].strip(')')
                        emp_id = next(e['id'] for e in unassigned_employees if e['email'] == emp_email)
                        db.update_user(emp_id, {'manager_id': selected_manager['id']})
                    
                    st.success(f"✅ Assigned {len(selected_employees)} employees to {selected_manager['name']}")
                    st.rerun()
            else:
                st.info("No unassigned employees available")
    
    # ===== VIEW ALL TEAMS TAB =====
    with tab4:
        st.subheader("All Teams Overview")
        
        managers = [u for u in db.get_all_users() if u['role'] == 'Manager']
        for manager in managers:
            with st.expander(f"👔 {manager['name']}'s Team ({manager.get('department', 'N/A')})"):
                team = db.get_team_members(manager['id'])
                
                if team:
                    team_data = []
                    for member in team:
                        team_data.append({
                            'Name': member['name'],
                            'Email': member['email'],
                            'Designation': member.get('designation', 'N/A'),
                            'Department': member.get('department', 'N/A')
                        })
                    
                    df = pd.DataFrame(team_data)
                    st.dataframe(df, use_container_width=True)
                    
                    # Option to remove members
                    remove_member = st.selectbox(
                        "Remove Member",
                        ["None"] + [m['name'] for m in team],
                        key=f"remove_{manager['id']}"
                    )
                    
                    if remove_member != "None" and st.button(f"Remove {remove_member}", key=f"btn_remove_{manager['id']}"):
                        member_id = next(m['id'] for m in team if m['name'] == remove_member)
                        db.update_user(member_id, {'manager_id': None})
                        st.success(f"✅ Removed {remove_member} from team")
                        st.rerun()
                else:
                    st.info("No team members yet")

    st.markdown("---")
    st.subheader("🧪 Email Testing")
    
    # Create tabs for different email tests
    test_tab1, test_tab2, test_tab3 = st.tabs([
        "📧 Old Email Tests", 
        "📊 Goal Sheet Reminders",
        "✉️ Custom Test"
    ])
    
    # ===== TAB 1: OLD EMAIL TESTS =====
    with test_tab1:
        st.markdown("**Legacy Email Tests**")
        
        col1, col2 = st.columns(2)

        with col1:
            if st.button("📧 Test Manager Reminder (All)", use_container_width=True, key="test_old_reminder"):
                with st.spinner("Sending test emails..."):
                    test_send_reminder(db)
                st.success("✅ Test emails sent to all managers!")

        with col2:
            if st.button("📧 Test Goal Completion", use_container_width=True, key="test_old_completion"):
                from monthly_reminder import send_goal_completion_email
                managers = [u for u in db.get_all_users() if u['role'] == 'Manager']
                if managers and managers[0].get('email'):
                    send_goal_completion_email(
                        managers[0]['email'],
                        "Test Employee",
                        "Test Goal - Achievement Unlocked",
                        completed=True
                    )
                    st.success(f"✅ Test completion email sent to {managers[0]['email']}!")
                else:
                    st.warning("No manager with email found")
    
    # ===== TAB 2: GOAL SHEET REMINDER TESTS =====
    with test_tab2:
        st.markdown("**Test New Goal Sheet Completion Reminders**")
        st.info("💡 These are the new reminders that will be sent daily from 26th-31st of each month")
        
        # Get all HR emails for display
        all_users = db.get_all_users()
        hr_users = [u for u in all_users if u['role'] == 'HR' and u.get('email')]
        hr_emails = [hr['email'] for hr in hr_users]
        
        if hr_emails:
            st.success(f"📋 HR emails that will be CC'd: {', '.join(hr_emails)}")
        else:
            st.warning("⚠️ No HR users with email found - no CC will be added")
        
        st.markdown("---")
        
        # Test 1: Employee Reminder
        st.markdown("#### 1️⃣ Test Employee Goal Sheet Reminder")
        st.caption("This email goes to employees who haven't completed their goal sheet")
        
        col_emp1, col_emp2 = st.columns([2, 1])
        
        with col_emp1:
            # Select employee to test
            employees = [u for u in all_users if u.get('email')]
            if employees:
                selected_emp = st.selectbox(
                    "Select Employee for Test",
                    [f"{e['name']} ({e['email']})" for e in employees],
                    key="test_emp_reminder_select"
                )
                
                emp_email = selected_emp.split('(')[1].strip(')')
                emp_obj = next(e for e in employees if e['email'] == emp_email)
        
        with col_emp2:
            if st.button("📧 Send Test", use_container_width=True, key="send_test_emp_reminder"):
                if employees:
                    from monthly_reminder import send_goal_sheet_reminder_email
                    
                    # Create sample incomplete goals
                    test_incomplete_goals = [
                        {
                            'goal_title': 'Sample Goal 1 - Q4 Sales Target',
                            'goal_id': 'test_goal_1',
                            'department': 'Sales',
                            'kpi': 'Revenue Growth',
                            'monthly_target': 100000,
                            'missing_items': [
                                'Week 1 achievement',
                                'Week 2 rating',
                                'Week 3 achievement',
                                'Week 4 rating'
                            ]
                        },
                        {
                            'goal_title': 'Sample Goal 2 - Customer Acquisition',
                            'goal_id': 'test_goal_2',
                            'department': 'Marketing',
                            'kpi': 'New Customers',
                            'monthly_target': 50,
                            'missing_items': [
                                'Week 3 achievement',
                                'Week 4 achievement',
                                'Monthly achievement'
                            ]
                        }
                    ]
                    
                    today = datetime.now(IST)
                    
                    with st.spinner("Sending employee reminder..."):
                        success = send_goal_sheet_reminder_email(
                            emp_obj['email'],
                            emp_obj['name'],
                            test_incomplete_goals,
                            today.month,
                            today.year,
                            hr_emails
                        )
                    
                    if success:
                        st.success(f"✅ Employee reminder sent to {emp_obj['email']}")
                        if hr_emails:
                            st.info(f"📋 CC sent to HR: {', '.join(hr_emails)}")
                    else:
                        st.error("❌ Failed to send email")
        
        st.markdown("---")
        
        # Test 2: Manager Alert
        st.markdown("#### 2️⃣ Test Manager Team Alert")
        st.caption("This email goes to managers about their team members' incomplete goal sheets")
        
        col_mgr1, col_mgr2 = st.columns([2, 1])
        
        with col_mgr1:
            # Select manager to test
            managers = [u for u in all_users if u['role'] == 'Manager' and u.get('email')]
            if managers:
                selected_mgr = st.selectbox(
                    "Select Manager for Test",
                    [f"{m['name']} ({m['email']})" for m in managers],
                    key="test_mgr_alert_select"
                )
                
                mgr_email = selected_mgr.split('(')[1].strip(')')
                mgr_obj = next(m for m in managers if m['email'] == mgr_email)
                
                # Select an employee from their team (or use test name)
                team_members = db.get_team_members(mgr_obj['id'])
                if team_members:
                    test_emp_name = st.selectbox(
                        "Select Team Member (for demo)",
                        [tm['name'] for tm in team_members],
                        key="test_team_member_select"
                    )
                else:
                    test_emp_name = st.text_input(
                        "Employee Name (no team members found - enter test name)",
                        value="John Doe",
                        key="test_emp_name_input"
                    )
        
        with col_mgr2:
            if st.button("📧 Send Test", use_container_width=True, key="send_test_mgr_alert"):
                if managers:
                    from monthly_reminder import send_manager_team_reminder_email
                    
                    # Create sample incomplete goals
                    test_incomplete_goals = [
                        {
                            'goal_title': 'Q4 Revenue Target',
                            'goal_id': 'test_goal_1',
                            'department': 'Sales',
                            'kpi': 'Monthly Revenue',
                            'monthly_target': 150000,
                            'missing_items': [
                                'Week 2 achievement',
                                'Week 3 rating',
                                'Week 4 achievement'
                            ]
                        },
                        {
                            'goal_title': 'Client Retention Goal',
                            'goal_id': 'test_goal_2',
                            'department': 'Customer Success',
                            'kpi': 'Retention Rate',
                            'monthly_target': 95,
                            'missing_items': [
                                'Week 1 rating',
                                'Week 4 achievement',
                                'Monthly achievement'
                            ]
                        }
                    ]
                    
                    today = datetime.now(IST)
                    
                    with st.spinner("Sending manager alert..."):
                        success = send_manager_team_reminder_email(
                            mgr_obj['email'],
                            mgr_obj['name'],
                            test_emp_name,
                            test_incomplete_goals,
                            today.month,
                            today.year,
                            hr_emails
                        )
                    
                    if success:
                        st.success(f"✅ Manager alert sent to {mgr_obj['email']}")
                        st.info(f"👤 About employee: {test_emp_name}")
                        if hr_emails:
                            st.info(f"📋 CC sent to HR: {', '.join(hr_emails)}")
                    else:
                        st.error("❌ Failed to send email")
        
        st.markdown("---")
        
        # Test 3: Run Full Check
        st.markdown("#### 3️⃣ Test Full System Check")
        st.caption("⚠️ This will check ALL users and send actual reminders for incomplete goal sheets")
        st.warning("**Warning:** This will send real emails to all employees and managers with incomplete goal sheets!")
        
        col_full1, col_full2 = st.columns([3, 1])
        
        with col_full1:
            confirm_full_test = st.checkbox(
                "I understand this will send real emails to users",
                key="confirm_full_test"
            )
        
        with col_full2:
            if st.button(
                "🚀 Run Check",
                disabled=not confirm_full_test,
                use_container_width=True,
                key="run_full_check"
            ):
                with st.spinner("Running full goal sheet completion check..."):
                    from monthly_reminder import check_and_send_reminders
                    check_and_send_reminders(db)
                st.success("✅ Full check completed! Check console/logs for details.")
    
    # ===== TAB 3: CUSTOM EMAIL TEST =====
    with test_tab3:
        st.markdown("**Send Test Email to Custom Address**")
        st.info("💡 Send any of the reminder emails to a custom email address for testing")
        
        custom_email = st.text_input(
            "📧 Enter Test Email Address",
            placeholder="test@example.com",
            key="custom_test_email"
        )
        
        test_type = st.radio(
            "Select Email Type to Test",
            [
                "Employee Goal Sheet Reminder",
                "Manager Team Alert"
            ],
            key="custom_test_type"
        )
        
        if st.button("📤 Send Custom Test", use_container_width=True, key="send_custom_test"):
            if custom_email and '@' in custom_email:
                today = datetime.now(IST)
                
                # Create sample data
                test_incomplete_goals = [
                    {
                        'goal_title': 'Test Goal - Monthly Sales Target',
                        'goal_id': 'test_1',
                        'department': 'Sales Department',
                        'kpi': 'Revenue Achievement',
                        'monthly_target': 100000,
                        'missing_items': [
                            'Week 1 achievement',
                            'Week 2 rating',
                            'Week 3 achievement',
                            'Week 4 rating'
                        ]
                    }
                ]
                
                with st.spinner(f"Sending {test_type} to {custom_email}..."):
                    if test_type == "Employee Goal Sheet Reminder":
                        from monthly_reminder import send_goal_sheet_reminder_email
                        success = send_goal_sheet_reminder_email(
                            custom_email,
                            "Test Employee",
                            test_incomplete_goals,
                            today.month,
                            today.year,
                            hr_emails
                        )
                    else:  # Manager Team Alert
                        from monthly_reminder import send_manager_team_reminder_email
                        success = send_manager_team_reminder_email(
                            custom_email,
                            "Test Manager",
                            "Test Employee Name",
                            test_incomplete_goals,
                            today.month,
                            today.year,
                            hr_emails
                        )
                
                if success:
                    st.success(f"✅ Test email sent to {custom_email}")
                    if hr_emails:
                        st.info(f"📋 CC sent to HR: {', '.join(hr_emails)}")
                else:
                    st.error("❌ Failed to send email")
            else:
                st.error("❌ Please enter a valid email address")
        
        st.markdown("---")
        st.caption("💡 **Tip:** Use your personal email to see how the emails look before sending to real users")

def display_approval_page():
    """Display pending goal and achievement approvals for managers"""
    user = st.session_state.user
    
    if user['role'] not in ['Manager', 'HR', 'VP']:
        st.warning("⚠️ Only Managers can access this page")
        return
    
    st.title("✅ Approval Dashboard")
    
    # Create tabs for different approval types
    tab1, tab2 = st.tabs(["📋 Goal Approvals", "📊 Achievement Approvals"])
    
    # ===== TAB 1: GOAL APPROVALS =====
    with tab1:
        st.subheader("Pending Goal Approvals")
        
        # Get pending goal approvals
        pending_goals = db.get_pending_approvals(user['id'])
        
        if not pending_goals:
            st.success("🎉 No pending goal approvals!")
        else:
            st.markdown(f"**You have {len(pending_goals)} goal(s) awaiting approval**")
            st.markdown("---")
            
            for goal in pending_goals:
                # Get employee info
                employee = db.get_user_by_id(goal['user_id'])
                employee_name = employee['name'] if employee else 'Unknown'
                
                with st.expander(f"📋 {goal['goal_title']} - {employee_name}"):
                    col1, col2 = st.columns([2, 1])
                    
                    with col1:
                        st.markdown(f"**Employee:** {employee_name}")
                        st.markdown(f"**Department:** {goal.get('department', 'N/A')}")
                        st.markdown(f"**KPI:** {goal.get('kpi', 'N/A')}")
                        st.markdown(f"**Monthly Target:** {goal.get('monthly_target', 0)}")
                        st.markdown(f"**Period:** {goal.get('year')}-Q{goal.get('quarter')}-M{goal.get('month')}")
                        
                        # Weekly Targets Section
                        st.markdown("**Weekly Breakdown:**")
                        col_w1, col_w2, col_w3, col_w4 = st.columns(4)
                        
                        with col_w1:
                            st.metric("Week 1", goal.get('week1_target', 0))
                        with col_w2:
                            st.metric("Week 2", goal.get('week2_target', 0))
                        with col_w3:
                            st.metric("Week 3", goal.get('week3_target', 0))
                        with col_w4:
                            st.metric("Week 4", goal.get('week4_target', 0))
                        
                        st.markdown(f"**Description:** {goal.get('goal_description', 'No description')}")
                    
                    with col2:
                        st.markdown("**Actions:**")
                        
                        if st.button("✅ Approve", key=f"approve_goal_{goal['goal_id']}", use_container_width=True):
                            if db.update_goal_approval(goal['goal_id'], 'approved', user['id']):
                                st.success("✅ Goal approved!")
                                
                                # Notify employee and HR
                                employee = db.get_user_by_id(goal['user_id'])
                                if employee:
                                    notify_goal_approved(goal, user, employee)
                                
                                st.rerun()
                            else:
                                st.error("❌ Failed to approve goal")
                        
                        if st.button("❌ Reject", key=f"reject_goal_{goal['goal_id']}", use_container_width=True):
                            st.session_state[f'rejecting_goal_{goal["goal_id"]}'] = True
                        
                        # Rejection reason form
                        if st.session_state.get(f'rejecting_goal_{goal["goal_id"]}'):
                            with st.form(f"reject_goal_form_{goal['goal_id']}"):
                                reason = st.text_area("Rejection Reason*", key=f"reason_goal_{goal['goal_id']}")
                                
                                col_r1, col_r2 = st.columns(2)
                                with col_r1:
                                    if st.form_submit_button("Confirm Reject"):
                                        if reason.strip():
                                            if db.update_goal_approval(
                                                goal['goal_id'], 
                                                'rejected', 
                                                user['id'],
                                                reason
                                            ):
                                                st.success("Goal rejected")
                                                del st.session_state[f'rejecting_goal_{goal["goal_id"]}']
                                                st.rerun()
                                        else:
                                            st.error("Please provide a reason")
                                
                                with col_r2:
                                    if st.form_submit_button("Cancel"):
                                        del st.session_state[f'rejecting_goal_{goal["goal_id"]}']
                                        st.rerun()
                    
                    st.markdown("---")
    
    # ===== TAB 2: ACHIEVEMENT APPROVALS =====
    with tab2:
        st.subheader("Pending Achievement Approvals")
        
        # Get team members
        team_members = db.get_team_members(user['id'])
        
        pending_achievements = []
        for member in team_members:
            goals = db.get_user_all_goals(member['id'])
            for goal in goals:
                if goal.get('achievement_approval_status') == 'pending':
                    goal['employee'] = member
                    pending_achievements.append(goal)
        
        if not pending_achievements:
            st.success("🎉 No pending achievement approvals!")
        else:
            st.markdown(f"**You have {len(pending_achievements)} achievement update(s) awaiting approval**")
            st.markdown("---")
            
            for goal in pending_achievements:
                employee = goal['employee']
                
                with st.expander(f"📊 {goal['goal_title']} - {employee['name']}"):
                    col1, col2 = st.columns([2, 1])
                    
                    with col1:
                        st.markdown(f"**Employee:** {employee['name']}")
                        st.markdown(f"**Goal:** {goal['goal_title']}")
                        st.markdown(f"**Department:** {goal.get('department', 'N/A')}")
                        st.markdown(f"**Monthly Target:** {goal.get('monthly_target', 0)}")
                        
                        # Show pending achievements
                        st.markdown("**Reported Achievements:**")
                        
                        col_w1, col_w2, col_w3, col_w4 = st.columns(4)
                        with col_w1:
                            w1_val = goal.get('week1_achievement_pending')
                            st.metric("Week 1", w1_val if w1_val is not None else '-')
                        with col_w2:
                            w2_val = goal.get('week2_achievement_pending')
                            st.metric("Week 2", w2_val if w2_val is not None else '-')
                        with col_w3:
                            w3_val = goal.get('week3_achievement_pending')
                            st.metric("Week 3", w3_val if w3_val is not None else '-')
                        with col_w4:
                            w4_val = goal.get('week4_achievement_pending')
                            st.metric("Week 4", w4_val if w4_val is not None else '-')
                        
                        monthly_pending = goal.get('monthly_achievement_pending', 0)
                        st.markdown(f"**Monthly Total:** {monthly_pending if monthly_pending is not None else 0}")
                        
                        # Show submission time
                        updated_at = goal.get('achievement_updated_at')
                        if updated_at:
                            st.caption(f"Submitted: {updated_at[:16]}")
                    
                    with col2:
                        st.markdown("**Actions:**")
                        
                        if st.button("✅ Approve", key=f"approve_ach_{goal['goal_id']}", use_container_width=True):
                            # Move pending to actual
                            updates = {
                                'week1_achievement': goal.get('week1_achievement_pending'),
                                'week2_achievement': goal.get('week2_achievement_pending'),
                                'week3_achievement': goal.get('week3_achievement_pending'),
                                'week4_achievement': goal.get('week4_achievement_pending'),
                                'monthly_achievement': goal.get('monthly_achievement_pending'),
                                'week1_rating': goal.get('week1_rating_pending'),
                                'week2_rating': goal.get('week2_rating_pending'),
                                'week3_rating': goal.get('week3_rating_pending'),
                                'week4_rating': goal.get('week4_rating_pending'),
                                'achievement_approval_status': 'approved',
                                'achievement_approved_by': user['id'],
                                'achievement_approved_at': datetime.now(IST).isoformat()
                            }
                            
                            if db.update_goal(goal['goal_id'], updates):
                                st.success("✅ Achievements approved!")
                                
                                # Notify employee
                                create_notification({
                                    'user_id': employee['id'],
                                    'action_by': user['id'],
                                    'action_by_name': user['name'],
                                    'action_type': 'achievement_approved',
                                    'details': f"Your achievements for '{goal['goal_title']}' have been approved",
                                    'is_read': False,
                                    'created_at': datetime.now(IST).isoformat()
                                })
                                
                                st.rerun()
                            else:
                                st.error("❌ Failed to approve achievements")
                        
                        if st.button("❌ Reject", key=f"reject_ach_{goal['goal_id']}", use_container_width=True):
                            st.session_state[f'rejecting_ach_{goal["goal_id"]}'] = True
                        
                        # Rejection form
                        if st.session_state.get(f'rejecting_ach_{goal["goal_id"]}'):
                            with st.form(f"reject_ach_form_{goal['goal_id']}"):
                                reason = st.text_area("Rejection Reason*", key=f"reason_ach_{goal['goal_id']}")
                                
                                col_r1, col_r2 = st.columns(2)
                                with col_r1:
                                    if st.form_submit_button("Confirm Reject"):
                                        if reason.strip():
                                            updates = {
                                                'achievement_approval_status': 'rejected',
                                                'achievement_rejection_reason': reason,
                                                'achievement_rejected_by': user['id'],
                                                'achievement_rejected_at': datetime.now(IST).isoformat()
                                            }
                                            
                                            if db.update_goal(goal['goal_id'], updates):
                                                st.success("Achievements rejected")
                                                
                                                # Notify employee
                                                create_notification({
                                                    'user_id': employee['id'],
                                                    'action_by': user['id'],
                                                    'action_by_name': user['name'],
                                                    'action_type': 'achievement_rejected',
                                                    'details': f"Your achievements for '{goal['goal_title']}' were rejected. Reason: {reason}",
                                                    'is_read': False,
                                                    'created_at': datetime.now(IST).isoformat()
                                                })
                                                
                                                del st.session_state[f'rejecting_ach_{goal["goal_id"]}']
                                                st.rerun()
                                        else:
                                            st.error("Please provide a reason")
                                
                                with col_r2:
                                    if st.form_submit_button("Cancel"):
                                        del st.session_state[f'rejecting_ach_{goal["goal_id"]}']
                                        st.rerun()
                    
                    st.markdown("---")
# ============================================
# SIDEBAR
# ============================================
def render_sidebar():
    """Render sidebar with navigation"""
    user = st.session_state.user
    role = user['role']
    
    with st.sidebar:
        # User profile with role indicator
        render_user_avatar(user)
        

        st.markdown("---")
        
        # Navigation Menu
        
        if st.button("Dashboard", use_container_width=True, key="nav_dashboard"):
            st.session_state.page = 'dashboard'
            st.session_state.pop('viewing_employee', None)
            st.session_state.pop('viewing_employee_year', None)
            save_session_to_storage()
            st.rerun()
        
        if st.button("My Goals", use_container_width=True, key="nav_my_goals"):
            st.session_state.page = 'my_goals'
            st.session_state.pop('viewing_employee', None)
            st.session_state.pop('viewing_employee_year', None)
            save_session_to_storage()
            st.rerun()
        
        
        # Role-specific navigation
        
        if role in ['CMD', 'VP','HR', 'Manager']:
            if st.button("My Team", use_container_width=True, key="nav_employees"):
                st.session_state.page = 'employees'
                st.session_state.pop('viewing_employee', None)
                st.session_state.pop('viewing_employee_year', None)
                save_session_to_storage()
                st.rerun()
        
        if role in ['Manager']:
            # Get pending count
            pending_count = len(db.get_pending_approvals(user['id']))
            
            approval_label = f" Approvals ({pending_count})" if pending_count > 0 else " Approvals"
            
            if st.button(approval_label, use_container_width=True, key="nav_approvals"):
                st.session_state.page = 'approvals'
                save_session_to_storage()
                st.rerun()

        if role in 'HR':
            if st.button("Employee Management", use_container_width=True, key="nav_emp_mgmt"):
                st.session_state.page = 'employee_management'
                st.session_state.pop('viewing_employee', None)
                st.session_state.pop('viewing_employee_year', None)
                save_session_to_storage()
                st.rerun()
            
            if st.button("Organization Info", use_container_width=True, key="nav_hr_info"):
                st.session_state.page = 'hr_info'
                st.session_state.pop('viewing_employee', None)
                st.session_state.pop('viewing_employee_year', None)
                save_session_to_storage()
                st.rerun()
        
        if st.button("Analytics", use_container_width=True, key="nav_analytics"):
            st.session_state.page = 'analytics'
            save_session_to_storage()
            st.rerun()
        
        if st.button("Feedback History", use_container_width=True, key="nav_feedback"):
            st.session_state.page = 'feedback_history'
            st.session_state.pop('viewing_employee', None)
            st.session_state.pop('viewing_employee_year', None)
            save_session_to_storage()
            st.rerun()
        
        
        # Settings
        st.markdown("---")
        st.markdown("**Settings**")
        
        if st.button("Profile", use_container_width=True, key="nav_profile"):
            st.session_state.page = 'profile'
            save_session_to_storage()
            st.rerun()
        
        # Logout
        st.markdown("---")
        if st.button("Logout", use_container_width=True):
            st.query_params.clear()
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.rerun()
# ============================================
# MAIN APPLICATION
# ============================================
def main():

    # ✅ NEW: Handle approval/rejection from email links
    query_params = st.query_params
    
    if 'action' in query_params and 'token' in query_params:
        action = query_params['action']
        token = query_params['token']
        
        # Get goal by token
        token_data = db.get_goal_by_token(token)
        
        if token_data:
            goal = token_data.get('goals')
            
            if action == 'approve':
                # Auto-login as manager or show login
                st.success(f"✅ Approving goal: {goal.get('goal_title')}")
                
                # You need to handle authentication here
                # Option 1: Auto-login with token
                # Option 2: Redirect to login page with action stored
                
                if db.update_goal_approval(goal['goal_id'], 'approved', token_data['approved_by']):
                    db.mark_token_used(token)
                    st.success("✅ Goal approved successfully!")
                    st.balloons()
                    
            elif action == 'reject':
                st.warning(f"❌ Rejecting goal: {goal.get('goal_title')}")
                
                reason = st.text_area("Please provide a rejection reason:")
                if st.button("Confirm Rejection"):
                    if reason.strip():
                        if db.update_goal_approval(goal['goal_id'], 'rejected', token_data['approved_by'], reason):
                            db.mark_token_used(token)
                            st.success("Goal rejected")
        else:
            st.error("❌ Invalid or expired approval link")
        
        # Clear query params
        st.query_params.clear()
        return
   
    if not st.session_state.user:
        login_page()
        return
    
    # Render sidebar
    render_sidebar()
    
    # Route to appropriate page
    page = st.session_state.get('page', 'dashboard')
    
    if page == 'dashboard':
        display_dashboard()
    elif page == 'my_goals':
        display_my_goals()
    elif page == 'view_all_goals':
        display_view_all_goals()
    elif page == 'quarters':
        display_quarter_selection()
    elif page == 'months':
        display_month_selection()
    elif page == 'month_goals':
        display_month_goals()
    elif page == 'employees':
        display_employees_page()
    elif page == 'vp_team_view':  # ADD THIS
        display_vp_team_view()
    elif page == 'hr_team_view':  # ADD THIS
        display_hr_team_view()
    elif page == 'manager_team_view':  # ADD THIS
        display_manager_team_view()
    elif page == 'employee_goals':
        display_employee_goals()
    elif page == 'employee_quarters':
        display_quarter_selection()
    elif page == 'employee_months':
        display_month_selection()
    elif page == 'employee_month_goals':
        display_month_goals()
    elif page == 'employee_management':
        display_employee_management()
    elif page == 'hr_info':
        display_hr_info()
    elif page == 'approvals':
        display_approval_page()
    elif page == 'analytics':  # ADD THIS
        display_analytics_page()
    elif page == 'feedback_history':
        display_feedback_history()
    elif page == 'profile':
        display_profile()
    elif page == 'permissions':
        display_permissions()


    else:
        # Default to dashboard
        st.session_state.page = 'dashboard'
        st.rerun()
        
if __name__ == "__main__":
    main()