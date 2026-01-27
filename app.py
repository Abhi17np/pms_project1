"""
Performance Management System - Main Application
A comprehensive PMS built with Streamlit and Supabase
"""
import sys
sys.setrecursionlimit(3000)

import streamlit as st
import pandas as pd
from datetime import datetime, date, timedelta
import pytz
import plotly.express as px
import plotly.graph_objects as go
import calendar
import io
from plotly.subplots import make_subplots
from datetime import datetime, timedelta
import numpy as np
from io import BytesIO
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from dotenv import load_dotenv
from dateutil.relativedelta import relativedelta
from database import get_supabase_client
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import secrets
import os
from datetime import datetime, timedelta
from monthly_reminder import start_reminder_scheduler, test_send_reminder
from monthly_reminder import send_goal_completion_email
import streamlit as st
import streamlit.components.v1 as components
import textwrap

def metric_card(
    value,
    label,
    color,
    shadow_rgba,
    icon=None,
    suffix="",
    label_bg=None
):
    import streamlit as st

    icon_html = f'<div style="font-size:34px; margin-bottom:10px;">{icon}</div>' if icon else ""
    label_bg_css = f"background:{label_bg};" if label_bg else ""

    html = (
        '<div style="padding:10px;">'
        '<div style="'
        'background:#FFFFFF;'
        'width:120%;'
        'height:170px;'
        'padding:20px;'
        'border-radius:14px;'
        'display:flex;'
        'text-align:center;'
        'flex-direction:column;'
        'justify-content:center;'
        'align-items:center;'
        'box-shadow:0 4px 10px rgba(0,0,0,0.06),'
        f'0 0 24px {shadow_rgba};'
        '">'
        f'{icon_html}'
        '<div style="'
        'font-size:36px;'
        'font-weight:700;'
        f'color:{color};'
        'margin-bottom:8px;'
        'text-align:center;'
        '">'
        f'{value}{suffix}'
        '</div>'
        '<div style="'
        'font-size:13px;'
        'font-weight:700;'
        'text-transform:uppercase;'
        'letter-spacing:0.4px;'
        f'color:{color};'
        f'{label_bg_css}'
        'padding:6px 14px;'
        'border-radius:8px;'
        'white-space:nowrap;'      # ✅ keeps text in one line
        'text-align:center;'       # ✅ centers text
        'overflow:hidden;'         # ✅ prevents wrap glitches
        '">'
        f'{label}'
        '</div>'
        '</div>'
        '</div>'
    )

    st.markdown(html, unsafe_allow_html=True)


def get_fiscal_quarter(month):
    """
    Calculate fiscal quarter based on month
    Q1: April-June (4-6)
    Q2: July-September (7-9)
    Q3: October-December (10-12)
    Q4: January-March (1-3)
    """
    if 4 <= month <= 6:
        return 1
    elif 7 <= month <= 9:
        return 2
    elif 10 <= month <= 12:
        return 3
    else:  # 1-3
        return 4
    
def safe_float(value, default=None):  # Changed default from 0 to None
    """Safely convert value to float, handling None and invalid types"""
    if value is None or value == '':
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default



def normalize_department(dept_name):
    """Normalize department name to be case-insensitive and handle variations"""
    if not dept_name or dept_name.strip() == '':
        return 'Unassigned'
    # Convert to uppercase and strip whitespace
    normalized = dept_name.strip().upper()
    # Handle common variations
    if normalized in ['N/A', 'NA', 'NONE']:
        return 'Unassigned'
    return normalized   
def get_feedback_giver_role(target_role):
    """Who is allowed to give official feedback to this role?"""
    mapping = {
        'VP': 'CMD',
        'HR': 'VP',
        'Manager': 'VP',
        'Employee': 'Manager',
        'CMD': None
    }
    return mapping.get(target_role)

def get_role_hierarchy():
    """Define role hierarchy levels"""
    return {
        'CMD': 5,
        'VP': 4,
        'HR': 3,
        'Manager': 2,
        'Employee': 1
    }

def create_notification(notification_data):
    """Create a new notification in the database"""
    try:
        # Ensure timezone-aware timestamp
        if 'created_at' not in notification_data:
            notification_data['created_at'] = datetime.now(IST).isoformat()
        
        result = supabase.table('notifications').insert(notification_data).execute()
        
        if result.data and len(result.data) > 0:
            print(f" Notification created: {notification_data.get('action_type')}")
            return result.data[0]
        else:
            print(f" Failed to create notification: No data returned")
            return None
    except Exception as e:
        print(f"  Error creating notification: {str(e)}")
        return None

def notify_goal_created(goal_data, creator_user):
    """Notify relevant users when a goal is created"""
    creator_role = creator_user['role']
    
    # Employee creates goal -> Notify Manager (pending approval)
    if creator_role == 'Employee':
        if creator_user.get('manager_id'):
            manager = db.get_user_by_id(creator_user['manager_id'])
            if manager:
                create_notification({
                    'user_id': manager['id'],
                    'action_by': creator_user['id'],
                    'action_by_name': creator_user['name'],
                    'action_type': 'goal_created',
                    'details': f"{creator_user['name']} created a new goal: '{goal_data['goal_title']}' (Pending Approval)",
                    'is_read': False,
                    'created_at': datetime.now(IST).isoformat()
                })
    
    # Manager creates goal -> Notify VP
    elif creator_role == 'Manager':
        all_users = db.get_all_users()
        vp_users = [u for u in all_users if u['role'] == 'VP']
        for vp in vp_users:
            create_notification({
                'user_id': vp['id'],
                'action_by': creator_user['id'],
                'action_by_name': creator_user['name'],
                'action_type': 'goal_created',
                'details': f"Manager {creator_user['name']} created goal: '{goal_data['goal_title']}'",
                'is_read': False,
                'created_at': datetime.now(IST).isoformat()
            })
    
    # HR creates goal -> Notify VP
    elif creator_role == 'HR':
        all_users = db.get_all_users()
        vp_users = [u for u in all_users if u['role'] == 'VP']
        for vp in vp_users:
            create_notification({
                'user_id': vp['id'],
                'action_by': creator_user['id'],
                'action_by_name': creator_user['name'],
                'action_type': 'goal_created',
                'details': f"HR {creator_user['name']} created goal: '{goal_data['goal_title']}'",
                'is_read': False,
                'created_at': datetime.now(IST).isoformat()
            })
    
    # VP creates goal -> Notify CMD
    elif creator_role == 'VP':
        all_users = db.get_all_users()
        cmd_users = [u for u in all_users if u['role'] == 'CMD']
        for cmd in cmd_users:
            create_notification({
                'user_id': cmd['id'],
                'action_by': creator_user['id'],
                'action_by_name': creator_user['name'],
                'action_type': 'goal_created',
                'details': f"VP {creator_user['name']} created goal: '{goal_data['goal_title']}'",
                'is_read': False,
                'created_at': datetime.now(IST).isoformat()
            })

def notify_goal_approved(goal, approver_user, employee_user):
    """Notify employee when their goal is approved"""
    create_notification({
        'user_id': employee_user['id'],
        'action_by': approver_user['id'],
        'action_by_name': approver_user['name'],
        'action_type': 'goal_approved',
        'details': f"Your goal '{goal['goal_title']}' has been approved by {approver_user['name']}",
        'is_read': False,
        'created_at': datetime.now(IST).isoformat()
    })
    
    # Also notify HR that an approved goal exists
    all_users = db.get_all_users()
    hr_users = [u for u in all_users if u['role'] == 'HR']
    for hr in hr_users:
        create_notification({
            'user_id': hr['id'],
            'action_by': employee_user['id'],
            'action_by_name': employee_user['name'],
            'action_type': 'goal_created',
            'details': f"{employee_user['name']}'s goal '{goal['goal_title']}' was approved by {approver_user['name']}",
            'is_read': False,
            'created_at': datetime.now(IST).isoformat()
        })
        
def notify_weekly_achievement_updated(goal, updater_user, week_num):
    """Notify relevant users when weekly achievement is updated"""
    updater_role = updater_user['role']
    goal_owner = db.get_user_by_id(goal['user_id'])
    
    # Employee updates -> Notify Manager AND HR
    if updater_role == 'Employee':
        # Notify Manager
        if goal_owner.get('manager_id'):
            manager = db.get_user_by_id(goal_owner['manager_id'])
            if manager:
                create_notification({
                    'user_id': manager['id'],
                    'action_by': updater_user['id'],
                    'action_by_name': updater_user['name'],
                    'action_type': 'weekly_achievement_updated',
                    'details': f"{updater_user['name']} updated Week {week_num} achievement for '{goal['goal_title']}'",
                    'is_read': False,
                    'created_at': datetime.now(IST).isoformat()
                })
        
        # ✅ ALWAYS notify HR
        all_users = db.get_all_users()
        hr_users = [u for u in all_users if u['role'] == 'HR']
        for hr in hr_users:
            create_notification({
                'user_id': hr['id'],
                'action_by': updater_user['id'],
                'action_by_name': updater_user['name'],
                'action_type': 'weekly_achievement_updated',
                'details': f"{updater_user['name']} updated Week {week_num} achievement for '{goal['goal_title']}'",
                'is_read': False,
                'created_at': datetime.now(IST).isoformat()
            })
    
    # Manager/HR updates -> Notify VP
    elif updater_role in ['Manager', 'HR']:
        all_users = db.get_all_users()
        vp_users = [u for u in all_users if u['role'] == 'VP']
        for vp in vp_users:
            create_notification({
                'user_id': vp['id'],
                'action_by': updater_user['id'],
                'action_by_name': updater_user['name'],
                'action_type': 'weekly_achievement_updated',
                'details': f"{updater_role} {updater_user['name']} updated Week {week_num} achievement for '{goal['goal_title']}'",
                'is_read': False,
                'created_at': datetime.now(IST).isoformat()
            })
    
    # VP updates -> Notify CMD
    elif updater_role == 'VP':
        all_users = db.get_all_users()
        cmd_users = [u for u in all_users if u['role'] == 'CMD']
        for cmd in cmd_users:
            create_notification({
                'user_id': cmd['id'],
                'action_by': updater_user['id'],
                'action_by_name': updater_user['name'],
                'action_type': 'weekly_achievement_updated',
                'details': f"VP {updater_user['name']} updated Week {week_num} achievement for '{goal['goal_title']}'",
                'is_read': False,
                'created_at': datetime.now(IST).isoformat()
            })

def render_clickable_metric_card(label, value, color, icon, key, detail_type):
    """Render a clickable metric card that acts as a button"""
    if st.button(
        f"{icon}\n\n**{value}**\n\n{label}",
        key=key,
        use_container_width=True,
        help=f"Click to view {label.lower()}"
    ):
        st.session_state.show_details = detail_type
        st.rerun()

def notify_goal_completed(goal, completer_user):
    """Notify relevant users when a goal is completed"""
    completer_role = completer_user['role']
    goal_owner = db.get_user_by_id(goal['user_id'])
    
    # Employee completes -> Notify Manager AND HR
    if completer_role == 'Employee':
        # Notify Manager
        if goal_owner.get('manager_id'):
            manager = db.get_user_by_id(goal_owner['manager_id'])
            if manager:
                create_notification({
                    'user_id': manager['id'],
                    'action_by': completer_user['id'],
                    'action_by_name': completer_user['name'],
                    'action_type': 'goal_completed',
                    'details': f" {completer_user['name']} completed goal: '{goal['goal_title']}'",
                    'is_read': False,
                    'created_at': datetime.now(IST).isoformat()
                })
        
        # ✅ ALWAYS notify HR
        all_users = db.get_all_users()
        hr_users = [u for u in all_users if u['role'] == 'HR']
        for hr in hr_users:
            create_notification({
                'user_id': hr['id'],
                'action_by': completer_user['id'],
                'action_by_name': completer_user['name'],
                'action_type': 'goal_completed',
                'details': f" {completer_user['name']} completed goal: '{goal['goal_title']}'",
                'is_read': False,
                'created_at': datetime.now(IST).isoformat()
            })
    
    # Manager/HR completes -> Notify VP
    elif completer_role in ['Manager', 'HR']:
        all_users = db.get_all_users()
        vp_users = [u for u in all_users if u['role'] == 'VP']
        for vp in vp_users:
            create_notification({
                'user_id': vp['id'],
                'action_by': completer_user['id'],
                'action_by_name': completer_user['name'],
                'action_type': 'goal_completed',
                'details': f" {completer_role} {completer_user['name']} completed goal: '{goal['goal_title']}'",
                'is_read': False,
                'created_at': datetime.now(IST).isoformat()
            })
    
    # VP completes -> Notify CMD
    elif completer_role == 'VP':
        all_users = db.get_all_users()
        cmd_users = [u for u in all_users if u['role'] == 'CMD']
        for cmd in cmd_users:
            create_notification({
                'user_id': cmd['id'],
                'action_by': completer_user['id'],
                'action_by_name': completer_user['name'],
                'action_type': 'goal_completed',
                'details': f" VP {completer_user['name']} completed goal: '{goal['goal_title']}'",
                'is_read': False,
                'created_at': datetime.now(IST).isoformat()
            })

def notify_goal_not_completed(goal, goal_owner):
    """Notify when goal is not completed by end date"""
    owner_role = goal_owner['role']
    
    # Notify the goal owner themselves
    create_notification({
        'user_id': goal_owner['id'],
        'action_by': goal_owner['id'],
        'action_by_name': 'System',
        'action_type': 'goal_not_completed',
        'details': f"⚠️ Your goal '{goal['goal_title']}' was not completed by the deadline",
        'is_read': False,
        'created_at': datetime.now(IST).isoformat()
    })
    
    # Employee -> Notify Manager
    if owner_role == 'Employee':
        if goal_owner.get('manager_id'):
            manager = db.get_user_by_id(goal_owner['manager_id'])
            if manager:
                create_notification({
                    'user_id': manager['id'],
                    'action_by': goal_owner['id'],
                    'action_by_name': goal_owner['name'],
                    'action_type': 'goal_not_completed',
                    'details': f"⚠️ {goal_owner['name']}'s goal '{goal['goal_title']}' was not completed by deadline",
                    'is_read': False,
                    'created_at': datetime.now(IST).isoformat()
                })
        
        # Notify HR
        all_users = db.get_all_users()
        hr_users = [u for u in all_users if u['role'] == 'HR']
        for hr in hr_users:
            create_notification({
                'user_id': hr['id'],
                'action_by': goal_owner['id'],
                'action_by_name': goal_owner['name'],
                'action_type': 'goal_not_completed',
                'details': f"⚠️ {goal_owner['name']}'s goal '{goal['goal_title']}' was not completed by deadline",
                'is_read': False,
                'created_at': datetime.now(IST).isoformat()
            })
    
    # Manager/HR -> Notify VP
    elif owner_role in ['Manager', 'HR']:
        all_users = db.get_all_users()
        vp_users = [u for u in all_users if u['role'] == 'VP']
        for vp in vp_users:
            create_notification({
                'user_id': vp['id'],
                'action_by': goal_owner['id'],
                'action_by_name': goal_owner['name'],
                'action_type': 'goal_not_completed',
                'details': f"⚠️ {owner_role} {goal_owner['name']}'s goal '{goal['goal_title']}' was not completed by deadline",
                'is_read': False,
                'created_at': datetime.now(IST).isoformat()
            })
    
    # VP -> Notify CMD
    elif owner_role == 'VP':
        all_users = db.get_all_users()
        cmd_users = [u for u in all_users if u['role'] == 'CMD']
        for cmd in cmd_users:
            create_notification({
                'user_id': cmd['id'],
                'action_by': goal_owner['id'],
                'action_by_name': goal_owner['name'],
                'action_type': 'goal_not_completed',
                'details': f"⚠️ VP {goal_owner['name']}'s goal '{goal['goal_title']}' was not completed by deadline",
                'is_read': False,
                'created_at': datetime.now(IST).isoformat()
            })

def notify_feedback_given(goal, feedback_giver, goal_owner):
    """Notify when feedback is given"""
    giver_role = feedback_giver['role']
    owner_role = goal_owner['role']
    
    # Always notify the goal owner
    create_notification({
        'user_id': goal_owner['id'],
        'action_by': feedback_giver['id'],
        'action_by_name': feedback_giver['name'],
        'action_type': 'feedback_received',
        'details': f"{feedback_giver['name']} gave feedback on your goal: '{goal['goal_title']}'",
        'is_read': False,
        'created_at': datetime.now(IST).isoformat()
    })
    
    # Manager gives feedback to Employee -> Notify HR
    if giver_role == 'Manager' and owner_role == 'Employee':
        all_users = db.get_all_users()
        hr_users = [u for u in all_users if u['role'] == 'HR']
        for hr in hr_users:
            create_notification({
                'user_id': hr['id'],
                'action_by': feedback_giver['id'],
                'action_by_name': feedback_giver['name'],
                'action_type': 'feedback_given',
                'details': f"Manager {feedback_giver['name']} gave feedback to {goal_owner['name']} on '{goal['goal_title']}'",
                'is_read': False,
                'created_at': datetime.now(IST).isoformat()
            })
    
    # VP gives feedback to Manager/HR -> Notify them
    elif giver_role == 'VP' and owner_role in ['Manager', 'HR']:
        create_notification({
            'user_id': goal_owner['id'],
            'action_by': feedback_giver['id'],
            'action_by_name': feedback_giver['name'],
            'action_type': 'feedback_received',
            'details': f"VP {feedback_giver['name']} gave feedback on your goal: '{goal['goal_title']}'",
            'is_read': False,
            'created_at': datetime.now(IST).isoformat()
        })
    
    # CMD gives feedback to VP -> Notify VP
    elif giver_role == 'CMD' and owner_role == 'VP':
        create_notification({
            'user_id': goal_owner['id'],
            'action_by': feedback_giver['id'],
            'action_by_name': feedback_giver['name'],
            'action_type': 'feedback_received',
            'details': f"CMD {feedback_giver['name']} gave feedback on your goal: '{goal['goal_title']}'",
            'is_read': False,
            'created_at': datetime.now(IST).isoformat()
        })

def notify_feedback_reply(feedback, replier_user, original_feedback_giver):
    """Notify when someone replies to feedback"""
    replier_role = replier_user['role']
    
    # Notify the original feedback giver
    create_notification({
        'user_id': original_feedback_giver['id'],
        'action_by': replier_user['id'],
        'action_by_name': replier_user['name'],
        'action_type': 'feedback_reply',
        'details': f"{replier_user['name']} replied to your feedback",
        'is_read': False,
        'created_at': datetime.now(IST).isoformat()
    })
    
    # Employee replies to Manager feedback -> Manager already notified above
    # VP replies to CMD feedback -> CMD already notified above
    # CMD replies to VP feedback -> VP already notified above
def cleanup_duplicate_notifications():
    """Remove duplicate notifications from database"""
    st.title("🧹 Clean Up Duplicate Notifications")
    
    user = st.session_state.user
    if user['role'] not in ['HR', 'CMD']:
        st.warning("⚠️ Only HR/CMD can clean up notifications")
        return
    
    st.warning("⚠️ This will remove duplicate notifications from the database")
    
    if st.button("🗑️ Clean Up Duplicates", type="primary"):
        try:
            # Get all notifications
            all_notifs = supabase.table('notifications').select('*').order(
                'created_at', desc=True
            ).execute()
            
            if all_notifs.data:
                seen = {}
                duplicates = []
                
                for notif in all_notifs.data:
                    # Create unique key
                    key = f"{notif['user_id']}_{notif['action_type']}_{notif.get('details', '')[:50]}"
                    
                    if key in seen:
                        duplicates.append(notif['id'])
                    else:
                        seen[key] = notif['id']
                
                st.info(f"Found {len(duplicates)} duplicate notifications")
                
                if duplicates:
                    # Delete duplicates
                    for dup_id in duplicates:
                        supabase.table('notifications').delete().eq('id', dup_id).execute()
                    
                    st.success(f"✅ Removed {len(duplicates)} duplicate notifications!")
                else:
                    st.success("✅ No duplicates found!")
            
        except Exception as e:
            st.error(f"Error: {str(e)}")
    
    # Show all notifications
    st.markdown("---")
    st.subheader(" All System Notifications")
    
    try:
        all_notifs = supabase.table('notifications').select('*').order(
            'created_at', desc=True
        ).limit(100).execute()
        
        if all_notifs.data:
            st.dataframe(pd.DataFrame(all_notifs.data), use_container_width=True, height=400)
        else:
            st.info("No notifications in database")
    except Exception as e:
        st.error(f"Error: {str(e)}")

def notify_goal_edited(goal, editor_user, goal_owner):
    """Notify when a goal is edited"""
    editor_role = editor_user['role']
    owner_role = goal_owner['role']
    
    # If editing someone else's goal
    if editor_user['id'] != goal_owner['id']:
        # Notify the goal owner
        create_notification({
            'user_id': goal_owner['id'],
            'action_by': editor_user['id'],
            'action_by_name': editor_user['name'],
            'action_type': 'goal_edited',
            'details': f"{editor_user['name']} edited your goal: '{goal['goal_title']}'",
            'is_read': False,
            'created_at': datetime.now(IST).isoformat()
        })
        
        # Notify HR when CMD/VP/Manager edits goals
        if editor_role in ['CMD', 'VP', 'Manager']:
            all_users = db.get_all_users()
            hr_users = [u for u in all_users if u['role'] == 'HR']
            for hr in hr_users:
                create_notification({
                    'user_id': hr['id'],
                    'action_by': editor_user['id'],
                    'action_by_name': editor_user['name'],
                    'action_type': 'goal_edited',
                    'details': f"{editor_role} {editor_user['name']} edited {goal_owner['name']}'s goal: '{goal['goal_title']}'",
                    'is_read': False,
                    'created_at': datetime.now(IST).isoformat()
                })
    
    # If Manager edits own goal -> Notify VP
    elif editor_role == 'Manager':
        all_users = db.get_all_users()
        vp_users = [u for u in all_users if u['role'] == 'VP']
        for vp in vp_users:
            create_notification({
                'user_id': vp['id'],
                'action_by': editor_user['id'],
                'action_by_name': editor_user['name'],
                'action_type': 'goal_edited',
                'details': f"Manager {editor_user['name']} edited their goal: '{goal['goal_title']}'",
                'is_read': False,
                'created_at': datetime.now(IST).isoformat()
            })
    
    # If VP edits own goal -> Notify CMD
    elif editor_role == 'VP':
        all_users = db.get_all_users()
        cmd_users = [u for u in all_users if u['role'] == 'CMD']
        for cmd in cmd_users:
            create_notification({
                'user_id': cmd['id'],
                'action_by': editor_user['id'],
                'action_by_name': editor_user['name'],
                'action_type': 'goal_edited',
                'details': f"VP {editor_user['name']} edited their goal: '{goal['goal_title']}'",
                'is_read': False,
                'created_at': datetime.now(IST).isoformat()
            })

def notify_goal_deleted(goal, deleter_user, goal_owner):
    """Notify when a goal is deleted"""
    deleter_role = deleter_user['role']
    owner_role = goal_owner['role']
    
    # If deleting someone else's goal
    if deleter_user['id'] != goal_owner['id']:
        # Notify the goal owner
        create_notification({
            'user_id': goal_owner['id'],
            'action_by': deleter_user['id'],
            'action_by_name': deleter_user['name'],
            'action_type': 'goal_deleted',
            'details': f"⚠️ {deleter_user['name']} deleted your goal: '{goal['goal_title']}'",
            'is_read': False,
            'created_at': datetime.now(IST).isoformat()
        })
        
        # Notify HR when CMD/VP/Manager deletes goals
        if deleter_role in ['CMD', 'VP', 'Manager']:
            all_users = db.get_all_users()
            hr_users = [u for u in all_users if u['role'] == 'HR']
            for hr in hr_users:
                create_notification({
                    'user_id': hr['id'],
                    'action_by': deleter_user['id'],
                    'action_by_name': deleter_user['name'],
                    'action_type': 'goal_deleted',
                    'details': f"⚠️ {deleter_role} {deleter_user['name']} deleted {goal_owner['name']}'s goal: '{goal['goal_title']}'",
                    'is_read': False,
                    'created_at': datetime.now(IST).isoformat()
                })
    
    # If Manager deletes own goal -> Notify VP
    elif deleter_role == 'Manager':
        all_users = db.get_all_users()
        vp_users = [u for u in all_users if u['role'] == 'VP']
        for vp in vp_users:
            create_notification({
                'user_id': vp['id'],
                'action_by': deleter_user['id'],
                'action_by_name': deleter_user['name'],
                'action_type': 'goal_deleted',
                'details': f"⚠️ Manager {deleter_user['name']} deleted their goal: '{goal['goal_title']}'",
                'is_read': False,
                'created_at': datetime.now(IST).isoformat()
            })
    
    # If VP deletes own goal -> Notify CMD
    elif deleter_role == 'VP':
        all_users = db.get_all_users()
        cmd_users = [u for u in all_users if u['role'] == 'CMD']
        for cmd in cmd_users:
            create_notification({
                'user_id': cmd['id'],
                'action_by': deleter_user['id'],
                'action_by_name': deleter_user['name'],
                'action_type': 'goal_deleted',
                'details': f"⚠️ VP {deleter_user['name']} deleted their goal: '{goal['goal_title']}'",
                'is_read': False,
                'created_at': datetime.now(IST).isoformat()
            })

def notify_goal_due_soon(goal, goal_owner, days_remaining):
    """Notify when goal is due in X days"""
    create_notification({
        'user_id': goal_owner['id'],
        'action_by': goal_owner['id'],
        'action_by_name': 'System',
        'action_type': 'goal_due_soon',
        'details': f" Your goal '{goal['goal_title']}' is due in {days_remaining} day(s)",
        'is_read': False,
        'created_at': datetime.now(IST).isoformat()
    })

def notify_goal_not_updated(goal, goal_owner, week_num):
    """Notify when weekly achievement is not updated (shows -)"""
    owner_role = goal_owner['role']
    
    # Notify the goal owner
    create_notification({
        'user_id': goal_owner['id'],
        'action_by': goal_owner['id'],
        'action_by_name': 'System',
        'action_type': 'goal_not_updated',
        'details': f"⚠️ Week {week_num} achievement not updated for goal: '{goal['goal_title']}'",
        'is_read': False,
        'created_at': datetime.now(IST).isoformat()
    })
    
    # Employee -> Notify Manager
    if owner_role == 'Employee':
        if goal_owner.get('manager_id'):
            manager = db.get_user_by_id(goal_owner['manager_id'])
            if manager:
                create_notification({
                    'user_id': manager['id'],
                    'action_by': goal_owner['id'],
                    'action_by_name': goal_owner['name'],
                    'action_type': 'goal_not_updated',
                    'details': f"⚠️ {goal_owner['name']} has not updated Week {week_num} for '{goal['goal_title']}'",
                    'is_read': False,
                    'created_at': datetime.now(IST).isoformat()
                })

def check_and_notify_due_dates_and_missing_updates():
    """Daily check for goals due soon and missing weekly updates"""
    today = date.today()
    all_users = db.get_all_users()
    
    # ✅ Track what we've already notified to prevent duplicates
    notified_goals = set()
    
    for user in all_users:
        goals = db.get_user_all_goals(user['id'])
        
        for goal in goals:
            goal_id = goal.get('goal_id')
            
            # ✅ Skip if we already notified about this goal
            if goal_id in notified_goals:
                continue
            
            if goal.get('status') == 'Active':
                # Check due dates
                end_date_str = goal.get('end_date')
                if end_date_str:
                    try:
                        end_date = datetime.strptime(str(end_date_str), '%Y-%m-%d').date()
                        days_remaining = (end_date - today).days
                        
                        # Notify if due in 5, 4, 3, 2, or 1 days
                        if 1 <= days_remaining <= 5:
                            # ✅ Check if we already sent this notification today
                            existing_notifs = supabase.table('notifications').select('*').eq(
                                'user_id', user['id']
                            ).eq('action_type', 'goal_due_soon').like(
                                'details', f"%{goal['goal_title']}%{days_remaining} day%"
                            ).gte('created_at', today.isoformat()).execute()
                            
                            if not existing_notifs.data:
                                notify_goal_due_soon(goal, user, days_remaining)
                                notified_goals.add(goal_id)
                        
                        # Check if goal is past due and not completed
                        if days_remaining < 0:
                            monthly_achievement = goal.get('monthly_achievement', 0)
                            monthly_target = goal.get('monthly_target', 1)
                            
                            if monthly_achievement is None:
                                monthly_achievement = 0
                            
                            progress = (monthly_achievement / monthly_target * 100) if monthly_target > 0 else 0
                            
                            if progress < 100:
                                # ✅ Check if we already sent this notification
                                existing_notifs = supabase.table('notifications').select('*').eq(
                                    'user_id', user['id']
                                ).eq('action_type', 'goal_not_completed').like(
                                    'details', f"%{goal['goal_title']}%"
                                ).gte('created_at', (today - timedelta(days=7)).isoformat()).execute()
                                
                                if not existing_notifs.data:
                                    notify_goal_not_completed(goal, user)
                                    notified_goals.add(goal_id)
                    except Exception as e:
                        print(f"Error checking deadline for goal {goal_id}: {str(e)}")
                        continue
                
                # Check for missing weekly updates (-)
                current_year = today.year
                current_month = today.month
                
                if goal['year'] == current_year and goal.get('month') == current_month:
                    # Check each week
                    for week_num in range(1, 5):
                        week_achievement = goal.get(f'week{week_num}_achievement')
                        
                        # If achievement is None (showing as '-'), notify
                        if week_achievement is None:
                            # Only notify if we're past that week's date
                            week_start, week_end = get_week_dates(current_year, current_month, week_num)
                            
                            if today > week_end:
                                # ✅ Check if we already sent this notification
                                existing_notifs = supabase.table('notifications').select('*').eq(
                                    'user_id', user['id']
                                ).eq('action_type', 'goal_not_updated').like(
                                    'details', f"%Week {week_num}%{goal['goal_title']}%"
                                ).gte('created_at', today.isoformat()).execute()
                                
                                if not existing_notifs.data:
                                    notify_goal_not_updated(goal, user, week_num)
                                    notified_goals.add(goal_id)
             # Check for all roles (Manager, HR, VP, CMD) about their own goals
        if user['role'] in ['Manager', 'HR', 'VP', 'CMD']:
            goals = db.get_user_all_goals(user['id'])
            
            for goal in goals:
                if goal.get('status') == 'Active':
                    end_date_str = goal.get('end_date')
                    if end_date_str:
                        try:
                            end_date = datetime.strptime(str(end_date_str), '%Y-%m-%d').date()
                            days_remaining = (end_date - today).days
                            
                            # Notify if due in 5, 4, 3, 2, or 1 days
                            if 1 <= days_remaining <= 5:
                                notify_goal_due_soon(goal, user, days_remaining)
                            
                            # Check if goal is past due and not completed
                            if days_remaining < 0:
                                monthly_achievement = goal.get('monthly_achievement', 0)
                                monthly_target = goal.get('monthly_target', 1)
                                
                                if monthly_achievement is None:
                                    monthly_achievement = 0
                                
                                progress = (monthly_achievement / monthly_target * 100) if monthly_target > 0 else 0
                                
                                if progress < 100:
                                    notify_goal_not_completed(goal, user)
                        except:
                            pass
                    
                    # Check for missing weekly updates
                    current_year = today.year
                    current_month = today.month
                    
                    if goal['year'] == current_year and goal.get('month') == current_month:
                        for week_num in range(1, 5):
                            week_achievement = goal.get(f'week{week_num}_achievement')
                            
                            if week_achievement is None:
                                week_start, week_end = get_week_dates(current_year, current_month, week_num)
                                
                                if today > week_end:
                                    notify_goal_not_updated(goal, user, week_num)

def get_user_notifications(user_id, limit=50):
    """Get notifications for a specific user"""
    try:
        result = supabase.table('notifications').select('*').eq(
            'user_id', user_id
        ).order('created_at', desc=True).limit(limit).execute()
        
        print(f" Query result for user {user_id}: {len(result.data) if result.data else 0} notifications")
        
        return result.data if result.data else []
    except Exception as e:
        print(f" Error getting notifications: {str(e)}")
        return []

def mark_notification_read(notification_id):
    """Mark a notification as read"""
    try:
        result = supabase.table('notifications').update({
            'is_read': True,
            'read_at': datetime.now(IST).isoformat()
        }).eq('id', notification_id).execute()
        return True
    except Exception as e:
        print(f"Error marking notification as read: {str(e)}")
        return False


def can_modify_user(current_user_role, target_user_role):
    """Check if current user can modify target user based on hierarchy"""
    hierarchy = get_role_hierarchy()
    current_level = hierarchy.get(current_user_role, 0)
    target_level = hierarchy.get(target_user_role, 0)
    
    # Can only modify users at lower hierarchy levels
    return current_level > target_level


def get_modifiable_roles(user_role):
    """Get list of roles that a user can modify"""
    hierarchy = get_role_hierarchy()
    user_level = hierarchy.get(user_role, 0)
    
    modifiable = []
    for role, level in hierarchy.items():
        if level < user_level:
            modifiable.append(role)
    
    return modifiable

def get_viewable_roles(user_role):
    """Get list of roles that a user can view"""
    # All roles can view users at same or lower level
    hierarchy = get_role_hierarchy()
    user_level = hierarchy.get(user_role, 0)
    
    viewable = []
    for role, level in hierarchy.items():
        if level <= user_level:
            viewable.append(role)
    
    return viewable

def check_password_strength(password):
    """Check password strength and return score, color, and feedback"""
    if not password:
        return 0, "#94a3b8", "No password entered", []
    
    score = 0
    feedback = []
    
    # Length check
    if len(password) >= 8:
        score += 25
    elif len(password) >= 6:
        score += 15
        feedback.append("🔸 Use at least 8 characters")
    else:
        feedback.append("❌ Too short (minimum 6 characters)")
    
    # Uppercase check
    if any(c.isupper() for c in password):
        score += 20
    else:
        feedback.append("🔸 Add uppercase letters (A-Z)")
    
    # Lowercase check
    if any(c.islower() for c in password):
        score += 20
    else:
        feedback.append("🔸 Add lowercase letters (a-z)")
    
    # Number check
    if any(c.isdigit() for c in password):
        score += 20
    else:
        feedback.append("🔸 Add numbers (0-9)")
    
    # Special character check
    special_chars = "!@#$%^&*()_+-=[]{}|;:,.<>?"
    if any(c in special_chars for c in password):
        score += 15
    else:
        feedback.append("🔸 Add special characters (!@#$%)")
    
    # Determine strength level and color
    if score >= 85:
        strength = "Very Strong "
        color = "#10b981"  # Green
    elif score >= 70:
        strength = "Strong "
        color = "#22c55e"  # Light green
    elif score >= 50:
        strength = "Medium "
        color = "#f59e0b"  # Orange
    elif score >= 30:
        strength = "Weak "
        color = "#fb923c"  # Light orange
    else:
        strength = "Very Weak "
        color = "#ef4444"  # Red
    
    return score, color, strength, feedback

def send_achievement_approval_email(manager_email, employee_name, goal_data, achievements):
    """Send achievement approval request email to manager"""
    try:
        smtp_server = os.getenv('SMTP_SERVER', 'smtp.gmail.com')
        smtp_port = int(os.getenv('SMTP_PORT', '587'))
        smtp_email = os.getenv('SMTP_EMAIL')
        smtp_password = os.getenv('SMTP_PASSWORD')
        
        if not smtp_email or not smtp_password:
            print("Email configuration not found")
            return False
        
        subject = f" Achievement Approval Request from {employee_name}"
        
        # Format achievements display
        achievements_html = ""
        for week, value in achievements.items():
            if week != 'monthly':
                display_value = value if value is not None else '-'
                achievements_html += f"""
                <tr>
                    <td style="padding: 8px; border: 1px solid #ddd;">{week.title()}</td>
                    <td style="padding: 8px; border: 1px solid #ddd; text-align: center;">{display_value}</td>
                </tr>
                """
        
        monthly_achievement = achievements.get('monthly', 0)
        if monthly_achievement is None:
            monthly_achievement = 0
        
        body = f"""
        <html>
            <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
                <div style="max-width: 600px; margin: 0 auto; padding: 20px; border: 1px solid #ddd; border-radius: 10px;">
                    <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                                padding: 20px; border-radius: 10px 10px 0 0; text-align: center;">
                        <h1 style="color: white; margin: 0; font-size: 28px;">📊 Achievement Approval Required</h1>
                    </div>
                    
                    <div style="padding: 30px 20px;">
                        <p style="font-size: 16px; margin-bottom: 20px;">
                            <strong>{employee_name}</strong> has updated their goal achievements and needs your approval:
                        </p>
                        
                        <div style="background: #f3f4f6; padding: 20px; border-radius: 8px; margin: 20px 0;">
                            <p style="margin: 0 0 10px 0;"><strong>Goal:</strong> {goal_data['goal_title']}</p>
                            <p style="margin: 0 0 10px 0;"><strong>Department:</strong> {goal_data.get('department', 'N/A')}</p>
                            <p style="margin: 0;"><strong>Monthly Target:</strong> {goal_data.get('monthly_target', 0)}</p>
                        </div>
                        
                        <h3 style="color: #1e40af; margin-top: 20px;">Reported Achievements:</h3>
                        <table style="width: 100%; border-collapse: collapse; margin: 15px 0;">
                            <thead>
                                <tr style="background: #3b82f6; color: white;">
                                    <th style="padding: 10px; border: 1px solid #ddd; text-align: left;">Week</th>
                                    <th style="padding: 10px; border: 1px solid #ddd; text-align: center;">Achievement</th>
                                </tr>
                            </thead>
                            <tbody>
                                {achievements_html}
                                <tr style="background: #dbeafe; font-weight: bold;">
                                    <td style="padding: 10px; border: 1px solid #ddd;">Monthly Total</td>
                                    <td style="padding: 10px; border: 1px solid #ddd; text-align: center;">{monthly_achievement}</td>
                                </tr>
                            </tbody>
                        </table>
                        
                        <p style="margin-top: 30px; padding-top: 20px; border-top: 1px solid #ddd; color: #64748b; font-size: 14px;">
                            Please log in to the Performance Management System to approve or reject these achievements.
                        </p>
                    </div>
                    
                    <div style="text-align: center; padding: 20px; background: #f9fafb; border-radius: 0 0 10px 10px;">
                        <p style="margin: 0; font-size: 12px; color: #64748b;">
                            This is an automated email from Performance Management System.<br>
                            Please do not reply to this email.
                        </p>
                    </div>
                </div>
            </body>
        </html>
        """
        
        message = MIMEMultipart("alternative")
        message["Subject"] = subject
        message["From"] = smtp_email
        message["To"] = manager_email
        
        html_part = MIMEText(body, "html")
        message.attach(html_part)
        
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(smtp_email, smtp_password)
            server.sendmail(smtp_email, manager_email, message.as_string())
        
        print(f"✅ Achievement approval email sent to {manager_email}")
        return True
    except Exception as e:
        print(f"❌ Failed to send achievement approval email: {str(e)}")
        return False
    
def send_goal_completion_email(manager_email, employee_name, goal_title, completed=True):
    """Send email to manager when employee completes or misses goal"""
    try:
        smtp_server = os.getenv('SMTP_SERVER', 'smtp.gmail.com')
        smtp_port = int(os.getenv('SMTP_PORT', '587'))
        smtp_email = os.getenv('SMTP_EMAIL')
        smtp_password = os.getenv('SMTP_PASSWORD')
        
        if not smtp_email or not smtp_password:
            print("Email configuration not found")
            return False
        
        if completed:
            subject = f"✅ Goal Completed: {employee_name}"
            status_color = "#10b981"
            status_icon = "✅"
            status_text = "COMPLETED"
            message_text = f"{employee_name} has successfully completed the goal"
        else:
            subject = f"⚠️ Goal Deadline Missed: {employee_name}"
            status_color = "#ef4444"
            status_icon = "⚠️"
            status_text = "DEADLINE MISSED"
            message_text = f"{employee_name} did not complete the goal by the deadline"
        
        body = f"""
        <html>
            <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
                <div style="max-width: 600px; margin: 0 auto; padding: 20px; border: 1px solid #ddd; border-radius: 10px;">
                    <div style="background: {status_color}; padding: 20px; border-radius: 10px 10px 0 0; text-align: center;">
                        <h1 style="color: white; margin: 0; font-size: 28px;">{status_icon} Goal {status_text}</h1>
                    </div>
                    
                    <div style="padding: 30px 20px;">
                        <p style="font-size: 16px; margin-bottom: 20px;">{message_text}:</p>
                        
                        <div style="background: #f3f4f6; padding: 20px; border-radius: 8px; margin: 20px 0;">
                            <p style="margin: 0 0 10px 0;"><strong>Employee:</strong> {employee_name}</p>
                            <p style="margin: 0;"><strong>Goal:</strong> {goal_title}</p>
                        </div>
                        
                        <p style="margin-top: 30px; padding-top: 20px; border-top: 1px solid #ddd; color: #64748b; font-size: 14px;">
                            Please review the goal details in the Performance Management System.
                        </p>
                    </div>
                    
                    <div style="text-align: center; padding: 20px; background: #f9fafb; border-radius: 0 0 10px 10px;">
                        <p style="margin: 0; font-size: 12px; color: #64748b;">
                            This is an automated email from Performance Management System.<br>
                            Please do not reply to this email.
                        </p>
                    </div>
                </div>
            </body>
        </html>
        """
        
        message = MIMEMultipart("alternative")
        message["Subject"] = subject
        message["From"] = smtp_email
        message["To"] = manager_email
        
        html_part = MIMEText(body, "html")
        message.attach(html_part)
        
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(smtp_email, smtp_password)
            server.sendmail(smtp_email, manager_email, message.as_string())
        
        print(f"Email sent to {manager_email}")
        return True
    except Exception as e:
        print(f"Failed to send email: {str(e)}")
        return False

def send_goal_approval_email(manager_email, employee_name, goal_data, goal_id):
    """Send approval request email to manager with approve/reject links"""
    try:
        smtp_server = os.getenv('SMTP_SERVER', 'smtp.gmail.com')
        smtp_port = int(os.getenv('SMTP_PORT', '587'))
        smtp_email = os.getenv('SMTP_EMAIL')
        smtp_password = os.getenv('SMTP_PASSWORD')
        
        if not smtp_email or not smtp_password:
            print("Email configuration not found")
            return False
        
        # Generate approval tokens
        import secrets
        approve_token = secrets.token_urlsafe(32)
        reject_token = secrets.token_urlsafe(32)
        
        # Store tokens in database (you'll need to create this table)
        db.store_approval_token(goal_id, approve_token, 'approve')
        db.store_approval_token(goal_id, reject_token, 'reject')
        
        # Create approval links (replace with your actual domain)
        app_url = os.getenv('APP_URL', 'http://localhost:8501')
        approve_link = f"{app_url}/?action=approve&token={approve_token}"
        reject_link = f"{app_url}/?action=reject&token={reject_token}"
        
        subject = f"🔔 Goal Approval Request from {employee_name}"
        
        body = f"""
        <html>
            <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
                <div style="max-width: 600px; margin: 0 auto; padding: 20px; border: 1px solid #ddd; border-radius: 10px;">
                    <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 20px; border-radius: 10px 10px 0 0; text-align: center;">
                        <h1 style="color: white; margin: 0; font-size: 28px;">🔔 Goal Approval Required</h1>
                    </div>
                    
                    <div style="padding: 30px 20px;">
                        <p style="font-size: 16px; margin-bottom: 20px;">
                            <strong>{employee_name}</strong> has created a new goal and needs your approval:
                        </p>
                        
                        <div style="background: #f3f4f6; padding: 20px; border-radius: 8px; margin: 20px 0;">
                            <p style="margin: 0 0 10px 0;"><strong>Goal Title:</strong> {goal_data['goal_title']}</p>
                            <p style="margin: 0 0 10px 0;"><strong>Department:</strong> {goal_data.get('department', 'N/A')}</p>
                            <p style="margin: 0 0 10px 0;"><strong>KPI:</strong> {goal_data.get('kpi', 'N/A')}</p>
                            <p style="margin: 0 0 10px 0;"><strong>Monthly Target:</strong> {goal_data.get('monthly_target', 0)}</p>
                            <p style="margin: 0 0 10px 0;"><strong>Period:</strong> {goal_data.get('year')}-Q{goal_data.get('quarter')}-M{goal_data.get('month')}</p>
                            <p style="margin: 0;"><strong>Description:</strong> {goal_data.get('goal_description', 'No description')}</p>
                        </div>
                        
                        <div style="text-align: center; margin: 30px 0;">
                            <a href="{approve_link}" style="display: inline-block; background: #10b981; color: white; padding: 15px 40px; text-decoration: none; border-radius: 8px; font-weight: bold; margin: 10px;">
                                ✅ Approve Goal
                            </a>
                            <a href="{reject_link}" style="display: inline-block; background: #ef4444; color: white; padding: 15px 40px; text-decoration: none; border-radius: 8px; font-weight: bold; margin: 10px;">
                                ❌ Reject Goal
                            </a>
                        </div>
                        
                        <p style="margin-top: 30px; padding-top: 20px; border-top: 1px solid #ddd; color: #64748b; font-size: 14px;">
                            You can also approve/reject this goal by logging into the Performance Management System.
                        </p>
                    </div>
                    
                    <div style="text-align: center; padding: 20px; background: #f9fafb; border-radius: 0 0 10px 10px;">
                        <p style="margin: 0; font-size: 12px; color: #64748b;">
                            This is an automated email from Performance Management System.<br>
                            Please do not reply to this email.
                        </p>
                    </div>
                </div>
            </body>
        </html>
        """
        
        message = MIMEMultipart("alternative")
        message["Subject"] = subject
        message["From"] = smtp_email
        message["To"] = manager_email
        
        html_part = MIMEText(body, "html")
        message.attach(html_part)
        
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(smtp_email, smtp_password)
            server.sendmail(smtp_email, manager_email, message.as_string())
        
        print(f"✅ Approval email sent to {manager_email}")
        return True
    except Exception as e:
        print(f"❌ Failed to send approval email: {str(e)}")
        return False
       
def calculate_performance_metrics(goals):
    """Calculate comprehensive performance metrics"""
    if not goals:
        return None
    
    total_goals = len(goals)
    completed = len([g for g in goals if g.get('status') == 'Completed'])
    active = len([g for g in goals if g.get('status') == 'Active'])
    on_hold = len([g for g in goals if g.get('status') == 'On Hold'])
    cancelled = len([g for g in goals if g.get('status') == 'Cancelled'])
    
    # ✅ FIXED: Calculate average progress with safe_float
    total_progress = 0
    goal_count = 0
    
    for goal in goals:
        achievement = safe_float(goal.get('monthly_achievement'), None)
        target = safe_float(goal.get('monthly_target'), 1)
        
        if achievement is not None and target > 0:
            progress = (achievement / target) * 100
            total_progress += progress
            goal_count += 1
    
    avg_progress = (total_progress / goal_count) if goal_count > 0 else 0
    
    # Calculate completion rate
    completion_rate = (completed / total_goals * 100) if total_goals > 0 else 0
    
    # ✅ FIXED: Calculate on-time completion rate correctly
    on_time = 0
    overdue = 0
    today = date.today()
    
    for goal in goals:
        if goal.get('status') == 'Completed':
            end_date_str = goal.get('end_date')
            completed_at_str = goal.get('completed_at')
            
            if end_date_str and completed_at_str:
                try:
                    end_date = datetime.strptime(str(end_date_str), '%Y-%m-%d').date()
                    
                    # Parse completion timestamp (ISO format)
                    completed_at = completed_at_str.replace('Z', '+00:00')
                    completed_date = datetime.fromisoformat(completed_at).date()
                    
                    # ✅ CORRECT: Compare completion date vs deadline
                    if completed_date <= end_date:
                        on_time += 1
                    else:
                        overdue += 1
                except Exception as e:
                    # If parsing fails, count as overdue
                    overdue += 1
            else:
                # If no completion date or end date, skip
                pass
    
    on_time_rate = (on_time / completed * 100) if completed > 0 else 0
    
    return {
        'total_goals': total_goals,
        'completed': completed,
        'active': active,
        'on_hold': on_hold,
        'cancelled': cancelled,
        'avg_progress': avg_progress,
        'completion_rate': completion_rate,
        'on_time': on_time,
        'overdue': overdue,
        'on_time_rate': on_time_rate
    }


def get_trend_data(user_id, months=6):
    """Get performance trend data for last N months"""
    trends = []
    today = date.today()
    
    for i in range(months):
        month_date = today - relativedelta(months=i)
        year = month_date.year
        month = month_date.month
        quarter = ((month - 1) // 3) + 1
        
        # Get goals for this month
        try:
            goals = db.get_month_goals(user_id, year, quarter, month)
            
            if goals and len(goals) > 0:
                metrics = calculate_performance_metrics(goals)
                trends.append({
                    'month': month_date.strftime('%b %Y'),
                    'completion_rate': metrics['completion_rate'],
                    'avg_progress': metrics['avg_progress'],
                    'total_goals': metrics['total_goals'],
                    'completed': metrics['completed']
                })
        except Exception as e:
            # Skip months with errors
            print(f"Error getting data for {month_date}: {str(e)}")
            continue
    
    return list(reversed(trends))


# ============================================
# ADVANCED CHARTS
# ============================================

def create_performance_gauge(value, title="Performance"):
    """Create a gauge chart for performance"""
    fig = go.Figure(go.Indicator(
        mode="gauge+number+delta",
        value=value,
        domain={'x': [0, 1], 'y': [0, 1]},
        title={'text': title, 'font': {'size': 24}},
        delta={'reference': 75, 'increasing': {'color': "green"}},
        gauge={
            'axis': {'range': [None, 100], 'tickwidth': 1, 'tickcolor': "darkblue"},
            'bar': {'color': "darkblue"},
            'bgcolor': "white",
            'borderwidth': 2,
            'bordercolor': "gray",
            'steps': [
                {'range': [0, 30], 'color': '#fee2e2'},
                {'range': [30, 60], 'color': '#fed7aa'},
                {'range': [60, 80], 'color': '#fef3c7'},
                {'range': [80, 100], 'color': '#d1fae5'}
            ],
            'threshold': {
                'line': {'color': "red", 'width': 4},
                'thickness': 0.75,
                'value': 90
            }
        }
    ))
    
    fig.update_layout(
        height=300,
        margin=dict(l=20, r=20, t=60, b=20),
        paper_bgcolor="white",
        font={'color': "darkblue", 'family': "Arial"}
    )
    
    return fig


def create_trend_chart(trend_data):
    """Create performance trend line chart"""
    if not trend_data:
        return None
    
    df = pd.DataFrame(trend_data)
    
    fig = make_subplots(
        rows=2, cols=1,
        subplot_titles=('Goal Completion Trend', 'Average Progress Trend'),
        vertical_spacing=0.15
    )
    
    # Completion rate trend
    fig.add_trace(
        go.Scatter(
            x=df['month'],
            y=df['completion_rate'],
            mode='lines+markers',
            name='Completion Rate',
            line=dict(color='#3b82f6', width=3),
            marker=dict(size=8)
        ),
        row=1, col=1
    )
    
    # Average progress trend
    fig.add_trace(
        go.Scatter(
            x=df['month'],
            y=df['avg_progress'],
            mode='lines+markers',
            name='Avg Progress',
            line=dict(color='#10b981', width=3),
            marker=dict(size=8)
        ),
        row=2, col=1
    )
    
    fig.update_xaxes(title_text="Month", row=2, col=1)
    fig.update_yaxes(title_text="Completion %", row=1, col=1)
    fig.update_yaxes(title_text="Progress %", row=2, col=1)
    
    fig.update_layout(
        height=600,
        showlegend=True,
        title_text="Performance Trends (Last 6 Months)",
        title_font_size=20
    )
    
    return fig


def create_status_distribution_chart(goals):
    """Create pie chart for goal status distribution"""
    status_counts = {}
    for goal in goals:
        status = goal.get('status', 'Active')
        status_counts[status] = status_counts.get(status, 0) + 1
    
    colors_map = {
        'Active': '#3b82f6',
        'Completed': '#10b981',
        'On Hold': '#f59e0b',
        'Cancelled': '#ef4444'
    }
    
    fig = go.Figure(data=[go.Pie(
        labels=list(status_counts.keys()),
        values=list(status_counts.values()),
        hole=0.4,
        marker_colors=[colors_map.get(k, '#64748b') for k in status_counts.keys()]
    )])
    
    fig.update_layout(
        title_text="Goal Status Distribution",
        title_font_size=20,
        height=400,
        showlegend=True
    )
    
    return fig


def create_department_performance_chart(goals):
    """Create bar chart for performance by department"""
    department_data = {}
    
    for goal in goals:
        # Normalize department name
        department = normalize_department(goal.get('department'))
        if department not in department_data:
            department_data[department] = {'total': 0, 'completed': 0, 'progress': []}
        
        department_data[department]['total'] += 1
        if goal.get('status') == 'Completed':
            department_data[department]['completed'] += 1
        
        progress = calculate_progress(
            goal.get('monthly_achievement', 0),
            goal.get('monthly_target', 1)
        )
        department_data[department]['progress'].append(progress)
    
    # ✅ FIX: Change variable name from 'department' to 'departments' (plural)
    departments = sorted(list(department_data.keys()))  # Sort alphabetically
    completion_rates = [
        (department_data[d]['completed'] / department_data[d]['total'] * 100) 
        if department_data[d]['total'] > 0 else 0 
        for d in departments
    ]
    avg_progress = [
        sum(department_data[d]['progress']) / len(department_data[d]['progress']) 
        if department_data[d]['progress'] else 0 
        for d in departments
    ]
    
    # ✅ FIX: Use 'departments' (plural) instead of 'department' (singular)
    fig = go.Figure(data=[
        go.Bar(name='Completion Rate', x=departments, y=completion_rates, marker_color='#3b82f6'),
        go.Bar(name='Avg Progress', x=departments, y=avg_progress, marker_color='#10b981')
    ])
    
    fig.update_layout(
        barmode='group',
        title_text="Performance by Department",
        title_font_size=20,
        xaxis_title="Department",
        yaxis_title="Percentage (%)",
        height=400
    )
    
    return fig


def create_heatmap_calendar(goals, year, month):
    """Create calendar heatmap for daily goal achievements"""
    # Get days in month
    days_in_month = calendar.monthrange(year, month)[1]
    
    # Initialize achievement data
    daily_achievements = {}
    for day in range(1, days_in_month + 1):
        daily_achievements[day] = 0
    
    # Calculate daily achievements
    for goal in goals:
        for week in range(1, 5):
            week_achievement = goal.get(f'week{week}_achievement', 0)
            if week_achievement > 0:
                # Distribute across week days
                week_start, week_end = get_week_dates(year, month, week)
                days_in_week = (week_end - week_start).days + 1
                daily_avg = week_achievement / days_in_week
                
                for day in range(week_start.day, min(week_end.day + 1, days_in_month + 1)):
                    daily_achievements[day] += daily_avg
    
    # Create matrix for heatmap
    weeks = 5
    days = 7
    matrix = np.zeros((weeks, days))
    day_labels = []
    
    first_day = date(year, month, 1)
    first_weekday = first_day.weekday()
    
    day_counter = 1
    for week in range(weeks):
        for day in range(days):
            if week == 0 and day < first_weekday:
                matrix[week][day] = np.nan
            elif day_counter > days_in_month:
                matrix[week][day] = np.nan
            else:
                matrix[week][day] = daily_achievements.get(day_counter, 0)
                day_counter += 1
    
    fig = go.Figure(data=go.Heatmap(
        z=matrix,
        x=['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun'],
        y=[f'Week {i+1}' for i in range(weeks)],
        colorscale='Blues',
        showscale=True
    ))
    
    fig.update_layout(
        title=f"Daily Achievement Heatmap - {get_month_name(month)} {year}",
        title_font_size=20,
        height=400
    )
    
    return fig


# ============================================
# PDF REPORT GENERATION
# ============================================

def generate_performance_report_pdf(user, goals, period="Monthly"):
    """Generate comprehensive PDF performance report"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    story = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1e40af'),
        alignment=TA_CENTER,
        spaceAfter=30
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#3b82f6'),
        spaceAfter=12
    )
    
    # Title
    story.append(Paragraph(f"Performance Report - {period}", title_style))
    story.append(Spacer(1, 0.2*inch))
    
    # User Information
    story.append(Paragraph("Employee Information", heading_style))
    user_data = [
        ['Name:', user['name']],
        ['Email:', user['email']],
        ['Role:', user['role']],
        ['Department:', user.get('department', 'N/A')],
        ['Report Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
    ]
    
    user_table = Table(user_data, colWidths=[2*inch, 4*inch])
    user_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e0e7ff')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey)
    ]))
    
    story.append(user_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Performance Metrics
    metrics = calculate_performance_metrics(goals)
    
    if metrics:
        story.append(Paragraph("Performance Summary", heading_style))
        
        metrics_data = [
            ['Metric', 'Value'],
            ['Total Goals', str(metrics['total_goals'])],
            ['Completed Goals', str(metrics['completed'])],
            ['Active Goals', str(metrics['active'])],
            ['Completion Rate', f"{metrics['completion_rate']:.1f}%"],
            ['Average Progress', f"{metrics['avg_progress']:.1f}%"],
            ['On-Time Completion Rate', f"{metrics['on_time_rate']:.1f}%"]
        ]
        
        metrics_table = Table(metrics_data, colWidths=[3*inch, 2*inch])
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(metrics_table)
        story.append(Spacer(1, 0.3*inch))
    
    # Goals Details
    story.append(Paragraph("Goals Details", heading_style))
    
    goals_data = [['Goal Title', 'Vertical', 'Target', 'Achievement', 'Progress', 'Status']]
    
    for goal in goals:
        progress = calculate_progress(
            goal.get('monthly_achievement', 0),
            goal.get('monthly_target', 1)
        )
        
        goals_data.append([
            goal['goal_title'][:30],
            goal.get('vertical', 'N/A'),
            str(goal.get('monthly_target', 0)),
            str(goal.get('monthly_achievement', 0)),
            f"{progress:.1f}%",
            goal.get('status', 'Active')
        ])
    
    goals_table = Table(goals_data, colWidths=[2*inch, 1.2*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch])
    goals_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    
    story.append(goals_table)
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    return buffer



# ============================================
# ANALYTICS PAGE
# ============================================

def display_analytics_page():
    """Display comprehensive analytics page"""
    user = st.session_state.user
    if not user:
        st.warning("⚠️ Session expired. Please login again.")
        st.rerun()
    role = user['role']
    
    st.title("Advanced Analytics & Reports")
    
    # Filters
    col_filter1, col_filter2, col_filter3 = st.columns(3)
    
    with col_filter1:
        if role in ['CMD', 'VP', 'HR']:
            all_users = db.get_all_users()
            selected_user = st.selectbox(
                "Select User",
                ["My Analytics"] + [f"{u['name']} ({u['email']})" for u in all_users]
            )
            if selected_user == "My Analytics":
                analysis_user = user
            else:
                user_email = selected_user.split('(')[1].strip(')')
                analysis_user = next(u for u in all_users if u['email'] == user_email)
        elif role == 'Manager':
            team_members = db.get_team_members(user['id'])
            selected_user = st.selectbox(
                "Select Team Member",
                ["My Analytics"] + [f"{m['name']} ({m['email']})" for m in team_members]
            )
            if selected_user == "My Analytics":
                analysis_user = user
            else:
                user_email = selected_user.split('(')[1].strip(')')
                analysis_user = next(m for m in team_members if m['email'] == user_email)
        else:
            analysis_user = user
    
    with col_filter2:
        analysis_period = st.selectbox(
            "Analysis Period",
            ["Current Month", "Current Quarter", "Current Year", "Last 6 Months", "All Time"]
        )
    
    with col_filter3:
        view_type = st.selectbox(
            "View Type",
            ["Overview", "Trends", "Comparisons", "Detailed"]
        )
    
    # Get goals based on period
    all_goals = db.get_user_all_goals(analysis_user['id'])
    today = date.today()
    
    if analysis_period == "Current Month":
        today = date.today()
        goals = [g for g in all_goals if g['year'] == today.year and g.get('month') == today.month]
    elif analysis_period == "Current Quarter":
        current_quarter = get_fiscal_quarter(today.month)
        goals = [g for g in all_goals 
                if g['year'] == today.year 
                and g.get('month') 
                and get_fiscal_quarter(g.get('month')) == current_quarter]
    elif analysis_period == "Current Year":
        today = date.today()
        goals = [g for g in all_goals if g['year'] == today.year]
    else:
        goals = all_goals
    
    if not goals:
        st.info(f"No goals found for {analysis_user['name']} in selected period")
        return
    
    st.markdown("---")
    
    # Calculate metrics
    metrics = calculate_performance_metrics(goals)
    
    # Overview Tab
    if view_type == "Overview":
        st.subheader(f" Performance Overview - {analysis_user['name']}")
        
        # KPI Cards
        col1, col2, col3, col4, col5 = st.columns(5)

        with col1:
            render_clickable_metric_card("Total Goals", str(metrics['total_goals']), "#3B82F6", "🎯", "analytics_card_total", 'analytics_total')

        with col2:
            render_clickable_metric_card("Completed", str(metrics['completed']), "#10B981", "✅", "analytics_card_completed", 'analytics_completed')

        with col3:
            render_clickable_metric_card("Active", str(metrics['active']), "#F5576C", "🔄", "analytics_card_active", 'analytics_active')

        with col4:
            render_clickable_metric_card("Completion Rate", f"{int(metrics['completion_rate'])}%", "#8b5cf6", "📊", "analytics_card_completion", 'analytics_completion')

        with col5:
            render_clickable_metric_card("Avg Progress", f"{int(metrics['avg_progress'])}%", "#ec4899", "📈", "analytics_card_progress", 'analytics_progress')
        
        # Analytics details modal
        if st.session_state.get('show_details') and st.session_state.get('show_details').startswith('analytics_'):
            st.markdown("---")
            detail_type = st.session_state.show_details
            
            col_header, col_close = st.columns([5, 1])
            
            with col_header:
                if detail_type == 'analytics_total':
                    st.subheader(f" All Goals ({metrics['total_goals']})")
                elif detail_type == 'analytics_completed':
                    st.subheader(f" Completed Goals ({metrics['completed']})")
                elif detail_type == 'analytics_active':
                    st.subheader(f" Active Goals ({metrics['active']})")
                elif detail_type == 'analytics_completion':
                    st.subheader(f" Goals by Completion Rate")
                elif detail_type == 'analytics_progress':
                    st.subheader(f" Goals by Progress")
            
            with col_close:
                if st.button(" Close", key="close_analytics_details"):
                    del st.session_state.show_details
                    st.rerun()
            
            # Filter goals based on detail type
            if detail_type == 'analytics_completed':
                display_goals = [g for g in goals if g.get('status') == 'Completed']
            elif detail_type == 'analytics_active':
                display_goals = [g for g in goals if g.get('status') == 'Active']
            elif detail_type == 'analytics_completion' or detail_type == 'analytics_progress':
                display_goals = goals.copy()
                for g in display_goals:
                    achievement = g.get('monthly_achievement')
                    target = g.get('monthly_target', 1)
                    achievement_val = 0 if achievement is None else achievement
                    target_val = 1 if target is None or target == 0 else target
                    g['_progress'] = calculate_progress(achievement_val, target_val)
                display_goals.sort(key=lambda x: x['_progress'], reverse=True)
            else:
                display_goals = goals
            
            # Display goals in table
            if display_goals:
                goal_data = []
                for goal in display_goals:
                    monthly_achievement = goal.get('monthly_achievement')
                    monthly_target = goal.get('monthly_target', 1)
                    achievement_value = 0 if monthly_achievement is None else monthly_achievement
                    target_value = 1 if monthly_target is None or monthly_target == 0 else monthly_target
                    progress = calculate_progress(achievement_value, target_value)
                    
                    goal_data.append({
                        'Goal Title': goal['goal_title'],
                        'Department': goal.get('department', 'N/A'),
                        'Month': get_month_name(goal.get('month', 1)),  # ADD THIS
                        'Year': str(goal['year']),  # ADD THIS
                        'Target': monthly_target if monthly_target is not None else 0,
                        'Achievement': monthly_achievement if monthly_achievement is not None else '-',
                        'Progress': f"{progress:.1f}%",
                        'Status': goal.get('status', 'Active')
                    })
                
                df_goals = pd.DataFrame(goal_data)
                st.dataframe(df_goals, use_container_width=True, height=400)
                
                st.download_button(
                    "📥 Export to CSV",
                    df_goals.to_csv(index=False).encode('utf-8'),
                    f"analytics_goals_{detail_type}.csv",
                    "text/csv",
                    key='download_analytics_details'
                )
            else:
                st.info("No goals found in this category")
            
            st.markdown("---")

        # Charts Row 1
        col_chart1, col_chart2 = st.columns(2)
        
        with col_chart1:
            # Performance Gauge
            gauge_fig = create_performance_gauge(metrics['avg_progress'], "Overall Performance")
            st.plotly_chart(gauge_fig, use_container_width=True)
        
        with col_chart2:
            # Status Distribution
            status_fig = create_status_distribution_chart(goals)
            st.plotly_chart(status_fig, use_container_width=True)
        
        st.markdown("---")
        
        # Charts Row 2
        col_chart3, col_chart4 = st.columns(2)
        
        with col_chart3:
            # Vertical Performance
            vertical_fig = create_department_performance_chart(goals)
            st.plotly_chart(vertical_fig, use_container_width=True)
        
        with col_chart4:
            # On-time vs Overdue
            on_time_data = pd.DataFrame({
                'Category': ['On-Time', 'Overdue'],
                'Count': [metrics['on_time'], metrics['overdue']]
            })
            fig_ontime = px.bar(on_time_data, x='Category', y='Count',
                               title="On-Time vs Overdue Completion",
                               color='Category',
                               color_discrete_map={'On-Time': '#10b981', 'Overdue': '#ef4444'})
            fig_ontime.update_layout(showlegend=False, height=400)
            st.plotly_chart(fig_ontime, use_container_width=True)
    
    # Trends Tab
    elif view_type == "Trends":
        st.subheader(f" Performance Trends - {analysis_user['name']}")
        
        trend_data = get_trend_data(analysis_user['id'], months=6)
        
        if trend_data and len(trend_data) > 0:
            # Main trend chart
            trend_fig = create_trend_chart(trend_data)
            st.plotly_chart(trend_fig, use_container_width=True)
            
            st.markdown("---")
            
            # Monthly Achievement Trend
            st.subheader("Monthly Goal Achievement Trend")
            
            monthly_data = []
            for month_info in trend_data:
                monthly_data.append({
                    'Month': month_info['month'],
                    'Goals': month_info['total_goals'],
                    'Completed': month_info['completed']
                })
            
            df_monthly = pd.DataFrame(monthly_data)
            
            if not df_monthly.empty:
                fig_monthly = px.line(df_monthly, x='Month', y=['Goals', 'Completed'],
                                     title="Goals Created vs Completed",
                                     markers=True)
                fig_monthly.update_layout(height=400)
                st.plotly_chart(fig_monthly, use_container_width=True)
            else:
                st.info("No monthly data available for visualization")
        else:
            st.info("Not enough data for trend analysis. Please ensure you have goals spanning multiple months.")
    
    # Comparisons Tab
    elif view_type == "Comparisons":
        st.subheader(f" Performance Comparisons - {analysis_user['name']}")
        
        if role in ['HR', 'Manager']:
            # Get comparison users based on the selected user's role
            if role == 'HR':
                # HR: Compare with same role users
                selected_role = analysis_user['role']
                all_users = db.get_all_users()
                
                if selected_role == 'Manager':
                    compare_users = [u for u in all_users if u['role'] == 'Manager']
                    comparison_title = "Compare with Other Managers"
                elif selected_role == 'Employee':
                    compare_users = [u for u in all_users if u['role'] == 'Employee']
                    comparison_title = "Compare with Other Employees"
                else:  # HR viewing another HR
                    compare_users = [u for u in all_users if u['role'] == 'HR']
                    comparison_title = "Compare with Other HR Members"
                    
            else:  # Manager
                # Manager compares team members only (not themselves)
                if analysis_user['id'] == user['id']:
                    # Manager viewing their own analytics - compare with team only
                    compare_users = db.get_team_members(user['id'])
                    comparison_title = "Compare Team Members"
                else:
                    # Manager viewing a team member - compare with other team members
                    team_members = db.get_team_members(user['id'])
                    compare_users = [m for m in team_members]
                    comparison_title = "Compare with Team Members"
            
            if compare_users:
                # Show appropriate info message
                if role == 'Manager' and analysis_user['id'] == user['id']:
                    st.info(f"**{comparison_title}** ({len(compare_users)} team members)\n\n💡 *Note: You are viewing your team's performance comparison (manager is not included in rankings)*")
                else:
                    st.info(f"**{comparison_title}** ({len(compare_users)} users)")
                
                comparison_data = []
                
                # Add all users (including current)
                for comp_user in compare_users:
                    comp_goals = db.get_user_all_goals(comp_user['id'])
                    if comp_goals:
                        comp_metrics = calculate_performance_metrics(comp_goals)
                        is_current = (comp_user['id'] == analysis_user['id'])
                        comparison_data.append({
                            'Name': f"⭐ {comp_user['name']}" if is_current else comp_user['name'],
                            'Display_Name': comp_user['name'],
                            'Total Goals': comp_metrics['total_goals'],
                            'Completion Rate': comp_metrics['completion_rate'],
                            'Avg Progress': comp_metrics['avg_progress'],
                            'Is_Current': is_current,
                            'User_ID': comp_user['id']
                        })
                
                if comparison_data and len(comparison_data) > 0:
                    df_compare = pd.DataFrame(comparison_data)
                    
                    # Highlight current user
                    st.success(f" **{analysis_user['name']}** is highlighted with ⭐ in the comparisons below")
                    
                    # Comparison bar chart
                    fig_compare = px.bar(
                        df_compare, 
                        x='Name', 
                        y=['Completion Rate', 'Avg Progress'],
                        title=f"{comparison_title} - Performance Metrics",
                        barmode='group',
                        height=500,
                        color_discrete_map={
                            'Completion Rate': '#3b82f6',
                            'Avg Progress': '#10b981'
                        }
                    )
                    fig_compare.update_layout(
                        xaxis_tickangle=-45,
                        xaxis_title="",
                        yaxis_title="Percentage (%)"
                    )
                    st.plotly_chart(fig_compare, use_container_width=True)
                    
                    st.markdown("---")
                    
                    # Comparison table with ranking
                    st.subheader("📊 Detailed Comparison Table")
                    
                    # Sort by completion rate for ranking
                    df_sorted = df_compare.sort_values('Completion Rate', ascending=False).reset_index(drop=True)
                    df_sorted.insert(0, 'Rank', range(1, len(df_sorted) + 1))
                    
                    # Create display dataframe
                    df_display = df_sorted[['Rank', 'Name', 'Total Goals', 'Completion Rate', 'Avg Progress']].copy()
                    df_display['Completion Rate'] = df_display['Completion Rate'].apply(lambda x: f"{x:.1f}%")
                    df_display['Avg Progress'] = df_display['Avg Progress'].apply(lambda x: f"{x:.1f}%")
                    
                    st.dataframe(df_display, use_container_width=True, height=400)
                    
                    # Performance insights
                    st.markdown("---")
                    st.subheader(" Performance Insights")
                    
                    # Calculate metrics correctly
                    avg_completion = df_compare['Completion Rate'].mean()
                    avg_progress = df_compare['Avg Progress'].mean()
                    
                    # Get current user's metrics
                    current_user_data = df_compare[df_compare['Is_Current'] == True]
                    
                    if not current_user_data.empty:
                        user_completion = current_user_data['Completion Rate'].iloc[0]
                        user_progress = current_user_data['Avg Progress'].iloc[0]
                        
                        # Get current user's rank
                        current_rank_row = df_sorted[df_sorted['Is_Current'] == True]
                        if not current_rank_row.empty:
                            current_rank = current_rank_row.iloc[0]['Rank']
                            
                            # Display rank
                            col_rank, col_space = st.columns([1, 3])
                            with col_rank:
                                st.metric("Current User's Rank", f"#{current_rank} out of {len(df_sorted)}")
                        
                        col_insight1, col_insight2 = st.columns(2)
                        
                        with col_insight1:
                            st.markdown("**Completion Rate Analysis:**")
                            diff = user_completion - avg_completion
                            if diff > 0:
                                st.success(f"✅ **Above Average**\n\n{analysis_user['name']}'s completion rate: **{user_completion:.1f}%**\n\nGroup average: **{avg_completion:.1f}%**\n\nDifference: **+{diff:.1f}%** higher")
                            elif diff < 0:
                                st.warning(f"⚠️ **Below Average**\n\n{analysis_user['name']}'s completion rate: **{user_completion:.1f}%**\n\nGroup average: **{avg_completion:.1f}%**\n\nDifference: **{abs(diff):.1f}%** lower")
                            else:
                                st.info(f"📊 **Average Performance**\n\n{analysis_user['name']}'s completion rate: **{user_completion:.1f}%**\n\nMatches group average: **{avg_completion:.1f}%**")
                        
                        with col_insight2:
                            st.markdown("**Progress Analysis:**")
                            prog_diff = user_progress - avg_progress
                            if prog_diff > 0:
                                st.success(f"✅ **Above Average**\n\n{analysis_user['name']}'s avg progress: **{user_progress:.1f}%**\n\nGroup average: **{avg_progress:.1f}%**\n\nDifference: **+{prog_diff:.1f}%** higher")
                            elif prog_diff < 0:
                                st.warning(f"⚠️ **Below Average**\n\n{analysis_user['name']}'s avg progress: **{user_progress:.1f}%**\n\nGroup average: **{avg_progress:.1f}%**\n\nDifference: **{abs(prog_diff):.1f}%** lower")
                            else:
                                st.info(f"📊 **Average Performance**\n\n{analysis_user['name']}'s avg progress: **{user_progress:.1f}%**\n\nMatches group average: **{avg_progress:.1f}%**")
                        
                        # Top performer
                        st.markdown("---")
                        top_performer = df_sorted.iloc[0]
                        if top_performer['Is_Current']:
                            st.success(f" **Congratulations!** {analysis_user['name']} has the highest completion rate ({top_performer['Completion Rate']:.1f}%) in the group!")
                        else:
                            st.info(f" **Top Performer:** {top_performer['Display_Name']} leads with **{top_performer['Completion Rate']:.1f}%** completion rate")
                    
                else:
                    st.info("No comparison data available - users need to have goals to be included in comparison")
            else:
                st.info(f"No {comparison_title.lower()} available for comparison")
        else:
            st.warning("⚠️ Comparison view is only available for HR and Managers")
    
    # Detailed Tab
    elif view_type == "Detailed":
        st.subheader(f"📋 Detailed Goal Analysis - {analysis_user['name']}")
        
        # Goals breakdown by status
        st.markdown("### Goals by Status")
        
        status_groups = {}
        for goal in goals:
            status = goal.get('status', 'Active')
            if status not in status_groups:
                status_groups[status] = []
            status_groups[status].append(goal)
        
        for status, status_goals in status_groups.items():
            with st.expander(f"{status} Goals ({len(status_goals)})"):
                for goal in status_goals:
                    progress = calculate_progress(
                        goal.get('monthly_achievement', 0),
                        goal.get('monthly_target', 1)
                    )
                    
                    col1, col2, col3 = st.columns([3, 1, 1])
                    with col1:
                        st.markdown(f"**{goal['goal_title']}**")
                        st.caption(f"Department: {goal.get('department', 'N/A')} | KPI: {goal.get('kpi', 'N/A')}")
                    with col2:
                        st.metric("Target", goal.get('monthly_target', 0))
                    with col3:
                        st.metric("Progress", f"{progress:.1f}%")
                    
                    st.markdown("---")
        
        st.markdown("---")
        
        # Weekly breakdown
        st.markdown("### Weekly Performance Breakdown")
        
        weekly_data = {
            'Week 1': {'target': 0, 'achievement': 0},
            'Week 2': {'target': 0, 'achievement': 0},
            'Week 3': {'target': 0, 'achievement': 0},
            'Week 4': {'target': 0, 'achievement': 0}
        }
        
        for goal in goals:
            for week in range(1, 5):
                # Handle None values properly
                week_target = goal.get(f'week{week}_target')
                week_achievement = goal.get(f'week{week}_achievement')
                
                # Add only if not None
                if week_target is not None:
                    weekly_data[f'Week {week}']['target'] += week_target
                
                if week_achievement is not None:
                    weekly_data[f'Week {week}']['achievement'] += week_achievement
        
        weekly_df = pd.DataFrame({
            'Week': list(weekly_data.keys()),
            'Target': [weekly_data[w]['target'] for w in weekly_data.keys()],
            'Achievement': [weekly_data[w]['achievement'] for w in weekly_data.keys()]
        })
        
        fig_weekly = px.bar(weekly_df, x='Week', y=['Target', 'Achievement'],
                           title="Weekly Targets vs Achievements",
                           barmode='group',
                           height=400)
        st.plotly_chart(fig_weekly, use_container_width=True)
    
        st.markdown("---")

        st.markdown("### Monthly Performance Breakdown")

        # Create monthly data structure
        monthly_data = {}

        for goal in goals:
            goal_month = goal.get('month')
            goal_year = goal.get('year', today.year)
            
            if goal_month:
                month_key = f"{get_month_name(goal_month)} {goal_year}"
                
                if month_key not in monthly_data:
                    monthly_data[month_key] = {'target': 0, 'achievement': 0, 'month_num': goal_month, 'year': goal_year}
                
                # Add targets and achievements (handle None values)
                monthly_target = safe_float(goal.get('monthly_target'), 0)
                monthly_achievement = safe_float(goal.get('monthly_achievement'), 0)
                
                monthly_data[month_key]['target'] += monthly_target
                monthly_data[month_key]['achievement'] += monthly_achievement

        if monthly_data:
            # Sort by year and month
            sorted_months = sorted(monthly_data.items(), key=lambda x: (x[1]['year'], x[1]['month_num']))
            
            # Create dataframe
            monthly_df = pd.DataFrame({
                'Month': [item[0] for item in sorted_months],
                'Target': [item[1]['target'] for item in sorted_months],
                'Achievement': [item[1]['achievement'] for item in sorted_months]
            })
            
            # Display as table
            st.dataframe(monthly_df, use_container_width=True)
            
            # Display as chart
            fig_monthly = px.bar(
                monthly_df, 
                x='Month', 
                y=['Target', 'Achievement'],
                title="Monthly Targets vs Achievements",
                barmode='group',
                height=400,
                color_discrete_map={'Target': '#3b82f6', 'Achievement': '#10b981'}
            )
            fig_monthly.update_layout(xaxis_tickangle=-45)
            st.plotly_chart(fig_monthly, use_container_width=True)
            
            # Summary metrics
            col_m1, col_m2, col_m3 = st.columns(3)
            
            total_monthly_target = monthly_df['Target'].sum()
            total_monthly_achievement = monthly_df['Achievement'].sum()
            monthly_completion_rate = (total_monthly_achievement / total_monthly_target * 100) if total_monthly_target > 0 else 0
            
            with col_m1:
                st.metric("Total Monthly Target", f"{total_monthly_target:.2f}")
            with col_m2:
                st.metric("Total Monthly Achievement", f"{total_monthly_achievement:.2f}")
            with col_m3:
                st.metric("Overall Completion Rate", f"{monthly_completion_rate:.1f}%")
        else:
            st.info("No monthly performance data available")

    # Export Report
    st.markdown("---")
    st.subheader("📥 Export Report")
    
    col_export1, col_export2, col_export3 = st.columns(3)
    
    with col_export1:
        if st.button("📄 Export to PDF", use_container_width=True):
            with st.spinner("Generating PDF report..."):
                pdf_buffer = generate_performance_report_pdf(analysis_user, goals, analysis_period)
                st.download_button(
                    label="📥 Download PDF Report",
                    data=pdf_buffer,
                    file_name=f"Performance_Report_{analysis_user['name']}_{datetime.now().strftime('%Y%m%d')}.pdf",
                    mime="application/pdf",
                    use_container_width=True
                )

    with col_export2:
        if st.button(" Export to Excel", use_container_width=True):
            # Get the first goal's date or use current date
            if goals:
                export_year = goals[0]['year']
                export_quarter = goals[0].get('quarter', 1)
                export_month = goals[0].get('month', 1)
            else:
                today = date.today()
                export_year = today.year
                export_quarter = ((today.month - 1) // 3) + 1
                export_month = today.month
            
            excel_buffer = export_goals_to_excel(
                analysis_user['id'], 
                export_year,
                export_quarter,
                export_month
            )
            
            if excel_buffer:
                st.download_button(
                    label="📥 Download Excel Report",
                    data=excel_buffer,
                    file_name=f"Goals_Report_{analysis_user['name']}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            else:
                st.warning("No data to export")
def get_completable_goals(user_id):
    """Get all goals that can be auto-completed"""
    all_goals = db.get_user_all_goals(user_id)
    
    completable = []
    for goal in all_goals:
        if goal.get('status') == 'Active':
            progress = calculate_progress(
                goal.get('monthly_achievement', 0),
                goal.get('monthly_target', 1)
            )
            if progress >= 100:
                completable.append(goal)
    
    return completable


def auto_complete_goal(goal_id, completed_by=None, remarks=None):
    """Auto-complete a goal that has reached 100%"""
    updates = {
        'status': 'Completed',
        'completed_at': datetime.now(IST).isoformat()
    }
    
    if completed_by:
        updates['completed_by'] = completed_by
    
    if remarks:
        updates['completion_remarks'] = remarks
    
    return db.update_goal(goal_id, updates)
def display_auto_complete_banner():
    """Show banner with auto-complete suggestions"""
    user = st.session_state.user
    
    completable_goals = get_completable_goals(user['id'])
    
    if completable_goals:
        with st.container():
            st.markdown("""
            <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                        padding: 20px; border-radius: 10px; margin-bottom: 20px;">
                <h3 style="color: white; margin: 0;">🎉 Goals Ready for Completion!</h3>
            </div>
            """, unsafe_allow_html=True)
            
            st.info(f"You have **{len(completable_goals)} goal(s)** at 100% progress")
            
            with st.expander(f" View {len(completable_goals)} Completable Goals", expanded=True):
                for goal in completable_goals:
                    col1, col2, col3 = st.columns([3, 1, 1])
                    
                    with col1:
                        st.markdown(f"**{goal['goal_title']}**")
                        st.caption(f"Target: {goal.get('monthly_target', 0)}")
                    
                    with col2:
                        st.success("100% ✅")
                    
                    with col3:
                        if st.button("✅ Complete", key=f"complete_{goal['goal_id']}", use_container_width=True):
                            if auto_complete_goal(goal['goal_id'], completed_by=user['id']):
                                st.success(f"Completed!")
                                st.rerun()
                
                st.markdown("---")
                
                col_bulk1, col_bulk2 = st.columns(2)
                
                with col_bulk1:
                    if st.button("✅ Complete All", type="primary", use_container_width=True):
                        for goal in completable_goals:
                            auto_complete_goal(goal['goal_id'], completed_by=user['id'])
                        st.success(f"✅ Completed {len(completable_goals)} goals!")
                        st.balloons()
                        st.rerun()
                
                with col_bulk2:
                    if st.button("⏭️ Remind Later", use_container_width=True):
                        st.rerun()

def calculate_performance_score(stats):
    """Calculate weighted performance score for ranking"""
    # Weighted scoring system
    completion_weight = 0.4  # 40%
    progress_weight = 0.3    # 30%
    total_goals_weight = 0.2  # 20%
    on_time_weight = 0.1     # 10%
    
    completion_score = stats.get('completion_rate', 0) * completion_weight
    progress_score = stats.get('avg_progress', 0) * progress_weight
    goals_score = min(stats.get('total_goals', 0) * 5, 100) * total_goals_weight  # Cap at 100
    on_time_score = stats.get('on_time_rate', 0) * on_time_weight
    
    total_score = completion_score + progress_score + goals_score + on_time_score
    
    return round(total_score, 2)


def get_current_team_rankings(manager_id, year=None, month=None):
    """Get current team rankings with scores"""
    if year is None:
        year = date.today().year
    if month is None:
        month = date.today().month
    
    team_members = db.get_team_members(manager_id)
    
    rankings = []
    for member in team_members:
        # Get goals for specific month
        quarter = ((month - 1) // 3) + 1
        goals = db.get_month_goals(member['id'], year, quarter, month)
        
        if goals:
            stats = calculate_performance_metrics(goals)
            score = calculate_performance_score(stats)
            
            rankings.append({
                'employee_id': member['id'],
                'name': member['name'],
                'email': member['email'],
                'designation': member.get('designation', 'N/A'),
                'department': member.get('department', 'N/A'),
                'total_goals': stats['total_goals'],
                'completed_goals': stats['completed'],
                'completion_rate': stats['completion_rate'],
                'avg_progress': stats['avg_progress'],
                'on_time_rate': stats.get('on_time_rate', 0),
                'score': score
            })
    
    # Sort by score (highest first)
    rankings.sort(key=lambda x: x['score'], reverse=True)
    
    # Assign ranks
    for idx, ranking in enumerate(rankings):
        ranking['rank'] = idx + 1
    
    return rankings


def save_monthly_rankings(manager_id, year, month):
    """Save monthly rankings to database"""
    rankings = get_current_team_rankings(manager_id, year, month)
    
    if not rankings:
        return False
    
    saved_count = 0
    for ranking in rankings:
        try:
            # Check if ranking already exists
            existing = supabase.table('team_rankings').select('id').eq(
                'manager_id', manager_id
            ).eq('employee_id', ranking['employee_id']).eq(
                'year', year
            ).eq('month', month).execute()
            
            ranking_data = {
                'manager_id': manager_id,
                'employee_id': ranking['employee_id'],
                'year': year,
                'month': month,
                'rank': ranking['rank'],
                'total_goals': ranking['total_goals'],
                'completed_goals': ranking['completed_goals'],
                'completion_rate': ranking['completion_rate'],
                'avg_progress': ranking['avg_progress'],
                'score': ranking['score']
            }
            
            if existing.data:
                # Update existing
                supabase.table('team_rankings').update(ranking_data).eq(
                    'id', existing.data[0]['id']
                ).execute()
            else:
                # Insert new
                supabase.table('team_rankings').insert(ranking_data).execute()
            
            saved_count += 1
        except Exception as e:
            print(f"Error saving ranking: {str(e)}")
            continue
    
    return saved_count > 0


def get_historical_rankings(manager_id, employee_id, months=6):
    """Get historical rankings for an employee"""
    try:
        result = supabase.table('team_rankings').select('*').eq(
            'manager_id', manager_id
        ).eq('employee_id', employee_id).order(
            'year', desc=True
        ).order('month', desc=True).limit(months).execute()
        
        return result.data if result.data else []
    except Exception as e:
        print(f"Error getting historical rankings: {str(e)}")
        return []


def get_average_ranking(manager_id, employee_id, months=6):
    """Calculate average ranking for an employee over last N months"""
    history = get_historical_rankings(manager_id, employee_id, months)
    
    if not history:
        return None
    
    total_rank = sum(r['rank'] for r in history)
    avg_rank = total_rank / len(history)
    
    return {
        'avg_rank': round(avg_rank, 1),
        'best_rank': min(r['rank'] for r in history),
        'worst_rank': max(r['rank'] for r in history),
        'months_tracked': len(history)
    }

def render_password_strength_meter(password, key_suffix=""):
    """Render password strength meter with visual feedback"""
    score, color, strength, feedback = check_password_strength(password)
    
    # Strength meter HTML
    st.markdown(f"""
    <div style="margin: 15px 0;">
        <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 8px;">
            <span style="font-size: 13px; font-weight: 600; color: #64748b;">Password Strength:</span>
            <span style="font-size: 14px; font-weight: bold; color: {color};">{strength}</span>
        </div>
        <div style="width: 100%; height: 8px; background: #e2e8f0; border-radius: 10px; overflow: hidden;">
            <div style="width: {score}%; height: 100%; background: {color}; transition: width 0.3s ease;"></div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # Show feedback if password is not very strong
    if score < 85 and feedback:
        with st.expander("💡 Tips to strengthen your password", expanded=False):
            for tip in feedback:
                st.markdown(f"- {tip}")
            
            st.markdown("""
            **Best Practices:**
            - Use a mix of uppercase and lowercase letters
            - Include numbers and special characters
            - Avoid common words or patterns
            - Use at least 8-12 characters
            - Don't reuse passwords from other accounts
            """)
def send_password_reset_email(email, reset_token):
    """Send password reset email"""
    try:
        smtp_server = os.getenv('SMTP_SERVER', 'smtp.gmail.com')
        smtp_port = int(os.getenv('SMTP_PORT', '587'))
        smtp_email = os.getenv('SMTP_EMAIL')
        smtp_password = os.getenv('SMTP_PASSWORD')
        
        if not smtp_email or not smtp_password:
            st.error("Email configuration not found. Please contact administrator.")
            return False
        
        # Create reset link (you'll need to handle this in your app)
        reset_link = f"http://your-app-url.com/reset-password?token={reset_token}"
        
        # Email content
        subject = "Password Reset Request - Performance Management System"
        body = f"""
        <html>
            <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
                <div style="max-width: 600px; margin: 0 auto; padding: 20px; border: 1px solid #ddd; border-radius: 10px;">
                    <h2 style="color: #3b82f6; text-align: center;">🎯 Password Reset Request</h2>
                    <p>Hello,</p>
                    <p>You requested to reset your password for the Performance Management System.</p>
                    <p><strong>Your Reset Token:</strong></p>
                    <div style="background: #f3f4f6; padding: 15px; border-radius: 5px; text-align: center; font-size: 24px; font-weight: bold; letter-spacing: 2px; color: #1e40af;">
                        {reset_token}
                    </div>
                    <p style="margin-top: 20px;">Please use this token to reset your password. This token will expire in 1 hour.</p>
                    <p style="color: #ef4444;"><strong>⚠️ Important:</strong> If you did not request this password reset, please ignore this email or contact your administrator.</p>
                    <hr style="margin: 30px 0; border: none; border-top: 1px solid #ddd;">
                    <p style="font-size: 12px; color: #64748b; text-align: center;">
                        This is an automated email from Performance Management System.<br>
                        Please do not reply to this email.
                    </p>
                </div>
            </body>
        </html>
        """
        
        # Create message
        message = MIMEMultipart("alternative")
        message["Subject"] = subject
        message["From"] = smtp_email
        message["To"] = email
        
        html_part = MIMEText(body, "html")
        message.attach(html_part)
        
        # Send email
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(smtp_email, smtp_password)
            server.sendmail(smtp_email, email, message.as_string())
        
        return True
    except Exception as e:
        st.error(f"Failed to send email: {str(e)}")
        return False


# ✅ PAGE CONFIG MUST BE FIRST STREAMLIT COMMAND
st.set_page_config(
    page_title="Performance Management System",
    page_icon="infopacee.jpg",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Load environment variables
load_dotenv()

# Initialize Supabase client ONCE
supabase = get_supabase_client()


# Import our modules
from database import Database
from helper import (
    apply_theme, init_session_state, get_quarter_months, get_month_name,
    get_quarter_name, calculate_progress, format_goal_table_data,
    render_user_avatar, render_card, render_metric_card, render_progress_bar,
    render_feedback_card, validate_goal_data
)



# Initialize session state and database
init_session_state()
db = Database()
apply_theme()

# IST Timezone
IST = pytz.timezone('Asia/Kolkata')

# Initialize session state and database
init_session_state()
db = Database()
apply_theme()

# IST Timezone
IST = pytz.timezone('Asia/Kolkata')

if 'reminder_scheduler_started' not in st.session_state:
    start_reminder_scheduler(db)
    st.session_state.reminder_scheduler_started = True

# Start daily notification checker
# ✅ DISABLE AUTO CHECKER - It's creating duplicate notifications
# Instead, we'll check only when goals are created/updated

# Start daily notification checker - DISABLED
# if 'notification_checker_started' not in st.session_state:
#     import threading
#     import time
#     
#     def run_daily_checker():
#         while True:
#             check_and_notify_due_dates_and_missing_updates()
#             # Run once per day (86400 seconds)
#             time.sleep(86400)
#     
#     checker_thread = threading.Thread(target=run_daily_checker, daemon=True)
#     checker_thread.start()
#     st.session_state.notification_checker_started = True

print("⚠️ Daily notification checker disabled to prevent duplicates")
def manually_check_notifications():
    """Manual notification check - for HR/Admin use"""
    st.title(" Manual Notification Check")
    
    user = st.session_state.user
    if user['role'] not in ['HR', 'CMD', 'VP']:
        st.warning(" Only HR/CMD/VP can run manual notification checks")
        return
    
    st.info("This will check all active goals and send notifications for:")
    st.markdown("""
    - Goals due in 1-5 days
    - Goals past deadline (not completed)
    - Missing weekly updates
    """)
    
    if st.button(" Run Notification Check", type="primary"):
        with st.spinner("Checking all goals..."):
            check_and_notify_due_dates_and_missing_updates()
            st.success("✅ Notification check completed!")
            st.info("Check the console/logs for details")
    
    st.markdown("---")
    
    # Show notification stats
    st.subheader(" Notification Statistics")
    
    all_users = db.get_all_users()
    total_notifs = 0
    unread_notifs = 0
    
    for u in all_users:
        user_notifs = get_user_notifications(u['id'], limit=1000)
        total_notifs += len(user_notifs)
        unread_notifs += len([n for n in user_notifs if not n.get('is_read')])
    
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Total Notifications", total_notifs)
    with col2:
        st.metric("Unread", unread_notifs)
    with col3:
        st.metric("Read", total_notifs - unread_notifs)
# ============================================
# SESSION PERSISTENCE
# ============================================
def save_session_to_storage():
    """Save session state to browser storage"""
    if st.session_state.user:
        try:
            # Store in Streamlit's query params (URL-based persistence)
            st.query_params['user_id'] = str(st.session_state.user['id'])
            st.query_params['page'] = st.session_state.page
            
            # ✅ Save navigation state
            if st.session_state.get('selected_year'):
                st.query_params['year'] = str(st.session_state.selected_year)
            if st.session_state.get('selected_quarter'):
                st.query_params['quarter'] = str(st.session_state.selected_quarter)
            if st.session_state.get('selected_month'):
                st.query_params['month'] = str(st.session_state.selected_month)
            
            # ✅ Save active tab states
            if st.session_state.get('active_month_tab'):
                st.query_params['month_tab'] = str(st.session_state.active_month_tab)
            if st.session_state.get('active_hr_tab'):
                st.query_params['hr_tab'] = str(st.session_state.active_hr_tab)
            
            # ✅ Save employee viewing state
            if st.session_state.get('viewing_employee'):
                st.query_params['viewing_emp_id'] = str(st.session_state.viewing_employee['id'])
            
            if st.session_state.get('viewing_employee_year'):
                st.query_params['viewing_emp_year'] = 'true'
            
        except Exception as e:
            print(f"Session save error: {str(e)}")


def restore_session_from_storage():
    """Restore session state from browser storage"""
    try:
        # Check URL params for user_id
        user_id = st.query_params.get('user_id')
        if user_id:
            # Restore user from database
            user = db.get_user_by_id(user_id)
            if user:
                st.session_state.user = user
                st.session_state.page = st.query_params.get('page', 'dashboard')
                
                # ✅ Restore navigation state
                year_param = st.query_params.get('year')
                if year_param:
                    st.session_state.selected_year = int(year_param)
                
                quarter_param = st.query_params.get('quarter')
                if quarter_param:
                    st.session_state.selected_quarter = int(quarter_param)
                
                month_param = st.query_params.get('month')
                if month_param:
                    st.session_state.selected_month = int(month_param)
                
                # ✅ Restore tab states
                month_tab = st.query_params.get('month_tab')
                if month_tab:
                    st.session_state.active_month_tab = int(month_tab)
                
                hr_tab = st.query_params.get('hr_tab')
                if hr_tab:
                    st.session_state.active_hr_tab = int(hr_tab)
                
                # ✅ Restore employee viewing state
                viewing_emp_id = st.query_params.get('viewing_emp_id')
                if viewing_emp_id:
                    viewing_employee = db.get_user_by_id(viewing_emp_id)
                    if viewing_employee:
                        st.session_state.viewing_employee = viewing_employee
                
                viewing_emp_year = st.query_params.get('viewing_emp_year')
                if viewing_emp_year == 'true':
                    st.session_state.viewing_employee_year = True
                
                return True
    except Exception as e:
        print(f"Session restore error: {str(e)}")
    
    return False


# ✅ RESTORE SESSION ON PAGE LOAD (if not already logged in)
if not st.session_state.user:
    restore_session_from_storage()

# ✅ SAVE SESSION ON CHANGES (if user is logged in)
if st.session_state.user:
    save_session_to_storage()



# ============================================
# WEEK DATE CALCULATION HELPERS
# ============================================
def get_week_dates(year, month, week_num):
    """Get start and end dates for a specific week in a month"""
    # Get first day of month
    first_day = date(year, month, 1)
    # Get last day of month
    last_day = date(year, month, calendar.monthrange(year, month)[1])
    
    # Calculate week ranges (approximate 7-day periods)
    if week_num == 1:
        start_date = first_day
        end_date = min(first_day + timedelta(days=6), last_day)
    elif week_num == 2:
        start_date = first_day + timedelta(days=7)
        end_date = min(first_day + timedelta(days=13), last_day)
    elif week_num == 3:
        start_date = first_day + timedelta(days=14)
        end_date = min(first_day + timedelta(days=20), last_day)
    else:  # week 4
        start_date = first_day + timedelta(days=21)
        end_date = last_day
    
    return start_date, end_date


def get_week_for_date(year, month, target_date):
    """Determine which week a date falls into"""
    first_day = date(year, month, 1)
    day_diff = (target_date - first_day).days
    
    if day_diff < 7:
        return 1
    elif day_diff < 14:
        return 2
    elif day_diff < 21:
        return 3
    else:
        return 4

def create_goal_with_date_based_placement(goal_data):
    """Create goal and place it in the correct month based on start/end dates"""
    start_date = datetime.strptime(goal_data['start_date'], '%Y-%m-%d').date()
    end_date = datetime.strptime(goal_data['end_date'], '%Y-%m-%d').date()
    
    # Determine the month based on start date
    goal_year = start_date.year
    goal_month = start_date.month
    goal_quarter = ((goal_month - 1) // 3) + 1
    
    # Update goal data with correct year, quarter, month
    goal_data['year'] = goal_year
    goal_data['quarter'] = goal_quarter
    goal_data['month'] = goal_month
    
    # Create the goal
    result = db.create_goal(goal_data)
    
    return result, goal_year, goal_quarter, goal_month

import base64

def get_base64_image(image_path):
    with open(image_path, "rb") as img_file:
        encoded = base64.b64encode(img_file.read()).decode()
    return encoded
# ============================================
# LOGIN PAGE
# ============================================
def login_page():
    """Display login page with forgot password option"""
    base64_img = get_base64_image("infopacee.jpg")

    st.markdown(f"""
    <div style="display:flex; justify-content:center; align-items:center; width:100%; margin-top:8px;">
        <img src="data:image/jpg;base64,{base64_img}" style="width:100px; height:100px;">
    </div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <style>
        .stApp {
            background-color: white !important;
        }
    </style>
    """, unsafe_allow_html=True)
    
    st.markdown(
    """
    <h1 style="
        text-align: center;
        font-size: 42px;
        font-weight: 800;
        margin-left: 56px;
        color: #0F172A;
    ">
        Performance Management System
    </h1>
    """,
    unsafe_allow_html=True
)

    st.markdown("<p style='text-align: center; color: #64748b;'>Sign in to continue</p>", unsafe_allow_html=True)
    
    # Check if showing forgot password or reset password
    if 'show_forgot_password' not in st.session_state:
        st.session_state.show_forgot_password = False
    
    if 'show_reset_password' not in st.session_state:
        st.session_state.show_reset_password = False
    
    col1, col2, col3 = st.columns([1, 2, 1])
    
    with col2:
        # ===== NORMAL LOGIN FORM =====
        if not st.session_state.show_forgot_password and not st.session_state.show_reset_password:
            with st.form("login_form"):
                email = st.text_input(" Email", placeholder="your@email.com")
                password = st.text_input(" Password", type="password", placeholder="••••••••")
                submit = st.form_submit_button("Sign In", use_container_width=True)
                
                if submit:
                    if email and password:
                        user = db.authenticate_user(email, password)
                        if user:
                            st.session_state.user = user
                            st.session_state.page = 'dashboard'  # ✅ Set default page
                            save_session_to_storage()  # ✅ Save session
                            st.success("✅ Login successful!")
                            st.rerun()
                        else:
                            st.error("❌ Invalid credentials")
                    else:
                        st.warning("⚠️ Please enter both email and password")
            
            # Forgot Password Link
            col_fp1, col_fp2, col_fp3 = st.columns([1, 2, 1])
            with col_fp2:
                if st.button("🔑 Forgot Password?", use_container_width=True):
                    st.session_state.show_forgot_password = True
                    st.rerun()
            
            
        
        # ===== FORGOT PASSWORD FORM =====
        elif st.session_state.show_forgot_password:
            st.markdown("### 🔑 Forgot Password")
            st.info("Enter your email address and we'll send you a reset token.")
            
            with st.form("forgot_password_form"):
                reset_email = st.text_input("📧 Email Address", placeholder="your@email.com")
                submit_reset = st.form_submit_button("Send Reset Token", use_container_width=True)
                
                if submit_reset:
                    if reset_email:
                        # Generate token
                        token = db.create_password_reset_token(reset_email)
                        
                        if token:
                            # Send email
                            if send_password_reset_email(reset_email, token):
                                st.success("✅ Password reset token sent to your email!")
                                st.info(f"**Your Reset Token:** `{token}`\n\nPlease check your email for the reset token. It will expire in 1 hour.")
                                st.session_state.show_forgot_password = False
                                st.session_state.show_reset_password = True
                                st.rerun()
                            else:
                                st.warning(f"⚠️ Could not send email, but here's your reset token: **{token}**")
                                st.session_state.show_forgot_password = False
                                st.session_state.show_reset_password = True
                                st.rerun()
                        else:
                            st.error("❌ Email not found in system")
                    else:
                        st.warning("⚠️ Please enter your email address")
            
            # Back to Login
            if st.button("← Back to Login", use_container_width=True):
                st.session_state.show_forgot_password = False
                st.rerun()
        
        # ===== RESET PASSWORD FORM =====
        elif st.session_state.show_reset_password:
            st.markdown("### 🔒 Reset Password")
            st.info("Enter the reset token you received via email and your new password.")
            
            with st.form("reset_password_form"):
                reset_token = st.text_input("🎫 Reset Token", placeholder="Enter 8-character token")
                new_password = st.text_input("🔒 New Password", type="password", placeholder="••••••••", key="login_new_pass")
                
                # Password strength meter
                if new_password:
                    render_password_strength_meter(new_password, "login_reset")
                
                confirm_password = st.text_input("🔒 Confirm Password", type="password", placeholder="••••••••", key="login_confirm_pass")
                submit_new_password = st.form_submit_button("Reset Password", use_container_width=True)
                
                if submit_new_password:
                    if reset_token and new_password and confirm_password:
                        if new_password == confirm_password:
                            if len(new_password) >= 6:
                                # Check password strength
                                score, _, strength, _ = check_password_strength(new_password)
                                
                                if score < 30:
                                    st.warning(f"⚠️ Your password is {strength}. Consider making it stronger for better security.")
                                
                                # Reset password
                                if db.reset_password_with_token(reset_token, new_password):
                                    st.success("✅ Password reset successful! You can now login with your new password.")
                                    st.balloons()
                                    st.session_state.show_reset_password = False
                                    st.query_params.clear()
                                    st.rerun()
                                else:
                                    st.error("❌ Invalid or expired token")
                            else:
                                st.error("❌ Password must be at least 6 characters")
                        else:
                            st.error("❌ Passwords don't match")
                    else:
                        st.warning("⚠️ Please fill all fields")
            
            # Back to Login
            if st.button("← Back to Login", use_container_width=True):
                st.session_state.show_reset_password = False
                st.rerun()

# ============================================
# ENHANCED DASHBOARD PAGE
# ============================================
def display_dashboard():
    """Display enhanced professional dashboard"""
    user = st.session_state.user
    role = user['role']

    if not user:
        st.warning("⚠️ Session expired. Please login again.")
        st.rerun()
    today = date.today()
    current_year = today.year
    current_month = today.month

    # ✅ INITIALIZE ALL VARIABLES AT THE START
    avg_progress = 0
    total_goals = 0
    completed_goals = 0
    active_goals = 0
    overdue_goals = 0
    user_goals = []

    # Welcome Header with gradient
    role_greetings = {
        'CMD': ' Chief Managing Director',
        'VP': ' Vice President',
        'HR': ' Human Resources',
        'Manager': ' Manager',
        'Employee': ' Employee'
    }

    st.markdown(f"""
    <div style="background: linear-gradient(135deg, #2DCCFF, #9BBCE0);
                padding: 30px; border-radius: 15px; margin-bottom: 30px; box-shadow: 0 4px 6px rgba(0,0,0,0.1);">
        <h1 style="color: white; margin: 0; font-size: 32px;">Welcome back, {user['name']}! </h1>
        <p style="color: rgba(255,255,255,0.9); margin: 10px 0 0 0; font-size: 16px;">
            {role_greetings.get(role, user.get('designation', 'Employee'))} || {datetime.now().strftime('%A, %B %d, %Y')}
        </p>
    </div>
    """, unsafe_allow_html=True)

    # Custom CSS for metric card buttons
    st.markdown("""
    <style>
    /* Style for metric card buttons */
    div[data-testid="column"] button {
        height: 120px !important;
        padding: 20px !important;
        font-size: 16px !important;
        white-space: pre-line !important;
        line-height: 1.4 !important;
    }
    
    div[data-testid="column"] button:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(0,0,0,0.15) !important;
        transition: all 0.3s ease;
    }
    
    div[data-testid="column"] button p {
        font-size: 14px !important;
    }
    </style>
    """, unsafe_allow_html=True)

    # Add CMD/VP/HR specific organization-wide metrics
    if role in ['CMD', 'VP', 'HR']:
        # ============================================
        # ORGANIZATION PERFORMANCE OVERVIEW (MONTHLY)
        # ============================================
        st.markdown("### Organization Overview")

        # Month/Year selector with current month default
        today = date.today()
        current_year = today.year
        current_month = today.month

        col_filter1, col_filter2, col_filter3 = st.columns([2, 2, 1])

        with col_filter1:
            # Team filter dropdown
            all_users = db.get_all_users()
            departments = list(set([normalize_department(u.get('department')) for u in all_users]))
            departments = ['All Teams'] + sorted([d for d in departments if d and d != 'UNASSIGNED'])
                    
            selected_team = st.selectbox(
                " Filter by Team",
                departments,
                key="team_filter_select"
            )

        with col_filter2:
            # User filter dropdown
            if selected_team == 'All Teams':
                user_options = ['All Users'] + [f"{u['name']} ({u['role']}) - {u['email']}" for u in all_users]
            else:
                filtered_users = [u for u in all_users if u.get('department') == selected_team]
                user_options = ['All Users'] + [f"{u['name']} ({u['role']}) - {u['email']}" for u in filtered_users]
            
            selected_user_option = st.selectbox(
                "👤 Filter by User",
                user_options,
                key="user_filter_select"
            )

        with col_filter3:
            selected_month = st.selectbox(
                "📅 Month",
                list(range(1, 13)),
                index=current_month - 1,
                format_func=lambda x: get_month_name(x),
                key="org_month_select"
            )

        # Use current year only
        selected_year = current_year

        # Build filter display text
        filter_parts = []
        if selected_team != 'All Teams':
            filter_parts.append(f"**{selected_team}**")
        if selected_user_option != 'All Users':
            user_name = selected_user_option.split(' (')[0]
            filter_parts.append(f"**{user_name}**")

        filter_text = " → ".join(filter_parts) if filter_parts else "**All Users**"
        st.caption(f"Showing: {get_month_name(selected_month)} {selected_year} • {filter_text}")

        # Get month goals (filtered by team and user)
        month_goals = get_organization_month_goals(selected_month, selected_year)

        # Apply team filter
        if selected_team != 'All Teams':
            month_goals = [g for g in month_goals if normalize_department(g.get('user_department')) == selected_team]

        # Apply user filter
        if selected_user_option != 'All Users':
            user_email = selected_user_option.split(' - ')[1]
            selected_user_obj = next(u for u in all_users if u['email'] == user_email)
            month_goals = [g for g in month_goals if g['user_id'] == selected_user_obj['id']]

        # Calculate metrics
        total_goals = len(month_goals)
        completed_goals = len([g for g in month_goals if g.get('status') == 'Completed'])
        active_goals = len([g for g in month_goals if g.get('status') == 'Active'])

        # Calculate average progress
        total_progress = 0
        goals_with_progress = 0
        for goal in month_goals:
            monthly_achievement = goal.get('monthly_achievement')
            if monthly_achievement is not None:
                monthly_target = goal.get('monthly_target', 1)
                if monthly_target > 0:
                    progress = (monthly_achievement / monthly_target * 100)
                    total_progress += progress
                    goals_with_progress += 1

        avg_progress = (total_progress / goals_with_progress) if goals_with_progress > 0 else 0

        # Count overdue goals
        overdue_goals = 0
        month_end = date(selected_year, selected_month, calendar.monthrange(selected_year, selected_month)[1])
        for goal in month_goals:
            if goal.get('status') == 'Active':
                end_date_str = goal.get('end_date')
                if end_date_str:
                    try:
                        end_date = datetime.strptime(str(end_date_str), '%Y-%m-%d').date()
                        if month_end > end_date:
                            overdue_goals += 1
                    except:
                        pass

        # Display clickable metric cards
        col_perf1, col_perf2, col_perf3, col_perf4, col_perf5 = st.columns(5)

        with col_perf1:
            metric_card(
                value=total_goals,
                label="Total Goals",
                color="#3B82F6",
                shadow_rgba="rgba(59,130,246,0.25)"
            )

        with col_perf2:
            metric_card(
                value=completed_goals,
                label="Completed",
                color="#10B981",
                shadow_rgba="rgba(16,185,129,0.22)",
            )

        with col_perf3:
            metric_card(
                value=active_goals,
                label="Active",
                color="#F5576C",
                shadow_rgba="rgba(245,87,108,0.23)",
            )

        with col_perf4:
            metric_card(
                value=f"{avg_progress:.1f}",
                label="Avg Progress",
                color="#00C9FF",
                shadow_rgba="rgba(0,201,255,0.23)",
                suffix="%"
            )

        with col_perf5:
            is_overdue = overdue_goals > 0

            metric_card(
                value=overdue_goals,
                label="Overdue",
                color="#EF4444" if is_overdue else "#10B981",
                shadow_rgba="rgba(239,68,68,0.23)" if is_overdue else "rgba(16,185,129,0.23)",
                icon="" if is_overdue else "",
                label_bg="#FEF2F2" if is_overdue else "#ECFDF5"
            )

        st.markdown("---")
        # ============================================
        # DETAILS MODAL (when card is clicked)
        # ============================================
        if st.session_state.get('show_details') and st.session_state.get('show_details').startswith('org_'):
            st.markdown("---")
            detail_type = st.session_state.show_details
            
            col_header, col_close = st.columns([5, 1])
            
            with col_header:
                if detail_type == 'org_total':
                    st.subheader(f" All Goals ({len(month_goals)}) - {filter_text}")
                elif detail_type == 'org_completed':
                    st.subheader(f" Completed Goals ({completed_goals})")
                elif detail_type == 'org_active':
                    st.subheader(f" Active Goals ({active_goals})")
                elif detail_type == 'org_progress':
                    st.subheader(f" Progress Breakdown")
                elif detail_type == 'org_overdue':
                    st.subheader(f" Overdue Goals ({overdue_goals})")
            
            with col_close:
                if st.button("✕ Close", key="close_org_details"):
                    del st.session_state.show_details
                    st.rerun()
            
            # Filter goals based on detail type
            if detail_type == 'org_completed':
                display_goals = [g for g in month_goals if g.get('status') == 'Completed']
            elif detail_type == 'org_active':
                display_goals = [g for g in month_goals if g.get('status') == 'Active']
            elif detail_type == 'org_overdue':
                display_goals = []
                for goal in month_goals:
                    if goal.get('status') == 'Active':
                        end_date_str = goal.get('end_date')
                        if end_date_str:
                            try:
                                end_date = datetime.strptime(str(end_date_str), '%Y-%m-%d').date()
                                if month_end > end_date:
                                    display_goals.append(goal)
                            except:
                                pass
            elif detail_type == 'org_progress':
                display_goals = month_goals.copy()
                for g in display_goals:
                    achievement = g.get('monthly_achievement')
                    target = g.get('monthly_target', 1)
                    achievement_val = 0 if achievement is None else achievement
                    target_val = 1 if target is None or target == 0 else target
                    g['_progress'] = calculate_progress(achievement_val, target_val)
                display_goals.sort(key=lambda x: x['_progress'], reverse=True)
            else:
                display_goals = month_goals
            
            # Display goals in table
            if display_goals:
                goal_data = []
                for goal in display_goals:
                    monthly_achievement = goal.get('monthly_achievement')
                    monthly_target = goal.get('monthly_target', 1)
                    achievement_value = 0 if monthly_achievement is None else monthly_achievement
                    target_value = 1 if monthly_target is None or monthly_target == 0 else monthly_target
                    progress = calculate_progress(achievement_value, target_value)
                    
                    goal_data.append({
                        'Employee': goal['user_name'],
                        'Role': goal['user_role'],
                        'Department': goal['user_department'],
                        'Goal Title': goal['goal_title'],
                        'Month': get_month_name(goal.get('month', 1)),  # ADD THIS
                        'Year': str(goal['year']),  # ADD THIS
                        'Target': monthly_target if monthly_target is not None else 0,
                        'Achievement': monthly_achievement if monthly_achievement is not None else '-',
                        'Progress': f"{progress:.1f}%",
                        'Status': goal.get('status', 'Active')
                    })
                
                df_goals = pd.DataFrame(goal_data)
                st.dataframe(df_goals, use_container_width=True, height=400)
                
                # Export button
                st.download_button(
                    "📥 Export to CSV",
                    df_goals.to_csv(index=False).encode('utf-8'),
                    f"org_goals_{detail_type}_{selected_month}_{selected_year}.csv",
                    "text/csv",
                    key='download_org_details'
                )
            else:
                st.info("No goals found in this category")
            
            st.markdown("---")

        
    
        # ============================================
        # TEAM PERFORMANCE (ROLE-SPECIFIC TEAMS)
        # ============================================
        if role == 'CMD':
            
            st.markdown("### 👥 My Team (VPs)")
            
            # Month selector for CMD's team
            col_cmd_team_filter1, col_cmd_team_filter2 = st.columns([4, 1])
            
            with col_cmd_team_filter1:
                st.caption("Filter VP team performance by month")
            
            with col_cmd_team_filter2:
                cmd_team_month = st.selectbox(
                    "Month",
                    list(range(1, 13)),
                    index=current_month - 1,
                    format_func=lambda x: get_month_name(x),
                    key="cmd_team_month_select"
                )
            
            st.caption(f"**VP Team Performance: {get_month_name(cmd_team_month)} {current_year}**")
            
            # Get all VPs (CMD's team)
            all_users = db.get_all_users()
            vp_team = [u for u in all_users if u['role'] == 'VP']
            
            # Calculate VP team metrics for selected month
            vp_team_total_goals = 0
            vp_team_completed_goals = 0
            vp_team_active_goals = 0
            vp_team_total_progress = 0
            vp_team_goals_count = 0
            vp_team_overdue_goals = 0
            
            today_date = date.today()
            
            for vp in vp_team:
                vp_month_goals = get_user_month_goals(vp['id'], cmd_team_month, current_year)
                vp_team_total_goals += len(vp_month_goals)
                vp_team_completed_goals += len([g for g in vp_month_goals if g.get('status') == 'Completed'])
                vp_team_active_goals += len([g for g in vp_month_goals if g.get('status') == 'Active'])
                
                for goal in vp_month_goals:
                    achievement = goal.get('monthly_achievement')
                    if achievement is not None:
                        target = goal.get('monthly_target', 1)
                        if target > 0:
                            progress = (achievement / target * 100)
                            vp_team_total_progress += progress
                            vp_team_goals_count += 1
                    
                    # Count overdue goals
                    if goal.get('status') == 'Active':
                        end_date_str = goal.get('end_date')
                        if end_date_str:
                            try:
                                end_date = datetime.strptime(str(end_date_str), '%Y-%m-%d').date()
                                if today_date > end_date:
                                    vp_team_overdue_goals += 1
                            except:
                                pass
            
            vp_team_avg_progress = (vp_team_total_progress / vp_team_goals_count) if vp_team_goals_count > 0 else 0
            
            col_cmd1, col_cmd2, col_cmd3, col_cmd4, col_cmd5 = st.columns(5)
            
            with col_cmd1:
                st.markdown(
                    f"""
                    <div style="
                        background:#FFFFFF;
                        width:100%;
                        height:160px;
                        padding:20px;
                        border-radius:10px;
                        text-align:center;
                        display:flex;
                        flex-direction:column;
                        justify-content:center;
                        box-shadow:
                            0 2px 4px rgba(0,0,0,0.08),
                            0 0 18px rgba(118,75,162,0.23);
                    ">
                        <div style="font-size:36px; margin-bottom:10px;"></div>

                        <div style="font-size:32px; font-weight:700; color:#764BA2; margin-bottom:6px;">
                            {len(vp_team)}
                        </div>

                        <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:#764BA2;
                            padding:4px 10px; border-radius:6px; display:inline-block;">
                            VPs
                        </div>
                    </div>
                    """,
                    unsafe_allow_html=True
                )

            
            with col_cmd2:
                st.markdown(
                    f"""
                    <div style="
                        background:#FFFFFF;
                        width:100%;
                        height:160px;
                        padding:20px;
                        border-radius:10px;
                        text-align:center;
                        display:flex;
                        flex-direction:column;
                        justify-content:center;
                        box-shadow:
                            0 2px 4px rgba(0,0,0,0.08),
                            0 0 18px rgba(56,249,215,0.23);
                    ">
                        <div style="font-size:36px; margin-bottom:10px;"></div>

                        <div style="font-size:32px; font-weight:700; color:#38F9D7; margin-bottom:6px;">
                            {vp_team_total_goals}
                        </div>

                        <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:#38F9D7;
                            padding:4px 10px; border-radius:6px; display:inline-block;">
                            VP Goals
                        </div>
                    </div>
                    """,
                    unsafe_allow_html=True
                )

            
            with col_cmd3:
                st.markdown(
                    f"""
                    <div style="
                        background:#FFFFFF;
                        width:100%;
                        height:160px;
                        padding:20px;
                        border-radius:10px;
                        text-align:center;
                        display:flex;
                        flex-direction:column;
                        justify-content:center;
                        box-shadow:
                            0 2px 4px rgba(0,0,0,0.08),
                            0 0 18px rgba(245,87,108,0.23);
                    ">
                        <div style="font-size:36px; margin-bottom:10px;"></div>

                        <div style="font-size:32px; font-weight:700; color:#F5576C; margin-bottom:6px;">
                            {vp_team_completed_goals}
                        </div>

                        <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:#F5576C;
                            padding:4px 10px; border-radius:6px; display:inline-block;">
                            Completed
                        </div>
                    </div>
                    """,
                    unsafe_allow_html=True
                )

            
            with col_cmd4:
                st.markdown(
                    f"""
                    <div style="
                        background:#FFFFFF;
                        width:100%;
                        height:160px;
                        padding:20px;
                        border-radius:10px;
                        text-align:center;
                        display:flex;
                        flex-direction:column;
                        justify-content:center;
                        box-shadow:
                            0 2px 4px rgba(0,0,0,0.08),
                            0 0 18px rgba(245,158,11,0.23);
                    ">
                        <div style="font-size:36px; margin-bottom:10px;"></div>

                        <div style="font-size:32px; font-weight:700; color:#F59E0B; margin-bottom:6px;">
                            {vp_team_avg_progress:.1f}%
                        </div>

                        <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:#F59E0B;
                            padding:4px 10px; border-radius:6px; display:inline-block;">
                            Avg Progress
                        </div>
                    </div>
                    """,
                    unsafe_allow_html=True
                )

                        
            with col_cmd5:
                overdue_color = "#EF4444" if vp_team_overdue_goals > 0 else "#10B981"
                label_bg = "#FEF2F2" if vp_team_overdue_goals > 0 else "#ECFDF5"
                glow = "rgba(239,68,68,0.23)" if vp_team_overdue_goals > 0 else "rgba(16,185,129,0.23)"
                icon = "" if vp_team_overdue_goals > 0 else ""

                st.markdown(
                    f"""
                    <div style="
                        background:#FFFFFF;
                        width:100%;
                        height:160px;
                        padding:20px;
                        border-radius:10px;
                        text-align:center;
                        display:flex;
                        flex-direction:column;
                        justify-content:center;
                        box-shadow:
                            0 2px 4px rgba(0,0,0,0.08),
                            0 0 18px {glow};
                    ">
                        <div style="font-size:36px; margin-bottom:10px;">{icon}</div>

                        <div style="font-size:32px; font-weight:700; color:{overdue_color}; margin-bottom:6px;">
                            {vp_team_overdue_goals}
                        </div>

                        <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:{overdue_color};
                            background:{label_bg}; padding:4px 10px; border-radius:6px; display:inline-block;">
                            Overdue
                        </div>
                    </div>
                    """,
                    unsafe_allow_html=True
                )

            st.markdown("---")
            # CMD Team details modal
            if st.session_state.get('show_details') and st.session_state.get('show_details').startswith('cmd_team_'):
                st.markdown("---")
                detail_type = st.session_state.show_details
                
                col_header, col_close = st.columns([5, 1])
                
                with col_header:
                    if detail_type == 'cmd_team_members':
                        st.subheader(f" VP Team Members ({len(vp_team)})")
                    elif detail_type == 'cmd_team_goals':
                        st.subheader(f" All VP Goals ({vp_team_total_goals}) - {get_month_name(cmd_team_month)} {current_year}")
                    elif detail_type == 'cmd_team_completed':
                        st.subheader(f" VP Completed Goals ({vp_team_completed_goals})")
                    elif detail_type == 'cmd_team_progress':
                        st.subheader(f" VP Progress Breakdown")
                    elif detail_type == 'cmd_team_overdue':
                        st.subheader(f" VP Overdue Goals ({vp_team_overdue_goals})")
                
                with col_close:
                    if st.button("✕ Close", key="close_cmd_team_details"):
                        del st.session_state.show_details
                        st.rerun()
                
                # Show CMD team details
                if detail_type == 'cmd_team_members':
                    # Display VP members list
                    members_data = []
                    for vp in vp_team:
                        vp_goals = get_user_month_goals(vp['id'], cmd_team_month, current_year)
                        completed = len([g for g in vp_goals if g.get('status') == 'Completed'])
                        
                        # Count VP's overdue goals
                        vp_overdue = 0
                        for goal in vp_goals:
                            if goal.get('status') == 'Active':
                                end_date_str = goal.get('end_date')
                                if end_date_str:
                                    try:
                                        end_date = datetime.strptime(str(end_date_str), '%Y-%m-%d').date()
                                        if today_date > end_date:
                                            vp_overdue += 1
                                    except:
                                        pass
                        
                        members_data.append({
                            'Name': vp['name'],
                            'Email': vp['email'],
                            'Department': vp.get('department', 'N/A'),
                            'Total Goals': len(vp_goals),
                            'Completed': completed,
                            'Overdue': vp_overdue
                        })
                    
                    df_members = pd.DataFrame(members_data)
                    st.dataframe(df_members, use_container_width=True, height=400)
                    
                    st.download_button(
                        "📥 Export to CSV",
                        df_members.to_csv(index=False).encode('utf-8'),
                        f"cmd_vp_team_{cmd_team_month}_{current_year}.csv",
                        "text/csv",
                        key='download_cmd_team_members'
                    )
                    
                else:
                    # Collect all VP team goals
                    all_vp_team_goals = []
                    for vp in vp_team:
                        vp_goals = get_user_month_goals(vp['id'], cmd_team_month, current_year)
                        for goal in vp_goals:
                            goal['member_name'] = vp['name']
                        all_vp_team_goals.extend(vp_goals)
                    
                    # Filter based on detail type
                    if detail_type == 'cmd_team_completed':
                        display_goals = [g for g in all_vp_team_goals if g.get('status') == 'Completed']
                    elif detail_type == 'cmd_team_overdue':
                        display_goals = []
                        for goal in all_vp_team_goals:
                            if goal.get('status') == 'Active':
                                end_date_str = goal.get('end_date')
                                if end_date_str:
                                    try:
                                        end_date = datetime.strptime(str(end_date_str), '%Y-%m-%d').date()
                                        if today_date > end_date:
                                            display_goals.append(goal)
                                    except:
                                        pass
                    elif detail_type == 'cmd_team_progress':
                        display_goals = all_vp_team_goals.copy()
                        for g in display_goals:
                            achievement = g.get('monthly_achievement')
                            target = g.get('monthly_target', 1)
                            achievement_val = 0 if achievement is None else achievement
                            target_val = 1 if target is None or target == 0 else target
                            g['_progress'] = calculate_progress(achievement_val, target_val)
                        display_goals.sort(key=lambda x: x['_progress'], reverse=True)
                    else:
                        display_goals = all_vp_team_goals
                    
                    if display_goals:
                        goal_data = []
                        for goal in display_goals:
                            monthly_achievement = goal.get('monthly_achievement')
                            monthly_target = goal.get('monthly_target', 1)
                            achievement_value = 0 if monthly_achievement is None else monthly_achievement
                            target_value = 1 if monthly_target is None or monthly_target == 0 else monthly_target
                            progress = calculate_progress(achievement_value, target_value)
                            
                            goal_data.append({
                                'VP': goal['member_name'],
                                'Goal Title': goal['goal_title'],
                                'Department': goal.get('department', 'N/A'),
                                'Month': get_month_name(goal.get('month', 1)),  # ADD THIS
                                'Year': str(goal['year']),  # ADD THIS
                                'Target': monthly_target if monthly_target is not None else 0,
                                'Achievement': monthly_achievement if monthly_achievement is not None else '-',
                                'Progress': f"{progress:.1f}%",
                                'Status': goal.get('status', 'Active')
                            })
                        
                        df_goals = pd.DataFrame(goal_data)
                        st.dataframe(df_goals, use_container_width=True, height=400)
                        
                        st.download_button(
                            "📥 Export to CSV",
                            df_goals.to_csv(index=False).encode('utf-8'),
                            f"cmd_vp_goals_{detail_type}_{cmd_team_month}_{current_year}.csv",
                            "text/csv",
                            key='download_cmd_team_modal_details'
                        )
                    else:
                        st.info("No goals found in this category")
                
                st.markdown("---")
        elif role == 'VP':
                
                st.markdown("### My Team (HR & Managers)")
                
                # Month selector for VP's team
                col_vp_team_filter1, col_vp_team_filter2 = st.columns([4, 1])
                
                with col_vp_team_filter1:
                    st.caption("Filter team performance by month")
                
                with col_vp_team_filter2:
                    vp_team_month = st.selectbox(
                        "Month",
                        list(range(1, 13)),
                        index=current_month - 1,
                        format_func=lambda x: get_month_name(x),
                        key="vp_team_month_select"
                    )
                
                st.caption(f"**Team Performance: {get_month_name(vp_team_month)} {current_year}**")
                
                # Get all HR and Managers (VP's team)
                all_users = db.get_all_users()
                vp_team = [u for u in all_users if u['role'] in ['HR', 'Manager']]
                
                # Calculate team metrics for selected month
                vp_team_total_goals = 0
                vp_team_completed_goals = 0
                vp_team_active_goals = 0
                vp_team_total_progress = 0
                vp_team_goals_count = 0
                vp_team_overdue_goals = 0
                
                today_date = date.today()
                
                for member in vp_team:
                    member_month_goals = get_user_month_goals(member['id'], vp_team_month, current_year)
                    vp_team_total_goals += len(member_month_goals)
                    vp_team_completed_goals += len([g for g in member_month_goals if g.get('status') == 'Completed'])
                    vp_team_active_goals += len([g for g in member_month_goals if g.get('status') == 'Active'])
                    
                    for goal in member_month_goals:
                        achievement = goal.get('monthly_achievement')
                        if achievement is not None:
                            target = goal.get('monthly_target', 1)
                            if target > 0:
                                progress = (achievement / target * 100)
                                vp_team_total_progress += progress
                                vp_team_goals_count += 1
                        
                        # Count overdue goals
                        if goal.get('status') == 'Active':
                            end_date_str = goal.get('end_date')
                            if end_date_str:
                                try:
                                    end_date = datetime.strptime(str(end_date_str), '%Y-%m-%d').date()
                                    if today_date > end_date:
                                        vp_team_overdue_goals += 1
                                except:
                                    pass
                
                vp_team_avg_progress = (vp_team_total_progress / vp_team_goals_count) if vp_team_goals_count > 0 else 0
                
                col_vp1, col_vp2, col_vp3, col_vp4, col_vp5 = st.columns(5)
                
                # ================= VP TEAM METRICS =================

# -------- col_vp1 : Team Members --------
                with col_vp1:
                    st.markdown(
                        f"""
                        <div style="
                            background:#FFFFFF;
                            width:100%;
                            height:160px;
                            padding:20px;
                            border-radius:10px;
                            text-align:center;
                            display:flex;
                            flex-direction:column;
                            justify-content:center;
                            box-shadow:
                                0 2px 4px rgba(0,0,0,0.08),
                                0 0 18px rgba(118,75,162,0.23);
                        ">
                            <div style="font-size:36px; margin-bottom:10px;"></div>
                            <div style="font-size:32px; font-weight:700; color:#764BA2; margin-bottom:6px;">
                                {len(vp_team)}
                            </div>
                            <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:#764BA2;
                                padding:4px 10px; border-radius:6px; display:inline-block;">
                                Team Members
                            </div>
                        </div>
                        """,
                        unsafe_allow_html=True
                    )


                # -------- col_vp2 : Team Goals --------
                with col_vp2:
                    st.markdown(
                        f"""
                        <div style="
                            background:#FFFFFF;
                            width:100%;
                            height:160px;
                            padding:20px;
                            border-radius:10px;
                            text-align:center;
                            display:flex;
                            flex-direction:column;
                            justify-content:center;
                            box-shadow:
                                0 2px 4px rgba(0,0,0,0.08),
                                0 0 18px rgba(56,249,215,0.23);
                        ">
                            <div style="font-size:36px; margin-bottom:10px;"></div>
                            <div style="font-size:32px; font-weight:700; color:#38F9D7; margin-bottom:6px;">
                                {vp_team_total_goals}
                            </div>
                            <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:#38F9D7;
                                padding:4px 10px; border-radius:6px; display:inline-block;">
                                Team Goals
                            </div>
                        </div>
                        """,
                        unsafe_allow_html=True
                    )


                # -------- col_vp3 : Completed --------
                with col_vp3:
                    st.markdown(
                        f"""
                        <div style="
                            background:#FFFFFF;
                            width:100%;
                            height:160px;
                            padding:20px;
                            border-radius:10px;
                            text-align:center;
                            display:flex;
                            flex-direction:column;
                            justify-content:center;
                            box-shadow:
                                0 2px 4px rgba(0,0,0,0.08),
                                0 0 18px rgba(245,87,108,0.23);
                        ">
                            <div style="font-size:36px; margin-bottom:10px;"></div>
                            <div style="font-size:32px; font-weight:700; color:#F5576C; margin-bottom:6px;">
                                {vp_team_completed_goals}
                            </div>
                            <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:#F5576C;
                                padding:4px 10px; border-radius:6px; display:inline-block;">
                                Completed
                            </div>
                        </div>
                        """,
                        unsafe_allow_html=True
                    )


                # -------- col_vp4 : Avg Progress --------
                with col_vp4:
                    st.markdown(
                        f"""
                        <div style="
                            background:#FFFFFF;
                            width:100%;
                            height:160px;
                            padding:20px;
                            border-radius:10px;
                            text-align:center;
                            display:flex;
                            flex-direction:column;
                            justify-content:center;
                            box-shadow:
                                0 2px 4px rgba(0,0,0,0.08),
                                0 0 18px rgba(245,158,11,0.23);
                        ">
                            <div style="font-size:36px; margin-bottom:10px;"></div>
                            <div style="font-size:32px; font-weight:700; color:#F59E0B; margin-bottom:6px;">
                                {vp_team_avg_progress:.1f}%
                            </div>
                            <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:#F59E0B;
                                padding:4px 10px; border-radius:6px; display:inline-block;">
                                Avg Progress
                            </div>
                        </div>
                        """,
                        unsafe_allow_html=True
                    )


                # -------- col_vp5 : Overdue (Conditional) --------
                with col_vp5:
                    overdue_color = "#EF4444" if vp_team_overdue_goals > 0 else "#10B981"
                    label_bg = "#FEF2F2" if vp_team_overdue_goals > 0 else "#ECFDF5"
                    glow = "rgba(239,68,68,0.23)" if vp_team_overdue_goals > 0 else "rgba(16,185,129,0.23)"
                    icon = "" if vp_team_overdue_goals > 0 else ""

                    st.markdown(
                        f"""
                        <div style="
                            background:#FFFFFF;
                            width:100%;
                            height:160px;
                            padding:20px;
                            border-radius:10px;
                            text-align:center;
                            display:flex;
                            flex-direction:column;
                            justify-content:center;
                            box-shadow:
                                0 2px 4px rgba(0,0,0,0.08),
                                0 0 18px {glow};
                        ">
                            <div style="font-size:36px; margin-bottom:10px;">{icon}</div>
                            <div style="font-size:32px; font-weight:700; color:{overdue_color}; margin-bottom:6px;">
                                {vp_team_overdue_goals}
                            </div>
                            <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:{overdue_color};
                                background:{label_bg}; padding:4px 10px; border-radius:6px; display:inline-block;">
                                Overdue
                            </div>
                        </div>
                        """,
                        unsafe_allow_html=True
                    )

                st.markdown("---")
                # VP Team details modal
                if st.session_state.get('show_details') and st.session_state.get('show_details').startswith('vp_team_'):
                    st.markdown("---")
                    detail_type = st.session_state.show_details
                    
                    col_header, col_close = st.columns([5, 1])
                    
                    with col_header:
                        if detail_type == 'vp_team_members':
                            st.subheader(f" Team Members ({len(vp_team)})")
                        elif detail_type == 'vp_team_goals':
                            st.subheader(f" All Team Goals ({vp_team_total_goals}) - {get_month_name(vp_team_month)} {current_year}")
                        elif detail_type == 'vp_team_completed':
                            st.subheader(f" Team Completed Goals ({vp_team_completed_goals})")
                        elif detail_type == 'vp_team_progress':
                            st.subheader(f" Team Progress Breakdown")
                        elif detail_type == 'vp_team_overdue':
                            st.subheader(f" Team Overdue Goals ({vp_team_overdue_goals})")
                    
                    with col_close:
                        if st.button("✕ Close", key="close_vp_team_details"):
                            del st.session_state.show_details
                            st.rerun()
                    
                    # Show VP team details
                    if detail_type == 'vp_team_members':
                        # Display team members list
                        members_data = []
                        for member in vp_team:
                            member_goals = get_user_month_goals(member['id'], vp_team_month, current_year)
                            completed = len([g for g in member_goals if g.get('status') == 'Completed'])
                            
                            # Count member's overdue goals
                            member_overdue = 0
                            for goal in member_goals:
                                if goal.get('status') == 'Active':
                                    end_date_str = goal.get('end_date')
                                    if end_date_str:
                                        try:
                                            end_date = datetime.strptime(str(end_date_str), '%Y-%m-%d').date()
                                            if today_date > end_date:
                                                member_overdue += 1
                                        except:
                                            pass
                            
                            members_data.append({
                                'Name': member['name'],
                                'Role': member['role'],
                                'Email': member['email'],
                                'Department': member.get('department', 'N/A'),
                                'Total Goals': len(member_goals),
                                'Completed': completed,
                                'Overdue': member_overdue
                            })
                        
                        df_members = pd.DataFrame(members_data)
                        st.dataframe(df_members, use_container_width=True, height=400)
                        
                        st.download_button(
                            "📥 Export to CSV",
                            df_members.to_csv(index=False).encode('utf-8'),
                            f"vp_team_{vp_team_month}_{current_year}.csv",
                            "text/csv",
                            key='download_vp_team_members'
                        )
                        
                    else:
                        # Collect all team goals
                        all_vp_team_goals = []
                        for member in vp_team:
                            member_goals = get_user_month_goals(member['id'], vp_team_month, current_year)
                            for goal in member_goals:
                                goal['member_name'] = member['name']
                                goal['member_role'] = member['role']
                            all_vp_team_goals.extend(member_goals)
                        
                        # Filter based on detail type
                        if detail_type == 'vp_team_completed':
                            display_goals = [g for g in all_vp_team_goals if g.get('status') == 'Completed']
                        elif detail_type == 'vp_team_overdue':
                            display_goals = []
                            for goal in all_vp_team_goals:
                                if goal.get('status') == 'Active':
                                    end_date_str = goal.get('end_date')
                                    if end_date_str:
                                        try:
                                            end_date = datetime.strptime(str(end_date_str), '%Y-%m-%d').date()
                                            if today_date > end_date:
                                                display_goals.append(goal)
                                        except:
                                            pass
                        elif detail_type == 'vp_team_progress':
                            display_goals = all_vp_team_goals.copy()
                            for g in display_goals:
                                achievement = g.get('monthly_achievement')
                                target = g.get('monthly_target', 1)
                                achievement_val = 0 if achievement is None else achievement
                                target_val = 1 if target is None or target == 0 else target
                                g['_progress'] = calculate_progress(achievement_val, target_val)
                            display_goals.sort(key=lambda x: x['_progress'], reverse=True)
                        else:
                            display_goals = all_vp_team_goals
                        
                        if display_goals:
                            goal_data = []
                            for goal in display_goals:
                                monthly_achievement = goal.get('monthly_achievement')
                                monthly_target = goal.get('monthly_target', 1)
                                achievement_value = 0 if monthly_achievement is None else monthly_achievement
                                target_value = 1 if monthly_target is None or monthly_target == 0 else monthly_target
                                progress = calculate_progress(achievement_value, target_value)
                                
                                goal_data.append({
                                    'Member': goal['member_name'],
                                    'Role': goal['member_role'],
                                    'Goal Title': goal['goal_title'],
                                    'Department': goal.get('department', 'N/A'),
                                    'Month': get_month_name(goal.get('month', 1)),  # ADD THIS
                                    'Year': str(goal['year']), # ADD THIS
                                    'Target': monthly_target if monthly_target is not None else 0,
                                    'Achievement': monthly_achievement if monthly_achievement is not None else '-',
                                    'Progress': f"{progress:.1f}%",
                                    'Status': goal.get('status', 'Active')
                                })
                            
                            df_goals = pd.DataFrame(goal_data)
                            st.dataframe(df_goals, use_container_width=True, height=400)
                            
                            st.download_button(
                                "📥 Export to CSV",
                                df_goals.to_csv(index=False).encode('utf-8'),
                                f"vp_team_goals_{detail_type}_{vp_team_month}_{current_year}.csv",
                                "text/csv",
                                key='download_vp_team_modal_details'
                            )
                        else:
                            st.info("No goals found in this category")
                    
                    st.markdown("---")
        elif role == 'HR':
            st.markdown("### 👥 HR Team Performance")
            
            # Get all HR members + employees with department='HR'
            all_users = db.get_all_users()
            hr_team = [u for u in all_users if u['role'] == 'HR' and u['id'] != user['id']]
            
            # ✅ NEW: Add employees with department='HR'
            hr_dept_employees = [u for u in all_users if u['role'] == 'Employee' and normalize_department(u.get('department')) == 'HR']
            
            # Combine for display
            all_hr = [user] + hr_team + hr_dept_employees
            
            if len(all_hr) > 1:  # More than just the current user
                # Month selector
                col_hr_filter1, col_hr_filter2 = st.columns([4, 1])
                
                with col_hr_filter1:
                    st.caption("Filter HR team performance by month")
                
                with col_hr_filter2:
                    hr_team_month = st.selectbox(
                        "Month",
                        list(range(1, 13)),
                        index=current_month - 1,
                        format_func=lambda x: get_month_name(x),
                        key="hr_team_month_select"
                    )
                
                st.caption(f"**HR Team Performance: {get_month_name(hr_team_month)} {current_year}**")
                
                # Calculate metrics including self and HR dept employees
                hr_team_total_goals = 0
                hr_team_completed_goals = 0
                hr_team_active_goals = 0
                hr_team_total_progress = 0
                hr_team_goals_count = 0
                hr_team_overdue_goals = 0
                
                today_date = date.today()
                
                for hr_member in all_hr:
                    member_month_goals = get_user_month_goals(hr_member['id'], hr_team_month, current_year)
                    hr_team_total_goals += len(member_month_goals)
                    hr_team_completed_goals += len([g for g in member_month_goals if g.get('status') == 'Completed'])
                    hr_team_active_goals += len([g for g in member_month_goals if g.get('status') == 'Active'])
                    
                    for goal in member_month_goals:
                        achievement = goal.get('monthly_achievement')
                        if achievement is not None:
                            target = goal.get('monthly_target', 1)
                            if target > 0:
                                progress = (achievement / target * 100)
                                hr_team_total_progress += progress
                                hr_team_goals_count += 1
                        
                        # Count overdue
                        if goal.get('status') == 'Active':
                            end_date_str = goal.get('end_date')
                            if end_date_str:
                                try:
                                    end_date = datetime.strptime(str(end_date_str), '%Y-%m-%d').date()
                                    if today_date > end_date:
                                        hr_team_overdue_goals += 1
                                except:
                                    pass
                
                hr_team_avg_progress = (hr_team_total_progress / hr_team_goals_count) if hr_team_goals_count > 0 else 0
                
                # Display metrics
                col_hr1, col_hr2, col_hr3, col_hr4, col_hr5 = st.columns(5)
                
                # ================= HR TEAM METRICS =================

# -------- col_hr1 : HR Team Members --------
                with col_hr1:
                    st.markdown(
                        f"""
                        <div style="
                            background:#FFFFFF;
                            width:100%;
                            height:160px;
                            padding:20px;
                            border-radius:10px;
                            text-align:center;
                            display:flex;
                            flex-direction:column;
                            justify-content:center;
                            box-shadow:
                                0 2px 4px rgba(0,0,0,0.08),
                                0 0 18px rgba(118,75,162,0.23);
                        ">
                            <div style="font-size:36px; margin-bottom:10px;"></div>
                            <div style="font-size:32px; font-weight:700; color:#764BA2; margin-bottom:6px;">
                                {len(all_hr)}
                            </div>
                            <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:#764BA2;
                                padding:4px 10px; border-radius:6px; display:inline-block;">
                                HR Team Members
                            </div>
                        </div>
                        """,
                        unsafe_allow_html=True
                    )


                # -------- col_hr2 : Team Goals --------
                with col_hr2:
                    st.markdown(
                        f"""
                        <div style="
                            background:#FFFFFF;
                            width:100%;
                            height:160px;
                            padding:20px;
                            border-radius:10px;
                            text-align:center;
                            display:flex;
                            flex-direction:column;
                            justify-content:center;
                            box-shadow:
                                0 2px 4px rgba(0,0,0,0.08),
                                0 0 18px rgba(56,249,215,0.23);
                        ">
                            <div style="font-size:36px; margin-bottom:10px;"></div>
                            <div style="font-size:32px; font-weight:700; color:#179eaf; margin-bottom:6px;">
                                {hr_team_total_goals}
                            </div>
                            <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:#179eaf;
                                padding:4px 10px; border-radius:6px; display:inline-block;">
                                Team Goals
                            </div>
                        </div>
                        """,
                        unsafe_allow_html=True
                    )


                # -------- col_hr3 : Completed --------
                with col_hr3:
                    st.markdown(
                        f"""
                        <div style="
                            background:#FFFFFF;
                            width:100%;
                            height:160px;
                            padding:20px;
                            border-radius:10px;
                            text-align:center;
                            display:flex;
                            flex-direction:column;
                            justify-content:center;
                            box-shadow:
                                0 2px 4px rgba(0,0,0,0.08),
                                0 0 18px rgba(245,87,108,0.23);
                        ">
                            <div style="font-size:36px; margin-bottom:10px;"></div>
                            <div style="font-size:32px; font-weight:700; color:#F5576C; margin-bottom:6px;">
                                {hr_team_completed_goals}
                            </div>
                            <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:#F5576C;
                                padding:4px 10px; border-radius:6px; display:inline-block;">
                                Completed
                            </div>
                        </div>
                        """,
                        unsafe_allow_html=True
                    )


                # -------- col_hr4 : Avg Progress --------
                with col_hr4:
                    st.markdown(
                        f"""
                        <div style="
                            background:#FFFFFF;
                            width:100%;
                            height:160px;
                            padding:20px;
                            border-radius:10px;
                            text-align:center;
                            display:flex;
                            flex-direction:column;
                            justify-content:center;
                            box-shadow:
                                0 2px 4px rgba(0,0,0,0.08),
                                0 0 18px rgba(245,158,11,0.23);
                        ">
                            <div style="font-size:36px; margin-bottom:10px;"></div>
                            <div style="font-size:32px; font-weight:700; color:#F59E0B; margin-bottom:6px;">
                                {hr_team_avg_progress:.1f}%
                            </div>
                            <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:#F59E0B;
                                padding:4px 10px; border-radius:6px; display:inline-block;">
                                Avg Progress
                            </div>
                        </div>
                        """,
                        unsafe_allow_html=True
                    )


                # -------- col_hr5 : Overdue (Conditional) --------
                with col_hr5:
                    overdue_color = "#EF4444" if hr_team_overdue_goals > 0 else "#10B981"
                    label_bg = "#FEF2F2" if hr_team_overdue_goals > 0 else "#ECFDF5"
                    glow = "rgba(239,68,68,0.23)" if hr_team_overdue_goals > 0 else "rgba(16,185,129,0.23)"
                    icon = "" if hr_team_overdue_goals > 0 else ""

                    st.markdown(
                        f"""
                        <div style="
                            background:#FFFFFF;
                            width:100%;
                            height:160px;
                            padding:20px;
                            border-radius:10px;
                            text-align:center;
                            display:flex;
                            flex-direction:column;
                            justify-content:center;
                            box-shadow:
                                0 2px 4px rgba(0,0,0,0.08),
                                0 0 18px {glow};
                        ">
                            <div style="font-size:36px; margin-bottom:10px;">{icon}</div>
                            <div style="font-size:32px; font-weight:700; color:{overdue_color}; margin-bottom:6px;">
                                {hr_team_overdue_goals}
                            </div>
                            <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:{overdue_color};
                                background:{label_bg}; padding:4px 10px; border-radius:6px; display:inline-block;">
                                Overdue
                            </div>
                        </div>
                        """,
                        unsafe_allow_html=True
                    )

                
                st.markdown("---")
            else:
                st.info(" You are the only HR member in the system.")           
                  
            # HR Team details modal
            if st.session_state.get('show_details') and st.session_state.get('show_details').startswith('hr_team_'):
                st.markdown("---")
                detail_type = st.session_state.show_details
                
                col_header, col_close = st.columns([5, 1])
                
                with col_header:
                    if detail_type == 'hr_team_members':
                        st.subheader(f" HR Team Members ({len(all_hr)})")
                    elif detail_type == 'hr_team_goals':
                        st.subheader(f" All HR Goals ({hr_team_total_goals})")
                    elif detail_type == 'hr_team_completed':
                        st.subheader(f" HR Completed Goals ({hr_team_completed_goals})")
                    elif detail_type == 'hr_team_progress':
                        st.subheader(f" HR Progress Breakdown")
                    elif detail_type == 'hr_team_overdue':
                        st.subheader(f" HR Overdue Goals ({hr_team_overdue_goals})")
                
                with col_close:
                    if st.button("✕ Close", key="close_hr_team_details"):
                        del st.session_state.show_details
                        st.rerun()
                


        # ✅ UPDATED: Organization Goals Distribution with Filters
        st.markdown("####  Organization Goals Distribution")

        # ===== STEP 1: ADD FILTERS (WITH USER FILTER) =====
        col_filter1, col_filter2, col_filter3 = st.columns(3)  # Changed to 3 columns

        with col_filter1:
            # Get unique departments
            all_depts = list(set([normalize_department(u.get('department')) for u in all_users]))
            departments = ['All Teams'] + sorted([d for d in all_depts if d and d != 'UNASSIGNED'])
            
            selected_team = st.selectbox(
                " Filter by Team",
                departments,
                key="org_goals_team_filter"
            )

        with col_filter2:
            # Filter users based on selected team
            if selected_team == 'All Teams':
                filtered_users_for_dropdown = all_users
            else:
                filtered_users_for_dropdown = [u for u in all_users if normalize_department(u.get('department')) == selected_team]
            
            user_options = ['All Users'] + [f"{u['name']} ({u['role']})" for u in filtered_users_for_dropdown]
            
            selected_user_option = st.selectbox(
                "👤 Filter by User",
                user_options,
                key="org_goals_user_filter"
            )

        with col_filter3:
            filter_month = st.selectbox(
                "📅 Filter by Month",
                ["All Months"] + [get_month_name(i) for i in range(1, 13)],
                key="org_goals_month_filter"
            )

        # Display filter status
        filter_display = []
        if selected_team != "All Teams":
            filter_display.append(f"**{selected_team}**")
        if selected_user_option != "All Users":
            user_name = selected_user_option.split(' (')[0]
            filter_display.append(f"**{user_name}**")
        if filter_month != "All Months":
            filter_display.append(f"**{filter_month}**")

        if filter_display:
            st.caption(f"Showing: {' | '.join(filter_display)}")
        else:
            st.caption("Showing: **All Goals**")

        st.markdown("---")

        # ===== STEP 2: CALCULATE FILTERED GOALS (WITH USER FILTER) =====
        org_total_goals = 0
        org_completed_goals = 0
        org_active_goals = 0
        org_overdue_goals = 0

        today_date = date.today()
        current_year = today_date.year

        # Get selected user ID if filtering by user
        selected_user_id = None
        if selected_user_option != 'All Users':
            selected_user_name = selected_user_option.split(' (')[0]
            selected_user_obj = next((u for u in all_users if u['name'] == selected_user_name), None)
            if selected_user_obj:
                selected_user_id = selected_user_obj['id']

        for u in all_users:
            # Apply team filter at user level first
            if selected_team != 'All Teams':
                if normalize_department(u.get('department')) != selected_team:
                    continue  # Skip this user entirely if not in selected team
            
            # Apply user filter
            if selected_user_id is not None:
                if u['id'] != selected_user_id:
                    continue  # Skip this user if not the selected user
            
            user_goals = db.get_user_all_goals(u['id'])
            
            for goal in user_goals:
                # Apply month filter
                if filter_month != 'All Months':
                    month_num = [get_month_name(i) for i in range(1, 13)].index(filter_month) + 1
                    if goal.get('month') != month_num or goal.get('year') != current_year:
                        continue
                
                org_total_goals += 1
                
                if goal.get('status') == 'Completed':
                    org_completed_goals += 1
                elif goal.get('status') == 'Active':
                    org_active_goals += 1
                    
                    # Check if overdue
                    end_date_str = goal.get('end_date')
                    if end_date_str:
                        try:
                            end_date = datetime.strptime(str(end_date_str), '%Y-%m-%d').date()
                            if today_date > end_date:
                                org_overdue_goals += 1
                        except:
                            pass

        # ===== STEP 3: DISPLAY CHARTS WITH NEW LAYOUT =====
        if org_total_goals > 0:
            col_pie1, col_pie2 = st.columns([2, 1])  # Changed ratio: bigger chart, smaller summary
            
            with col_pie1:
                # Goals by status (excluding overdue count from active)
                org_on_track_active = org_active_goals - org_overdue_goals
                
                pie_data = {
                    'Status': ['Completed', 'Active (On Track)', 'Overdue', 'Other'],
                    'Count': [
                        org_completed_goals,
                        org_on_track_active,
                        org_overdue_goals,
                        org_total_goals - org_completed_goals - org_active_goals
                    ]
                }
                
                # Remove categories with 0 count
                pie_df = pd.DataFrame(pie_data)
                pie_df = pie_df[pie_df['Count'] > 0]
                
                fig_pie = px.pie(
                    pie_df,
                    values='Count',
                    names='Status',
                    title='Goals by Status',
                    color='Status',
                    color_discrete_map={
                        'Completed': '#10b981',
                        'Active (On Track)': '#3b82f6',
                        'Overdue': '#ef4444',
                        'Other': '#94a3b8'
                    },
                    hole=0.4  # Makes it a donut chart
                )
                fig_pie.update_traces(textposition='inside', textinfo='percent+label')
                fig_pie.update_layout(height=450)  # Increased from 350 for bigger chart
                st.plotly_chart(fig_pie, use_container_width=True)
            
            # ===== STEP 4: COMPACT SUMMARY =====
            with col_pie2:
                st.markdown("##### Summary")
                
                # Compact metrics with custom styling
                st.markdown(f"""
                <div style='background: #f0f9ff; padding: 15px; border-radius: 8px; margin-bottom: 10px;'>
                    <p style='margin: 0; font-size: 14px; color: #64748b;'>Total Goals</p>
                    <h2 style='margin: 5px 0; color: #1e40af;'>{org_total_goals}</h2>
                </div>
                
                <div style='background: #d1fae5; padding: 10px; border-radius: 8px; margin-bottom: 8px;'>
                    <p style='margin: 0; font-size: 12px; color: #065f46;'>Completed</p>
                    <p style='margin: 5px 0; font-size: 18px; font-weight: bold; color: #059669;'>
                        {org_completed_goals} ({org_completed_goals/org_total_goals*100:.1f}%)
                    </p>
                </div>
                
                <div style='background: #dbeafe; padding: 10px; border-radius: 8px; margin-bottom: 8px;'>
                    <p style='margin: 0; font-size: 12px; color: #1e40af;'>Active</p>
                    <p style='margin: 5px 0; font-size: 18px; font-weight: bold; color: #2563eb;'>
                        {org_active_goals} ({org_active_goals/org_total_goals*100:.1f}%)
                    </p>
                </div>
                
                <div style='background: #fee2e2; padding: 10px; border-radius: 8px; margin-bottom: 10px;'>
                    <p style='margin: 0; font-size: 12px; color: #991b1b;'>Overdue</p>
                    <p style='margin: 5px 0; font-size: 18px; font-weight: bold; color: #dc2626;'>
                        {org_overdue_goals} ({org_overdue_goals/org_total_goals*100:.1f}%)
                    </p>
                </div>
                """, unsafe_allow_html=True)
                
                # Compact completion rate
                completion_rate = (org_completed_goals / org_total_goals * 100) if org_total_goals > 0 else 0
                st.markdown("**Completion Rate:**")
                st.progress(completion_rate / 100)
                st.caption(f"{completion_rate:.1f}%")
        else:
            st.info("No organization goals found matching the selected filters")

        st.markdown("---")       
                

            

        # ✅ Top Performers - Role-based display
        st.markdown("####  Top Performers (This Month)")
        rankings = []

        if role == 'Manager':
            team_members = db.get_team_members(user['id'])
            employees_to_rank = [m for m in team_members if m['role'] == 'Employee']
            top_count = 3
            performer_scope = "Team"
        else:
            employees_to_rank = [u for u in all_users if u['role'] == 'Employee']
            top_count = 5
            performer_scope = "Organization"

        for emp in employees_to_rank:
            emp_stats = db.get_user_goal_stats(emp['id'])
            if emp_stats.get('total_goals', 0) > 0:
                rankings.append({
                    'Name': emp['name'],
                    'Department': emp.get('department', 'N/A'),
                    'Designation': emp.get('designation', 'N/A'),
                    'Total Goals': emp_stats.get('total_goals', 0),
                    'Completed': emp_stats.get('completed_goals', 0),
                    'Progress %': f"{emp_stats.get('avg_progress', 0):.1f}%",
                    'Progress_Val': emp_stats.get('avg_progress', 0)
                })

        if rankings:
            df_rank = pd.DataFrame(rankings)
            df_rank = df_rank.sort_values('Progress_Val', ascending=False).head(top_count)
            df_rank = df_rank.drop('Progress_Val', axis=1)
            df_rank.insert(0, 'Rank', range(1, len(df_rank) + 1))

            st.caption(f" Showing Top {min(len(df_rank), top_count)} of {len(rankings)} {performer_scope} Performers")
            st.dataframe(df_rank, use_container_width=True, height=300)
        else:
            st.info(f"No performance data available for {performer_scope.lower()}")

    elif role == 'Manager':
    # Personal + Team view for Manager
        st.markdown("###  Your Performance Overview")

        # Month selector for manager's personal performance
        col_mgr_filter1, col_mgr_filter2 = st.columns([4, 1])

        with col_mgr_filter1:
            st.caption("Filter your personal goals by month")

        with col_mgr_filter2:
            mgr_selected_month = st.selectbox(
                "Month",
                list(range(1, 13)),
                index=current_month - 1,
                format_func=lambda x: get_month_name(x),
                key="mgr_month_select"
            )

        st.caption(f"**Your Performance: {get_month_name(mgr_selected_month)} {current_year}**")

        # Get user goals for selected month ONLY
        user_goals = get_user_month_goals(user['id'], mgr_selected_month, current_year)

        total_goals = len(user_goals)
        completed_goals = len([g for g in user_goals if g.get('status') == 'Completed'])
        active_goals = len([g for g in user_goals if g.get('status') == 'Active'])

        total_progress = 0
        goals_with_progress = 0
        for goal in user_goals:
            monthly_achievement = goal.get('monthly_achievement')
            if monthly_achievement is not None:
                monthly_target = goal.get('monthly_target', 1)
                if monthly_target > 0:
                    progress = (monthly_achievement / monthly_target * 100)
                    total_progress += progress
                    goals_with_progress += 1

        avg_progress = (total_progress / goals_with_progress) if goals_with_progress > 0 else 0

        overdue_goals = 0
        today_date = date.today()
        for goal in user_goals:
            if goal.get('status') == 'Active':
                end_date_str = goal.get('end_date')
                if end_date_str:
                    try:
                        end_date = datetime.strptime(str(end_date_str), '%Y-%m-%d').date()
                        if today_date > end_date:
                            overdue_goals += 1
                    except:
                        pass

        col_perf1, col_perf2, col_perf3, col_perf4, col_perf5 = st.columns(5)

        # ================= MANAGER PERSONAL METRICS =================

# -------- col_perf1 : Total Goals --------
        with col_perf1:
            st.markdown(
                f"""
                <div style="
                    background:#FFFFFF;
                    width:100%;
                    height:160px;
                    padding:20px;
                    border-radius:10px;
                    text-align:center;
                    display:flex;
                    flex-direction:column;
                    justify-content:center;
                    box-shadow:
                        0 2px 4px rgba(0,0,0,0.08),
                        0 0 18px rgba(59,130,246,0.25);
                ">
                    <div style="font-size:36px; margin-bottom:10px;"></div>
                    <div style="font-size:32px; font-weight:700; color:#3B82F6; margin-bottom:6px;">
                        {total_goals}
                    </div>
                    <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:#3B82F6;
                        padding:4px 10px; border-radius:6px; display:inline-block;">
                        Total Goals
                    </div>
                </div>
                """,
                unsafe_allow_html=True
            )


        # -------- col_perf2 : Completed --------
        with col_perf2:
            st.markdown(
                f"""
                <div style="
                    background:#FFFFFF;
                    width:100%;
                    height:160px;
                    padding:20px;
                    border-radius:10px;
                    text-align:center;
                    display:flex;
                    flex-direction:column;
                    justify-content:center;
                    box-shadow:
                        0 2px 4px rgba(0,0,0,0.08),
                        0 0 18px rgba(16,185,129,0.22);
                ">
                    <div style="font-size:36px; margin-bottom:10px;"></div>
                    <div style="font-size:32px; font-weight:700; color:#10B981; margin-bottom:6px;">
                        {completed_goals}
                    </div>
                    <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:#10B981;
                        padding:4px 10px; border-radius:6px; display:inline-block;">
                        Completed
                    </div>
                </div>
                """,
                unsafe_allow_html=True
            )


        # -------- col_perf3 : Active --------
        with col_perf3:
            st.markdown(
                f"""
                <div style="
                    background:#FFFFFF;
                    width:100%;
                    height:160px;
                    padding:20px;
                    border-radius:10px;
                    text-align:center;
                    display:flex;
                    flex-direction:column;
                    justify-content:center;
                    box-shadow:
                        0 2px 4px rgba(0,0,0,0.08),
                        0 0 18px rgba(245,87,108,0.23);
                ">
                    <div style="font-size:36px; margin-bottom:10px;"></div>
                    <div style="font-size:32px; font-weight:700; color:#F5576C; margin-bottom:6px;">
                        {active_goals}
                    </div>
                    <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:#F5576C;
                        padding:4px 10px; border-radius:6px; display:inline-block;">
                        Active
                    </div>
                </div>
                """,
                unsafe_allow_html=True
            )


        # -------- col_perf4 : Avg Progress --------
        with col_perf4:
            st.markdown(
                f"""
                <div style="
                    background:#FFFFFF;
                    width:100%;
                    height:160px;
                    padding:20px;
                    border-radius:10px;
                    text-align:center;
                    display:flex;
                    flex-direction:column;
                    justify-content:center;
                    box-shadow:
                        0 2px 4px rgba(0,0,0,0.08),
                        0 0 18px rgba(0,201,255,0.23);
                ">
                    <div style="font-size:36px; margin-bottom:10px;"></div>
                    <div style="font-size:32px; font-weight:700; color:#00C9FF; margin-bottom:6px;">
                        {avg_progress:.1f}%
                    </div>
                    <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:#00C9FF;
                        padding:4px 10px; border-radius:6px; display:inline-block;">
                        Avg Progress
                    </div>
                </div>
                """,
                unsafe_allow_html=True
            )


        # -------- col_perf5 : Overdue (Conditional) --------
        with col_perf5:
            overdue_color = "#EF4444" if overdue_goals > 0 else "#10B981"
            label_bg = "#FEF2F2" if overdue_goals > 0 else "#ECFDF5"
            glow = "rgba(239,68,68,0.23)" if overdue_goals > 0 else "rgba(16,185,129,0.23)"
            icon = "" if overdue_goals > 0 else ""

            st.markdown(
                f"""
                <div style="
                    background:#FFFFFF;
                    width:100%;
                    height:160px;
                    padding:20px;
                    border-radius:10px;
                    text-align:center;
                    display:flex;
                    flex-direction:column;
                    justify-content:center;
                    box-shadow:
                        0 2px 4px rgba(0,0,0,0.08),
                        0 0 18px {glow};
                ">
                    <div style="font-size:36px; margin-bottom:10px;">{icon}</div>
                    <div style="font-size:32px; font-weight:700; color:{overdue_color}; margin-bottom:6px;">
                        {overdue_goals}
                    </div>
                    <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:{overdue_color};
                        background:{label_bg}; padding:4px 10px; border-radius:6px; display:inline-block;">
                        Overdue
                    </div>
                </div>
                """,
                unsafe_allow_html=True
            )


        # Manager personal details modal
        if st.session_state.get('show_details') and st.session_state.get('show_details').startswith('mgr_personal_'):
            st.markdown("---")
            detail_type = st.session_state.show_details
            
            col_header, col_close = st.columns([5, 1])
            
            with col_header:
                if detail_type == 'mgr_personal_total':
                    st.subheader(f" Your Goals ({total_goals}) - {get_month_name(mgr_selected_month)} {current_year}")
                elif detail_type == 'mgr_personal_completed':
                    st.subheader(f" Your Completed Goals ({completed_goals})")
                elif detail_type == 'mgr_personal_active':
                    st.subheader(f" Your Active Goals ({active_goals})")
                elif detail_type == 'mgr_personal_progress':
                    st.subheader(f" Your Progress Breakdown")
                elif detail_type == 'mgr_personal_overdue':
                    st.subheader(f" Your Overdue Goals ({overdue_goals})")
            
            with col_close:
                if st.button("✕ Close", key="close_mgr_personal_details"):
                    del st.session_state.show_details
                    st.rerun()
            
            # Filter goals
            if detail_type == 'mgr_personal_completed':
                display_goals = [g for g in user_goals if g.get('status') == 'Completed']
            elif detail_type == 'mgr_personal_active':
                display_goals = [g for g in user_goals if g.get('status') == 'Active']
            elif detail_type == 'mgr_personal_overdue':
                display_goals = []
                for goal in user_goals:
                    if goal.get('status') == 'Active':
                        end_date_str = goal.get('end_date')
                        if end_date_str:
                            try:
                                end_date = datetime.strptime(str(end_date_str), '%Y-%m-%d').date()
                                if today_date > end_date:
                                    display_goals.append(goal)
                            except:
                                pass
            elif detail_type == 'mgr_personal_progress':
                display_goals = user_goals.copy()
                for g in display_goals:
                    achievement = g.get('monthly_achievement')
                    target = g.get('monthly_target', 1)
                    achievement_val = 0 if achievement is None else achievement
                    target_val = 1 if target is None or target == 0 else target
                    g['_progress'] = calculate_progress(achievement_val, target_val)
                display_goals.sort(key=lambda x: x['_progress'], reverse=True)
            else:
                display_goals = user_goals
            
            # Display goals table
            if display_goals:
                goal_data = []
                for goal in display_goals:
                    monthly_achievement = goal.get('monthly_achievement')
                    monthly_target = goal.get('monthly_target', 1)
                    achievement_value = 0 if monthly_achievement is None else monthly_achievement
                    target_value = 1 if monthly_target is None or monthly_target == 0 else monthly_target
                    progress = calculate_progress(achievement_value, target_value)
                    
                    goal_data.append({
                        'Goal Title': goal['goal_title'],
                        'Department': goal.get('department', 'N/A'),
                        'KPI': goal.get('kpi', 'N/A'),
                        'Month': get_month_name(goal.get('month', 1)),  # ADD THIS
                        'Year': str(goal['year']),  # ADD THIS
                        'Target': monthly_target if monthly_target is not None else 0,
                        'Achievement': monthly_achievement if monthly_achievement is not None else '-',
                        'Progress': f"{progress:.1f}%",
                        'Status': goal.get('status', 'Active')
                    })
                
                df_goals = pd.DataFrame(goal_data)
                st.dataframe(df_goals, use_container_width=True, height=400)
                
                st.download_button(
                    "📥 Export to CSV",
                    df_goals.to_csv(index=False).encode('utf-8'),
                    f"mgr_personal_goals_{detail_type}_{mgr_selected_month}_{current_year}.csv",
                    "text/csv",
                    key='download_mgr_personal_modal_details'
                )
            else:
                st.info("No goals found in this category")
            
            st.markdown("---")

        st.markdown("---")

        # Team Overview
        team_members = db.get_team_members(user['id'])
        st.markdown("###  Team Overview")

        # Month selector for team performance
        col_team_filter1, col_team_filter2 = st.columns([4, 1])

        with col_team_filter1:
            st.caption("Filter team performance by month")

        with col_team_filter2:
            team_selected_month = st.selectbox(
                "Month",
                list(range(1, 13)),
                index=current_month - 1,
                format_func=lambda x: get_month_name(x),
                key="team_month_select"
            )

        st.caption(f"**Team Performance: {get_month_name(team_selected_month)} {current_year}**")

        # Calculate team metrics for selected month
        team_total_goals = 0
        team_completed_goals = 0
        team_active_goals = 0
        team_total_progress = 0
        team_goals_count = 0
        team_overdue_goals = 0

        today_date = date.today()

        for member in team_members:
            member_month_goals = get_user_month_goals(member['id'], team_selected_month, current_year)
            team_total_goals += len(member_month_goals)
            team_completed_goals += len([g for g in member_month_goals if g.get('status') == 'Completed'])
            team_active_goals += len([g for g in member_month_goals if g.get('status') == 'Active'])
            
            for goal in member_month_goals:
                achievement = goal.get('monthly_achievement')
                if achievement is not None:
                    target = goal.get('monthly_target', 1)
                    if target > 0:
                        progress = (achievement / target * 100)
                        team_total_progress += progress
                        team_goals_count += 1
                
                # Count overdue goals
                if goal.get('status') == 'Active':
                    end_date_str = goal.get('end_date')
                    if end_date_str:
                        try:
                            end_date = datetime.strptime(str(end_date_str), '%Y-%m-%d').date()
                            if today_date > end_date:
                                team_overdue_goals += 1
                        except:
                            pass

        team_avg_progress = (team_total_progress / team_goals_count) if team_goals_count > 0 else 0

        col1, col2, col3, col4, col5 = st.columns(5)

        # ================= TEAM METRICS =================

# -------- col1 : Team Members --------
        with col1:
            st.markdown(
                f"""
                <div style="
                    background:#FFFFFF;
                    width:100%;
                    height:160px;
                    padding:20px;
                    border-radius:10px;
                    text-align:center;
                    display:flex;
                    flex-direction:column;
                    justify-content:center;
                    box-shadow:
                        0 2px 4px rgba(0,0,0,0.08),
                        0 0 18px rgba(118,75,162,0.23);
                ">
                    <div style="font-size:36px; margin-bottom:10px;"></div>
                    <div style="font-size:32px; font-weight:700; color:#764BA2; margin-bottom:6px;">
                        {len(team_members)}
                    </div>
                    <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:#764BA2;
                        padding:4px 10px; border-radius:6px; display:inline-block;">
                        Team Members
                    </div>
                </div>
                """,
                unsafe_allow_html=True
            )


        # -------- col2 : Team Goals --------
        with col2:
            st.markdown(
                f"""
                <div style="
                    background:#FFFFFF;
                    width:100%;
                    height:160px;
                    padding:20px;
                    border-radius:10px;
                    text-align:center;
                    display:flex;
                    flex-direction:column;
                    justify-content:center;
                    box-shadow:
                        0 2px 4px rgba(0,0,0,0.08),
                        0 0 18px rgba(56,249,215,0.23);
                ">
                    <div style="font-size:36px; margin-bottom:10px;"></div>
                    <div style="font-size:32px; font-weight:700; color:#38F9D7; margin-bottom:6px;">
                        {team_total_goals}
                    </div>
                    <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:#38F9D7;
                        padding:4px 10px; border-radius:6px; display:inline-block;">
                        Team Goals
                    </div>
                </div>
                """,
                unsafe_allow_html=True
            )


        # -------- col3 : Completed --------
        with col3:
            st.markdown(
                f"""
                <div style="
                    background:#FFFFFF;
                    width:100%;
                    height:160px;
                    padding:20px;
                    border-radius:10px;
                    text-align:center;
                    display:flex;
                    flex-direction:column;
                    justify-content:center;
                    box-shadow:
                        0 2px 4px rgba(0,0,0,0.08),
                        0 0 18px rgba(245,87,108,0.23);
                ">
                    <div style="font-size:36px; margin-bottom:10px;"></div>
                    <div style="font-size:32px; font-weight:700; color:#F5576C; margin-bottom:6px;">
                        {team_completed_goals}
                    </div>
                    <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:#F5576C;
                        padding:4px 10px; border-radius:6px; display:inline-block;">
                        Completed
                    </div>
                </div>
                """,
                unsafe_allow_html=True
            )


        # -------- col4 : Avg Progress --------
        with col4:
            st.markdown(
                f"""
                <div style="
                    background:#FFFFFF;
                    width:100%;
                    height:160px;
                    padding:20px;
                    border-radius:10px;
                    text-align:center;
                    display:flex;
                    flex-direction:column;
                    justify-content:center;
                    box-shadow:
                        0 2px 4px rgba(0,0,0,0.08),
                        0 0 18px rgba(245,158,11,0.23);
                ">
                    <div style="font-size:36px; margin-bottom:10px;"></div>
                    <div style="font-size:32px; font-weight:700; color:#F59E0B; margin-bottom:6px;">
                        {team_avg_progress:.1f}%
                    </div>
                    <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:#F59E0B;
                        padding:4px 10px; border-radius:6px; display:inline-block;">
                        Avg Progress
                    </div>
                </div>
                """,
                unsafe_allow_html=True
            )


        # -------- col5 : Overdue (Conditional) --------
        with col5:
            overdue_color = "#EF4444" if team_overdue_goals > 0 else "#10B981"
            label_bg = "#FEF2F2" if team_overdue_goals > 0 else "#ECFDF5"
            glow = "rgba(239,68,68,0.23)" if team_overdue_goals > 0 else "rgba(16,185,129,0.23)"
            icon = "" if team_overdue_goals > 0 else ""

            st.markdown(
                f"""
                <div style="
                    background:#FFFFFF;
                    width:100%;
                    height:160px;
                    padding:20px;
                    border-radius:10px;
                    text-align:center;
                    display:flex;
                    flex-direction:column;
                    justify-content:center;
                    box-shadow:
                        0 2px 4px rgba(0,0,0,0.08),
                        0 0 18px {glow};
                ">
                    <div style="font-size:36px; margin-bottom:10px;">{icon}</div>
                    <div style="font-size:32px; font-weight:700; color:{overdue_color}; margin-bottom:6px;">
                        {team_overdue_goals}
                    </div>
                    <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:{overdue_color};
                        background:{label_bg}; padding:4px 10px; border-radius:6px; display:inline-block;">
                        Overdue
                    </div>
                </div>
                """,
                unsafe_allow_html=True
            )

        st.markdown("---")
        # Team details modal
        if st.session_state.get('show_details') and st.session_state.get('show_details').startswith('team_'):
            st.markdown("---")
            detail_type = st.session_state.show_details
            
            col_header, col_close = st.columns([5, 1])
            
            with col_header:
                if detail_type == 'team_members':
                    st.subheader(f" Team Members ({len(team_members)})")
                elif detail_type == 'team_goals':
                    st.subheader(f" All Team Goals ({team_total_goals}) - {get_month_name(team_selected_month)} {current_year}")
                elif detail_type == 'team_completed':
                    st.subheader(f" Team Completed Goals ({team_completed_goals})")
                elif detail_type == 'team_progress':
                    st.subheader(f" Team Progress Breakdown")
                elif detail_type == 'team_overdue':
                    st.subheader(f" Team Overdue Goals ({team_overdue_goals})")
            
            with col_close:
                if st.button("✕ Close", key="close_team_details"):
                    del st.session_state.show_details
                    st.rerun()
            
            # Show team details
            if detail_type == 'team_members':
                # Display team members list
                members_data = []
                for member in team_members:
                    member_goals = get_user_month_goals(member['id'], team_selected_month, current_year)
                    completed = len([g for g in member_goals if g.get('status') == 'Completed'])
                    
                    # Count member's overdue goals
                    member_overdue = 0
                    for goal in member_goals:
                        if goal.get('status') == 'Active':
                            end_date_str = goal.get('end_date')
                            if end_date_str:
                                try:
                                    end_date = datetime.strptime(str(end_date_str), '%Y-%m-%d').date()
                                    if today_date > end_date:
                                        member_overdue += 1
                                except:
                                    pass
                    
                    members_data.append({
                        'Name': member['name'],
                        'Email': member['email'],
                        'Department': member.get('department', 'N/A'),
                        'Total Goals': len(member_goals),
                        'Completed': completed,
                        'Overdue': member_overdue
                    })
                
                df_members = pd.DataFrame(members_data)
                st.dataframe(df_members, use_container_width=True, height=400)
                
            else:
                # Collect all team goals
                all_team_goals = []
                for member in team_members:
                    member_goals = get_user_month_goals(member['id'], team_selected_month, current_year)
                    for goal in member_goals:
                        goal['member_name'] = member['name']
                    all_team_goals.extend(member_goals)
                
                # Filter based on detail type
                if detail_type == 'team_completed':
                    display_goals = [g for g in all_team_goals if g.get('status') == 'Completed']
                elif detail_type == 'team_overdue':
                    display_goals = []
                    for goal in all_team_goals:
                        if goal.get('status') == 'Active':
                            end_date_str = goal.get('end_date')
                            if end_date_str:
                                try:
                                    end_date = datetime.strptime(str(end_date_str), '%Y-%m-%d').date()
                                    if today_date > end_date:
                                        display_goals.append(goal)
                                except:
                                    pass
                elif detail_type == 'team_progress':
                    display_goals = all_team_goals.copy()
                    for g in display_goals:
                        achievement = g.get('monthly_achievement')
                        target = g.get('monthly_target', 1)
                        achievement_val = 0 if achievement is None else achievement
                        target_val = 1 if target is None or target == 0 else target
                        g['_progress'] = calculate_progress(achievement_val, target_val)
                    display_goals.sort(key=lambda x: x['_progress'], reverse=True)
                else:
                    display_goals = all_team_goals
                
                if display_goals:
                    goal_data = []
                    for goal in display_goals:
                        monthly_achievement = goal.get('monthly_achievement')
                        monthly_target = goal.get('monthly_target', 1)
                        achievement_value = 0 if monthly_achievement is None else monthly_achievement
                        target_value = 1 if monthly_target is None or monthly_target == 0 else monthly_target
                        progress = calculate_progress(achievement_value, target_value)
                        
                        goal_data.append({
                            'Employee': goal['member_name'],
                            'Goal Title': goal['goal_title'],
                            'Department': goal.get('department', 'N/A'),
                            'Month': get_month_name(goal.get('month', 1)),  # ADD THIS
                            'Year': str(goal['year']), # ADD THIS
                            'Target': monthly_target if monthly_target is not None else 0,
                            'Achievement': monthly_achievement if monthly_achievement is not None else '-',
                            'Progress': f"{progress:.1f}%",
                            'Status': goal.get('status', 'Active')
                        })
                    
                    df_goals = pd.DataFrame(goal_data)
                    st.dataframe(df_goals, use_container_width=True, height=400)
                    
                    st.download_button(
                        "📥 Export to CSV",
                        df_goals.to_csv(index=False).encode('utf-8'),
                        f"team_goals_{detail_type}_{team_selected_month}_{current_year}.csv",
                        "text/csv",
                        key='download_team_modal_details'
                    )
                else:
                    st.info("No goals found in this category")
            
            st.markdown("---")
            
            st.markdown("---")
        # Top Performers from Team
        st.markdown("###  Top 3 Team Performers (This Month)")
        team_rankings = []

        for member in team_members:
            member_stats = db.get_user_goal_stats(member['id'])
            if member_stats.get('total_goals', 0) > 0:
                team_rankings.append({
                    'Name': member['name'],
                    'Department': member.get('department', 'N/A'),
                    'Designation': member.get('designation', 'N/A'),
                    'Total Goals': member_stats.get('total_goals', 0),
                    'Completed': member_stats.get('completed_goals', 0),
                    'Progress %': f"{member_stats.get('avg_progress', 0):.1f}%",
                    'Progress_Val': member_stats.get('avg_progress', 0)
                })

        if team_rankings:
            df_team_rank = pd.DataFrame(team_rankings)
            df_team_rank = df_team_rank.sort_values('Progress_Val', ascending=False).head(3)
            df_team_rank = df_team_rank.drop('Progress_Val', axis=1)
            df_team_rank.insert(0, 'Rank', range(1, len(df_team_rank) + 1))
            st.dataframe(df_team_rank, use_container_width=True, height=200)
        else:
            st.info("No team performance data available yet")

    else:
    # Employee view
        st.markdown("###  Your Performance Overview")

        # Month selector for employee's personal performance
        col_emp_filter1, col_emp_filter2 = st.columns([4, 1])

        with col_emp_filter1:
            st.caption("Filter your personal goals by month")

        with col_emp_filter2:
            emp_selected_month = st.selectbox(
                "Month",
                list(range(1, 13)),
                index=current_month - 1,
                format_func=lambda x: get_month_name(x),
                key="emp_month_select"
            )

        st.caption(f"**Your Performance: {get_month_name(emp_selected_month)} {current_year}**")

        # Get all user goals
        all_user_goals = db.get_user_all_goals(user['id'])

        # Only approved goals count for employee metrics
        if user['role'] == 'Employee':
            approved_goals = [g for g in all_user_goals if g.get('approval_status') == 'approved']
            # Filter by selected month
            user_goals = [g for g in approved_goals if g['year'] == current_year and g.get('month') == emp_selected_month]
            
            pending_count = len([g for g in all_user_goals if g.get('approval_status') == 'pending'])
            if pending_count > 0:
                st.info(f"ℹ️ You have {pending_count} goal(s) pending manager approval (not shown in metrics)")
        else:
            # Filter by selected month
            user_goals = [g for g in all_user_goals if g['year'] == current_year and g.get('month') == emp_selected_month]

        total_goals = len(user_goals)
        completed_goals = len([g for g in user_goals if g.get('status') == 'Completed'])
        active_goals = len([g for g in user_goals if g.get('status') == 'Active'])

        total_progress = 0
        goals_with_progress = 0
        for goal in user_goals:
            monthly_achievement = goal.get('monthly_achievement')
            if monthly_achievement is not None:
                monthly_target = goal.get('monthly_target', 1)
                if monthly_target > 0:
                    progress = (monthly_achievement / monthly_target * 100)
                    total_progress += progress
                    goals_with_progress += 1

        avg_progress = (total_progress / goals_with_progress) if goals_with_progress > 0 else 0

        overdue_goals = 0
        today_date = date.today()
        for goal in user_goals:
            if goal.get('status') == 'Active':
                end_date_str = goal.get('end_date')
                if end_date_str:
                    try:
                        end_date = datetime.strptime(str(end_date_str), '%Y-%m-%d').date()
                        if today_date > end_date:
                            overdue_goals += 1
                    except:
                        pass

        col_perf1, col_perf2, col_perf3, col_perf4, col_perf5 = st.columns(5)

        # ================= EMPLOYEE PERSONAL METRICS =================

# -------- col_perf1 : Total Goals --------
        with col_perf1:
            st.markdown(
                f"""
                <div style="
                    background:#FFFFFF;
                    width:100%;
                    height:160px;
                    padding:20px;
                    border-radius:10px;
                    text-align:center;
                    display:flex;
                    flex-direction:column;
                    justify-content:center;
                    box-shadow:
                        0 2px 4px rgba(0,0,0,0.08),
                        0 0 18px rgba(59,130,246,0.25);
                ">
                    <div style="font-size:36px; margin-bottom:10px;"></div>
                    <div style="font-size:32px; font-weight:700; color:#3B82F6; margin-bottom:6px;">
                        {total_goals}
                    </div>
                    <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:#3B82F6;
                        padding:4px 10px; border-radius:6px; display:inline-block;">
                        Total Goals
                    </div>
                </div>
                """,
                unsafe_allow_html=True
            )


        # -------- col_perf2 : Completed --------
        with col_perf2:
            st.markdown(
                f"""
                <div style="
                    background:#FFFFFF;
                    width:100%;
                    height:160px;
                    padding:20px;
                    border-radius:10px;
                    text-align:center;
                    display:flex;
                    flex-direction:column;
                    justify-content:center;
                    box-shadow:
                        0 2px 4px rgba(0,0,0,0.08),
                        0 0 18px rgba(16,185,129,0.22);
                ">
                    <div style="font-size:36px; margin-bottom:10px;"></div>
                    <div style="font-size:32px; font-weight:700; color:#10B981; margin-bottom:6px;">
                        {completed_goals}
                    </div>
                    <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:#10B981;
                        padding:4px 10px; border-radius:6px; display:inline-block;">
                        Completed
                    </div>
                </div>
                """,
                unsafe_allow_html=True
            )


        # -------- col_perf3 : Active --------
        with col_perf3:
            st.markdown(
                f"""
                <div style="
                    background:#FFFFFF;
                    width:100%;
                    height:160px;
                    padding:20px;
                    border-radius:10px;
                    text-align:center;
                    display:flex;
                    flex-direction:column;
                    justify-content:center;
                    box-shadow:
                        0 2px 4px rgba(0,0,0,0.08),
                        0 0 18px rgba(245,87,108,0.23);
                ">
                    <div style="font-size:36px; margin-bottom:10px;"></div>
                    <div style="font-size:32px; font-weight:700; color:#F5576C; margin-bottom:6px;">
                        {active_goals}
                    </div>
                    <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:#F5576C;
                        padding:4px 10px; border-radius:6px; display:inline-block;">
                        Active
                    </div>
                </div>
                """,
                unsafe_allow_html=True
            )


        # -------- col_perf4 : Avg Progress --------
        with col_perf4:
            st.markdown(
                f"""
                <div style="
                    background:#FFFFFF;
                    width:100%;
                    height:160px;
                    padding:20px;
                    border-radius:10px;
                    text-align:center;
                    display:flex;
                    flex-direction:column;
                    justify-content:center;
                    box-shadow:
                        0 2px 4px rgba(0,0,0,0.08),
                        0 0 18px rgba(0,201,255,0.23);
                ">
                    <div style="font-size:36px; margin-bottom:10px;"></div>
                    <div style="font-size:32px; font-weight:700; color:#00C9FF; margin-bottom:6px;">
                        {avg_progress:.1f}%
                    </div>
                    <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:#00C9FF;
                        padding:4px 10px; border-radius:6px; display:inline-block;">
                        Avg Progress
                    </div>
                </div>
                """,
                unsafe_allow_html=True
            )


        # -------- col_perf5 : Overdue (Conditional) --------
        with col_perf5:
            overdue_color = "#EF4444" if overdue_goals > 0 else "#10B981"
            label_bg = "#FEF2F2" if overdue_goals > 0 else "#ECFDF5"
            glow = "rgba(239,68,68,0.23)" if overdue_goals > 0 else "rgba(16,185,129,0.23)"
            icon = "" if overdue_goals > 0 else ""

            st.markdown(
                f"""
                <div style="
                    background:#FFFFFF;
                    width:100%;
                    height:160px;
                    padding:20px;
                    border-radius:10px;
                    text-align:center;
                    display:flex;
                    flex-direction:column;
                    justify-content:center;
                    box-shadow:
                        0 2px 4px rgba(0,0,0,0.08),
                        0 0 18px {glow};
                ">
                    <div style="font-size:36px; margin-bottom:10px;">{icon}</div>
                    <div style="font-size:32px; font-weight:700; color:{overdue_color}; margin-bottom:6px;">
                        {overdue_goals}
                    </div>
                    <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:{overdue_color};
                        background:{label_bg}; padding:4px 10px; border-radius:6px; display:inline-block;">
                        Overdue
                    </div>
                </div>
                """,
                unsafe_allow_html=True
            )


        # Employee personal details modal
        if st.session_state.get('show_details') and st.session_state.get('show_details').startswith('emp_personal_'):
            st.markdown("---")
            detail_type = st.session_state.show_details
            
            col_header, col_close = st.columns([5, 1])
            
            with col_header:
                if detail_type == 'emp_personal_total':
                    st.subheader(f" Your Goals ({total_goals}) - {get_month_name(emp_selected_month)} {current_year}")
                elif detail_type == 'emp_personal_completed':
                    st.subheader(f" Your Completed Goals ({completed_goals})")
                elif detail_type == 'emp_personal_active':
                    st.subheader(f" Your Active Goals ({active_goals})")
                elif detail_type == 'emp_personal_progress':
                    st.subheader(f" Your Progress Breakdown")
                elif detail_type == 'emp_personal_overdue':
                    st.subheader(f" Your Overdue Goals ({overdue_goals})")
            
            with col_close:
                if st.button("✕ Close", key="close_emp_personal_details"):
                    del st.session_state.show_details
                    st.rerun()
            
            # Filter goals
            if detail_type == 'emp_personal_completed':
                display_goals = [g for g in user_goals if g.get('status') == 'Completed']
            elif detail_type == 'emp_personal_active':
                display_goals = [g for g in user_goals if g.get('status') == 'Active']
            elif detail_type == 'emp_personal_overdue':
                display_goals = []
                for goal in user_goals:
                    if goal.get('status') == 'Active':
                        end_date_str = goal.get('end_date')
                        if end_date_str:
                            try:
                                end_date = datetime.strptime(str(end_date_str), '%Y-%m-%d').date()
                                if today_date > end_date:
                                    display_goals.append(goal)
                            except:
                                pass
            elif detail_type == 'emp_personal_progress':
                display_goals = user_goals.copy()
                for g in display_goals:
                    achievement = g.get('monthly_achievement')
                    target = g.get('monthly_target', 1)
                    achievement_val = 0 if achievement is None else achievement
                    target_val = 1 if target is None or target == 0 else target
                    g['_progress'] = calculate_progress(achievement_val, target_val)
                display_goals.sort(key=lambda x: x['_progress'], reverse=True)
            else:
                display_goals = user_goals
            
            # Display goals table
            if display_goals:
                goal_data = []
                for goal in display_goals:


                    monthly_achievement = goal.get('monthly_achievement')
                    monthly_target = goal.get('monthly_target', 1)
                    achievement_value = 0 if monthly_achievement is None else monthly_achievement
                    target_value = 1 if monthly_target is None or monthly_target == 0 else monthly_target
                    progress = calculate_progress(achievement_value, target_value)
                    
                    goal_data.append({
                        'Goal Title': goal['goal_title'],
                        'Department': goal.get('department', 'N/A'),
                        'KPI': goal.get('kpi', 'N/A'),
                        'Month': get_month_name(goal.get('month', 1)),  # ADD THIS
                        'Year': str(goal['year']),  # ADD THIS
                        'Target': monthly_target if monthly_target is not None else 0,
                        'Achievement': monthly_achievement if monthly_achievement is not None else '-',
                        'Progress': f"{progress:.1f}%",
                        'Status': goal.get('status', 'Active')
                    })
                
                df_goals = pd.DataFrame(goal_data)
                st.dataframe(df_goals, use_container_width=True, height=400)
                
                st.download_button(
                    "📥 Export to CSV",
                    df_goals.to_csv(index=False).encode('utf-8'),
                    f"emp_personal_goals_{detail_type}_{emp_selected_month}_{current_year}.csv",
                    "text/csv",
                    key='download_emp_personal_modal_details'
                )
            else:
                st.info("No goals found in this category")
            
            st.markdown("---")

    # Notifications Section (common to all roles)
    st.markdown("---")

    col_notif_header, col_spacer, col_notif_actions, col_notif_clear = st.columns([4, 2, 1.2, 1.6])

    with col_notif_header:
        st.markdown("### 🔔 Recent Activity & Reminders")

    notifications = get_enhanced_notifications(user)
    unread_count = len([n for n in notifications if not n.get('is_read', False)])

    # Count by priority
    notif_counts = {'critical': 0, 'important': 0, 'normal': 0}
    for notif in notifications:
        if not notif.get('is_read', False):
            notif_type = notif.get('type', '')
            if notif_type in ['goal_not_completed', 'overdue', 'goal_not_updated']:
                notif_counts['critical'] += 1
            elif notif_type in ['deadline', 'feedback_received', 'goal_approved']:
                notif_counts['important'] += 1
            else:
                notif_counts['normal'] += 1

    with col_notif_actions:
        st.markdown("<div style='margin-top:8px'></div>", unsafe_allow_html=True)
        if st.button("View All"):
            st.session_state.show_all_notifications = True

    with col_notif_clear:
        st.markdown("<div style='margin-top:8px'></div>", unsafe_allow_html=True)
        if st.button("✓ Mark All Read"):
            try:
                supabase.table('notifications').update({
                    'is_read': True,
                    'read_at': datetime.now(IST).isoformat()
                }).eq('user_id', user['id']).eq('is_read', False).execute()
                st.success("✅ All notifications marked as read!")
                st.rerun()
            except Exception as e:
                st.error(f"Error: {str(e)}")

    st.caption(f"**{unread_count} unread notifications**")

    # Show only unread notifications
    unread_notifications = [n for n in notifications if not n.get('is_read', False)]

    if unread_notifications:
        display_count = len(unread_notifications) if st.session_state.get('show_all_notifications') else 5

        for notif in unread_notifications[:display_count]:
            notif_type = notif['type']

            if notif_type == 'goal_created':
                icon, color = "📝", "#3b82f6"
            elif notif_type == 'goal_approved':
                icon, color = "✅", "#10b981"
            elif notif_type == 'goal_edited':
                icon, color = "✏️", "#f59e0b"
            elif notif_type == 'goal_deleted':
                icon, color = "🗑️", "#ef4444"
            elif notif_type == 'achievement':
                icon, color = "🎉", "#10b981"
            elif notif_type == 'goal_not_completed':
                icon, color = "❌", "#ef4444"
            elif notif_type == 'update':
                icon, color = "📊", "#3b82f6"
            elif notif_type == 'goal_not_updated':
                icon, color = "⚠️", "#f97316"
            elif notif_type == 'feedback':
                icon, color = "💬", "#8b5cf6"
            elif notif_type == 'feedback_given':
                icon, color = "✍️", "#6366f1"
            elif notif_type == 'feedback_reply':
                icon, color = "↩️", "#8b5cf6"
            elif notif_type == 'deadline':
                icon, color = "⏰", "#f59e0b"
            elif notif_type == 'overdue':
                icon, color = "🚨", "#ef4444"
            elif notif_type == 'assignment':
                icon, color = "📬", "#3b82f6"
            else:
                icon, color = "ℹ️", "#64748b"

            col_notif, col_tick = st.columns([20, 1])
            with col_notif:
                st.markdown(f'''
                <div style="background: #ffffff; padding: 12px 16px; border-radius: 8px; border-left: 4px solid {color}; box-shadow: 0 1px 3px rgba(0,0,0,0.1);">
                    <div style="display: flex; justify-content: space-between; align-items: start;">
                        <div style="flex: 1;">
                            <div style="font-weight: 600; color: #1e293b; margin-bottom: 4px;">
                                {icon} {notif['title']}
                            </div>
                            <div style="color: #475569; font-size: 14px;">
                                {notif['message']}
                            </div>
                            <div style="font-size: 11px; color: #94a3b8; margin-top: 6px;">
                                {notif.get('time', '')}
                            </div>
                        </div>
                    </div>
                </div>
                ''', unsafe_allow_html=True)

            with col_tick:
                if notif.get('notification_id'):
                    if st.button("✓", key=f"mark_read_{notif.get('notification_id')}", help="Mark as read"):
                        if mark_notification_read(notif.get('notification_id')):
                            st.rerun()

        # Show more / show less button
        if len(unread_notifications) > 5 and not st.session_state.get('show_all_notifications'):
            if st.button(" Show All Notifications", use_container_width=True):
                st.session_state.show_all_notifications = True
                st.rerun()
        elif st.session_state.get('show_all_notifications'):
            if st.button(" Show Less", use_container_width=True):
                st.session_state.show_all_notifications = False
                st.rerun()
    else:
        st.info(" All caught up! No new notifications.")

def get_enhanced_notifications(user):
    """Get enhanced notifications from database only"""
    notifications = []
    role = user['role']
    
    # ✅ Fetch database notifications
    db_notifications = get_user_notifications(user['id'], limit=100)
    
    print(f" Fetched {len(db_notifications)} notifications for {user['name']}")
    
    if not db_notifications:
        print(f"⚠️ No notifications found in database for user {user['id']}")
        return []
    
    for notif in db_notifications:
        try:
            created_at = notif.get('created_at', 'Unknown time')
            
            # ✅ ROBUST TIME CALCULATION
            time_ago = "Recently"
            try:
                if isinstance(created_at, str) and created_at:
                    # Remove timezone indicator
                    timestamp_str = created_at.replace('Z', '').replace('+00:00', '')
                    
                    # Handle microseconds - normalize to 6 digits or remove entirely
                    if '.' in timestamp_str:
                        date_part, micro_part = timestamp_str.rsplit('.', 1)
                        
                        # Clean microseconds - keep only digits
                        micro_digits = ''.join(c for c in micro_part if c.isdigit())
                        
                        if len(micro_digits) > 6:
                            micro_digits = micro_digits[:6]
                        elif len(micro_digits) < 6:
                            micro_digits = micro_digits.ljust(6, '0')
                        
                        timestamp_str = f"{date_part}.{micro_digits}"
                        format_str = '%Y-%m-%dT%H:%M:%S.%f'
                    else:
                        format_str = '%Y-%m-%dT%H:%M:%S'
                    
                    # Parse timestamp
                    notif_dt = datetime.strptime(timestamp_str, format_str)
                    notif_dt = pytz.utc.localize(notif_dt)
                    
                    # Calculate time difference
                    now_dt = datetime.now(pytz.utc)
                    time_diff = now_dt - notif_dt
                    hours = time_diff.total_seconds() / 3600
                    
                    if hours < 0:
                        time_ago = "Just now"
                    elif hours < 1:
                        minutes = max(0, int(hours * 60))
                        time_ago = f"{minutes} min ago" if minutes > 0 else "Just now"
                    elif hours < 24:
                        hours_int = int(hours)
                        time_ago = f"{hours_int} hour{'s' if hours_int != 1 else ''} ago"
                    else:
                        days = int(hours / 24)
                        time_ago = f"{days} day{'s' if days != 1 else ''} ago"
                        
            except Exception as e:
                # Silently default to "Recently" - don't spam logs
                time_ago = "Recently"
            
            action_type = notif.get('action_type', 'update')
            
            # Map action types to notification types
            type_map = {
                'goal_created': ('goal_created', '📝'),
                'goal_approved': ('goal_approved', '✅'),
                'goal_edited': ('goal_edited', '✏️'),
                'goal_deleted': ('goal_deleted', '🗑️'),
                'goal_completed': ('achievement', '🎉'),
                'goal_not_completed': ('goal_not_completed', '❌'),
                'goal_assigned': ('assignment', '🎯'),
                'weekly_achievement_updated': ('update', '📊'),
                'goal_not_updated': ('goal_not_updated', '⚠️'),
                'feedback_received': ('feedback', '💬'),
                'feedback_given': ('feedback_given', '✍️'),
                'feedback_reply': ('feedback_reply', '↩️'),
                'goal_due_soon': ('deadline', '⏰'),
                'goal_overdue': ('overdue', '🚨'),
                'achievement_approved': ('goal_approved', '✅'),
                'achievement_rejected': ('goal_not_completed', '❌'),
                'test_notification': ('update', '🧪')
            }
            
            notif_type, icon = type_map.get(action_type, ('update', 'ℹ️'))
            
            notifications.append({
                'type': notif_type,
                'title': notif.get('action_type', 'Update').replace('_', ' ').title(),
                'message': notif.get('details', 'No details available'),
                'time': time_ago,
                'timestamp': created_at,
                'priority': 0 if not notif.get('is_read') else 5,
                'is_read': notif.get('is_read', False),
                'notification_id': notif.get('id')
            })
            
        except Exception as e:
            # Skip problematic notifications silently
            continue
    
    print(f"✅ Processed {len(notifications)} valid notifications")
    
    # Sort by priority and read status
    priority_weights = {
        'goal_not_completed': 1,
        'overdue': 1,
        'goal_not_updated': 2,
        'deadline': 3,
        'feedback_received': 4,
        'goal_approved': 5,
        'achievement': 5,
        'goal_created': 6,
        'update': 7,
        'feedback_reply': 8,
        'goal_edited': 9,
        'goal_deleted': 10
    }

    notifications.sort(key=lambda x: (
        x.get('priority', 999),  # Unread first (priority 0)
        priority_weights.get(x.get('type', ''), 999)
    ))
    
    return notifications[:50]

def debug_notifications():
    """Debug notification database"""
    st.title("🔍 Notification Debugger")
    
    user = st.session_state.user
    
    st.markdown(f"**Current User:** {user['name']} (ID: {user['id']})")
    
    # Check raw database
    st.subheader(" Raw Database Query")
    
    try:
        # Get ALL notifications
        all_notifs = supabase.table('notifications').select('*').order(
            'created_at', desc=True
        ).limit(20).execute()
        
        if all_notifs.data:
            st.success(f"✅ Found {len(all_notifs.data)} total notifications in database")
            
            # Show raw data
            df = pd.DataFrame(all_notifs.data)
            st.dataframe(df, use_container_width=True, height=300)
            
            # Check for current user
            user_notifs = [n for n in all_notifs.data if n.get('user_id') == user['id']]
            st.info(f"📍 {len(user_notifs)} notifications belong to current user")
            
            if user_notifs:
                st.markdown("**Your Notifications:**")
                for n in user_notifs[:5]:
                    st.json(n)
            else:
                st.warning("⚠️ No notifications found for current user!")
                st.info("Check if user_id in notifications matches your user ID")
                
        else:
            st.error("❌ No notifications found in database")
            
    except Exception as e:
        st.error(f"Database error: {str(e)}")
    
    st.markdown("---")
    
    # Test creating a notification
    st.subheader(" Test Create Notification")
    
    if st.button("Create Test Notification"):
        test_notif = {
            'user_id': user['id'],
            'action_by': user['id'],
            'action_by_name': user['name'],
            'action_type': 'test_notification',
            'details': f'Test notification created at {datetime.now(IST).strftime("%H:%M:%S")}',
            'is_read': False,
            'created_at': datetime.now(IST).isoformat()
        }
        
        result = create_notification(test_notif)
        
        if result:
            st.success("✅ Test notification created!")
            st.json(result)
            
            # Try to fetch it back
            fetched = get_user_notifications(user['id'], limit=5)
            st.info(f" Fetched {len(fetched)} notifications after creation")
            
            if fetched:
                st.success("✅ Notification successfully retrieved!")
                st.json(fetched[0])
            else:
                st.error("❌ Could not retrieve the notification we just created!")
        else:
            st.error("❌ Failed to create test notification")
def check_and_notify_missed_deadlines():
    """Check for goals past their deadline and notify managers"""
    try:
        today = date.today()
        all_users = db.get_all_users()
        
        for user in all_users:
            if user['role'] in ['Employee', 'Manager', 'HR', 'VP']:
                goals = db.get_user_all_goals(user['id'])
                
                for goal in goals:
                    if goal.get('status') == 'Active':
                        end_date_str = goal.get('end_date')
                        
                        if end_date_str:
                            try:
                                end_date = datetime.strptime(str(end_date_str), '%Y-%m-%d').date()
                                
                                # Check if deadline was yesterday (to send notification once)
                                if end_date == today - timedelta(days=1):
                                    # Check if goal was completed
                                    monthly_achievement = goal.get('monthly_achievement', 0)
                                    monthly_target = goal.get('monthly_target', 1)
                                    
                                    if monthly_achievement is None:
                                        monthly_achievement = 0
                                    
                                    progress = (monthly_achievement / monthly_target * 100) if monthly_target > 0 else 0
                                    
                                    # If not completed (less than 100%), notify manager
                                    if progress < 100:
                                        if user.get('manager_id'):
                                            manager = db.get_user_by_id(user['manager_id'])
                                            if manager and manager.get('email'):
                                                send_goal_completion_email(
                                                    manager['email'],
                                                    user['name'],
                                                    goal['goal_title'],
                                                    completed=False
                                                )
                                                
                                                # Also create in-app notification
                                                notification_data = {
                                                    'user_id': manager['id'],
                                                    'action_by': user['id'],
                                                    'action_by_name': user['name'],
                                                    'action_type': 'goal_deadline_missed',
                                                    'details': f"{user['name']}'s goal '{goal['goal_title']}' deadline was missed ({progress:.1f}% completed)",
                                                    'is_read': False,
                                                    'created_at': datetime.now(IST).isoformat()
                                                }
                                                create_notification(notification_data)
                            except Exception as e:
                                print(f"Error checking deadline for goal {goal['goal_id']}: {str(e)}")
                                continue
    except Exception as e:
        print(f"Error in check_and_notify_missed_deadlines: {str(e)}")

# ============================================
# VIEW ALL GOALS PAGE (NEW)
# ============================================
def display_view_all_goals():
    """View all goals of a user with edit/delete options"""
    user = st.session_state.user
    role = user['role']
    
    if not user:
        st.warning("⚠️ Session expired. Please login again.")
        st.rerun()

    # Check if viewing organization-wide goals
    view_mode = st.session_state.get('view_all_goals_mode', 'personal')
    # In display_dashboard() function, add at the very top:
    import time

    # ✅ Force notification refresh
    if 'last_notif_check' not in st.session_state:
        st.session_state.last_notif_check = time.time()

    # Check every 30 seconds
    current_time = time.time()
    if current_time - st.session_state.last_notif_check > 30:
        st.session_state.last_notif_check = current_time
        st.rerun()
    # Back button - different destinations based on mode
    col1, col2 = st.columns([1, 5])
    with col1:
        if view_mode == 'organization':
            if st.button("← Back "):
                st.session_state.page = 'hr_info'
                st.session_state.pop('view_all_goals_mode', None)
                st.rerun()
        else:
            if st.button("← Back"):
                st.session_state.page = 'my_goals'
                st.rerun()
    
    with col2:
        if view_mode == 'organization':
            st.title(" All Organization Goals")
        else:
            st.title(" View All Goals")
    
    # Select user based on mode
    if view_mode == 'organization' and role in ['CMD', 'VP', 'HR']:
        # Organization-wide view - show all users
        all_users = db.get_all_users()
        
        # Filter users based on role permissions
        if role == 'HR':
            viewable_users = all_users
        elif role == 'VP':
            viewable_users = [u for u in all_users if u['role'] in ['VP', 'HR', 'Manager', 'Employee']]
        else:  # CMD
            viewable_users = all_users
        
        user_options = ["All Users"] + [f"{u['name']} ({u['role']}) - {u['email']}" for u in viewable_users]
        selected_user_option = st.selectbox("Select User", user_options)
        
        if selected_user_option == "All Users":
            # Get all goals from all viewable users
            all_goals = []
            for u in viewable_users:
                user_goals = db.get_user_all_goals(u['id'])
                for g in user_goals:
                    g['user_name'] = u['name']
                    g['user_role'] = u['role']
                    g['user_department'] = u.get('department', 'N/A')
                all_goals.extend(user_goals)
            view_user_id = None  # Special case for all users
        else:
            user_email = selected_user_option.split(' - ')[1]
            selected_user_obj = next(u for u in viewable_users if u['email'] == user_email)
            view_user_id = selected_user_obj['id']
            all_goals = db.get_user_all_goals(view_user_id)
            # Add user info to goals
            for g in all_goals:
                g['user_name'] = selected_user_obj['name']
                g['user_role'] = selected_user_obj['role']
                g['user_department'] = selected_user_obj.get('department', 'N/A')
    
    else:
        # Personal/Team view (existing logic)
        all_goals = []  # Initialize empty list
        
        if role == 'HR':
            all_users = db.get_all_users()
            selected_user = st.selectbox(
                "Select User",
                [f"{u['name']} ({u['email']})" for u in all_users],
                key="hr_user_select"
            )
            user_email = selected_user.split('(')[1].strip(')')
            selected_user_obj = next(u for u in all_users if u['email'] == user_email)
            view_user_id = selected_user_obj['id']
            all_goals = db.get_user_all_goals(view_user_id)
            # Add user info
            for g in all_goals:
                g['user_name'] = selected_user_obj['name']
                g['user_role'] = selected_user_obj['role']
                g['user_department'] = selected_user_obj.get('department', 'N/A')
        
        elif role == 'CMD':
            # CMD can view all users
            all_users = db.get_all_users()
            selected_user = st.selectbox(
                "Select User",
                [f"{u['name']} ({u['role']}) - {u['email']}" for u in all_users],
                key="cmd_user_select"
            )
            user_email = selected_user.split(' - ')[1]
            selected_user_obj = next(u for u in all_users if u['email'] == user_email)
            view_user_id = selected_user_obj['id']
            all_goals = db.get_user_all_goals(view_user_id)
            # Add user info
            for g in all_goals:
                g['user_name'] = selected_user_obj['name']
                g['user_role'] = selected_user_obj['role']
                g['user_department'] = selected_user_obj.get('department', 'N/A')
        
        elif role == 'VP':
            # VP can view VP, HR, Manager, Employee (not CMD)
            all_users = db.get_all_users()
            viewable_users = [u for u in all_users if u['role'] in ['VP', 'HR', 'Manager', 'Employee']]
            selected_user = st.selectbox(
                "Select User",
                [f"{u['name']} ({u['role']}) - {u['email']}" for u in viewable_users],
                key="vp_user_select"
            )
            user_email = selected_user.split(' - ')[1]
            selected_user_obj = next(u for u in viewable_users if u['email'] == user_email)
            view_user_id = selected_user_obj['id']
            all_goals = db.get_user_all_goals(view_user_id)
            # Add user info
            for g in all_goals:
                g['user_name'] = selected_user_obj['name']
                g['user_role'] = selected_user_obj['role']
                g['user_department'] = selected_user_obj.get('department', 'N/A')
        
        elif role == 'Manager':
            team_members = db.get_team_members(user['id'])
            if team_members:
                selected_user = st.selectbox(
                    "Select Team Member",
                    [user['name']] + [f"{m['name']} ({m['email']})" for m in team_members],
                    key="manager_user_select"
                )
                if selected_user == user['name']:
                    view_user_id = user['id']
                    all_goals = db.get_user_all_goals(view_user_id)
                    # Add user info
                    for g in all_goals:
                        g['user_name'] = user['name']
                        g['user_role'] = user['role']
                        g['user_department'] = user.get('department', 'N/A')
                else:
                    user_email = selected_user.split('(')[1].strip(')')
                    selected_user_obj = next(m for m in team_members if m['email'] == user_email)
                    view_user_id = selected_user_obj['id']
                    all_goals = db.get_user_all_goals(view_user_id)
                    # Add user info
                    for g in all_goals:
                        g['user_name'] = selected_user_obj['name']
                        g['user_role'] = selected_user_obj['role']
                        g['user_department'] = selected_user_obj.get('department', 'N/A')
            else:
                view_user_id = user['id']
                all_goals = db.get_user_all_goals(view_user_id)
                # Add user info
                for g in all_goals:
                    g['user_name'] = user['name']
                    g['user_role'] = user['role']
                    g['user_department'] = user.get('department', 'N/A')
        
        else:  # Employee
            view_user_id = user['id']
            all_goals = db.get_user_all_goals(view_user_id)
            # Add user info
            for g in all_goals:
                g['user_name'] = user['name']
                g['user_role'] = user['role']
                g['user_department'] = user.get('department', 'N/A')
    
    if not all_goals:
        st.info("No goals found for this user")
        return
    
    # Filter options
    col1, col2, col3 = st.columns(3)
    with col1:
        filter_year = st.selectbox("Filter by Year", ["All"] + sorted(list(set([g['year'] for g in all_goals])), reverse=True))
    with col2:
        filter_status = st.selectbox("Filter by Status", ["All", "Active", "Completed", "On Hold", "Cancelled"])
    with col3:
        filter_month = st.selectbox(
            "Filter by Month",
            ["All"] + [get_month_name(i) for i in range(1, 13)],
            key="view_all_goals_month_filter"
        )
    
    # Apply filters
    filtered_goals = all_goals
    if filter_year != "All":
        filtered_goals = [g for g in filtered_goals if g['year'] == filter_year]
    if filter_status != "All":
        filtered_goals = [g for g in filtered_goals if g.get('status') == filter_status]
    if filter_month != "All":
        month_num = [get_month_name(i) for i in range(1, 13)].index(filter_month) + 1
        filtered_goals = [g for g in filtered_goals if g.get('month') == month_num]
    
    st.markdown(f"**Showing {len(filtered_goals)} of {len(all_goals)} goals**")
    
    # Display goals in expandable cards
    for goal in filtered_goals:
        monthly_achievement = goal.get('monthly_achievement')
        monthly_achievement = 0 if monthly_achievement is None else monthly_achievement
        monthly_target = goal.get('monthly_target', 1)
        progress = calculate_progress(monthly_achievement, monthly_target)
        
        # Determine status color
        if progress >= 100:
            status_color = "#10b981"
            status_text = "Completed"
        elif progress >= 60:
            status_color = "#f59e0b"
            status_text = "On Track"
        else:
            status_color = "#ef4444"
            status_text = "At Risk"
        
        if view_mode == 'organization':
            expander_title = f" {goal.get('user_name', 'Unknown')} ({goal.get('user_role', 'N/A')}) | 🎯 {goal['goal_title']} - {goal['year']}/Q{goal.get('quarter', 'N/A')}/M{goal.get('month', 'N/A')}"
        else:
            expander_title = f"🎯 {goal['goal_title']} - {goal['year']}/Q{goal.get('quarter', 'N/A')}/M{goal.get('month', 'N/A')}"
        
        with st.expander(expander_title):
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                # Show employee info in organization mode
                if view_mode == 'organization':
                    st.markdown(f"**Employee:** {goal.get('user_name', 'Unknown')}")
                    st.markdown(f"**Role:** {goal.get('user_role', 'N/A')}")
                    st.markdown(f"**Department:** {goal.get('user_department', 'N/A')}")
                else:
                    st.markdown(f"**Department:** {goal.get('department', 'N/A')}")
                st.markdown(f"**KPI:** {goal.get('kpi', 'N/A')}")
            
            with col2:
                st.markdown(f"**Start:** {goal['start_date']}")
                st.markdown(f"**End:** {goal['end_date']}")
            
            with col3:
                st.markdown(f"**Target:** {goal.get('monthly_target', 0)}")
                achievement_display = goal.get('monthly_achievement', 0) if goal.get('monthly_achievement') is not None else '-'
                st.markdown(f"**Achievement:** {achievement_display}")
            
            with col4:
                st.markdown(f"**Status:** <span style='color: {status_color}; font-weight: bold;'>{goal.get('status', 'Active')}</span>", unsafe_allow_html=True)
                st.markdown(f"**Progress:** <span style='color: {status_color}; font-weight: bold;'>{progress:.1f}%</span>", unsafe_allow_html=True)
            
            st.markdown(f"**Description:** {goal.get('goal_description', 'No description')}")
            
            # Edit and Delete buttons
            col_edit, col_delete, col_space = st.columns([1, 1, 3])
            
            with col_edit:
                if st.button("✏️ Edit", key=f"edit_goal_{goal['goal_id']}", use_container_width=True):
                    st.session_state.editing_goal = goal
                    st.rerun()
            
            with col_delete:
                if st.button("🗑️ Delete", key=f"delete_goal_{goal['goal_id']}", use_container_width=True):
                    goal_owner = db.get_user_by_id(goal['user_id'])
                    if db.delete_goal(goal['goal_id']):
                        if goal_owner:
                            notify_goal_deleted(goal, user, goal_owner)
                        st.success("Goal deleted!")
                        st.rerun()
    
    # Edit Goal Modal
    if 'editing_goal' in st.session_state:
        st.markdown("---")
        st.subheader("✏️ Edit Goal")
        
        edit_goal = st.session_state.editing_goal
        
        with st.form("edit_all_goals_form"):
            col1, col2 = st.columns(2)
            
            with col1:
                new_department = st.text_input("Department*", value=edit_goal.get('department', ''))
                new_title = st.text_input("Goal Title*", value=edit_goal['goal_title'])
                new_kpi = st.text_input("KPI*", value=edit_goal.get('kpi', ''))
                new_monthly_target = st.number_input("Monthly Target", min_value=0.0, value=float(edit_goal.get('monthly_target', 0)))
            
            with col2:
                new_status = st.selectbox("Status", ['Active', 'Completed', 'On Hold', 'Cancelled'],
                                         index=['Active', 'Completed', 'On Hold', 'Cancelled'].index(edit_goal.get('status', 'Active')))
                achievement_val = edit_goal.get('monthly_achievement')
                achievement_val = 0.0 if achievement_val is None else float(achievement_val)
                new_monthly_achievement = st.number_input("Monthly Achievement", min_value=0.0, value=achievement_val)
                new_description = st.text_area("Description", value=edit_goal.get('goal_description', ''))
            
            col_save, col_cancel = st.columns(2)
            
            with col_save:
                if st.form_submit_button("💾 Save Changes", use_container_width=True):
                    updates = {
                        'department': new_department,
                        'goal_title': new_title,
                        'kpi': new_kpi,
                        'monthly_target': new_monthly_target,
                        'monthly_achievement': new_monthly_achievement,
                        'goal_description': new_description,
                        'status': new_status
                    }
                    if db.update_goal(edit_goal['goal_id'], updates):
                        goal_owner = db.get_user_by_id(edit_goal['user_id'])
                        if goal_owner:
                            notify_goal_edited(edit_goal, user, goal_owner)
                        st.success("✅ Goal updated!")
                        del st.session_state.editing_goal
                        st.rerun()
            
            with col_cancel:
                if st.form_submit_button("❌ Cancel", use_container_width=True):
                    del st.session_state.editing_goal
                    st.rerun()



# ============================================
# HR INFO PAGE (FIXED)
# ============================================
def display_hr_info():
    """Display all HR information with delete user option"""
    user = st.session_state.user
    
    if not user:
        st.warning("⚠️ Session expired. Please login again.")
        st.rerun()

    # Allow CMD, VP, and HR to access
    if user['role'] not in ['CMD', 'VP', 'HR']:
        st.warning("⚠️ You don't have permission to access this page")
        return
    
    # Header with three-dot menu
    col_title, col_menu = st.columns([10, 1])
    with col_title:
        st.title(" Organization Info")
    
    with col_menu:
        # Three-dot menu using popover
        with st.popover("⋮", use_container_width=True):
            if st.button(" View All Goals", use_container_width=True, key="view_all_goals_hr_menu"):
                st.session_state.page = 'view_all_goals'
                st.session_state.view_all_goals_mode = 'organization'
                st.rerun()
    
    all_users = db.get_all_users()

    # Calculate all metrics first
    total_users = len(all_users)
    total_goals = sum([len(db.get_user_all_goals(u['id'])) for u in all_users])
    total_feedback = len(db.get_all_feedback())
    active_goals = len(db.get_all_active_goals())

    # Calculate overdue goals
    overdue_count = 0
    today = date.today()
    for u in all_users:
        user_goals = db.get_user_all_goals(u['id'])
        for goal in user_goals:
            if goal.get('status') == 'Active':
                end_date_str = goal.get('end_date')
                if end_date_str:
                    try:
                        end_date = datetime.strptime(str(end_date_str), '%Y-%m-%d').date()
                        if today > end_date:
                            overdue_count += 1
                    except:
                        pass

    # Display clickable metric cards
    col1, col2, col3, col4, col5 = st.columns(5)

    # ================= HR SUMMARY METRICS =================

# -------- col1 : Total Users --------
    with col1:
        st.markdown(
            f"""
            <div style="
                background:#FFFFFF;
                width:100%;
                height:160px;
                padding:20px;
                border-radius:10px;
                text-align:center;
                display:flex;
                flex-direction:column;
                justify-content:center;
                box-shadow:
                    0 2px 4px rgba(0,0,0,0.08),
                    0 0 18px rgba(59,130,246,0.25);
            ">
                <div style="font-size:36px; margin-bottom:10px;"></div>
                <div style="font-size:32px; font-weight:700; color:#3B82F6; margin-bottom:6px;">
                    {total_users}
                </div>
                <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:#3B82F6;
                    padding:4px 10px; border-radius:6px; display:inline-block;">
                    Total Users
                </div>
            </div>
            """,
            unsafe_allow_html=True
        )


    # -------- col2 : Total Goals --------
    with col2:
        st.markdown(
            f"""
            <div style="
                background:#FFFFFF;
                width:100%;
                height:160px;
                padding:20px;
                border-radius:10px;
                text-align:center;
                display:flex;
                flex-direction:column;
                justify-content:center;
                box-shadow:
                    0 2px 4px rgba(0,0,0,0.08),
                    0 0 18px rgba(139,92,246,0.25);
            ">
                <div style="font-size:36px; margin-bottom:10px;"></div>
                <div style="font-size:32px; font-weight:700; color:#8B5CF6; margin-bottom:6px;">
                    {total_goals}
                </div>
                <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:#8B5CF6;
                    padding:4px 10px; border-radius:6px; display:inline-block;">
                    Total Goals
                </div>
            </div>
            """,
            unsafe_allow_html=True
        )


    # -------- col3 : Total Feedback --------
    with col3:
        st.markdown(
            f"""
            <div style="
                background:#FFFFFF;
                width:100%;
                height:160px;
                padding:20px;
                border-radius:10px;
                text-align:center;
                display:flex;
                flex-direction:column;
                justify-content:center;
                box-shadow:
                    0 2px 4px rgba(0,0,0,0.08),
                    0 0 18px rgba(16,185,129,0.22);
            ">
                <div style="font-size:36px; margin-bottom:10px;"></div>
                <div style="font-size:32px; font-weight:700; color:#10B981; margin-bottom:6px;">
                    {total_feedback}
                </div>
                <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:#10B981;
                    padding:4px 10px; border-radius:6px; display:inline-block;">
                    Total Feedback
                </div>
            </div>
            """,
            unsafe_allow_html=True
        )


    # -------- col4 : Active Goals --------
    with col4:
        st.markdown(
            f"""
            <div style="
                background:#FFFFFF;
                width:100%;
                height:160px;
                padding:20px;
                border-radius:10px;
                text-align:center;
                display:flex;
                flex-direction:column;
                justify-content:center;
                box-shadow:
                    0 2px 4px rgba(0,0,0,0.08),
                    0 0 18px rgba(245,158,11,0.23);
            ">
                <div style="font-size:36px; margin-bottom:10px;"></div>
                <div style="font-size:32px; font-weight:700; color:#F59E0B; margin-bottom:6px;">
                    {active_goals}
                </div>
                <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:#F59E0B;
                    padding:4px 10px; border-radius:6px; display:inline-block;">
                    Active Goals
                </div>
            </div>
            """,
            unsafe_allow_html=True
        )


    # -------- col5 : Overdue Goals --------
    with col5:
        overdue_color = "#EF4444" if overdue_count > 0 else "#10B981"
        label_bg = "#FEF2F2" if overdue_count > 0 else "#ECFDF5"
        glow = "rgba(239,68,68,0.23)" if overdue_count > 0 else "rgba(16,185,129,0.23)"
        icon = "" if overdue_count > 0 else ""

        st.markdown(
            f"""
            <div style="
                background:#FFFFFF;
                width:100%;
                height:160px;
                padding:20px;
                border-radius:10px;
                text-align:center;
                display:flex;
                flex-direction:column;
                justify-content:center;
                box-shadow:
                    0 2px 4px rgba(0,0,0,0.08),
                    0 0 18px {glow};
            ">
                <div style="font-size:36px; margin-bottom:10px;">{icon}</div>
                <div style="font-size:32px; font-weight:700; color:{overdue_color}; margin-bottom:6px;">
                    {overdue_count}
                </div>
                <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:{overdue_color};
                    background:{label_bg}; padding:4px 10px; border-radius:6px; display:inline-block;">
                    Overdue Goals
                </div>
            </div>
            """,
            unsafe_allow_html=True
        )

    st.markdown("---")
    # HR Dashboard Details Modal
    if st.session_state.get('show_details') and st.session_state.get('show_details').startswith('hr_') and not st.session_state.get('show_details').startswith('hr_goal'):
        st.markdown("---")
        detail_type = st.session_state.show_details
        
        col_header, col_close = st.columns([5, 1])
        
        with col_header:
            if detail_type == 'hr_users':
                st.subheader(f" All Users ({total_users})")
            elif detail_type == 'hr_goals':
                st.subheader(f" All Goals ({total_goals})")
            elif detail_type == 'hr_feedback':
                st.subheader(f" All Feedback ({total_feedback})")
            elif detail_type == 'hr_active_goals':
                st.subheader(f" Active Goals ({active_goals})")
            elif detail_type == 'hr_overdue_goals':
                st.subheader(f" Overdue Goals ({overdue_count})")
        
        with col_close:
            if st.button("✕ Close", key="close_hr_dashboard_details"):
                del st.session_state.show_details
                st.rerun()
        
        # Display details based on type
        if detail_type == 'hr_users':
            # Show all users table
            users_data = []
            for u in all_users:
                stats = db.get_user_goal_stats(u['id'])
                manager_name = "N/A"
                if u.get('manager_id'):
                    manager = db.get_user_by_id(u['manager_id'])
                    if manager:
                        manager_name = manager['name']
                
                users_data.append({
                    'Name': u['name'],
                    'Email': u['email'],
                    'Role': u['role'],
                    'Department': u.get('department', 'N/A'),
                    'Manager': manager_name,
                    'Total Goals': stats.get('total_goals', 0),
                    'Completed': stats.get('completed_goals', 0),
                    'Progress %': f"{stats.get('avg_progress', 0):.1f}%"
                })
            
            df_users = pd.DataFrame(users_data)
            st.dataframe(df_users, use_container_width=True, height=400)
            
            st.download_button(
                "📥 Export to CSV",
                df_users.to_csv(index=False).encode('utf-8'),
                "all_users.csv",
                "text/csv",
                key='download_hr_users'
            )
        
        elif detail_type == 'hr_goals':
            # Show all goals
            all_goals_list = []
            for u in all_users:
                user_goals = db.get_user_all_goals(u['id'])
                for goal in user_goals:
                    goal['user_name'] = u['name']
                    goal['user_role'] = u['role']
                    goal['user_department'] = u.get('department', 'N/A')
                all_goals_list.extend(user_goals)
            
            if all_goals_list:
                goal_data = []
                for goal in all_goals_list:
                    monthly_achievement = goal.get('monthly_achievement')
                    monthly_target = goal.get('monthly_target', 1)
                    achievement_value = 0 if monthly_achievement is None else monthly_achievement
                    target_value = 1 if monthly_target is None or monthly_target == 0 else monthly_target
                    progress = calculate_progress(achievement_value, target_value)
                    
                    goal_data.append({
                        'Employee': goal['user_name'],
                        'Role': goal['user_role'],
                        'Department': goal['user_department'],
                        'Goal Title': goal['goal_title'],
                        'Month': get_month_name(goal.get('month', 1)),  # ADD THIS
                        'Year': str(goal['year']),  # ADD THIS
                        'Target': monthly_target if monthly_target is not None else 0,
                        'Achievement': monthly_achievement if monthly_achievement is not None else '-',
                        'Progress': f"{progress:.1f}%",
                        'Status': goal.get('status', 'Active'),
                        'Year': goal['year']
                    })
                
                df_goals = pd.DataFrame(goal_data)
                st.dataframe(df_goals, use_container_width=True, height=400)
                
                st.download_button(
                    "📥 Export to CSV",
                    df_goals.to_csv(index=False).encode('utf-8'),
                    "all_goals.csv",
                    "text/csv",
                    key='download_hr_all_goals'
                )
            else:
                st.info("No goals found")
        
        elif detail_type == 'hr_feedback':
            # Show all feedback
            all_feedback = db.get_all_feedback()
            
            if all_feedback:
                feedback_data = []
                for fb in all_feedback:
                    # Get user info
                    fb_user = db.get_user_by_id(fb.get('user_id'))
                    user_name = fb_user['name'] if fb_user else 'Unknown'
                    
                    feedback_data.append({
                        'Employee': user_name,
                        'Goal': fb.get('goal_title', 'N/A'),
                        'Type': fb.get('feedback_type', 'N/A'),
                        'Rating': fb.get('rating', 0),
                        'Comment': fb.get('comment', '')[:50] + '...' if len(fb.get('comment', '')) > 50 else fb.get('comment', ''),
                        'Date': fb.get('created_at', 'N/A')[:10] if fb.get('created_at') else 'N/A'
                    })
                
                df_feedback = pd.DataFrame(feedback_data)
                st.dataframe(df_feedback, use_container_width=True, height=400)
                
                st.download_button(
                    "📥 Export to CSV",
                    df_feedback.to_csv(index=False).encode('utf-8'),
                    "all_feedback.csv",
                    "text/csv",
                    key='download_hr_feedback'
                )
            else:
                st.info("No feedback found")
        
        elif detail_type == 'hr_active_goals':
            # Show active goals
            active_goals_list = []
            for u in all_users:
                user_goals = db.get_user_all_goals(u['id'])
                for goal in user_goals:
                    if goal.get('status') == 'Active':
                        goal['user_name'] = u['name']
                        goal['user_role'] = u['role']
                        goal['user_department'] = u.get('department', 'N/A')
                        active_goals_list.append(goal)
            
            if active_goals_list:
                goal_data = []
                for goal in active_goals_list:
                    monthly_achievement = goal.get('monthly_achievement')
                    monthly_target = goal.get('monthly_target', 1)
                    achievement_value = 0 if monthly_achievement is None else monthly_achievement
                    target_value = 1 if monthly_target is None or monthly_target == 0 else monthly_target
                    progress = calculate_progress(achievement_value, target_value)
                    
                    goal_data.append({
                        'Employee': goal['user_name'],
                        'Role': goal['user_role'],
                        'Department': goal['user_department'],
                        'Goal Title': goal['goal_title'],
                        'Month': get_month_name(goal.get('month', 1)),  # ADD THIS
                        'Year': str(goal['year']),  # ADD THIS
                        'Target': monthly_target if monthly_target is not None else 0,
                        'Achievement': monthly_achievement if monthly_achievement is not None else '-',
                        'Progress': f"{progress:.1f}%",
                        'End Date': goal.get('end_date', 'N/A')
                    })
                
                df_goals = pd.DataFrame(goal_data)
                st.dataframe(df_goals, use_container_width=True, height=400)
                
                st.download_button(
                    "📥 Export to CSV",
                    df_goals.to_csv(index=False).encode('utf-8'),
                    "active_goals.csv",
                    "text/csv",
                    key='download_hr_active_goals'
                )
            else:
                st.info("No active goals found")
        
        elif detail_type == 'hr_overdue_goals':
            # Show overdue goals
            overdue_goals_list = []
            today = date.today()
            
            for u in all_users:
                user_goals = db.get_user_all_goals(u['id'])
                for goal in user_goals:
                    if goal.get('status') == 'Active':
                        end_date_str = goal.get('end_date')
                        if end_date_str:
                            try:
                                end_date = datetime.strptime(str(end_date_str), '%Y-%m-%d').date()
                                if today > end_date:
                                    goal['user_name'] = u['name']
                                    goal['user_role'] = u['role']
                                    goal['user_department'] = u.get('department', 'N/A')
                                    goal['days_overdue'] = (today - end_date).days
                                    overdue_goals_list.append(goal)
                            except:
                                pass
            
            if overdue_goals_list:
                goal_data = []
                for goal in overdue_goals_list:
                    monthly_achievement = goal.get('monthly_achievement')
                    monthly_target = goal.get('monthly_target', 1)
                    achievement_value = 0 if monthly_achievement is None else monthly_achievement
                    target_value = 1 if monthly_target is None or monthly_target == 0 else monthly_target
                    progress = calculate_progress(achievement_value, target_value)
                    
                    goal_data.append({
                        'Employee': goal['user_name'],
                        'Role': goal['user_role'],
                        'Department': goal['user_department'],
                        'Goal Title': goal['goal_title'],
                        'Month': get_month_name(goal.get('month', 1)),  # ADD THIS
                        'Year': str(goal['year']),  # ADD THIS
                        'Target': monthly_target if monthly_target is not None else 0,
                        'Achievement': monthly_achievement if monthly_achievement is not None else '-',
                        'Progress': f"{progress:.1f}%",
                        'End Date': goal.get('end_date', 'N/A'),
                        'Days Overdue': goal['days_overdue']
                    })
                
                # Sort by days overdue (most overdue first)
                df_goals = pd.DataFrame(goal_data)
                df_goals = df_goals.sort_values('Days Overdue', ascending=False)
                st.dataframe(df_goals, use_container_width=True, height=400)
                
                st.download_button(
                    "📥 Export to CSV",
                    df_goals.to_csv(index=False).encode('utf-8'),
                    "overdue_goals.csv",
                    "text/csv",
                    key='download_hr_overdue_goals'
                )
            else:
                st.info("No overdue goals found")
        
        st.markdown("---")
     # Department-wise breakdown
    col_left, col_right = st.columns([1, 1])

    with col_left:
        st.markdown("####  Department Distribution")
        dept_count = {}
        for u in all_users:
            dept = normalize_department(u.get('department'))
            dept_count[dept] = dept_count.get(dept, 0) + 1

        if dept_count:
            fig_dept = px.pie(
                values=list(dept_count.values()),
                names=list(dept_count.keys()),
                title="Employees by Department",
                hole=0.4,
                color_discrete_sequence=px.colors.qualitative.Set3
            )
            fig_dept.update_traces(textposition='inside', textinfo='percent+label')
            fig_dept.update_layout(height=350, showlegend=True)
            st.plotly_chart(fig_dept, use_container_width=True)

    with col_right:
        st.markdown("####  Goals Performance")
        total_goals_org = sum([len(db.get_user_all_goals(u['id'])) for u in all_users])
        completed_goals_org = sum([db.get_user_goal_stats(u['id']).get('completed_goals', 0) for u in all_users])
        active_goals_org = sum([db.get_user_goal_stats(u['id']).get('active_goals', 0) for u in all_users])

        fig_goals = go.Figure(data=[
            go.Bar(name='Total', x=['Goals'], y=[total_goals_org], marker_color='#3b82f6'),
            go.Bar(name='Completed', x=['Goals'], y=[completed_goals_org], marker_color='#10b981'),
            go.Bar(name='Active', x=['Goals'], y=[active_goals_org], marker_color='#f59e0b')
        ])
        fig_goals.update_layout(
            title="Organization Goals Overview",
            barmode='group',
            height=350,
            showlegend=True
        )
        st.plotly_chart(fig_goals, use_container_width=True)

    st.markdown("---")
    # Tabs for different views
    tab1, tab2, tab3, tab4 = st.tabs(["👥 All Users", " Department Stats", "🎯 Goal Summary", "💬 Feedback Summary"])
    
    with tab1:
        st.subheader("All Users in System")
        
        users_data = []
        for u in all_users:
            stats = db.get_user_goal_stats(u['id'])
            manager_name = "N/A"
            if u.get('manager_id'):
                manager = db.get_user_by_id(u['manager_id'])
                if manager:
                    manager_name = manager['name']
            
            users_data.append({
                'Name': u['name'],
                'Email': u['email'],
                'Role': u['role'],
                'Department': u.get('department', 'N/A'),
                'Manager': manager_name,
                'Joining Date': u.get('joining_date', 'N/A'),
                'End Date': u.get('end_date', 'N/A'),
                'Total Goals': stats.get('total_goals', 0),
                'Completed': stats.get('completed_goals', 0),
                'Progress %': f"{stats.get('avg_progress', 0):.1f}%",
                'User_ID': u['id']
            })
        
        df_users = pd.DataFrame(users_data)
        
        # Display without User_ID column
        display_df = df_users.drop('User_ID', axis=1)
        st.dataframe(display_df, use_container_width=True, height=500)
        
        # Delete User option - only show deletable users based on hierarchy
        st.markdown("---")
        st.subheader("⚠️ Delete User")
        
        # Get deletable users based on current user's role
        deletable_users = [u for u in all_users 
                          if u['id'] != user['id'] 
                          and can_modify_user(user['role'], u['role'])]
        
        if deletable_users:
            col_del1, col_del2 = st.columns([2, 1])
            with col_del1:
                delete_user_select = st.selectbox(
                    "Select User to Delete",
                    [f"{u['name']} ({u['role']}) - {u['email']}" for u in deletable_users]
                )
            
            with col_del2:
                if st.button("🗑️ Delete Selected User", type="primary", use_container_width=True):
                    user_email = delete_user_select.split(' - ')[1]
                    delete_user_obj = next(u for u in deletable_users if u['email'] == user_email)
                    st.session_state['user_to_delete'] = delete_user_obj
            
            # Confirmation prompt
            if 'user_to_delete' in st.session_state:
                delete_user_obj = st.session_state['user_to_delete']
                
                st.warning(f"Are you sure you want to delete **{delete_user_obj['name']}** ({delete_user_obj['role']})?")
                st.error("⚠️ This will also delete all their goals and feedback!")
                
                confirm_col1, confirm_col2 = st.columns(2)
                with confirm_col1:
                    if st.button("Yes, Delete", key="confirm_delete_user"):
                        result = db.delete_user(delete_user_obj['id'])
                        if result:
                            st.success(f"✅ User {delete_user_obj['name']} deleted successfully!")
                            del st.session_state['user_to_delete']
                            st.rerun()
                        else:
                            st.error("❌ Failed to delete user")
                
                with confirm_col2:
                    if st.button("Cancel", key="cancel_delete_user"):
                        del st.session_state['user_to_delete']
                        st.info("❎ Deletion cancelled.")
        else:
            st.info("No users available for deletion based on your permissions")

    with tab2:
        st.subheader("Department-wise Statistics")
        
        dept_stats = {}
        for u in all_users:
            dept = normalize_department(u.get('department'))
            if dept not in dept_stats:
                dept_stats[dept] = {'users': 0, 'goals': 0, 'completed': 0}
            dept_stats[dept]['users'] += 1
            stats = db.get_user_goal_stats(u['id'])
            dept_stats[dept]['goals'] += stats.get('total_goals', 0)
            dept_stats[dept]['completed'] += stats.get('completed_goals', 0)
        
        dept_data = []
        for dept, stats in dept_stats.items():
            dept_data.append({
                'Department': dept,
                'Total Users': stats['users'],
                'Total Goals': stats['goals'],
                'Completed Goals': stats['completed'],
                'Completion Rate': f"{(stats['completed'] / stats['goals'] * 100) if stats['goals'] > 0 else 0:.1f}%"
            })
        
        df_dept = pd.DataFrame(dept_data)
        st.dataframe(df_dept, use_container_width=True)
    
    with tab3:
        st.subheader("Goal Summary")
        
        # Add filter options
        col_filter1, col_filter2 = st.columns(2)
        
        with col_filter1:
            filter_type = st.selectbox(
                "Filter by:",
                ["All Users", "Specific User", "By Department"],
                key="goal_summary_filter"
            )
        
        # ✅ FIXED: Define all_goals BEFORE using it
        all_goals = []
        selected_user_obj = None
        selected_dept = None
        dept_users = []
        
        # Get goals based on filter
        if filter_type == "All Users":
            for u in all_users:
                all_goals.extend(db.get_user_all_goals(u['id']))
        
        elif filter_type == "Specific User":
            with col_filter2:
                selected_user = st.selectbox(
                    "Select User:",
                    [f"{u['name']} ({u['role']}) - {u['email']}" for u in all_users],
                    key="goal_summary_user"
                )
            
            user_email = selected_user.split(' - ')[1]
            selected_user_obj = next(u for u in all_users if u['email'] == user_email)
            all_goals = db.get_user_all_goals(selected_user_obj['id'])
        
        else:  # By Department
            with col_filter2:
                departments = list(set([u.get('department', 'N/A') for u in all_users]))
                departments = sorted([d for d in departments if d and d != 'N/A'])
                selected_dept = st.selectbox(
                    "Select Department:",
                    departments,
                    key="goal_summary_dept"
                )
            
            # Get users from selected department
            dept_users = [u for u in all_users if u.get('department') == selected_dept]
            for u in dept_users:
                all_goals.extend(db.get_user_all_goals(u['id']))
        
        # Display filter info
        if filter_type == "All Users":
            st.info(f" Showing goals for **all {len(all_users)} users**")
        elif filter_type == "Specific User":
            st.info(f" Showing goals for **{selected_user_obj['name']}** ({selected_user_obj['role']})")
        else:
            st.info(f" Showing goals for **{selected_dept}** department ({len(dept_users)} users)")

        st.markdown("---")

        # Status breakdown with clickable cards
        status_count = {}
        for goal in all_goals:
            status = goal.get('status', 'Active')
            status_count[status] = status_count.get(status, 0) + 1
        
        col_status1, col_status2, col_status3, col_status4 = st.columns(4)

        with col_status1:
            render_clickable_metric_card("Active", str(status_count.get('Active', 0)), "#3B82F6", "🔄", "hr_goal_active", 'hr_active')

        with col_status2:
            render_clickable_metric_card("Completed", str(status_count.get('Completed', 0)), "#10B981", "✅", "hr_goal_completed", 'hr_completed')

        with col_status3:
            render_clickable_metric_card("On Hold", str(status_count.get('On Hold', 0)), "#F59E0B", "⏸️", "hr_goal_onhold", 'hr_onhold')

        with col_status4:
            render_clickable_metric_card("Cancelled", str(status_count.get('Cancelled', 0)), "#EF4444", "❌", "hr_goal_cancelled", 'hr_cancelled')
        
        # Goal details modal
        if st.session_state.get('show_details') and st.session_state.get('show_details').startswith('hr_'):
            st.markdown("---")
            detail_type = st.session_state.show_details
            
            col_header, col_close = st.columns([5, 1])
            
            with col_header:
                status_map = {
                    'hr_active': 'Active',
                    'hr_completed': 'Completed',
                    'hr_onhold': 'On Hold',
                    'hr_cancelled': 'Cancelled'
                }
                status = status_map.get(detail_type, 'All')
                st.subheader(f"📋 {status} Goals ({status_count.get(status, 0)})")
            
            with col_close:
                if st.button("✕ Close", key="close_hr_goal_details"):
                    del st.session_state.show_details
                    st.rerun()
            
            # Filter goals by status
            filtered_goals = [g for g in all_goals if g.get('status') == status]
            
            if filtered_goals:
                goal_data = []
                for goal in filtered_goals:
                    # Get user info
                    goal_user = db.get_user_by_id(goal['user_id'])
                    user_name = goal_user['name'] if goal_user else 'Unknown'
                    
                    monthly_achievement = goal.get('monthly_achievement')
                    monthly_target = goal.get('monthly_target', 1)
                    achievement_value = 0 if monthly_achievement is None else monthly_achievement
                    target_value = 1 if monthly_target is None or monthly_target == 0 else monthly_target
                    progress = calculate_progress(achievement_value, target_value)
                    
                    goal_data.append({
                        'Employee': user_name,
                        'Goal Title': goal['goal_title'],
                        'Department': goal.get('department', 'N/A'),
                        'Target': monthly_target if monthly_target is not None else 0,
                        'Achievement': monthly_achievement if monthly_achievement is not None else '-',
                        'Progress': f"{progress:.1f}%",
                        'Year': str(goal['year']),
                        'Month': get_month_name(goal.get('month', 1))
                    })
                
                df_goals = pd.DataFrame(goal_data)
                st.dataframe(df_goals, use_container_width=True, height=400)
                
                st.download_button(
                    "📥 Export to CSV",
                    df_goals.to_csv(index=False).encode('utf-8'),
                    f"hr_{status.lower()}_goals.csv",
                    "text/csv",
                    key='download_hr_goal_details'
                )
            else:
                st.info(f"No {status} goals found")
            
            st.markdown("---")
        
        # Year-wise breakdown
        st.markdown("---")
        st.subheader("Year-wise Goal Distribution")
        
        year_stats = {}
        for goal in all_goals:
            year = goal['year']
            if year not in year_stats:
                year_stats[year] = 0
            year_stats[year] += 1
        
        if year_stats:
            # Create dataframe for proper plotting
            df_year = pd.DataFrame({
                'Year': list(year_stats.keys()),
                'Goals': list(year_stats.values())
            })
            
            fig = px.bar(
                df_year,
                x='Year',
                y='Goals',
                title='Goals by Year',
                color='Goals',
                color_continuous_scale='Blues'
            )
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("No goals data available")
    
    with tab4:
        st.subheader("Feedback Summary")
        
        # User filter
        user_filter_options = ["All Users"] + [f"{u['name']} ({u['email']})" for u in all_users]
        selected_user_filter = st.selectbox("Filter by User", user_filter_options)
        
        # Get feedback based on filter
        if selected_user_filter == "All Users":
            all_feedback = db.get_all_feedback()
            filter_display = "All Users"
        else:
            user_email = selected_user_filter.split('(')[1].strip(')')
            filtered_user = next(u for u in all_users if u['email'] == user_email)
            all_feedback = db.get_user_all_feedback(filtered_user['id'])
            filter_display = filtered_user['name']
        
        st.markdown(f"**Showing feedback for: {filter_display}**")
        st.markdown("---")
        
        # Average ratings
        st.subheader("Average Ratings")
        
        if all_feedback:
            total_rating = sum([fb.get('rating', 0) for fb in all_feedback])
            avg_rating = total_rating / len(all_feedback) if all_feedback else 0
            
            st.metric(f"Average Rating for {filter_display}", f"{avg_rating:.2f} ⭐")
            
            # Rating distribution
            st.markdown("---")
            st.subheader("Rating Distribution")
            
            rating_count = {}
            for fb in all_feedback:
                rating = fb.get('rating', 0)
                rating_count[rating] = rating_count.get(rating, 0) + 1
            
            if rating_count:
                df_rating = pd.DataFrame({
                    'Rating': list(rating_count.keys()),
                    'Count': list(rating_count.values())
                })
                
                fig2 = px.bar(
                    df_rating,
                    x='Rating',
                    y='Count',
                    title=f'Feedback Rating Distribution - {filter_display}',
                    color='Rating',
                    color_continuous_scale='RdYlGn'
                )
                st.plotly_chart(fig2, use_container_width=True)
            else:
                st.info("No rating data available")
        else:
            st.info(f"No feedback data available for {filter_display}")
# ============================================
# EMPLOYEES PAGE
# ============================================
def display_employees_page():
    """Display employees with hierarchical drill-down"""
    user = st.session_state.user
    role = user['role']
    
    if not user:
        st.warning("⚠️ Session expired. Please login again.")
        st.rerun()

    # Get all users
    all_users = db.get_all_users()
    
    # Determine what to show based on role
    if role == 'CMD':
        # CMD sees VP, HR, and Manager cards at top level
        employees = [u for u in all_users if u['role'] in ['VP', 'HR', 'Manager']]
        view_title = "Organization Overview"
    elif role == 'VP':
        # VP sees HR and Manager cards
        employees = [u for u in all_users if u['role'] in ['HR', 'Manager']]
        view_title = "HR & Managers"
    elif role == 'HR':
        # HR sees only Manager cards
        employees = [u for u in all_users if u['role'] == 'Manager' or 
                    (u['role'] == 'Employee' and normalize_department(u.get('department')) == 'HR')]
        view_title = "Managers & HR Team"
    elif role == 'Manager':
        # Manager sees their team employees only
        employees = db.get_team_members(user['id'])
        view_title = "My Team"
    else:
        st.warning("⚠️ You don't have permission to view this page")
        return
    
    if not employees:
        st.info("No employees found")
        return
    
    st.title(f"👥 {view_title}")
    
    # Assign Goal Section (for users who can modify)
    if role in ['CMD', 'VP', 'HR', 'Manager']:
        with st.expander("➕ Assign Goal to Employee"):
            display_quick_assign_goal_form(user, employees)
    
    # Search and Filter for EMPLOYEES (not goals!)
    col1, col2 = st.columns([2, 1])
    with col1:
        search = st.text_input("🔍 Search by name or email", "")
    with col2:
        # Get unique departments from employees
        all_depts = list(set([normalize_department(e.get('department', '')) for e in employees if e.get('department')]))
        filter_dept = st.selectbox("Filter by Department", ["All"] + sorted(all_depts))
    
    # Filter employees (not goals!)
    filtered_employees = employees
    if search:
        filtered_employees = [e for e in filtered_employees if search.lower() in e['name'].lower() or search.lower() in e['email'].lower()]
    if filter_dept != "All":
        filtered_employees = [e for e in filtered_employees if normalize_department(e.get('department', '')) == filter_dept]
    
    # Display employee cards
    st.markdown("---")
    cols = st.columns(3)
    for idx, emp in enumerate(filtered_employees):
        with cols[idx % 3]:
            stats = db.get_user_goal_stats(emp['id'])
            
            # Role-based badge colors
            role_colors = {
                'CMD': '#8B0000',
                'VP': '#FF4500',
                'HR': '#4facfe',
                'Manager': '#f093fb',
                'Employee': '#dbeafe'
            }
            role_color = role_colors.get(emp['role'], '#dbeafe')

            st.markdown(f"""
            <div class='hierarchy-card'>
                <div style='text-align: center;'>
                    <div style='width: 60px; height: 60px; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                                border-radius: 50%; margin: 0 auto 10px; display: flex; align-items: center; 
                                justify-content: center; color: white; font-size: 24px; font-weight: bold;'>
                        {emp['name'][0].upper()}
                    </div>
                    <h3 style='margin: 5px 0;'>{emp['name']}</h3>
                    <p style='color: #64748b; font-size: 14px;'>{emp.get('designation', 'Employee')}</p>
                    <p style='color: #64748b; font-size: 12px;'>{emp.get('department', 'N/A')}</p>
                    <div style='margin-top: 10px;'>
                        <span style='background: {role_color}; color: white; padding: 3px 10px; 
                                     border-radius: 10px; font-size: 11px; font-weight: bold;'>
                            {emp['role']}
                        </span>
                    </div>
                    <div style='margin-top: 15px; font-size: 13px;'>
                        <p>Goals: {stats.get('total_goals', 0)} | Progress: {stats.get('avg_progress', 0):.1f}%</p>
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)
            
            # Determine button actions based on role and employee type
            if emp['role'] == 'VP':
                col_goals, col_team = st.columns(2)
                with col_goals:
                    if st.button("📊", key=f"view_vp_goals_{emp['id']}_{idx}", use_container_width=True, help="View Goals"):
                        st.session_state.viewing_employee = emp
                        st.session_state.viewing_employee_year = True
                        st.session_state.selected_year = date.today().year
                        st.session_state.page = 'employee_quarters'
                        st.rerun()
                with col_team:
                    if st.button("👥", key=f"view_vp_team_{emp['id']}_{idx}", use_container_width=True, help="View Team"):
                        st.session_state.viewing_vp_team = emp
                        st.session_state.page = 'vp_team_view'
                        st.rerun()
            
            elif emp['role'] == 'HR':
                # HR card - show their goals and team button
                col_goals, col_team = st.columns(2)
                with col_goals:
                    if st.button("📊", key=f"view_hr_goals_{emp['id']}_{idx}", use_container_width=True, help="View Goals"):
                        st.session_state.viewing_employee = emp
                        st.session_state.viewing_employee_year = True
                        st.session_state.selected_year = date.today().year
                        st.session_state.page = 'employee_quarters'
                        st.rerun()
                with col_team:
                    if st.button("👥", key=f"view_hr_team_{emp['id']}_{idx}", use_container_width=True, help="View Team"):
                        st.session_state.viewing_hr_team = emp
                        st.session_state.page = 'hr_team_view'
                        st.rerun()
            
            elif emp['role'] == 'Manager':
                # Manager card - show their goals and team button
                col_goals, col_team = st.columns(2)
                with col_goals:
                    if st.button("📊", key=f"view_mgr_goals_{emp['id']}_{idx}", use_container_width=True, help="View Goals"):
                        st.session_state.viewing_employee = emp
                        st.session_state.viewing_employee_year = True
                        st.session_state.selected_year = date.today().year
                        st.session_state.page = 'employee_quarters'
                        st.rerun()
                with col_team:
                    if st.button("👥", key=f"view_mgr_team_{emp['id']}_{idx}", use_container_width=True, help="View Team"):
                        st.session_state.viewing_manager_team = emp
                        st.session_state.page = 'manager_team_view'
                        st.rerun()
            
            else:  # Employee
                # Employee card - only goals button
                if st.button("👁️ View Goals", key=f"view_emp_{emp['id']}_{idx}", use_container_width=True):
                    st.session_state.viewing_employee = emp
                    st.session_state.viewing_employee_year = True
                    st.session_state.selected_year = date.today().year
                    st.session_state.page = 'employee_quarters'
                    st.rerun()
                    
                    # Get current year as default
                    today = date.today()
                    st.session_state.selected_year = today.year
                    
                    st.session_state.page = 'employee_quarters'  # ✅ GO DIRECTLY TO QUARTERS
                    st.rerun()

    
    # Edit Employee Modal - Show at top with expander
    if 'editing_employee' in st.session_state:
        st.markdown("---")
        
        
        edit_emp = st.session_state.editing_employee
        
        # Prominent header
        st.markdown(f"""
        <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                    padding: 20px; border-radius: 10px; margin-bottom: 20px;">
            <h2 style="color: white; margin: 0;">✏️ Editing: {edit_emp['name']}</h2>
        </div>
        """, unsafe_allow_html=True)
        
        with st.form("edit_employee_form"):
            col1, col2 = st.columns(2)
            
            with col1:
                new_name = st.text_input("Full Name*", value=edit_emp['name'])
                new_email = st.text_input("Email*", value=edit_emp['email'])
                new_designation = st.text_input("Designation", value=edit_emp.get('designation', ''))
            
            with col2:
                new_role = st.selectbox(
                    "Role*", 
                    ["Employee", "Manager", "HR"],
                    index=["Employee", "Manager", "HR"].index(edit_emp['role'])
                )
                new_department = st.text_input("Department", value=edit_emp.get('department', ''))
                
                # Manager assignment
                managers = [u for u in db.get_all_users() if u['role'] == 'Manager' and u['id'] != edit_emp['id']]
                manager_options = ["None"] + [f"{m['name']} ({m['email']})" for m in managers]
                
                current_manager_idx = 0
                if edit_emp.get('manager_id'):
                    current_manager = db.get_user_by_id(edit_emp['manager_id'])
                    if current_manager:
                        current_manager_str = f"{current_manager['name']} ({current_manager['email']})"
                        if current_manager_str in manager_options:
                            current_manager_idx = manager_options.index(current_manager_str)
                
                selected_manager = st.selectbox("Assign to Manager", manager_options, index=current_manager_idx)
            
            col_save, col_cancel = st.columns(2)
            
            with col_save:
                if st.form_submit_button("💾 Save Changes", use_container_width=True):
                    if new_name and new_email and new_role:
                        manager_id = None
                        if selected_manager != "None":
                            manager_email = selected_manager.split('(')[1].strip(')')
                            manager_id = next((m['id'] for m in managers if m['email'] == manager_email), None)
                        
                        updates = {
                            'name': new_name,
                            'email': new_email.lower().strip(),
                            'designation': new_designation,
                            'role': new_role,
                            'department': new_department,
                            'manager_id': manager_id
                        }
                        
                        if db.update_user(edit_emp['id'], updates):
                            st.success(f"✅ Employee {new_name} updated successfully!")
                            # No notification needed for employee creation per new requirements
                            del st.session_state.editing_employee
                            st.rerun()
                        else:
                            st.error("❌ Failed to update employee")
                    else:
                        st.error("❌ Please fill all required fields")
            
            with col_cancel:
                if st.form_submit_button("❌ Cancel", use_container_width=True):
                    del st.session_state.editing_employee
                    st.rerun()
    
    # Delete Employee Modal - Show at top with expander
    if 'deleting_employee' in st.session_state:
        st.markdown("---")
        st.markdown("---")
        
        del_emp = st.session_state.deleting_employee
        
        # Prominent warning header
        st.markdown(f"""
        <div style="background: linear-gradient(135deg, #ef4444 0%, #dc2626 100%); 
                    padding: 20px; border-radius: 10px; margin-bottom: 20px;">
            <h2 style="color: white; margin: 0;">⚠️ Delete Employee: {del_emp['name']}</h2>
        </div>
        """, unsafe_allow_html=True)
        
        st.warning(f"**Are you sure you want to delete {del_emp['name']}?**")
        st.error("⚠️ This will also delete all their goals and feedback! This action cannot be undone.")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.info(f"**Name:** {del_emp['name']}")
        with col2:
            st.info(f"**Email:** {del_emp['email']}")
        with col3:
            st.info(f"**Role:** {del_emp['role']}")
        
        # Get stats
        emp_goals = db.get_user_all_goals(del_emp['id'])
        st.warning(f" This will delete **{len(emp_goals)} goals** associated with this employee")
        
        confirm = st.checkbox("I understand this action cannot be undone", key="confirm_emp_delete")
        
        col_del1, col_del2, col_del3 = st.columns([1, 1, 1])
        
        with col_del2:
            if st.button(
                "🗑️ Delete Employee", 
                disabled=not confirm,
                use_container_width=True,
                type="primary",
                key="execute_emp_delete"
            ):
                if db.delete_user(del_emp['id']):
                    st.success(f"✅ Employee {del_emp['name']} deleted successfully!")
                    # No notification needed for employee creation per new requirements
                    del st.session_state.deleting_employee
                    st.balloons()
                    st.rerun()
                else:
                    st.error("❌ Failed to delete employee")
        
        with col_del3:
            if st.button("❌ Cancel", use_container_width=True, key="cancel_emp_delete"):
                del st.session_state.deleting_employee
                st.rerun()

# ============================================
# VP TEAM VIEW
# ============================================
def display_vp_team_view():
    """Show employees under a VP"""
    if not st.session_state.get('viewing_vp_team'):
        st.warning("⚠️ VP data lost. Returning to employees page...")
        st.session_state.page = 'employees'
        st.rerun()
    
    vp = st.session_state.viewing_vp_team
    user = st.session_state.user
    
    col1, col2 = st.columns([1, 5])
    with col1:
        if st.button("← Back"):
            st.session_state.pop('viewing_vp_team', None)
            st.session_state.page = 'employees'
            st.rerun()
    with col2:
        st.title(f"👥 {vp['name']}'s Team")
    
    # Get all employees under this VP (all employees managed by managers under this VP)
    all_users = db.get_all_users()
    
    # Get all managers under this VP
    vp_managers = [u for u in all_users if u['role'] == 'Manager' and u.get('manager_id') == vp['id']]
    
    # Get all employees under those managers
    vp_employees = []
    for manager in vp_managers:
        team_members = db.get_team_members(manager['id'])
        vp_employees.extend(team_members)
    
    if not vp_employees:
        st.info(f"No employees found under {vp['name']}")
        return
    
    st.markdown(f"**Total Team Members: {len(vp_employees)}**")
    st.markdown("---")
    
    # Display employee cards
    cols = st.columns(3)
    for idx, emp in enumerate(vp_employees):
        with cols[idx % 3]:
            stats = db.get_user_goal_stats(emp['id'])
            
            st.markdown(f"""
            <div class='hierarchy-card'>
                <div style='text-align: center;'>
                    <div style='width: 60px; height: 60px; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                                border-radius: 50%; margin: 0 auto 10px; display: flex; align-items: center; 
                                justify-content: center; color: white; font-size: 24px; font-weight: bold;'>
                        {emp['name'][0].upper()}
                    </div>
                    <h3 style='margin: 5px 0;'>{emp['name']}</h3>
                    <p style='color: #64748b; font-size: 14px;'>{emp.get('designation', 'Employee')}</p>
                    <p style='color: #64748b; font-size: 12px;'>{emp.get('department', 'N/A')}</p>
                    <div style='margin-top: 10px;'>
                        <span style='background: #dbeafe; color: #1e40af; padding: 3px 10px; 
                                     border-radius: 10px; font-size: 11px; font-weight: bold;'>
                            Employee
                        </span>
                    </div>
                    <div style='margin-top: 15px; font-size: 13px;'>
                        <p>Goals: {stats.get('total_goals', 0)} | Progress: {stats.get('avg_progress', 0):.1f}%</p>
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)
            
            if st.button("👁️ View Goals", key=f"view_vp_emp_{emp['id']}_{idx}", use_container_width=True):
                st.session_state.viewing_employee = emp
                st.session_state.viewing_employee_year = True
                st.session_state.selected_year = date.today().year
                st.session_state.page = 'employee_quarters'
                st.rerun()


# ============================================
# HR TEAM VIEW
# ============================================
def display_hr_team_view():
    """Show employees under an HR"""
    if not st.session_state.get('viewing_hr_team'):
        st.warning("⚠️ HR data lost. Returning to employees page...")
        st.session_state.page = 'employees'
        st.rerun()
    
    hr = st.session_state.viewing_hr_team
    user = st.session_state.user
    
    col1, col2 = st.columns([1, 5])
    with col1:
        if st.button("← Back"):
            st.session_state.pop('viewing_hr_team', None)
            st.session_state.page = 'employees'
            st.rerun()
    with col2:
        st.title(f"👥 {hr['name']}'s Team")
    
    # Get all employees under this HR (employees managed by this HR directly)
    hr_employees = db.get_team_members(hr['id'])
    
    if not hr_employees:
        st.info(f"No employees found under {hr['name']}")
        return
    
    st.markdown(f"**Total Team Members: {len(hr_employees)}**")
    st.markdown("---")
    
    # Display employee cards
    cols = st.columns(3)
    for idx, emp in enumerate(hr_employees):
        with cols[idx % 3]:
            stats = db.get_user_goal_stats(emp['id'])
            
            st.markdown(f"""
            <div class='hierarchy-card'>
                <div style='text-align: center;'>
                    <div style='width: 60px; height: 60px; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                                border-radius: 50%; margin: 0 auto 10px; display: flex; align-items: center; 
                                justify-content: center; color: white; font-size: 24px; font-weight: bold;'>
                        {emp['name'][0].upper()}
                    </div>
                    <h3 style='margin: 5px 0;'>{emp['name']}</h3>
                    <p style='color: #64748b; font-size: 14px;'>{emp.get('designation', 'Employee')}</p>
                    <p style='color: #64748b; font-size: 12px;'>{emp.get('department', 'N/A')}</p>
                    <div style='margin-top: 10px;'>
                        <span style='background: #dbeafe; color: #1e40af; padding: 3px 10px; 
                                     border-radius: 10px; font-size: 11px; font-weight: bold;'>
                            Employee
                        </span>
                    </div>
                    <div style='margin-top: 15px; font-size: 13px;'>
                        <p>Goals: {stats.get('total_goals', 0)} | Progress: {stats.get('avg_progress', 0):.1f}%</p>
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)
            
            if st.button("👁️ View Goals", key=f"view_hr_emp_{emp['id']}_{idx}", use_container_width=True):
                st.session_state.viewing_employee = emp
                st.session_state.viewing_employee_year = True
                st.session_state.selected_year = date.today().year
                st.session_state.page = 'employee_quarters'
                st.rerun()


# ============================================
# MANAGER TEAM VIEW
# ============================================
def display_manager_team_view():
    """Show employees under a Manager"""
    if not st.session_state.get('viewing_manager_team'):
        st.warning("⚠️ Manager data lost. Returning to employees page...")
        st.session_state.page = 'employees'
        st.rerun()
    
    manager = st.session_state.viewing_manager_team
    user = st.session_state.user
    
    col1, col2 = st.columns([1, 5])
    with col1:
        if st.button("← Back"):
            st.session_state.pop('viewing_manager_team', None)
            st.session_state.page = 'employees'
            st.rerun()
    with col2:
        st.title(f"👥 {manager['name']}'s Team")
    
    # Get team members
    team_members = db.get_team_members(manager['id'])
    
    if not team_members:
        st.info(f"No team members found under {manager['name']}")
        return
    
    st.markdown(f"**Total Team Members: {len(team_members)}**")
    st.markdown("---")
    
    # Display employee cards
    cols = st.columns(3)
    for idx, emp in enumerate(team_members):
        with cols[idx % 3]:
            stats = db.get_user_goal_stats(emp['id'])
            
            st.markdown(f"""
            <div class='hierarchy-card'>
                <div style='text-align: center;'>
                    <div style='width: 60px; height: 60px; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                                border-radius: 50%; margin: 0 auto 10px; display: flex; align-items: center; 
                                justify-content: center; color: white; font-size: 24px; font-weight: bold;'>
                        {emp['name'][0].upper()}
                    </div>
                    <h3 style='margin: 5px 0;'>{emp['name']}</h3>
                    <p style='color: #64748b; font-size: 14px;'>{emp.get('designation', 'Employee')}</p>
                    <p style='color: #64748b; font-size: 12px;'>{emp.get('department', 'N/A')}</p>
                    <div style='margin-top: 10px;'>
                        <span style='background: #dbeafe; color: #1e40af; padding: 3px 10px; 
                                     border-radius: 10px; font-size: 11px; font-weight: bold;'>
                            Employee
                        </span>
                    </div>
                    <div style='margin-top: 15px; font-size: 13px;'>
                        <p>Goals: {stats.get('total_goals', 0)} | Progress: {stats.get('avg_progress', 0):.1f}%</p>
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)
            
            if st.button("👁️ View Goals", key=f"view_mgr_emp_{emp['id']}_{idx}", use_container_width=True):
                st.session_state.viewing_employee = emp
                st.session_state.viewing_employee_year = True
                st.session_state.selected_year = date.today().year
                st.session_state.page = 'employee_quarters'
                st.rerun()

def display_quick_assign_goal_form(user, employees):
    """Quick assign goal form in employees page"""
    # Filter based on role hierarchy
    if user['role'] == 'HR':
        employees = [e for e in employees if e.get('manager_id') == user['id']]
        if not employees:
            st.info("You can only assign goals to employees in your team.")
            return
    elif user['role'] == 'VP':
        # VP can only assign to HR and Manager
        employees = [e for e in employees if e['role'] in ['HR', 'Manager']]
        if not employees:
            st.info("You can only assign goals to HR and Manager roles.")
            return
    elif user['role'] == 'CMD':
        # CMD can only assign to VP
        employees = [e for e in employees if e['role'] == 'VP']
        if not employees:
            st.info("You can only assign goals to VP roles.")
            return
    
    with st.form("quick_assign_goal"):
        st.subheader("Quick Assign Goal")
        
        col1, col2 = st.columns(2)
        
        with col1:
            selected_emp = st.selectbox(
                "Assign To*",
                [f"{e['name']} ({e['email']})" for e in employees]
            )
            
            emp_email = selected_emp.split('(')[1].strip(')')
            emp_id = next(e['id'] for e in employees if e['email'] == emp_email)
            
            year = st.number_input("Year", min_value=2020, max_value=2100, value=datetime.now().year)
            quarter = st.selectbox("Quarter", [1, 2, 3, 4])
            month = st.selectbox("Month", list(range(1, 13)), index=datetime.now().month - 1)
        
        with col2:
            department = st.text_input("Department*")
            title = st.text_input("Goal Title*")
            kpi = st.text_input("KPI*")
        
        # Auto-fill dates based on month
            month_start = date(year, month, 1)
            month_end = date(year, month, calendar.monthrange(year, month)[1])
            
            col_date1, col_date2 = st.columns(2)
            with col_date1:
                start_date = st.date_input("Start Date", value=month_start)
            with col_date2:
                end_date = st.date_input("End Date", value=month_end)
            
            description = st.text_area("Description")
            monthly_target = st.number_input("Monthly Target*", min_value=0.0)
            
            st.markdown("**Weekly Targets**")
            auto_divide = st.checkbox("Auto-divide monthly target equally into 4 weeks", value=True, key="quick_auto_divide")
            
            if auto_divide and monthly_target > 0:
                week_target = monthly_target / 4
            else:
                week_target = 0.0
            
            col3, col4, col5, col6 = st.columns(4)
            with col3:
                w1_t = st.number_input("Week 1 Target", min_value=0.0, value=week_target, key="quick_w1")
            with col4:
                w2_t = st.number_input("Week 2 Target", min_value=0.0, value=week_target, key="quick_w2")
            with col5:
                w3_t = st.number_input("Week 3 Target", min_value=0.0, value=week_target, key="quick_w3")
            with col6:
                w4_t = st.number_input("Week 4 Target", min_value=0.0, value=week_target, key="quick_w4")
        # ========== ADD THIS ENTIRE SECTION ==========
        st.markdown("**Weekly Ratings (Optional)**")
        col_r1, col_r2, col_r3, col_r4 = st.columns(4)
        with col_r1:
            w1_rating = st.selectbox("Week 1 Rating", [0, 1, 2, 3, 4], key="quick_w1_rating")
        with col_r2:
            w2_rating = st.selectbox("Week 2 Rating", [0, 1, 2, 3, 4], key="quick_w2_rating")
        with col_r3:
            w3_rating = st.selectbox("Week 3 Rating", [0, 1, 2, 3, 4], key="quick_w3_rating")
        with col_r4:
            w4_rating = st.selectbox("Week 4 Rating", [0, 1, 2, 3, 4], key="quick_w4_rating")
        # ========== END OF ADDED SECTION ==========
        if st.form_submit_button("✅ Assign Goal", use_container_width=True):
            if department and title and kpi:
                goal_data = {
                    'user_id': emp_id,
                    'year': year,
                    'quarter': quarter,
                    'month': month,
                    'department': department,
                    'goal_title': title,
                    'goal_description': description,
                    'kpi': kpi,
                    'monthly_target': monthly_target,
                    'week1_target': w1_t,
                    'week2_target': w2_t,
                    'week3_target': w3_t,
                    'week4_target': w4_t,
                    'week1_remarks': w1_rating,  # ✅ ADD
                    'week2_remarks': w2_rating,  # ✅ ADD
                    'week3_remarks': w3_rating,  # ✅ ADD
                    'week4_remarks': w4_rating,
                    'start_date': str(start_date),
                    'end_date': str(end_date),
                    'created_by': user['id']
                }
                
                if db.create_goal(goal_data):
                    # Get assigned employee
                    assigned_emp = db.get_user_by_id(emp_id)
                    if assigned_emp:
                        # Create goal data for notification
                        assigned_goal_data = {
                            'goal_title': title,
                            'user_id': emp_id
                        }
                        notify_goal_created(assigned_goal_data, user)
                    st.rerun()
            else:
                st.error("❌ Please fill all required fields")

def get_organization_month_goals(month=None, year=None):
    """Get all organization goals for a specific month"""
    if month is None:
        today = date.today()
        month = today.month
        year = today.year
    
    all_users = db.get_all_users()
    month_goals = []
    
    for user in all_users:
        user_goals = db.get_user_all_goals(user['id'])
        for goal in user_goals:
            if goal['year'] == year and goal.get('month') == month:
                goal['user_name'] = user['name']
                goal['user_role'] = user['role']
                goal['user_department'] = user.get('department', 'N/A')
                month_goals.append(goal)
    
    return month_goals

def get_user_month_goals(user_id, month=None, year=None):
    """Get user's goals for a specific month"""
    if month is None:
        today = date.today()
        month = today.month
        year = today.year
    
    all_goals = db.get_user_all_goals(user_id)
    month_goals = [g for g in all_goals if g['year'] == year and g.get('month') == month]
    
    return month_goals

# ============================================
# EMPLOYEE GOALS VIEW
# ============================================
def display_employee_goals():
    """Display goals for a specific employee"""
    if not st.session_state.get('viewing_employee'):
        st.warning("⚠️ Employee data lost. Returning to employees page...")
        st.session_state.page = 'employees'
        st.rerun()
    emp = st.session_state.viewing_employee
    user = st.session_state.user
    
    col1, col2 = st.columns([1, 5])
    with col1:
        if st.button("← Back"):
            st.session_state.page = 'employees'
            st.rerun()
    with col2:
        st.title(f" {emp['name']}'s Goals")
    
    # Show year selection for this employee
    years = db.get_years(emp['id'])
    
    if not years:
        st.info(f"No goals found for {emp['name']}")
        return
    
    st.subheader(" Select Year")
    
    # Display years in rows with goal counts
    sorted_years = sorted(years.items(), reverse=True)
    for year, summary in sorted_years:
        # Get goal count for this year
        year_goals = [g for g in db.get_user_all_goals(emp['id']) if g['year'] == year]
        goal_count = len(year_goals)
        
        st.markdown(f"""
        <div class='hierarchy-card' style='cursor: pointer;'>
            <h2 style='margin:0;'> {year} <span style='color: #64748b; font-size: 16px;'>({goal_count} goals)</span></h2>
            <p style='color: #64748b; margin-top: 8px;'>{summary[:80] if summary else 'Click to view quarters'}</p>
        </div>
        """, unsafe_allow_html=True)
        
        if st.button(f"View {year}", key=f"emp_year_{year}", use_container_width=True):
            st.session_state.selected_year = year
            st.session_state.viewing_employee_year = True
            st.session_state.page = 'employee_quarters'
            st.rerun()
        
        st.markdown("<br>", unsafe_allow_html=True)


# ============================================
# MY GOALS PAGE (UPDATED)
# ============================================
def display_my_goals():
    """Display user's own goals with month search"""
    user = st.session_state.user
    if not user:
        st.warning("⚠️ Session expired. Please login again.")
        st.session_state.page = 'login'
        st.rerun()
    role = user['role']
    
    # Header with three-dot menu
    col_title, col_menu = st.columns([10, 1])
    with col_title:
        st.title(f" My Goals - {user['name']}")
        st.caption(f"{user.get('designation', 'Employee')} • {user['role']}")

    with col_menu:
        # Three-dot menu using popover
        with st.popover("⋮", use_container_width=True):
            if st.button(" View All Goals", use_container_width=True, key="view_all_goals_menu"):
                st.session_state.page = 'view_all_goals'
                st.rerun()
            
            # Show Create/Delete Year options ONLY for HR
            if role == 'HR':
                st.markdown("---")
                if st.button("➕ Create New Year", use_container_width=True, key="create_year_menu"):
                    st.session_state.creating_new_year = True
                    st.rerun()
                
                if st.button("🗑️ Delete Year", use_container_width=True, key="delete_year_menu"):
                    st.session_state.show_delete_year_selector = True
                    st.rerun()
        
        
    
    # Get years data
    years = db.get_years(user['id'])
    current_year = datetime.now().year
    if current_year not in years:
        years[current_year] = ""


    # Add this after the title/caption and before year selection
    if role in ['HR', 'CMD', 'VP']:
        st.markdown("---")
        st.markdown("### 👤 Your Personal Performance")
        
        # Add month filter
        col_perf_title, col_perf_filter = st.columns([3, 1])
        with col_perf_title:
            st.caption("Filter your personal goals by month")
        with col_perf_filter:
            today = date.today()
            current_year = today.year
            personal_filter_month = st.selectbox(
                "Month",
                list(range(1, 13)),
                index=today.month - 1,
                format_func=lambda x: get_month_name(x),
                key="personal_perf_month_filter"
            )
        
        st.caption(f"**Your Performance: {get_month_name(personal_filter_month)} {current_year}**")
        
        personal_goals = get_user_month_goals(user['id'], personal_filter_month, current_year)
            
        # Calculate personal metrics
        personal_total = len(personal_goals)
        personal_completed = len([g for g in personal_goals if g.get('status') == 'Completed'])
        personal_active = len([g for g in personal_goals if g.get('status') == 'Active'])
        
        # Calculate average progress
        total_progress = 0
        goals_with_progress = 0
        for goal in personal_goals:
            monthly_achievement = goal.get('monthly_achievement')
            if monthly_achievement is not None:
                monthly_target = goal.get('monthly_target', 1)
                if monthly_target > 0:
                    progress = (monthly_achievement / monthly_target * 100)
                    total_progress += progress
                    goals_with_progress += 1
        
        personal_avg_progress = (total_progress / goals_with_progress) if goals_with_progress > 0 else 0
        
        # Count overdue goals
        personal_overdue = 0
        today_date = date.today()
        for goal in personal_goals:
            if goal.get('status') == 'Active':
                end_date_str = goal.get('end_date')
                if end_date_str:
                    try:
                        end_date = datetime.strptime(str(end_date_str), '%Y-%m-%d').date()
                        if today_date > end_date:
                            personal_overdue += 1
                    except:
                        pass
        
        # Display clickable metric cards
        col_p1, col_p2, col_p3, col_p4, col_p5 = st.columns(5)
        
        # ================= MY GOALS – PERSONAL METRICS =================

# -------- col_p1 : Total Goals --------
        with col_p1:
            st.markdown(
                f"""
                <div style="
                    background:#FFFFFF;
                    width:100%;
                    height:160px;
                    padding:20px;
                    border-radius:10px;
                    text-align:center;
                    display:flex;
                    flex-direction:column;
                    justify-content:center;
                    box-shadow:
                        0 2px 4px rgba(0,0,0,0.08),
                        0 0 18px rgba(59,130,246,0.25);
                ">
                    <div style="font-size:36px; margin-bottom:10px;"></div>
                    <div style="font-size:32px; font-weight:700; color:#3B82F6; margin-bottom:6px;">
                        {personal_total}
                    </div>
                    <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:#3B82F6;
                        padding:4px 10px; border-radius:6px; display:inline-block;">
                        Total Goals
                    </div>
                </div>
                """,
                unsafe_allow_html=True
            )


        # -------- col_p2 : Completed --------
        with col_p2:
            st.markdown(
                f"""
                <div style="
                    background:#FFFFFF;
                    width:100%;
                    height:160px;
                    padding:20px;
                    border-radius:10px;
                    text-align:center;
                    display:flex;
                    flex-direction:column;
                    justify-content:center;
                    box-shadow:
                        0 2px 4px rgba(0,0,0,0.08),
                        0 0 18px rgba(16,185,129,0.22);
                ">
                    <div style="font-size:36px; margin-bottom:10px;"></div>
                    <div style="font-size:32px; font-weight:700; color:#10B981; margin-bottom:6px;">
                        {personal_completed}
                    </div>
                    <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:#10B981;
                        padding:4px 10px; border-radius:6px; display:inline-block;">
                        Completed
                    </div>
                </div>
                """,
                unsafe_allow_html=True
            )


        # -------- col_p3 : Active --------
        with col_p3:
            st.markdown(
                f"""
                <div style="
                    background:#FFFFFF;
                    width:100%;
                    height:160px;
                    padding:20px;
                    border-radius:10px;
                    text-align:center;
                    display:flex;
                    flex-direction:column;
                    justify-content:center;
                    box-shadow:
                        0 2px 4px rgba(0,0,0,0.08),
                        0 0 18px rgba(245,87,108,0.23);
                ">
                    <div style="font-size:36px; margin-bottom:10px;"></div>
                    <div style="font-size:32px; font-weight:700; color:#F5576C; margin-bottom:6px;">
                        {personal_active}
                    </div>
                    <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:#F5576C;
                        padding:4px 10px; border-radius:6px; display:inline-block;">
                        Active
                    </div>
                </div>
                """,
                unsafe_allow_html=True
            )


        # -------- col_p4 : Avg Progress --------
        with col_p4:
            st.markdown(
                f"""
                <div style="
                    background:#FFFFFF;
                    width:100%;
                    height:160px;
                    padding:20px;
                    border-radius:10px;
                    text-align:center;
                    display:flex;
                    flex-direction:column;
                    justify-content:center;
                    box-shadow:
                        0 2px 4px rgba(0,0,0,0.08),
                        0 0 18px rgba(0,201,255,0.23);
                ">
                    <div style="font-size:36px; margin-bottom:10px;"></div>
                    <div style="font-size:32px; font-weight:700; color:#00C9FF; margin-bottom:6px;">
                        {personal_avg_progress:.1f}%
                    </div>
                    <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:#00C9FF;
                        padding:4px 10px; border-radius:6px; display:inline-block;">
                        Avg Progress
                    </div>
                </div>
                """,
                unsafe_allow_html=True
            )


        # -------- col_p5 : Overdue (Conditional) --------
        with col_p5:
            overdue_color = "#EF4444" if personal_overdue > 0 else "#10B981"
            label_bg = "#FEF2F2" if personal_overdue > 0 else "#ECFDF5"
            glow = "rgba(239,68,68,0.23)" if personal_overdue > 0 else "rgba(16,185,129,0.23)"
            icon = "" if personal_overdue > 0 else ""

            st.markdown(
                f"""
                <div style="
                    background:#FFFFFF;
                    width:100%;
                    height:160px;
                    padding:20px;
                    border-radius:10px;
                    text-align:center;
                    display:flex;
                    flex-direction:column;
                    justify-content:center;
                    box-shadow:
                        0 2px 4px rgba(0,0,0,0.08),
                        0 0 18px {glow};
                ">
                    <div style="font-size:36px; margin-bottom:10px;">{icon}</div>
                    <div style="font-size:32px; font-weight:700; color:{overdue_color}; margin-bottom:6px;">
                        {personal_overdue}
                    </div>
                    <div style="font-size:13px; font-weight:700; text-transform:uppercase; color:{overdue_color};
                        background:{label_bg}; padding:4px 10px; border-radius:6px; display:inline-block;">
                        Overdue
                    </div>
                </div>
                """,
                unsafe_allow_html=True
            )

        
        # Personal details modal (similar to dashboard)
        if st.session_state.get('show_details') and st.session_state.get('show_details').startswith('my_goals_personal_'):
            st.markdown("---")
            detail_type = st.session_state.show_details
            
            col_header, col_close = st.columns([5, 1])
            
            with col_header:
                if detail_type == 'my_goals_personal_total':
                    st.subheader(f" Your Goals ({personal_total}) - {get_month_name(personal_filter_month)} {current_year}")
                elif detail_type == 'my_goals_personal_completed':
                    st.subheader(f" Your Completed Goals ({personal_completed}) - {get_month_name(personal_filter_month)} {current_year}")
                elif detail_type == 'my_goals_personal_active':
                    st.subheader(f"🔄 Your Active Goals ({personal_active}) - {get_month_name(personal_filter_month)} {current_year}")
                elif detail_type == 'my_goals_personal_progress':
                    st.subheader(f"📈 Your Progress Breakdown - {get_month_name(personal_filter_month)} {current_year}")
                elif detail_type == 'my_goals_personal_overdue':
                    st.subheader(f"🚨 Your Overdue Goals ({personal_overdue}) - {get_month_name(personal_filter_month)} {current_year}")
            
            with col_close:
                if st.button("✕ Close", key="close_my_goals_personal_details"):
                    del st.session_state.show_details
                    st.rerun()
            
            # Filter goals
            if detail_type == 'my_goals_personal_completed':
                display_goals = [g for g in personal_goals if g.get('status') == 'Completed']
            elif detail_type == 'my_goals_personal_active':
                display_goals = [g for g in personal_goals if g.get('status') == 'Active']
            elif detail_type == 'my_goals_personal_overdue':
                display_goals = []
                for goal in personal_goals:
                    if goal.get('status') == 'Active':
                        end_date_str = goal.get('end_date')
                        if end_date_str:
                            try:
                                end_date = datetime.strptime(str(end_date_str), '%Y-%m-%d').date()
                                if today_date > end_date:
                                    display_goals.append(goal)
                            except:
                                pass
            elif detail_type == 'my_goals_personal_progress':
                display_goals = personal_goals.copy()
                for g in display_goals:
                    achievement = g.get('monthly_achievement')
                    target = g.get('monthly_target', 1)
                    achievement_val = 0 if achievement is None else achievement
                    target_val = 1 if target is None or target == 0 else target
                    g['_progress'] = calculate_progress(achievement_val, target_val)
                display_goals.sort(key=lambda x: x['_progress'], reverse=True)
            else:
                display_goals = personal_goals
            
            # Display goals table
            if display_goals:
                goal_data = []
                for goal in display_goals:
                    monthly_achievement = goal.get('monthly_achievement')
                    monthly_target = goal.get('monthly_target', 1)
                    achievement_value = 0 if monthly_achievement is None else monthly_achievement
                    target_value = 1 if monthly_target is None or monthly_target == 0 else monthly_target
                    progress = calculate_progress(achievement_value, target_value)
                    
                    goal_data.append({
                        'Goal Title': goal['goal_title'],
                        'Department': goal.get('department', 'N/A'),
                        'KPI': goal.get('kpi', 'N/A'),
                        'Month': get_month_name(goal.get('month', 1)),  # ADD THIS
                        'Year': str(goal['year']), # ADD THIS
                        'Target': monthly_target if monthly_target is not None else 0,
                        'Achievement': monthly_achievement if monthly_achievement is not None else '-',
                        'Progress': f"{progress:.1f}%",
                        'Status': goal.get('status', 'Active')
                    })
                
                df_goals = pd.DataFrame(goal_data)
                st.dataframe(df_goals, use_container_width=True, height=400)
                
                st.download_button(
                    "📥 Export to CSV",
                    df_goals.to_csv(index=False).encode('utf-8'),
                    f"my_goals_personal_{detail_type}_{personal_filter_month}_{current_year}.csv",
                    "text/csv",
                    key='download_my_goals_personal_modal_details'
                )
            else:
                st.info("No goals found in this category")
            
            st.markdown("---")
        
        st.markdown("---")
    # Month Quick Search and Year Browser - Side by Side
    col_search1, col_search2 = st.columns(2)
    
    with col_search1:
        search_month = st.selectbox(
            " Quick Search by Month",
            ["None"] + [get_month_name(i) for i in range(1, 13)]
        )
    
    with col_search2:
        sorted_years = sorted(years.keys(), reverse=True) if years else []
        if sorted_years:
            selected_year = st.selectbox(
                " Browse by Year",
                sorted_years,
                key="my_goals_year_select_top"
            )
        else:
            selected_year = None
    st.markdown("---")
    # If month search is active
    if search_month != "None":
        month_num = [get_month_name(i) for i in range(1, 13)].index(search_month) + 1
        
        st.markdown("---")
        st.markdown(f"### 📅 {search_month} Goals ")
        
        all_goals = db.get_user_all_goals(user['id'])
        month_goals = [g for g in all_goals if g.get('month') == month_num]
        
        if month_goals:
            # Group by year
            year_groups = {}
            for goal in month_goals:
                year = goal['year']
                if year not in year_groups:
                    year_groups[year] = []
                year_groups[year].append(goal)
            
            for year in sorted(year_groups.keys(), reverse=True):
                with st.expander(f"📅 {search_month} {year} ({len(year_groups[year])} goals)"):
                    for goal in year_groups[year]:
                        progress = calculate_progress(
                            goal.get('monthly_achievement', 0),
                            goal.get('monthly_target', 1)
                        )
                        
                        col1, col2, col3 = st.columns([2, 1, 1])
                        with col1:
                            st.markdown(f"**{goal['goal_title']}**")
                            st.caption(f"Department: {goal.get('department', 'N/A')} | KPI: {goal.get('kpi', 'N/A')}")
                        with col2:
                            st.metric("Target", goal.get('monthly_target', 0))
                        with col3:
                            st.metric("Progress", f"{progress:.1f}%")
                        
                        if st.button(f"View Goal", key=f"view_month_goal_{goal['goal_id']}"):
                            st.session_state.selected_year = year
                            st.session_state.selected_quarter = goal.get('quarter')
                            st.session_state.selected_month = month_num
                            st.session_state.page = 'month_goals'
                            st.rerun()
                        
                        st.markdown("---")
        else:
            st.info(f"No goals found for {search_month}")
        
        st.markdown("---")
    
    # If year is selected from top, show quarters directly
    if selected_year:
        st.session_state.selected_year = selected_year
        display_quarter_selection()
        return
    
    st.markdown("---")
    
    
    
    # Show create year form if button clicked
    if st.session_state.get('creating_new_year'):
        with st.expander("➕ Create New Year", expanded=True):
            with st.form("create_new_year_form"):
                new_year = st.number_input("Year*", min_value=2020, max_value=2100, value=datetime.now().year + 1)
                new_year_summary = st.text_area("Year Summary (Optional)", placeholder="Enter goals/plans for this year...")
                
                col_create1, col_create2 = st.columns(2)
                with col_create1:
                    if st.form_submit_button("✅ Create Year", use_container_width=True):
                        if new_year:
                            # Check if year already exists
                            if new_year in years:
                                st.error(f"❌ Year {new_year} already exists!")
                            else:
                                # Create year entry
                                if db.update_year_summary(user['id'], new_year, new_year_summary):
                                    st.success(f"✅ Year {new_year} created successfully!")
                                    st.session_state.creating_new_year = False
                                    st.rerun()
                                else:
                                    st.error("❌ Failed to create year")
                        else:
                            st.error("❌ Please enter a valid year")
                
                with col_create2:
                    if st.form_submit_button("❌ Cancel", use_container_width=True):
                        st.session_state.creating_new_year = False
                        st.rerun()
        st.markdown("---")
    
    # Show delete year selector and confirmation
    if st.session_state.get('show_delete_year_selector'):
        st.markdown("### 🗑️ Delete Year")
        
        delete_year = st.selectbox(
            "Select Year to Delete",
            sorted_years,
            key="delete_year_select"
        )
        
        if delete_year:
            st.warning(f"⚠️ Are you sure you want to delete year **{delete_year}**?")
            st.error("This will delete ALL goals and data for this year. This action cannot be undone!")
            
            # Show year stats
            year_goals = [g for g in db.get_user_all_goals(user['id']) if g['year'] == delete_year]
            st.info(f" This year has **{len(year_goals)} goal(s)** that will be deleted.")
            
            confirm_delete = st.checkbox("I understand this action cannot be undone", key="confirm_year_delete")
            
            col_del1, col_del2, col_del3 = st.columns([1, 1, 1])
            
            with col_del2:
                if st.button("🗑️ Delete Year", disabled=not confirm_delete, use_container_width=True, type="primary"):
                    if db.delete_year(user['id'], delete_year):
                        st.success(f"✅ Year {delete_year} deleted!")
                        st.session_state.show_delete_year_selector = False
                        st.rerun()
                    else:
                        st.error("❌ Failed to delete year")
            
            with col_del3:
                if st.button("❌ Cancel", use_container_width=True):
                    st.session_state.show_delete_year_selector = False
                    st.rerun()
        
        st.markdown("---")
    
    # Show rejected goals for employees
    if user['role'] == 'Employee':
        all_user_goals = db.get_user_all_goals(user['id'])
        rejected_goals = [g for g in all_user_goals if g.get('approval_status') == 'rejected']
        
        if rejected_goals:
            st.error(f"⚠️ You have {len(rejected_goals)} rejected goal(s)")
            
            with st.expander("View Rejected Goals"):
                for goal in rejected_goals:
                    st.markdown(f"**{goal['goal_title']}**")
                    st.caption(f"Period: {goal['year']}-Q{goal['quarter']}-M{goal['month']}")
                    st.warning(f"**Reason:** {goal.get('rejection_reason', 'No reason provided')}")
                    
                    if st.button(f"Revise & Resubmit", key=f"revise_{goal['goal_id']}"):
                        st.session_state.revising_goal = goal
                        st.rerun()
                    
                    st.markdown("---")
            
            st.markdown("---")
    
    # If no year selected, show message
    if not sorted_years:
        st.info(" No years found. Create a new year to get started.")
    else:
        st.info(" Please select a year from the dropdown above to view your goals.")

# ============================================
# QUARTER SELECTION PAGE (UPDATED WITH GOAL COUNT)
# ============================================
def display_quarter_selection():
    """Display quarter selection page with goal counts"""
    user = st.session_state.user
    year = st.session_state.selected_year
    
    if not year:
        st.warning("⚠️ Navigation state lost. Returning to My Goals...")
        st.session_state.page = 'my_goals'
        st.rerun()
    
    # Determine user_id first
    if st.session_state.get('viewing_employee_year'):
        if not st.session_state.get('viewing_employee'):
            st.warning("⚠️ Employee data lost. Returning to employees page...")
            st.session_state.page = 'employees'
            st.rerun()
        emp = st.session_state.viewing_employee
        user_id = emp['id']
        
        # ✅ SINGLE HEADER WITH BACK BUTTON
        if emp['role'] == 'Manager':
            col1, col2, col_menu = st.columns([1, 7, 1])
            with col1:
                if st.button("← Back "):
                    st.session_state.viewing_employee_year = False
                    st.session_state.pop('viewing_employee', None)
                    st.session_state.pop('selected_year', None)
                    st.session_state.page = 'employees'
                    st.rerun()
            with col2:
                st.title(f" {emp['name']}'s Year {year} - Quarters")
            with col_menu:
                with st.popover("⋮", use_container_width=True):
                    st.markdown("**View Team Member:**")
                    team_members = db.get_team_members(emp['id'])
                    if team_members:
                        for member in team_members:
                            if st.button(
                                f"👤 {member['name']}", 
                                key=f"switch_quarter_{member['id']}", 
                                use_container_width=True
                            ):
                                st.session_state.previous_manager = emp
                                st.session_state.viewing_employee = member
                                st.rerun()
                    else:
                        st.caption("No team members")
        else:
            # Regular header for non-managers
            col1, col2 = st.columns([1, 5])
            with col1:
                if st.session_state.get('previous_manager'):
                    if st.button("← Back "):
                        st.session_state.viewing_employee = st.session_state.previous_manager
                        st.session_state.pop('previous_manager', None)
                        st.rerun()
                else:
                    if st.button("← Back "):
                        st.session_state.viewing_employee_year = False
                        st.session_state.pop('viewing_employee', None)
                        st.session_state.pop('selected_year', None)
                        st.session_state.page = 'employees'
                        st.rerun()
            with col2:
                st.title(f" {emp['name']}'s Year {year} - Quarters")
        
        # ✅ YEAR FILTER - ALWAYS SHOW
        st.markdown("---")
        years = db.get_years(user_id)
        
        if years:
            col_year_filter, col_space = st.columns([1, 4])
            with col_year_filter:
                year_options = sorted(years.keys(), reverse=True)
                
                # Add current year if not in list
                if year not in year_options:
                    year_options = sorted([year] + year_options, reverse=True)
                
                selected_year_filter = st.selectbox(
                    "📅 Filter by Year",
                    year_options,
                    index=year_options.index(year) if year in year_options else 0,
                    key="employee_year_filter"
                )
                
                if selected_year_filter != year:
                    st.session_state.selected_year = selected_year_filter
                    st.rerun()
        else:
            # If no years found, still show current year option
            col_year_filter, col_space = st.columns([1, 4])
            with col_year_filter:
                st.selectbox(
                    "📅 Filter by Year",
                    [year],
                    index=0,
                    key="employee_year_filter_single",
                    disabled=True
                )
        
    else:
        # User viewing their own goals
        st.title(f" Year {year} - Quarters")
        user_id = user['id']
    
    # ✅ REST OF THE FUNCTION CONTINUES HERE
    quarters = db.get_quarters(user_id, year)
    for q in [1, 2, 3, 4]:
        if q not in quarters:
            quarters[q] = ""
    
    cols = st.columns(2)
    for idx, (quarter, summary) in enumerate(sorted(quarters.items())):
        with cols[idx % 2]:
            quarter_goals = [g for g in db.get_user_all_goals(user_id) if g['year'] == year and g.get('quarter') == quarter]
            goal_count = len(quarter_goals)
            
            st.markdown(f"""
            <div class='hierarchy-card'>
                <h2>📈 Quarter {quarter} <span style='color: #64748b; font-size: 14px;'>({goal_count} goals)</span></h2>
                <p style='color: #64748b;'>{get_quarter_name(quarter)}</p>
                <p style='margin-top: 8px;'>{summary[:80] if summary else 'Click to view months'}</p>
            </div>
            """, unsafe_allow_html=True)
            
            if st.button(f"Open Q{quarter}", key=f"q_{quarter}", use_container_width=True):
                st.session_state.selected_quarter = quarter
                st.session_state.page = 'employee_months' if st.session_state.get('viewing_employee_year') else 'months'
                st.rerun()
            
            # Only show edit for own goals
            if not st.session_state.get('viewing_employee_year'):
                with st.expander(f"✏️ Edit Q{quarter} Summary"):
                    with st.form(f"edit_q{quarter}"):
                        new_summary = st.text_area("Summary", value=summary, key=f"qsum_{quarter}")
                        if st.form_submit_button("Save"):
                            if db.update_quarter_summary(user_id, year, quarter, new_summary):
                                st.success("✅ Saved!")
                                st.rerun()


# ============================================
# MONTH SELECTION PAGE (UPDATED WITH GOAL COUNT)
# ============================================
def display_month_selection():
    """Display month selection page with goal counts"""
    user = st.session_state.user
    year = st.session_state.selected_year
    quarter = st.session_state.selected_quarter
    
    if not year or not quarter:
        st.warning("⚠️ Navigation state lost. Returning to My Goals...")
        st.session_state.page = 'my_goals'
        st.rerun()
    
    # Check if viewing employee goals
    if st.session_state.get('viewing_employee_year'):
        if not st.session_state.get('viewing_employee'):
            st.warning("⚠️ Employee data lost. Returning to employees page...")
            st.session_state.page = 'employees'
            st.rerun()
        emp = st.session_state.viewing_employee
        
        # Check if viewing a Manager's goals
        if emp['role'] == 'Manager':
            # Show header with team dropdown
            col1, col2, col_menu = st.columns([1, 7, 1])
            with col1:
                if st.button("← Back "):
                    st.session_state.page = 'employee_quarters'  # ✅ GO BACK TO QUARTERS
                    st.rerun()
            with col2:
                st.title(f"📅 {emp['name']}'s Year {year} - Q{quarter} - Months")
            with col_menu:
                # Three-dot menu for team members
                with st.popover("⋮", use_container_width=True):
                    st.markdown("**View Team Member:**")
                    
                    # Get team members under this manager
                    team_members = db.get_team_members(emp['id'])
                    
                    if team_members:
                        for member in team_members:
                            if st.button(
                                f"👤 {member['name']}", 
                                key=f"switch_month_{member['id']}", 
                                use_container_width=True
                            ):
                                # Store current manager for back navigation
                                st.session_state.previous_manager = emp
                                # Switch to viewing this team member
                                st.session_state.viewing_employee = member
                                st.rerun()
                    else:
                        st.caption("No team members")
        else:
            # Regular header for non-managers
            col1, col2 = st.columns([1, 5])
            with col1:
                # Check if we came from a manager's view
                if st.session_state.get('previous_manager'):
                    if st.button("← Back "):
                        # Go back to the manager's goal sheet
                        st.session_state.viewing_employee = st.session_state.previous_manager
                        st.session_state.pop('previous_manager', None)
                        st.rerun()
                else:
                    if st.button("← Back "):
                        st.session_state.page = 'employee_quarters'  # ✅ GO BACK TO QUARTERS
                        st.rerun()
            with col2:
                st.title(f"📅 {emp['name']}'s Year {year} - Q{quarter} - Months")
        
        user_id = emp['id']
    else:
        col1, col2 = st.columns([1, 5])
        with col1:
            if st.button("← Back "):
                st.session_state.page = 'quarters'
                st.rerun()
        with col2:
            st.title(f"📅 Year {year} - Q{quarter} - Months")
        user_id = user['id']
    
    quarter_month_nums = get_quarter_months(quarter)
    months = db.get_months(user_id, year, quarter)
    
    cols = st.columns(3)
    for idx, month_num in enumerate(quarter_month_nums):
        with cols[idx]:
            month_name = get_month_name(month_num)
            summary = months.get(month_num, "")
            month_goals = db.get_month_goals(user_id, year, quarter, month_num)
            goal_count = len(month_goals)

            st.markdown(f"""
            <div class='month-card'>
                <div class='month-card-content'>
                    <h2>📆 {month_name}</h2>
                    <p>{goal_count} goals</p>
                </div>
            </div>
            """, unsafe_allow_html=True)

            if st.button(
                f"Open {month_name}",
                key=f"m_{user_id}_{year}_{quarter}_{month_num}", 
                use_container_width=True):
                st.session_state.selected_month = month_num
                st.session_state.page = 'employee_month_goals' if st.session_state.get('viewing_employee_year') else 'month_goals'
                st.rerun()
            
            # Only show edit for own goals
            if not st.session_state.get('viewing_employee_year'):
                with st.expander(f"✏️ Edit {month_name} Summary"):
                    with st.form(f"edit_m{month_num}"):
                        new_summary = st.text_area("Summary", value=summary, key=f"msum_{month_num}")
                        if st.form_submit_button("Save"):
                            if db.update_month_summary(user_id, year, quarter, month_num, new_summary):
                                st.success("✅ Saved!")
                                st.rerun()


# ============================================
# MONTH GOALS PAGE (WITH ASSIGN IN MONTHLY VIEW)
# ============================================
def display_month_goals():
    """Display month goals with week tabs"""
    user = st.session_state.user
    year = st.session_state.selected_year
    quarter = st.session_state.selected_quarter
    month = st.session_state.selected_month

    if not year or not quarter or not month:
        st.warning("⚠️ Navigation state lost. Returning to dashboard...")
        st.session_state.page = 'dashboard'
        st.rerun()

    month_name = get_month_name(month)
    
    # ✅ CRITICAL: Determine viewing permissions based on role hierarchy
    is_read_only = False
    viewing_employee = st.session_state.get('viewing_employee')
    
    if st.session_state.get('viewing_employee_year') and viewing_employee:
        viewing_employee_role = viewing_employee['role']
        
        # HR can only edit their own team members
        if user['role'] == 'HR':
            if viewing_employee_role == 'CMD':
                is_read_only = True
            else:
                is_read_only = False 
                
        # Manager can only edit their team members (already handled by employees page access)
        elif user['role'] == 'Manager':
            if viewing_employee.get('manager_id') != user['id']:
                is_read_only = True
                
        # VP can edit HR and Manager only
        elif user['role'] == 'VP':
            if viewing_employee_role not in ['HR', 'Manager']:
                is_read_only = True
                
        # CMD can edit VP only
        elif user['role'] == 'CMD':
            if viewing_employee_role != 'VP':
                is_read_only = True
    
    # ===== HEADER WITH TEAM DROPDOWN FOR MANAGER =====
    if st.session_state.get('viewing_employee_year'):
        # Safety check for employee
        if not viewing_employee:
            st.warning("⚠️ Employee data lost. Returning to employees page...")
            st.session_state.page = 'employees'
            st.rerun()
        
        # Check if viewing a Manager's goals
        if viewing_employee['role'] == 'Manager':
            # Show header with team dropdown
            col1, col2, col_menu = st.columns([1, 7, 1])
            with col1:
                if st.button("← Back"):
                    st.session_state.page = 'employee_months'  # ✅ GO BACK TO MONTH SELECTION
                    st.rerun()
            with col2:
                st.title(f" {viewing_employee['name']}'s {month_name} {year} Goals")
            with col_menu:
                # Three-dot menu for team members
                with st.popover("⋮", use_container_width=True):
                    st.markdown("**View Team Member:**")
                    
                    # Get team members under this manager
                    team_members = db.get_team_members(viewing_employee['id'])
                    
                    if team_members:
                        for member in team_members:
                            if st.button(
                                f"👤 {member['name']}", 
                                key=f"switch_to_{member['id']}", 
                                use_container_width=True
                            ):
                                # Store current manager for back navigation
                                st.session_state.previous_manager = viewing_employee
                                # Switch to viewing this team member
                                st.session_state.viewing_employee = member
                                st.rerun()
                    else:
                        st.caption("No team members")
        else:
            # Regular header for non-managers
            col1, col2 = st.columns([1, 5])
            with col1:
                # Check if we came from a manager's view
                if st.session_state.get('previous_manager'):
                    if st.button("← Back"):
                        # Go back to the manager's goal sheet
                        st.session_state.viewing_employee = st.session_state.previous_manager
                        st.session_state.pop('previous_manager', None)
                        st.rerun()
                else:
                    if st.button("← Back"):
                        st.session_state.page = 'employee_months'
                        st.rerun()
            with col2:
                st.title(f" {viewing_employee['name']}'s {month_name} {year} Goals")
        
        display_user = viewing_employee
    else:
        # User viewing their own goals
        col1, col2 = st.columns([1, 5])
        with col1:
            if st.button("← Back"):
                st.session_state.page = 'months'
                st.rerun()
        with col2:
            st.title(f" {month_name} {year} Goal Sheet")
        display_user = user
        is_read_only = False  # Always False for own goals
    
    # Rest of the function remains the same...
    # Create tabs
    tab_all, tab_w1, tab_w2, tab_w3, tab_w4 = st.tabs([
        "📋 Monthly View",
        "📅 Week 1",
        "📅 Week 2",
        "📅 Week 3",
        "📅 Week 4"
    ])

    
    # Monthly view - PASS THE FLAG
    with tab_all:
        display_monthly_view(display_user, year, quarter, month,is_read_only)
    
    # Week views
    for week_num, tab in enumerate([tab_w1, tab_w2, tab_w3, tab_w4], 1):
        with tab:
            display_week_view(display_user, year, quarter, month, week_num)

def export_goals_to_excel(user_id, year, quarter, month):
    """Export goals to Excel with proper formatting including Monthly Achievement"""
    
    goals = db.get_month_goals(user_id, year, quarter, month)
    
    if not goals:
        st.warning("No goals to export")
        return None
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"{get_month_name(month)} {year}"
    
    # Define styles
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    subheader_fill = PatternFill(start_color="93C5FD", end_color="93C5FD", fill_type="solid")
    subheader_font = Font(bold=True, color="FFFFFF", size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "Department", "Goal Title", "KPI", "Monthly Target", "Start Date", "End Date"
    ]
    
    # Weekly headers
    weekly_headers = ["Week 1", "Week 2", "Week 3", "Week 4"]
    
    # Row 1: Main headers
    current_col = 1
    for header in headers:
        cell = ws.cell(row=1, column=current_col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
        current_col += 1
    
    # Add "Weekly Target" merged header
    ws.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col + 3)
    cell = ws.cell(row=1, column=current_col)
    cell.value = "Weekly Target"
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border
    current_col += 4
    
    # Add "Weekly Achievement" merged header
    ws.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col + 3)
    cell = ws.cell(row=1, column=current_col)
    cell.value = "Weekly Achievement"
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border
    current_col += 4
    
    # Add "Monthly Achievement" header (single column)
    cell = ws.cell(row=1, column=current_col, value="Monthly Achievement")
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border
    ws.merge_cells(start_row=1, start_column=current_col, end_row=2, end_column=current_col)
    current_col += 1
    
    # Add "Rating" merged header
    ws.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col + 3)
    cell = ws.cell(row=1, column=current_col)
    cell.value = "Rating"
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border
    current_col += 4
    
    # Row 2: Week subheaders
    current_col = 7  # Start after main headers
    
    # Week numbers under "Weekly Target"
    for week in weekly_headers:
        cell = ws.cell(row=2, column=current_col)
        cell.value = week
        cell.fill = subheader_fill
        cell.font = subheader_font
        cell.alignment = center_align
        cell.border = border
        current_col += 1
    
    # Week numbers under "Weekly Achievement"
    for week in weekly_headers:
        cell = ws.cell(row=2, column=current_col)
        cell.value = week
        cell.fill = subheader_fill
        cell.font = subheader_font
        cell.alignment = center_align
        cell.border = border
        current_col += 1
    
    # Skip Monthly Achievement column (already merged)
    current_col += 1
    
    # Week numbers under "Rating"
    for week in weekly_headers:
        cell = ws.cell(row=2, column=current_col)
        cell.value = week
        cell.fill = subheader_fill
        cell.font = subheader_font
        cell.alignment = center_align
        cell.border = border
        current_col += 1
    
    # Fill main header cells in row 2
    for col in range(1, 7):
        cell = ws.cell(row=2, column=col)
        cell.fill = header_fill
        cell.border = border
    
    # Data rows
    row_num = 3
    for goal in goals:
        col_num = 1
        
        # Main goal info
        ws.cell(row=row_num, column=col_num, value=goal.get('department', '')).border = border
        col_num += 1
        ws.cell(row=row_num, column=col_num, value=goal['goal_title']).border = border
        col_num += 1
        ws.cell(row=row_num, column=col_num, value=goal.get('kpi', '')).border = border
        col_num += 1
        ws.cell(row=row_num, column=col_num, value=goal.get('monthly_target', 0)).border = border
        col_num += 1
        ws.cell(row=row_num, column=col_num, value=goal.get('start_date', '')).border = border
        col_num += 1
        ws.cell(row=row_num, column=col_num, value=goal.get('end_date', '')).border = border
        col_num += 1
        
        # Weekly targets
        for week in range(1, 5):
            ws.cell(row=row_num, column=col_num, value=goal.get(f'week{week}_target', 0)).border = border
            col_num += 1
        
        # Weekly achievements
        for week in range(1, 5):
            achievement_val = goal.get(f'week{week}_achievement')
            cell_value = achievement_val if achievement_val is not None else '-'
            ws.cell(row=row_num, column=col_num, value=cell_value).border = border
            col_num += 1

        # Monthly achievement
        monthly_ach = goal.get('monthly_achievement')
        cell = ws.cell(row=row_num, column=col_num, value=monthly_ach if monthly_ach is not None else '-')
        cell.border = border
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        col_num += 1
        
        # ✅ FIX: Weekly ratings INSIDE the goal loop
        for week in range(1, 5):
            rating_value = goal.get(f'week{week}_rating', 0)
            cell = ws.cell(row=row_num, column=col_num, value=rating_value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            
            # Color based on rating
            if rating_value == 1:
                cell.fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")  # Light red
            elif rating_value == 2:
                cell.fill = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")  # Light amber
            elif rating_value == 3:
                cell.fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")  # Light yellow
            elif rating_value == 4:
                cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
            
            col_num += 1
        
        row_num += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15  # Department
    ws.column_dimensions['B'].width = 30  # Goal Title
    ws.column_dimensions['C'].width = 15  # KPI
    ws.column_dimensions['D'].width = 15  # Monthly Target
    ws.column_dimensions['E'].width = 12  # Start Date
    ws.column_dimensions['F'].width = 12  # End Date
    
    # Weekly columns (targets and achievements)
    for col in range(7, 15):  # Weeks 1-4 for targets and achievements
        ws.column_dimensions[chr(64 + col)].width = 12
    
    # Monthly Achievement column
    ws.column_dimensions['O'].width = 15
    
    # Rating columns
    for col in range(16, 20):
        ws.column_dimensions[chr(64 + col)].width = 12
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

def check_and_auto_complete_goal(goal_id):
    """Check if goal should be auto-completed based on monthly achievement"""
    goal = db.get_goal_by_id(goal_id)
    
    if not goal or goal.get('status') != 'Active':
        return False
    
    monthly_achievement = goal.get('monthly_achievement')
    monthly_target = goal.get('monthly_target', 0)
    
    if monthly_achievement is None or monthly_target == 0:
        return False
    
    # Check if monthly target is achieved (100% or more)
    progress = (monthly_achievement / monthly_target * 100) if monthly_target > 0 else 0
    
    if progress >= 100:
        # Auto-complete the goal
        updates = {
            'status': 'Completed',
            'completed_at': datetime.now(IST).isoformat(),
            'completion_remarks': 'Auto-completed: Monthly target achieved'
        }
        
        result = db.update_goal(goal_id, updates)
        
        if result:
            # Get user info for notification
            user = db.get_user_by_id(goal['user_id'])
            if user:
                notify_goal_completed(goal, user)
        
        return result
    
    return False

# ============================================
# MONTHLY VIEW (WITH ASSIGN GOAL)
# ============================================
def display_monthly_view(user, year, quarter, month, is_read_only=False):
    """Display monthly goals view with Excel-like format and assign goal option"""
    
    # Don't show any title - just the content
    if st.session_state.get('show_create_goal_form'):
        if not st.session_state.get('viewing_employee_year') or not is_read_only:
            with st.expander("➕ Create New Monthly Goal", expanded=True):
                display_add_goal_form_inline(user, year, quarter, month)
            st.markdown("---")
    
    # Assign Goal Section (for HR and Manager) - NOW IN MONTHLY VIEW
    if user['role'] in ['HR', 'Manager'] and not st.session_state.get('viewing_employee_year'):
        with st.expander("➕ Assign Goal to Employee"):
            display_assign_goal_form_monthly(user, year, quarter, month)
    
    # Get goals - only show approved goals for employees
    all_goals = db.get_month_goals(user['id'], year, quarter, month)

    if user['role'] == 'Employee':
        # Show only approved goals
        goals = [g for g in all_goals if g.get('approval_status') == 'approved']
        
        # Show pending count
        pending_count = len([g for g in all_goals if g.get('approval_status') == 'pending'])
        if pending_count > 0:
            st.info(f"ℹ️ You have {pending_count} goal(s) pending manager approval")
    else:
        # Managers/HR/VP/CMD see all goals
        goals = all_goals
    
    # Create & Export Buttons - Only show for own goals or if not read-only
    if not st.session_state.get('viewing_employee_year') or not is_read_only:
        col1, col2, col3 = st.columns([3, 1, 1])

        with col2:
            if st.button("➕ Create New Goal", key="create_goal_month", use_container_width=True):
                st.session_state.show_create_goal_form = True

        with col3:
            if st.button("📥 Export", key="export_excel_month", use_container_width=True):
                excel_file = export_goals_to_excel(user['id'], year, quarter, month)
                if excel_file:
                    st.download_button(
                        label="📄 Download",
                        data=excel_file,
                        file_name=f"Goals_{get_month_name(month)}_{year}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
        
    else:
        # Read-only view - only show export button
        col_space, col_export = st.columns([5, 1])
        with col_export:
            if st.button("📥 Export", key="export_excel_month_readonly", use_container_width=True):
                excel_file = export_goals_to_excel(user['id'], year, quarter, month)
                if excel_file:
                    st.download_button(
                        label="📄 Download",
                        data=excel_file,
                        file_name=f"Goals_{get_month_name(month)}_{year}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
        
    
    if goals:
        # Display goals in Excel-like format
        st.markdown("### 📋 Monthly Goal Sheet")
        
        # Create the formatted table data
        table_data = []
        for goal in goals:
            # ✅ FIX: Properly handle None values and convert to '-'
            def format_achievement(value):
                if value is None:
                    return '-'
                return value
            
            row = {
                'Department': goal.get('department', ''),
                'Goal Title': goal['goal_title'],
                'KPI': goal.get('kpi', ''),
                'Monthly Target': goal.get('monthly_target', 0),
                'Start Date': goal.get('start_date', ''),
                'End Date': goal.get('end_date', ''),
                # Weekly Targets
                'Week 1 Target': goal.get('week1_target', 0),
                'Week 2 Target': goal.get('week2_target', 0),
                'Week 3 Target': goal.get('week3_target', 0),
                'Week 4 Target': goal.get('week4_target', 0),
                # ✅ Weekly Achievements - properly formatted
                'Week 1 Achievement': format_achievement(goal.get('week1_achievement')),
                'Week 2 Achievement': format_achievement(goal.get('week2_achievement')),
                'Week 3 Achievement': format_achievement(goal.get('week3_achievement')),
                'Week 4 Achievement': format_achievement(goal.get('week4_achievement')),
                # ✅ Monthly Achievement - properly formatted
                'Monthly Achievement': format_achievement(goal.get('monthly_achievement')),
                # Weekly Remarks
                'Week 1 Remarks': goal.get('week1_remarks', ''),
                'Week 2 Remarks': goal.get('week2_remarks', ''),
                'Week 3 Remarks': goal.get('week3_remarks', ''),
                'Week 4 Remarks': goal.get('week4_remarks', '')
            }
            table_data.append(row)

        df = pd.DataFrame(table_data)
        
        # Display with better styling using custom HTML/CSS
        st.markdown("""
        <style>
        .excel-table {
            width: 100%;
            overflow-x: auto;
            overflow-y: auto;
            max-height: 600px;
            display: block;             /* ✅ Required for overflow to work */
        }

        /* Table Base */
        .excel-table table {
            border-collapse: collapse;
            width: 100%;
            min-width: 2000px;  /* Increased to ensure all columns fit */
            table-layout: fixed;          /* ✅ Force horizontal scroll for wide tables */
            font-size: 13px;
            background-color: #ffffff;
            color: #222222;
            border: 1px solid #ddd;
        }

        /* Unified Header Styling */
        .excel-table th,
        .excel-table .section-header {
            background-color: #3B82F6;
            color: #ffffff;
            padding: 10px;
            text-align: center;
            border: 1px solid #cfcfcf;
            font-weight: 600;
            position: sticky;
            top: 0;
            z-index: 10;
        }

        /* Table Cells */
        .excel-table td {
            border: 1px solid #e0e0e0;
            padding: 8px;
            text-align: center;
            background-color: #ffffff;
            vertical-align: middle;
        }

        /* Target columns */
        .excel-table .target-col {
            background-color: #F3F8FF;
        }

        /* Achievement columns */
        .excel-table .achievement-col {
            background-color: #F6FFF8;
        }

        /* Monthly Achievement */
        .excel-table .monthly-achievement-col {
            background-color: #E9FCEB;
            font-weight: bold;
        }

        /* Remarks columns */
        .excel-table .remarks-col {
            background-color: #FFFDF4;
            color: #444;
            word-wrap: break-word;
            max-width: 180px;
        }

        /* ✅ CUSTOM COLUMN WIDTHS - FIXED FOR ALL ROLES */

        /* Department column */
        .excel-table td:nth-child(1),
        .excel-table th:nth-child(1) {
            min-width: 120px;
            max-width: 120px;
            word-wrap: break-word;
            white-space: normal;
            text-align: center;
            line-height: 1.4;
            font-size: 12px;
        }

        /* Goal Title - FIXED WIDTH */
        .excel-table td:nth-child(2),
        .excel-table th:nth-child(2) {
            min-width: 200px;
            max-width: 200px;
            word-wrap: break-word;
            white-space: normal;
            text-align: left;
            line-height: 1.4;
            padding: 8px;
            font-size: 12px;
        }

        /* KPI - FIXED WIDTH */
        .excel-table th:nth-child(3),
        .excel-table td:nth-child(3) {
            min-width: 150px;
            max-width: 150px;
            word-wrap: break-word;
            white-space: normal;
            text-align: center;
            line-height: 1.4;
            font-size: 12px;
        }

        /* Monthly Target column */
        .excel-table td:nth-child(4),
        .excel-table th:nth-child(4) {
            min-width: 100px;
            max-width: 100px;
            text-align: center;
            font-size: 12px;
        }

        /* Start Date - FIXED WIDTH */
        .excel-table td:nth-child(5),
        .excel-table th:nth-child(5) {
            min-width: 100px;
            max-width: 100px;
            text-align: center;
            white-space: nowrap;
            font-size: 11px;
            padding: 4px 6px;
        }

        /* End Date - FIXED WIDTH */
        .excel-table td:nth-child(6),
        .excel-table th:nth-child(6) {
            min-width: 100px;
            max-width: 100px;
            text-align: center;
            white-space: nowrap;
            font-size: 11px;
            padding: 4px 6px;
        }

        /* Weekly Target columns (7-10) */
        .excel-table td:nth-child(7),
        .excel-table td:nth-child(8),
        .excel-table td:nth-child(9),
        .excel-table td:nth-child(10),
        .excel-table th:nth-child(7),
        .excel-table th:nth-child(8),
        .excel-table th:nth-child(9),
        .excel-table th:nth-child(10) {
            min-width: 90px;
            max-width: 90px;
            text-align: center;
            font-size: 12px;
        }

        /* Weekly Achievement columns (11-14) */
        .excel-table td:nth-child(11),
        .excel-table td:nth-child(12),
        .excel-table td:nth-child(13),
        .excel-table td:nth-child(14),
        .excel-table th:nth-child(11),
        .excel-table th:nth-child(12),
        .excel-table th:nth-child(13),
        .excel-table th:nth-child(14) {
            min-width: 90px;
            max-width: 90px;
            text-align: center;
            font-size: 12px;
        }

        /* Monthly Achievement column (15) */
        .excel-table td:nth-child(15),
        .excel-table th:nth-child(15) {
            min-width: 110px;
            max-width: 110px;
            text-align: center;
            font-weight: bold;
            font-size: 12px;
        }

        /* Rating columns (16-19) */
        .excel-table td:nth-child(16),
        .excel-table td:nth-child(17),
        .excel-table td:nth-child(18),
        .excel-table td:nth-child(19),
        .excel-table th:nth-child(16),
        .excel-table th:nth-child(17),
        .excel-table th:nth-child(18),
        .excel-table th:nth-child(19) {
            min-width: 80px;
            max-width: 80px;
            text-align: center;
            font-size: 12px;
        }

        /* Hover effect */
        .excel-table tr:hover td {
            background-color: #F9FAFB;
        }

        /* Bold key columns */
        .excel-table th,
        .excel-table td:first-child,
        .excel-table td:nth-child(2) {
            font-weight: 500;
        }

        /* Clean borders */
        .excel-table table,
        .excel-table th,
        .excel-table td {
            border: 1px solid #e6e6e6;
        }

        /* Goal title button styling */
        .goal-title-btn {
            background: none;
            border: none;
            color: #3B82F6;
            text-decoration: underline;
            font-weight: bold;
            cursor: pointer;
            padding: 0;
            font-size: inherit;
        }

        .goal-title-btn:hover {
            color: #1E40AF;
        }

        /* Modal styling */
        .goal-modal {
            display: none;
            position: fixed;
            z-index: 10000;
            left: 0;
            top: 0;
            width: 100%;
            height: 100%;
            background-color: rgba(0,0,0,0.5);
            overflow: auto;
        }

        .modal-content {
            background-color: #fefefe;
            margin: 5% auto;
            padding: 20px;
            border-radius: 10px;
            width: 80%;
            max-width: 600px;
            box-shadow: 0 4px 20px rgba(0,0,0,0.3);
            position: relative;
            animation: slideIn 0.3s ease-out;
        }

        @keyframes slideIn {
            from { transform: translateY(-50px); opacity: 0; }
            to { transform: translateY(0); opacity: 1; }
        }

        .close-btn {
            position: absolute;
            right: 15px;
            top: 15px;
            background: #ef4444;
            color: white;
            border: none;
            border-radius: 5px;
            padding: 8px 15px;
            cursor: pointer;
            font-weight: bold;
            font-size: 18px;
        }

        .close-btn:hover {
            background: #dc2626;
        }

        /* Responsive styling */
        @media (max-width: 768px) {
            .excel-table table {
                font-size: 11px;
            }
            .excel-table th,
            .excel-table td {
                padding: 6px;
            }
        }
        </style>
        """, unsafe_allow_html=True)

        # Store the complete HTML in a variable for components
        complete_html = f'''
        <!DOCTYPE html>
        <html>
        <head>
        <style>
        .excel-table {{
            width: 100%;
            overflow-x: auto;
        }}

        .excel-table table {{
            border-collapse: collapse;
            width: 100%;
            min-width: 1800px;
            font-size: 13px;
            background-color: #ffffff;
            color: #222222;
            border: 1px solid #ddd;
        }}

        .excel-table th {{
            background-color: #3B82F6;
            color: #ffffff;
            padding: 10px;
            text-align: center;
            border: 1px solid #cfcfcf;
            font-weight: 600;
        }}

        .excel-table td {{
            border: 1px solid #e0e0e0;
            padding: 8px;
            text-align: center;
            background-color: #ffffff;
            vertical-align: middle;
        }}

        .target-col {{ background-color: #F3F8FF; }}
        .achievement-col {{ background-color: #F6FFF8; }}
        .monthly-achievement-col {{ background-color: #E9FCEB; font-weight: bold; }}
        .remarks-col {{ background-color: #FFFDF4; max-width: 180px; word-wrap: break-word; }}

        .goal-title-btn {{
            background: none;
            border: none;
            color: #3B82F6;
            text-decoration: underline;
            font-weight: bold;
            cursor: pointer;
            padding: 0;
            font-size: inherit;
        }}

        .goal-title-btn:hover {{
            color: #1E40AF;
        }}

        .modal {{
            display: none;
            position: fixed;
            z-index: 10000;
            left: 0;
            top: 0;
            width: 100%;
            height: 100%;
            background-color: rgba(0,0,0,0.5);
            overflow: auto;
        }}

        .modal-content {{
            background-color: #fefefe;
            margin: 5% auto;
            padding: 20px;
            border-radius: 10px;
            width: 80%;
            max-width: 600px;
            box-shadow: 0 4px 20px rgba(0,0,0,0.3);
            position: relative;
            animation: slideIn 0.3s ease-out;
        }}

        @keyframes slideIn {{
            from {{ transform: translateY(-50px); opacity: 0; }}
            to {{ transform: translateY(0); opacity: 1; }}
        }}

        .close-btn {{
            position: absolute;
            right: 15px;
            top: 15px;
            background: #ef4444;
            color: white;
            border: none;
            border-radius: 5px;
            padding: 8px 15px;
            cursor: pointer;
            font-weight: bold;
            font-size: 18px;
        }}

        .close-btn:hover {{
            background: #dc2626;
        }}
        </style>
        </head>
        <body>

        <div class="excel-table">
        <table>
        <thead>
        <tr>
            <th rowspan="2">Department</th>
            <th rowspan="2">Goal Title</th>
            <th rowspan="2">KPI</th>
            <th rowspan="2">Monthly Target</th>
            <th rowspan="2">Start Date</th>
            <th rowspan="2">End Date</th>
            <th colspan="4">Weekly Target</th>
            <th colspan="4">Weekly Achievement</th>
            <th rowspan="2">Monthly Achievement</th>
            <th colspan="4">Rating</th>
        </tr>
        <tr>
            <th>Week 1</th><th>Week 2</th><th>Week 3</th><th>Week 4</th>
            <th>Week 1</th><th>Week 2</th><th>Week 3</th><th>Week 4</th>
            <th>Week 1</th><th>Week 2</th><th>Week 3</th><th>Week 4</th>
        </tr>
        </thead>
        <tbody>
        '''

        # Add data rows
        for goal_idx, (goal, (_, row)) in enumerate(zip(goals, df.iterrows())):
            goal_desc = goal.get('goal_description', 'No description available').replace("'", "\\'").replace('"', '\\"').replace('\n', '<br>')
            
            complete_html += f'''
        <tr>
            <td>{row["Department"]}</td>
            <td>
                <button class="goal-title-btn" onclick="openModal({goal_idx})">
                    {row["Goal Title"]}
                </button>
            </td>
            <td>{row["KPI"]}</td>
            <td>{row["Monthly Target"]}</td>
            <td>{row["Start Date"]}</td>
            <td>{row["End Date"]}</td>
        '''
            
            # Weekly targets
            for week in range(1, 5):
                complete_html += f'<td class="target-col">{row[f"Week {week} Target"]}</td>'
        
            # ✅ Weekly achievements - check for '-' string
            for week in range(1, 5):
                achievement = row[f"Week {week} Achievement"]
                # Check if it's the '-' string
                if str(achievement) == '-':
                    complete_html += f'<td class="achievement-col" style="color: #94a3b8; font-style: italic;">-</td>'
                else:
                    target = row[f"Week {week} Target"]
                    try:
                        achievement_num = float(achievement)
                        progress = (achievement_num / target * 100) if target > 0 else 0
                        color = '#4CAF50' if progress >= 100 else '#FFC107' if progress >= 60 else '#F44336'
                        complete_html += f'<td class="achievement-col" style="color: {color}; font-weight: bold;">{achievement}</td>'
                    except (ValueError, TypeError):
                        # If conversion fails, show as '-'
                        complete_html += f'<td class="achievement-col" style="color: #94a3b8; font-style: italic;">-</td>'

            # ✅ Monthly achievement - check for '-' string
            monthly_achievement = row["Monthly Achievement"]
            if str(monthly_achievement) == '-':
                complete_html += f'<td class="monthly-achievement-col" style="color: #94a3b8; font-weight: bold; font-style: italic;">-</td>'
            else:
                monthly_target = row["Monthly Target"]
                try:
                    monthly_ach_num = float(monthly_achievement)
                    monthly_progress = (monthly_ach_num / monthly_target * 100) if monthly_target > 0 else 0
                    monthly_color = '#4CAF50' if monthly_progress >= 100 else '#FFC107' if monthly_progress >= 60 else '#F44336'
                    complete_html += f'<td class="monthly-achievement-col" style="color: {monthly_color}; font-weight: bold;">{monthly_achievement}</td>'
                except (ValueError, TypeError):
                    complete_html += f'<td class="monthly-achievement-col" style="color: #94a3b8; font-weight: bold; font-style: italic;">-</td>'
            
            # Weekly ratings (unchanged)
            for week in range(1, 5):
                rating_value = goal.get(f'week{week}_rating', 0)
                
                if rating_value == 1:
                    bg_color = '#FFCCCB'
                elif rating_value == 2:
                    bg_color = '#FFD580'
                elif rating_value == 3:
                    bg_color = '#FFFFE0'
                elif rating_value == 4:
                    bg_color = '#90EE90'
                else:
                    bg_color = '#ffffff'
                
                complete_html += f'<td class="remarks-col" style="background-color: {bg_color}; font-weight: bold;">{rating_value if rating_value else "-"}</td>'
            
            complete_html += '</tr>'
            
            # Add modal
            complete_html += f'''
        <div id="modal{goal_idx}" class="modal">
            <div class="modal-content">
                <button class="close-btn" onclick="closeModal({goal_idx})">✕</button>
                <h3 style="margin: 0 0 15px 0; color: #1E3A8A; padding-right: 40px;">Goal Description</h3>
                <div style="color: #333; line-height: 1.6; max-height: 400px; overflow-y: auto; 
                    padding: 15px; background: #f9fafb; border-radius: 5px;">
                    {goal_desc}
                </div>
            </div>
        </div>
        '''

        complete_html += '''
        </tbody>
        </table>
        </div>

        <script>
        function openModal(index) {
            document.getElementById('modal' + index).style.display = 'block';
        }

        function closeModal(index) {
            document.getElementById('modal' + index).style.display = 'none';
        }

        window.onclick = function(event) {
            const modals = document.querySelectorAll('.modal');
            modals.forEach(modal => {
                if (event.target == modal) {
                    modal.style.display = 'none';
                }
            });
        }

        document.addEventListener('keydown', function(event) {
            if (event.key === 'Escape') {
                const modals = document.querySelectorAll('.modal');
                modals.forEach(modal => {
                    modal.style.display = 'none';
                });
            }
        });
        </script>

        </body>
        </html>
        '''

        # Render using components
        import streamlit.components.v1 as components
        # Calculate dynamic height based on number of goals (tight fit)
        dynamic_height = 150 + (len(goals) * 50)  # Tight calculation: header + rows
        components.html(complete_html, height=dynamic_height, scrolling=False)
        
       # Calculate summary metrics
        # ✅ FIX: Handle None values properly using safe_float helper
        # Calculate summary metrics (find this section around line 3700)
        # ✅ FIX: Handle '-' string in calculations
        total_target = sum(safe_float(g.get('monthly_target'), 0) for g in goals)

        # Calculate achievement - skip '-' values
        total_achievement = 0
        for g in goals:
            ach = g.get('monthly_achievement')
            if ach is not None and str(ach) != '-':
                total_achievement += safe_float(ach, 0)

        avg_progress = calculate_progress(total_achievement, total_target)
        # Display metrics
        st.markdown("---")
        col1, col2, col3 = st.columns(3)
        with col1:
            render_metric_card("Total Target", f"{total_target:.2f}")
        with col2:
            render_metric_card("Total Achievement", f"{total_achievement:.2f}")
        with col3:
            render_metric_card("Progress", f"{avg_progress:.1f}%")
        
        st.markdown("---")
        
        
       # Action tabs: Update, Edit, Delete
        # ✅ NEW: Everyone can update achievements for their own goals
        # Only viewing_employee_year with read_only restrictions apply to edit/delete
        if st.session_state.get('viewing_employee_year') and is_read_only:
            # Read-only view - no tabs
            action_tab1, action_tab2, action_tab3 = None, None, None
        elif st.session_state.get('viewing_employee_year'):
            # Viewing employee (with permissions) - all tabs
            action_tab1, action_tab2, action_tab3 = st.tabs([
                "📝 Update Achievements",
                "✏️ Edit Goal",
                "🗑️ Delete Goal"
            ])
        elif user['role'] in ['HR', 'Manager']:
            # HR/Manager viewing own goals - all tabs
            action_tab1, action_tab2, action_tab3 = st.tabs([
                "📝 Update Achievements",
                "✏️ Edit Goal",
                "🗑️ Delete Goal"
            ])
        else:
            # ✅ CHANGE: Regular employees can now update achievements, but not edit/delete
            action_tab1 = st.tabs(["📝 Update Achievements"])[0]
            action_tab2, action_tab3 = None, None

        # ===== UPDATE ACHIEVEMENTS TAB =====
        
        if action_tab1:
                st.subheader("Update Weekly Achievements")
                
                selected_goal_title = st.selectbox(
                    "Select Goal to Update", 
                    [g['goal_title'] for g in goals],
                    key="update_goal_select"
                )
                selected_goal = next(g for g in goals if g['goal_title'] == selected_goal_title)
                
                # Show current achievement status
                current_monthly = selected_goal.get('monthly_achievement')
                if current_monthly is None:
                    st.info("💡 This goal has no achievements recorded yet. Enter values below to start tracking.")
                
                # ===== SIMPLE VERSION =====
                col1, col2 = st.columns(2)
                
                with col1:
                    current_w1 = selected_goal.get('week1_achievement')
                    w1_input = st.text_input(
                        "Week 1 Achievement (type number or 'none')", 
                        value=str(current_w1) if current_w1 is not None else "none",
                        key="w1_update",
                        help="Enter a number or type 'none' to clear"
                    )
                    
                    current_w2 = selected_goal.get('week2_achievement')
                    w2_input = st.text_input(
                        "Week 2 Achievement (type number or 'none')", 
                        value=str(current_w2) if current_w2 is not None else "none",
                        key="w2_update",
                        help="Enter a number or type 'none' to clear"
                    )
                
                with col2:
                    current_w3 = selected_goal.get('week3_achievement')
                    w3_input = st.text_input(
                        "Week 3 Achievement (type number or 'none')", 
                        value=str(current_w3) if current_w3 is not None else "none",
                        key="w3_update",
                        help="Enter a number or type 'none' to clear"
                    )
                    
                    current_w4 = selected_goal.get('week4_achievement')
                    w4_input = st.text_input(
                        "Week 4 Achievement (type number or 'none')", 
                        value=str(current_w4) if current_w4 is not None else "none",
                        key="w4_update",
                        help="Enter a number or type 'none' to clear"
                    )
                
                # Parse inputs
                def parse_achievement(value_str):
                    """Parse achievement input - returns None or float"""
                    value_str = value_str.strip().lower()
                    if value_str in ['none', 'null', '', '-']:
                        return None
                    try:
                        return float(value_str)
                    except ValueError:
                        return None
                
                w1 = parse_achievement(w1_input)
                w2 = parse_achievement(w2_input)
                w3 = parse_achievement(w3_input)
                w4 = parse_achievement(w4_input)
                
                # Calculate total
                w1_calc = 0 if w1 is None else w1
                w2_calc = 0 if w2 is None else w2
                w3_calc = 0 if w3 is None else w3
                w4_calc = 0 if w4 is None else w4
                total_monthly = w1_calc + w2_calc + w3_calc + w4_calc
                
                st.markdown("---")
                st.markdown(f"**📊 Monthly Achievement (Auto-calculated):** `{total_monthly:.2f}`")
                st.progress(min(total_monthly / selected_goal.get('monthly_target', 1), 1.0))
                
                st.markdown("---")
                st.markdown("**Weekly Ratings (1-4)**")
                col_r1, col_r2 = st.columns(2)
                with col_r1:
                    w1_rating_value = selected_goal.get('week1_rating') or 0
                    w1_rating = st.selectbox("Week 1 Rating", [0, 1, 2, 3, 4], 
                                            index=int(w1_rating_value) if w1_rating_value in [0, 1, 2, 3, 4] else 0, 
                                            key="w1_rating_update",
                                            help="1=Poor (Red), 2=Fair (Amber), 3=Good (Yellow), 4=Excellent (Green)")
                    w2_rating_value = selected_goal.get('week2_rating') or 0
                    w2_rating = st.selectbox("Week 2 Rating", [0, 1, 2, 3, 4], 
                                            index=int(w2_rating_value) if w2_rating_value in [0, 1, 2, 3, 4] else 0, 
                                            key="w2_rating_update",
                                            help="1=Poor (Red), 2=Fair (Amber), 3=Good (Yellow), 4=Excellent (Green)")
                with col_r2:
                    w3_rating_value = selected_goal.get('week3_rating') or 0
                    w3_rating = st.selectbox("Week 3 Rating", [0, 1, 2, 3, 4], 
                                            index=int(w3_rating_value) if w3_rating_value in [0, 1, 2, 3, 4] else 0, 
                                            key="w3_rating_update",
                                            help="1=Poor (Red), 2=Fair (Amber), 3=Good (Yellow), 4=Excellent (Green)")
                    w4_rating_value = selected_goal.get('week4_rating') or 0
                    w4_rating = st.selectbox("Week 4 Rating", [0, 1, 2, 3, 4], 
                                            index=int(w4_rating_value) if w4_rating_value in [0, 1, 2, 3, 4] else 0, 
                                            key="w4_rating_update",
                                            help="1=Poor (Red), 2=Fair (Amber), 3=Good (Yellow), 4=Excellent (Green)")
                

                if st.button("💾 Save Achievements", use_container_width=True, key="save_achievements"):
                    # Determine which week was updated FIRST (before any DB operations)
                    updated_week = 0
                    if w4 is not None:
                        updated_week = 4
                    elif w3 is not None:
                        updated_week = 3
                    elif w2 is not None:
                        updated_week = 2
                    elif w1 is not None:
                        updated_week = 1
                    
                    if user['role'] == 'Employee':
                        updates = {
                            'week1_achievement_pending': w1,
                            'week2_achievement_pending': w2,
                            'week3_achievement_pending': w3,
                            'week4_achievement_pending': w4,
                            'monthly_achievement_pending': total_monthly if total_monthly > 0 else None,
                            'week1_rating_pending': w1_rating,
                            'week2_rating_pending': w2_rating,
                            'week3_rating_pending': w3_rating,
                            'week4_rating_pending': w4_rating,
                            'achievement_approval_status': 'pending',
                            'achievement_updated_at': datetime.now(IST).isoformat()
                        }
                        
                        if db.update_goal(selected_goal['goal_id'], updates):
                            st.success("✅ Achievements submitted for manager approval!")
                            st.info("💡 Your manager will review these achievements before they appear in the goal sheet")
                            
                            # Send email to manager
                            if user.get('manager_id'):
                                manager = db.get_user_by_id(user['manager_id'])
                                if manager and manager.get('email'):
                                    send_achievement_approval_email(
                                        manager['email'],
                                        user['name'],
                                        selected_goal,
                                        {
                                            'week1': w1,
                                            'week2': w2,
                                            'week3': w3,
                                            'week4': w4,
                                            'monthly': total_monthly
                                        }
                                    )
                                    st.success("📧 Approval request sent to your manager")
                    else:
                        updates = {
                            'week1_achievement': w1,
                            'week2_achievement': w2,
                            'week3_achievement': w3,
                            'week4_achievement': w4,
                            'monthly_achievement': total_monthly if total_monthly > 0 else None,
                            'week1_rating': w1_rating,
                            'week2_rating': w2_rating,
                            'week3_rating': w3_rating,
                            'week4_rating': w4_rating
                        }
                        
                        if db.update_goal(selected_goal['goal_id'], updates):
                            st.success("✅ Achievements and ratings saved to monthly goal sheet!")
                            if check_and_auto_complete_goal(selected_goal['goal_id']):
                                st.success("🎉 Goal completed - Monthly target achieved!")
                            st.info("💡 Ratings are now available in the respective week views")
                            
                            # Send notification if any week was updated
                            if updated_week > 0:
                                notify_weekly_achievement_updated(selected_goal, user, updated_week)
                            
                            st.rerun()
                        else:
                            st.error("❌ Failed to save achievements")
        # ===== EDIT GOAL TAB =====
        if not is_read_only and action_tab2:
            if action_tab2:
                with action_tab2:
                    st.subheader("Edit Goal Details")
                    
                    edit_goal_title = st.selectbox(
                        "Select Goal to Edit", 
                        [g['goal_title'] for g in goals],
                        key="edit_goal_select"
                    )
                    edit_goal = next(g for g in goals if g['goal_title'] == edit_goal_title)
                    
                    with st.form("edit_goal_form"):
                        st.markdown("**Basic Information**")
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            new_department = st.text_input("Department*", value=edit_goal.get('department', ''))
                            new_title = st.text_input("Goal Title*", value=edit_goal['goal_title'])
                            new_kpi = st.text_input("KPI*", value=edit_goal.get('kpi', ''))
                        
                        with col2:
                            from datetime import datetime as dt
                            start_date_str = edit_goal.get('start_date', str(date.today()))
                            end_date_str = edit_goal.get('end_date', str(date.today()))
                            
                            if isinstance(start_date_str, str):
                                start_date_val = dt.strptime(start_date_str, '%Y-%m-%d').date()
                            else:
                                start_date_val = start_date_str
                            
                            if isinstance(end_date_str, str):
                                end_date_val = dt.strptime(end_date_str, '%Y-%m-%d').date()
                            else:
                                end_date_val = end_date_str
                            
                            new_start_date = st.date_input("Start Date", value=start_date_val)
                            new_end_date = st.date_input("End Date", value=end_date_val)
                            new_status = st.selectbox(
                                "Status", 
                                ['Active', 'Completed', 'On Hold', 'Cancelled'],
                                index=['Active', 'Completed', 'On Hold', 'Cancelled'].index(edit_goal.get('status', 'Active'))
                            )
                        
                        new_description = st.text_area("Description", value=edit_goal.get('goal_description', ''))
                        
                        st.markdown("**Targets**")
                        col3, col4 = st.columns(2)
                        
                        with col3:
                            new_monthly_target = st.number_input(
                                "Monthly Target*", 
                                min_value=0.0, 
                                value=float(edit_goal.get('monthly_target', 0))
                            )
                        
                        st.markdown("**Weekly Targets**")
                        col5, col6, col7, col8 = st.columns(4)
                        
                        with col5:
                            new_w1_target = st.number_input(
                                "Week 1 Target", 
                                min_value=0.0, 
                                value=float(edit_goal.get('week1_target', 0)),
                                key="edit_w1_target"
                            )
                        with col6:
                            new_w2_target = st.number_input(
                                "Week 2 Target", 
                                min_value=0.0, 
                                value=float(edit_goal.get('week2_target', 0)),
                                key="edit_w2_target"
                            )
                        with col7:
                            new_w3_target = st.number_input(
                                "Week 3 Target", 
                                min_value=0.0, 
                                value=float(edit_goal.get('week3_target', 0)),
                                key="edit_w3_target"
                            )
                        with col8:
                            new_w4_target = st.number_input(
                                "Week 4 Target", 
                                min_value=0.0, 
                                value=float(edit_goal.get('week4_target', 0)),
                                key="edit_w4_target"
                            )
                        
                        st.markdown("**Weekly Remarks**")
                        col_r5, col_r6, col_r7, col_r8 = st.columns(4)

                        with col_r5:
                            new_w1_remarks = st.text_area(
                                "Week 1 Remarks",
                                value=edit_goal.get('week1_remarks', ''),
                                key="edit_w1_remarks",
                                height=80
                            )
                        with col_r6:
                            new_w2_remarks = st.text_area(
                                "Week 2 Remarks",
                                value=edit_goal.get('week2_remarks', ''),
                                key="edit_w2_remarks",
                                height=80
                            )
                        with col_r7:
                            new_w3_remarks = st.text_area(
                                "Week 3 Remarks",
                                value=edit_goal.get('week3_remarks', ''),
                                key="edit_w3_remarks",
                                height=80
                            )
                        with col_r8:
                            new_w4_remarks = st.text_area(
                                "Week 4 Remarks",
                                value=edit_goal.get('week4_remarks', ''),
                                key="edit_w4_remarks",
                                height=80
                            )
                        
                        submitted = st.form_submit_button("💾 Save Changes", use_container_width=True)
                        
                        if submitted:
                            if new_department and new_title and new_kpi:
                                updates = {
                                    'department': new_department,
                                    'goal_title': new_title,
                                    'goal_description': new_description,
                                    'kpi': new_kpi,
                                    'monthly_target': new_monthly_target,
                                    'week1_target': new_w1_target,
                                    'week2_target': new_w2_target,
                                    'week3_target': new_w3_target,
                                    'week4_target': new_w4_target,
                                    'week1_remarks': new_w1_remarks,
                                    'week2_remarks': new_w2_remarks,
                                    'week3_remarks': new_w3_remarks,
                                    'week4_remarks': new_w4_remarks,
                                    'start_date': str(new_start_date),
                                    'end_date': str(new_end_date),
                                    'status': new_status
                                }
                                
                                if db.update_goal(edit_goal['goal_id'], updates):
                                    st.success("✅ Goal updated successfully!")
                                    goal_owner = db.get_user_by_id(edit_goal['user_id'])
                                    if goal_owner:
                                        notify_goal_edited(edit_goal, user, goal_owner)
                                    st.rerun()
                                else:
                                    st.error("❌ Failed to update goal")
                            else:
                                st.error("❌ Please fill all required fields (Department, Title, KPI)")
        
        # ===== DELETE GOAL TAB =====
        if not is_read_only and action_tab3:
            if action_tab3:
                with action_tab3:
                    st.subheader("⚠️ Delete Goal")
                    st.warning("**Warning:** Deleting a goal will also delete all associated feedback. This action cannot be undone!")
                    
                    delete_goal_title = st.selectbox(
                        "Select Goal to Delete", 
                        [g['goal_title'] for g in goals],
                        key="delete_goal_select"
                    )
                    delete_goal = next(g for g in goals if g['goal_title'] == delete_goal_title)
                    
                    st.markdown("**Goal Details:**")
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.info(f"**Vertical:** {delete_goal.get('vertical', 'N/A')}")
                    with col2:
                        st.info(f"**KPI:** {delete_goal.get('kpi', 'N/A')}")
                    with col3:
                        st.info(f"**Target:** {delete_goal.get('monthly_target', 0)}")
                    
                    st.markdown(f"**Description:** {delete_goal.get('goal_description', 'No description')}")
                    
                    st.markdown("---")
                    confirm_delete = st.checkbox("I understand this action cannot be undone", key="confirm_delete")
                    
                    col_del1, col_del2, col_del3 = st.columns([1, 1, 1])
                    
                    with col_del2:
                        if st.button(
                            "🗑️ Delete Goal", 
                            disabled=not confirm_delete,
                            use_container_width=True,
                            type="primary"
                        ):
                            if db.delete_goal(delete_goal['goal_id']):
                                st.success("✅ Goal deleted successfully!")
                                goal_owner = db.get_user_by_id(delete_goal['user_id'])
                                if goal_owner:
                                    notify_goal_deleted(delete_goal, user, goal_owner)
                                st.rerun()
                            else:
                                st.error("❌ Failed to delete goal")
        
            
                        # Add new goal (only for own goals)
                    if not st.session_state.get('viewing_employee_year')and not goals:
                        st.markdown("---")
                        display_add_goal_form(user, year, quarter, month)
        
        # Feedback section
    if goals:
        st.markdown("---")
        display_feedback_section(goals, 'month')

    
def display_assign_goal_form_monthly(user, year, quarter, month):
    """Form to assign goals in monthly view"""
    with st.form("assign_goal_monthly"):
        st.markdown(f"**Assigning for:** {get_month_name(month)} {year}, Q{quarter}")
        
        # Get employees based on role
        if user['role'] == 'HR':
            # HR can only assign to their own team members
            employees = db.get_team_members(user['id'])
        elif user['role'] == 'VP':
            # VP can only assign to HR and Manager
            all_users = db.get_all_users()
            employees = [u for u in all_users if u['role'] in ['HR', 'Manager']]
        elif user['role'] == 'CMD':
            # CMD can only assign to VP
            all_users = db.get_all_users()
            employees = [u for u in all_users if u['role'] == 'VP']
        else:  # Manager
            employees = db.get_team_members(user['id'])
        
        if not employees:
            st.info("No employees available")
            return
        
        selected_emp = st.selectbox(
            "Assign To*",
            [f"{e['name']} ({e['email']})" for e in employees]
        )
        
        emp_email = selected_emp.split('(')[1].strip(')')
        emp_id = next(e['id'] for e in employees if e['email'] == emp_email)
        
        col1, col2 = st.columns(2)
        with col1:
            department = st.text_input("Department*")
            title = st.text_input("Goal Title*")
            kpi = st.text_input("KPI*")
        
        with col2:
            description = st.text_area("Description")
            monthly_target = st.number_input("Monthly Target*", min_value=0.0)
        
        st.markdown("**Weekly Targets**")
        auto_divide = st.checkbox("Auto-divide monthly target equally into 4 weeks", value=True, key="monthly_assign_auto")
        
        if auto_divide and monthly_target > 0:
            weekly_target = monthly_target / 4
        else:
            weekly_target = 0.0
        
        col3, col4, col5, col6 = st.columns(4)
        with col3:
            w1_t = st.number_input("Week 1", min_value=0.0, value=weekly_target, key="monthly_assign_w1")
        with col4:
            w2_t = st.number_input("Week 2", min_value=0.0, value=weekly_target, key="monthly_assign_w2")
        with col5:
            w3_t = st.number_input("Week 3", min_value=0.0, value=weekly_target, key="monthly_assign_w3")
        with col6:
            w4_t = st.number_input("Week 4", min_value=0.0, value=weekly_target, key="monthly_assign_w4")
        
        # ========== ADD THIS ENTIRE SECTION ==========
        st.markdown("**Weekly Ratings (Optional)**")
        col_r1, col_r2, col_r3, col_r4 = st.columns(4)
        with col_r1:
            w1_rating = st.selectbox("Week 1 Rating", [0, 1, 2, 3, 4], key="quick_w1_rating")
        with col_r2:
            w2_rating = st.selectbox("Week 2 Rating", [0, 1, 2, 3, 4], key="quick_w2_rating")
        with col_r3:
            w3_rating = st.selectbox("Week 3 Rating", [0, 1, 2, 3, 4], key="quick_w3_rating")
        with col_r4:
            w4_rating = st.selectbox("Week 4 Rating", [0, 1, 2, 3, 4], key="quick_w4_rating")
        # ========== END OF ADDED SECTION ==========

        if st.form_submit_button("✅ Assign Goal", use_container_width=True):
            if department and title and kpi:
                goal_data = {
                    'user_id': emp_id,
                    'year': year,
                    'quarter': quarter,
                    'month': month,
                    'department': department,
                    'goal_title': title,
                    'goal_description': description,
                    'kpi': kpi,
                    'monthly_target': monthly_target,
                    'week1_target': w1_t,
                    'week2_target': w2_t,
                    'week3_target': w3_t,
                    'week4_target': w4_t,
                    'week1_remarks': w1_rating,  # ✅ ADD
                    'week2_remarks': w2_rating,  # ✅ ADD
                    'week3_remarks': w3_rating,  # ✅ ADD
                    'week4_remarks': w4_rating,
                    'start_date': f'{year}-{month:02d}-01',
                    'end_date': f'{year}-{month:02d}-28',
                    'created_by': user['id']
                }
                
                if db.create_goal(goal_data):
                    st.success(f"✅ Goal assigned to {selected_emp}")
    
                    # Notify about goal creation
                    assigned_emp = db.get_user_by_id(emp_id)
                    if assigned_emp:
                        notify_goal_created(goal_data, user)
                    st.rerun()
            else:
                st.error("❌ Please fill all required fields")


# ============================================
# ENHANCED FEEDBACK HISTORY (WITH USER INFO)
# ============================================
def display_feedback_history():
    """Display feedback history with reply and new feedback options"""
    user = st.session_state.user

    if not user:
        st.warning("⚠️ Session expired. Please login again.")
        st.rerun()

    role = user['role']
    
    st.title("💬 Feedback History")
    
    # Add New Feedback Button
    if st.button("➕ Add New Feedback", use_container_width=True):
        st.session_state.adding_new_feedback = True
    
    # New Feedback Modal
    if st.session_state.get('adding_new_feedback'):
        st.markdown("---")
        st.subheader("➕ Add New Feedback")
        
        with st.form("add_new_feedback_form"):
            # Get all goals based on role
            if role == 'HR':
                all_users = db.get_all_users()
                all_goals = []
                for u in all_users:
                    user_goals = db.get_user_all_goals(u['id'])
                    for g in user_goals:
                        g['user_name'] = u['name']
                    all_goals.extend(user_goals)
            elif role == 'Manager':
                team_members = db.get_team_members(user['id'])
                all_goals = []
                for m in team_members:
                    member_goals = db.get_user_all_goals(m['id'])
                    for g in member_goals:
                        g['user_name'] = m['name']
                    all_goals.extend(member_goals)
                # Add own goals
                own_goals = db.get_user_all_goals(user['id'])
                for g in own_goals:
                    g['user_name'] = user['name']
                all_goals.extend(own_goals)
            else:
                all_goals = db.get_user_all_goals(user['id'])
                for g in all_goals:
                    g['user_name'] = user['name']
            
            if all_goals:
                selected_goal_str = st.selectbox(
                    "Select Goal*",
                    [f"{g.get('user_name', 'Unknown')} - {g['goal_title']} ({g['year']}/Q{g.get('quarter', 'N/A')}/M{g.get('month', 'N/A')})" for g in all_goals]
                )
                
                selected_goal = all_goals[[f"{g.get('user_name', 'Unknown')} - {g['goal_title']} ({g['year']}/Q{g.get('quarter', 'N/A')}/M{g.get('month', 'N/A')})" for g in all_goals].index(selected_goal_str)]
                
                # Determine feedback type
                if role == 'HR':
                    fb_types = ["HR", "Manager"] if selected_goal['user_id'] != user['id'] else ["Self Appraisal", "HR"]
                elif role == 'Manager':
                    fb_types = ["Manager"] if selected_goal['user_id'] != user['id'] else ["Self Appraisal", "Manager"]
                else:
                    fb_types = ["Self Appraisal"]
                
                fb_type = st.selectbox("Feedback Type*", fb_types)
                rating = st.slider("Rating", 1, 5, 3)
                comment = st.text_area("Comment*")
                
                col_submit, col_cancel = st.columns(2)
                
                with col_submit:
                    if st.form_submit_button("Submit Feedback", use_container_width=True):
                        if comment.strip():
                            feedback_data = {
                                'goal_id': selected_goal['goal_id'],
                                'user_id': selected_goal['user_id'],
                                'feedback_by': user['id'],
                                'feedback_type': fb_type,
                                'rating': rating,
                                'comment': comment.strip(),
                                'level': 'general'
                            }
                            if db.create_feedback(feedback_data):
                                st.success("✅ Feedback submitted!")
                                del st.session_state.adding_new_feedback
                                st.rerun()
                        else:
                            st.error("❌ Please enter a comment")
                
                with col_cancel:
                    if st.form_submit_button("Cancel", use_container_width=True):
                        del st.session_state.adding_new_feedback
                        st.rerun()
            else:
                st.info("No goals available for feedback")
                if st.form_submit_button("Cancel", use_container_width=True):
                    del st.session_state.adding_new_feedback
                    st.rerun()
        
        st.markdown("---")
    
    # Get feedback based on role
    if role == 'HR':
        all_feedbacks = db.get_all_feedback()
        st.subheader("All Feedback (System-wide)")
    elif role == 'Manager':
        my_feedback = db.get_user_all_feedback(user['id'])
        team_feedback = []
        for member in db.get_team_members(user['id']):
            team_feedback.extend(db.get_user_all_feedback(member['id']))
        all_feedbacks = my_feedback + team_feedback
        st.subheader("My Feedback & Team Feedback")
    else:
        all_feedbacks = db.get_user_all_feedback(user['id'])
        st.subheader("My Feedback")
    
    all_feedbacks.sort(key=lambda x: x.get('created_at', ''), reverse=True)

    if not all_feedbacks:
        st.info("No feedback history found")
        return
    
    # Filters
    col1, col2, col3 = st.columns(3)
    with col1:
        filter_type = st.selectbox("Filter by Type", ["All", "Self Appraisal", "Manager", "HR"])
    with col2:
        filter_rating = st.selectbox("Filter by Rating", ["All", "5", "4", "3", "2", "1"])
    with col3:
        search_goal = st.text_input("Search Goal")
    
    # Filter feedbacks
    filtered = all_feedbacks
    if filter_type != "All":
        filtered = [f for f in filtered if f.get('feedback_type') == filter_type]
    if filter_rating != "All":
        filtered = [f for f in filtered if f.get('rating') == int(filter_rating)]
    if search_goal:
        filtered = [f for f in filtered if search_goal.lower() in f.get('goal_title', '').lower()]
    
    # Display feedback
    st.markdown(f"**Showing {len(filtered)} of {len(all_feedbacks)} feedback entries**")
    
    for feedback in filtered:
        # Get user who received feedback
        feedback_user = db.get_user_by_id(feedback.get('user_id'))
        feedback_user_name = feedback_user['name'] if feedback_user else 'Unknown'
        
        is_new = False
        created_at_str = feedback.get('created_at', '')
        if created_at_str:
            try:
                utc_time = datetime.strptime(created_at_str[:19], '%Y-%m-%dT%H:%M:%S')
                utc_time = pytz.utc.localize(utc_time)
                ist_time = utc_time.astimezone(IST)
                hours_ago = (datetime.now(IST) - ist_time).total_seconds() / 3600
                is_new = hours_ago <= 24
            except:
                pass

        with st.container():
            # Header with user info
            st.markdown(f"### 💬 Feedback for: **{feedback_user_name}**")
            
            col_fb1, col_fb2 = st.columns([3, 1])
            
            with col_fb1:
                st.markdown(f"**Goal:** {feedback.get('goal_title', 'N/A')}")
                st.markdown(f"**Type:** {feedback.get('feedback_type')} | **Rating:** {'⭐' * feedback.get('rating', 0)}")
                st.markdown(f"**Comment:** {feedback.get('comment')}")
                
                # Reply Section
                replies = db.get_feedback_replies(feedback.get('feedback_id'))
                if replies:
                    st.markdown("**Replies:**")
                    for reply in replies:
                        st.info(f"↳ **{reply.get('reply_by_name', 'Unknown')}:** {reply.get('reply_text')} ({reply.get('created_at', '')[:16]})")
                
                # Add Reply Button
                if st.button("💬 Reply", key=f"reply_btn_{feedback.get('feedback_id')}"):
                    st.session_state[f"replying_to_{feedback.get('feedback_id')}"] = True
                
                # Reply Form
                if st.session_state.get(f"replying_to_{feedback.get('feedback_id')}"):
                    with st.form(f"reply_form_{feedback.get('feedback_id')}"):
                        reply_text = st.text_area("Your Reply*", key=f"reply_text_{feedback.get('feedback_id')}")
                        
                        col_reply1, col_reply2 = st.columns(2)
                        with col_reply1:
                            if st.form_submit_button("Send Reply"):
                                if reply_text.strip():
                                    reply_data = {
                                        'feedback_id': feedback.get('feedback_id'),
                                        'reply_by': user['id'],
                                        'reply_text': reply_text.strip()
                                    }
                                    if db.create_feedback_reply(reply_data):
                                        # Notify the original feedback giver
                                        feedback_giver = db.get_user_by_id(feedback.get('feedback_by'))
                                        if feedback_giver:
                                            notify_feedback_reply(feedback, user, feedback_giver)
                                        
                                        st.success("✅ Reply added!")
                                        del st.session_state[f"replying_to_{feedback.get('feedback_id')}"]
                                        st.rerun()
                                else:
                                    st.error("❌ Please enter a reply")
                        
                        with col_reply2:
                            if st.form_submit_button("Cancel"):
                                del st.session_state[f"replying_to_{feedback.get('feedback_id')}"]
                                st.rerun()
            
            with col_fb2:
                st.markdown(f"**Given By:** {feedback.get('feedback_by_name', 'Unknown')}")
                
                # Convert UTC to IST
                created_at_str = feedback.get('created_at', '')
                if created_at_str:
                    try:
                        utc_time = datetime.strptime(created_at_str[:19], '%Y-%m-%dT%H:%M:%S')
                        utc_time = pytz.utc.localize(utc_time)
                        ist_time = utc_time.astimezone(IST)
                        st.markdown(f"**Date:** {ist_time.strftime('%Y-%m-%d')}")
                        st.markdown(f"**Time:** {ist_time.strftime('%I:%M %p IST')}")
                    except:
                        st.markdown(f"**Date:** {feedback.get('date', 'N/A')}")
                else:
                    st.markdown(f"**Date:** {feedback.get('date', 'N/A')}")
                
                # Delete button for HR
                if role == 'HR':
                    if st.button("🗑️ Delete", key=f"del_fb_{feedback.get('feedback_id')}", use_container_width=True):
                        if db.delete_feedback(feedback.get('feedback_id')):
                            st.success("✅ Feedback deleted!")
                            st.rerun()
            
            st.markdown("---")




# ============================================
# WEEK VIEW (Keep existing)
# ============================================
def display_week_view(user, year, quarter, month, week_num):
    """Display week-specific view with remarks from monthly goals"""
    if not year or not quarter or not month or not week_num:
        st.warning("⚠️ Invalid week view parameters. Returning to dashboard...")
        st.session_state.page = 'dashboard'
        st.rerun()
    
    month_name = get_month_name(month)
    st.subheader(f"📅 Week {week_num} - {month_name} {year}")
    
    # Get monthly goals to show breakdown
    monthly_goals = db.get_month_goals(user['id'], year, quarter, month)
    
    if monthly_goals:
        st.markdown("**Weekly Breakdown from Monthly Goals**")
        
        table_data = []
        for goal in monthly_goals:
            week_target = goal.get(f'week{week_num}_target', 0)
            week_achievement_raw = goal.get(f'week{week_num}_achievement')
            week_remarks_raw = goal.get(f'week{week_num}_remarks')
            
            # ✅ FIX: Separate display values from calculation values
            week_achievement_display = '-' if week_achievement_raw is None else str(week_achievement_raw)
            week_achievement_numeric = 0 if week_achievement_raw is None else float(week_achievement_raw)
            week_remarks = '-' if not week_remarks_raw else str(week_remarks_raw)
            
            # ✅ FIX: Use numeric value for progress calculation
            progress = calculate_progress(week_achievement_numeric, week_target)
            
            table_data.append({
                'Department': str(goal.get('department', '')),
                'Title': str(goal['goal_title']),
                'KPI': str(goal.get('kpi', '')),
                'Target': float(week_target),
                'Achievement': week_achievement_display,  # Display string
                'Progress %': f"{progress:.1f}%",
                'Remarks': week_remarks
            })
        
        df = pd.DataFrame(table_data)
        
        # ✅ Ensure proper data types for Arrow serialization
        df['Department'] = df['Department'].astype(str)
        df['Title'] = df['Title'].astype(str)
        df['KPI'] = df['KPI'].astype(str)
        df['Target'] = pd.to_numeric(df['Target'], errors='coerce').fillna(0)
        df['Achievement'] = df['Achievement'].astype(str)
        df['Progress %'] = df['Progress %'].astype(str)
        df['Remarks'] = df['Remarks'].astype(str)
        
        st.dataframe(df, use_container_width=True)
        
        # Show detailed goal cards with remarks
        st.markdown("---")
        st.markdown("** Goal Details**")
        
        for goal in monthly_goals:
            week_target = goal.get(f'week{week_num}_target', 0)
            week_achievement_raw = goal.get(f'week{week_num}_achievement')
            week_remarks = goal.get(f'week{week_num}_remarks', '')
            
            # ✅ FIX: Use 0 instead of None for calculations
            week_achievement_numeric = 0 if week_achievement_raw is None else float(week_achievement_raw)
            progress = calculate_progress(week_achievement_numeric, week_target)
            
            # Display value
            week_achievement_display = '-' if week_achievement_raw is None else week_achievement_raw
            
            with st.expander(f"📌 {goal['goal_title']}"):
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Target", week_target)
                with col2:
                    st.metric("Achievement", week_achievement_display)
                with col3:
                    st.metric("Progress", f"{progress:.1f}%")
                
                render_progress_bar(progress, goal['goal_title'])
                
                st.markdown(f"**Department:** {goal.get('department', 'N/A')}")
                st.markdown(f"**KPI:** {goal.get('kpi', 'N/A')}")
                st.markdown(f"**Description:** {goal.get('goal_description', 'No description')}")
                
                # Show remarks
                st.markdown("---")
                if week_remarks:
                    st.markdown(f"**📝 Week {week_num} Remarks:**")
                    st.info(week_remarks)
                else:
                    st.caption(f"No remarks added for Week {week_num} yet")

    
    # Week-specific goals with management
    st.markdown("---")
    week_goals = db.get_week_goals(user['id'], year, quarter, month, week_num)
    
    if week_goals:
        st.markdown("**Week-Specific Goals**")
        
        # Tabs for managing week goals
        tab1, tab2, tab3 = st.tabs(["📋 View Goals", "✏️ Edit Goal", "🗑️ Delete Goal"])
        
        with tab1:
            for goal in week_goals:
                with st.expander(f"📌 {goal['goal_title']}"):
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Target", goal.get('weekly_target', 0))
                    with col2:
                        st.metric("Achievement", goal.get('weekly_achievement', 0))
                    with col3:
                        progress = calculate_progress(
                            goal.get('weekly_achievement', 0),
                            goal.get('weekly_target', 0)
                        )
                        st.metric("Progress", f"{progress:.1f}%")
                    
                    st.markdown(f"**Department:** {goal.get('department', 'N/A')}")
                    st.markdown(f"**KPI:** {goal.get('kpi', 'N/A')}")
                    st.markdown(f"**Description:** {goal.get('goal_description', 'No description')}")
                    
                    week_rating = goal.get(f'week{week_num}_rating', 0)

                    # Show rating with color
                    st.markdown("---")
                    if week_rating:
                        rating_colors = {1: '🔴', 2: '🟠', 3: '🟡', 4: '🟢'}
                        rating_text = {1: 'Poor', 2: 'Fair', 3: 'Good', 4: 'Excellent'}
                        st.markdown(f"**📊 Week {week_num} Rating:** {rating_colors.get(week_rating, '')} {rating_text.get(week_rating, 'N/A')} ({week_rating}/4)")
                    else:
                        st.caption(f"No rating added for Week {week_num} yet")

        with tab2:
            st.subheader("Edit Week-Specific Goal")
            
            edit_week_goal_title = st.selectbox(
                "Select Goal to Edit",
                [g['goal_title'] for g in week_goals],
                key=f"edit_week_goal_{week_num}"
            )
            edit_week_goal = next(g for g in week_goals if g['goal_title'] == edit_week_goal_title)
            
            with st.form(f"edit_week_goal_form_{week_num}"):
                col1, col2 = st.columns(2)
                with col1:
                    new_department = st.text_input("Department", value=edit_week_goal.get('department', ''))
                    new_title = st.text_input("Goal Title*", value=edit_week_goal['goal_title'])
                with col2:
                    new_kpi = st.text_input("KPI", value=edit_week_goal.get('kpi', ''))
                    new_status = st.selectbox(
                        "Status",
                        ['Active', 'Completed', 'On Hold', 'Cancelled'],
                        index=['Active', 'Completed', 'On Hold', 'Cancelled'].index(edit_week_goal.get('status', 'Active'))
                    )
                
                new_description = st.text_area("Description", value=edit_week_goal.get('goal_description', ''))
                
                col3, col4 = st.columns(2)
                with col3:
                    new_target = st.number_input("Weekly Target", min_value=0.0, value=float(edit_week_goal.get('weekly_target', 0)))
                with col4:
                    new_achievement = st.number_input("Weekly Achievement", min_value=0.0, value=float(edit_week_goal.get('weekly_achievement', 0)))
                
                # Add remarks field for week-specific goals
                new_remarks = st.text_area(f"Week {week_num} Remarks", value=edit_week_goal.get(f'week{week_num}_remarks', ''), height=100)
                
                if st.form_submit_button("💾 Save Changes", use_container_width=True):
                    if new_title:
                        updates = {
                            'department': new_department,
                            'goal_title': new_title,
                            'goal_description': new_description,
                            'kpi': new_kpi,
                            'weekly_target': new_target,
                            'weekly_achievement': new_achievement,
                            'status': new_status,
                            f'week{week_num}_remarks': new_remarks
                        }
                        
                        if db.update_goal(edit_week_goal['goal_id'], updates):
                            goal_owner = db.get_user_by_id(edit_week_goal['user_id'])
                            if goal_owner:
                                notify_goal_edited(edit_week_goal, user, goal_owner)
                            st.success("✅ Week goal updated!")
                            st.rerun()
                    else:
                        st.error("❌ Goal title is required")
        
        with tab3:
            st.subheader("⚠️ Delete Week-Specific Goal")
            st.warning("This action cannot be undone!")
            
            delete_week_goal_title = st.selectbox(
                "Select Goal to Delete",
                [g['goal_title'] for g in week_goals],
                key=f"delete_week_goal_{week_num}"
            )
            delete_week_goal = next(g for g in week_goals if g['goal_title'] == delete_week_goal_title)
            
            st.info(f"**Goal:** {delete_week_goal['goal_title']}")
            st.info(f"**Target:** {delete_week_goal.get('weekly_target', 0)}")
            
            confirm = st.checkbox("I understand this cannot be undone", key=f"confirm_week_del_{week_num}")
            
            if st.button("🗑️ Delete Goal", disabled=not confirm, use_container_width=True):
                goal_owner = db.get_user_by_id(delete_week_goal['user_id'])
                if db.delete_goal(delete_week_goal['goal_id']):
                    if goal_owner:
                        notify_goal_deleted(delete_week_goal, user, goal_owner)
                    st.success("✅ Week goal deleted!")
                    st.rerun()
    
    # Add week-specific goal (only for own goals)
    if not st.session_state.get('viewing_employee_year'):
        with st.expander("➕ Add Week-Specific Goal"):
            with st.form(f"add_week{week_num}_goal"):
                department = st.text_input("Department")
                title = st.text_input("Goal Title*")
                kpi = st.text_input("KPI")
                target = st.number_input("Weekly Target", min_value=0.0)
                description = st.text_area("Description")
                remarks = st.text_area(f"Week {week_num} Remarks", height=100)
                
                if st.form_submit_button("Create Goal"):
                    if title:
                        goal_data = {
                            'user_id': user['id'],
                            'year': year,
                            'quarter': quarter,
                            'month': month,
                            'week': week_num,
                            'department': department,
                            'goal_title': title,
                            'goal_description': description,
                            'kpi': kpi,
                            'weekly_target': target,
                            f'week{week_num}_remarks': remarks,
                            'start_date': str(date.today()),
                            'end_date': str(date.today())
                        }
                        if db.create_goal(goal_data):
                            st.success("✅ Goal created!")
                            st.rerun()
                    else:
                        st.error("❌ Please enter a goal title")
    
    # Week feedback
    st.markdown("---")
    all_week_goals = monthly_goals + week_goals
    if all_week_goals:
        display_feedback_section(all_week_goals, f'week{week_num}')

def display_add_goal_form(user, year, quarter, month):
    """Display form to add new goal"""
    with st.expander("➕ Add New Monthly Goal"):
        with st.form("add_goal"):
            col1, col2 = st.columns(2)
            
            with col1:
                department = st.text_input("Department*")
                title = st.text_input("Goal Title*")
                kpi = st.text_input("KPI*")
                monthly_target = st.number_input("Monthly Target*", min_value=0.0)
            
            with col2:
                start_date = st.date_input("Start Date", value=date.today())
                end_date = st.date_input("End Date", value=date.today())
                description = st.text_area("Description")
            
            st.markdown("**Weekly Targets**")
            auto_divide = st.checkbox("Auto-divide monthly target equally into 4 weeks", value=True)
            
            if auto_divide and monthly_target > 0:
                weekly_target = monthly_target / 4
            else:
                weekly_target = 0.0
            
            col3, col4, col5, col6 = st.columns(4)
            with col3:
                w1_t = st.number_input("Week 1 Target", min_value=0.0, value=weekly_target)
            with col4:
                w2_t = st.number_input("Week 2 Target", min_value=0.0, value=weekly_target)
            with col5:
                w3_t = st.number_input("Week 3 Target", min_value=0.0, value=weekly_target)
            with col6:
                w4_t = st.number_input("Week 4 Target", min_value=0.0, value=weekly_target)

            # ADD THIS NEW SECTION
            st.markdown("**Weekly Ratings (Optional)**")
            col_r1, col_r2, col_r3, col_r4 = st.columns(4)
            with col_r1:
                w1_rating = st.selectbox("Week 1 Rating", [0, 1, 2, 3, 4], key="quick_w1_rating")
            with col_r2:
                w2_rating = st.selectbox("Week 2 Rating", [0, 1, 2, 3, 4], key="quick_w2_rating")
            with col_r3:
                w3_rating = st.selectbox("Week 3 Rating", [0, 1, 2, 3, 4], key="quick_w3_rating")
            with col_r4:
                w4_rating = st.selectbox("Week 4 Rating", [0, 1, 2, 3, 4], key="quick_w4_rating")
                        
            if st.form_submit_button("Create Goal", use_container_width=True):
                if department and title and kpi:
                    goal_data = {
                        'user_id': user['id'],
                        'year': year,
                        'quarter': quarter,
                        'month': month,
                        'department': normalize_department(department),
                        'goal_title': title,
                        'goal_description': description,
                        'kpi': kpi,
                        'monthly_target': monthly_target,
                        'week1_target': w1_t,
                        'week2_target': w2_t,
                        'week3_target': w3_t,
                        'week4_target': w4_t,
                        'week1_remarks': w1_rating,  # ADD
                        'week2_remarks': w2_rating,  # ADD
                        'week3_remarks': w3_rating,  # ADD
                        'week4_remarks': w4_rating,
                        'start_date': str(start_date),
                        'end_date': str(end_date)
                    }
                    
                    is_valid, error_msg = validate_goal_data(goal_data)
                    if is_valid:
                        result, goal_year, goal_quarter, goal_month = create_goal_with_date_based_placement(goal_data)
                        if result:
                            st.success(f"✅ Goal created successfully in {get_month_name(goal_month)} {goal_year}!")
                            notify_goal_created(goal_data, user)
                            st.rerun()
                    else:
                        st.error(f"❌ {error_msg}")
                else:
                    st.error("❌ Please fill all required fields (Department, Title, KPI)")

def display_add_goal_form_inline(user, year, quarter, month):
    """Inline form to add new goal at the top"""
    with st.form("add_goal_inline_top"):
        col1, col2 = st.columns(2)
        
        with col1:
            department = st.text_input("Department*")
            title = st.text_input("Goal Title*")
            kpi = st.text_input("KPI*")
            monthly_target = st.number_input("Monthly Target*", min_value=0.0)
        
        with col2:
            start_date = st.date_input("Start Date", value=date(year, month, 1))
            end_date = st.date_input("End Date", value=date(year, month, calendar.monthrange(year, month)[1]))
            description = st.text_area("Description", height=100)
        
        st.markdown("**Weekly Targets**")
        auto_divide = st.checkbox("Auto-divide monthly target equally into 4 weeks", value=True, key="inline_auto_divide")
        
        if auto_divide and monthly_target > 0:
            weekly_target = monthly_target / 4
        else:
            weekly_target = 0.0
        
        col3, col4, col5, col6 = st.columns(4)
        with col3:
            w1_t = st.number_input("Week 1", min_value=0.0, value=weekly_target, key="inline_top_w1")
        with col4:
            w2_t = st.number_input("Week 2", min_value=0.0, value=weekly_target, key="inline_top_w2")
        with col5:
            w3_t = st.number_input("Week 3", min_value=0.0, value=weekly_target, key="inline_top_w3")
        with col6:
            w4_t = st.number_input("Week 4", min_value=0.0, value=weekly_target, key="inline_top_w4")
        
        st.markdown("**Weekly Ratings (Optional)**")
        col_r1, col_r2, col_r3, col_r4 = st.columns(4)
        with col_r1:
            w1_rating = st.selectbox("Week 1 Rating", [0, 1, 2, 3, 4], key="quick_w1_rating")
        with col_r2:
            w2_rating = st.selectbox("Week 2 Rating", [0, 1, 2, 3, 4], key="quick_w2_rating")
        with col_r3:
            w3_rating = st.selectbox("Week 3 Rating", [0, 1, 2, 3, 4], key="quick_w3_rating")
        with col_r4:
            w4_rating = st.selectbox("Week 4 Rating", [0, 1, 2, 3, 4], key="quick_w4_rating")
        
        col_submit, col_cancel = st.columns(2)
        
        with col_submit:
            if st.form_submit_button("✅ Create Goal", use_container_width=True):
                if department and title and kpi:
                    goal_data = {
                        'user_id': user['id'],
                        'year': year,
                        'quarter': quarter,
                        'month': month,
                        'department': department,
                        'goal_title': title,
                        'goal_description': description,
                        'kpi': kpi,
                        'monthly_target': monthly_target,
                        'week1_target': w1_t,
                        'week2_target': w2_t,
                        'week3_target': w3_t,
                        'week4_target': w4_t,
                        'week1_remarks': w1_rating,
                        'week2_remarks': w2_rating,
                        'week3_remarks': w3_rating,
                        'week4_remarks': w4_rating,
                        'start_date': str(start_date),
                        'end_date': str(end_date),
                        'created_by': user['id']  # Track who created the goal
                    }
                    
                    if db.create_goal(goal_data):
                        if user['role'] == 'Employee' and user.get('manager_id'):
                            manager = db.get_user_by_id(user['manager_id'])
                            if manager and manager.get('email'):
                                send_goal_approval_email(
                                    manager['email'],
                                    user['name'],
                                    goal_data,
                                    goal_data.get('goal_id')  # You'll need to get this from create_goal return
                                )
                                st.success(f"✅ Goal created! Approval request sent to your manager.")
                            else:
                                st.warning(f"✅ Goal created! But manager email not found.")
                        else:
                            st.success(f"✅ Goal created successfully!")
                        
                        notify_goal_created(goal_data, user)
                        st.session_state.show_create_goal_form = False
                        st.rerun()
                    else:
                        st.error("❌ Failed to create goal")
                else:
                    st.error("❌ Please fill all required fields (Department, Title, KPI)")
        
        with col_cancel:
            if st.form_submit_button("❌ Cancel", use_container_width=True):
                st.session_state.show_create_goal_form = False
                st.rerun()

def display_feedback_section(goals, level):
    """Display feedbacks with replies - Official feedback only"""
    if not goals:
        return

    user = st.session_state.user
    target_user_id = goals[0]['user_id']
    target_user = db.get_user_by_id(target_user_id)
    if not target_user:
        st.error("User not found")
        return

    target_role = target_user['role']
    st.subheader("💬 Feedback & Replies")

    # Goal selector
    selected_goal_title = st.selectbox(
        "Select Goal",
        [g['goal_title'] for g in goals],
        key=f"fb_goal_{level}_{target_user_id}"
    )
    selected_goal = next(g for g in goals if g['goal_title'] == selected_goal_title)

    # ──────────────────────────────────────────────
    # 1. Who is allowed to give feedback to this person?
    # ──────────────────────────────────────────────
    feedback_giver_role = get_feedback_giver_role(target_role)   # returns 'CMD', 'VP', 'Manager' or None

    if feedback_giver_role and user['role'] == feedback_giver_role:
        feedback_label = {
            'CMD':     'CMD Feedback',
            'VP':      'VP Feedback',
            'Manager': 'Manager Feedback'
        }[feedback_giver_role]
    else:
        feedback_label = None

    # ──────────────────────────────────────────────
    # 2. Show existing feedback WITH REPLIES
    # ──────────────────────────────────────────────
    feedbacks = db.get_goal_feedback(selected_goal['goal_id'])

    # Ensure feedback_by_name is populated
    for fb in feedbacks:
        if not fb.get('feedback_by_name'):
            feedback_by_id = fb.get('feedback_by')
            if feedback_by_id:
                giver = db.get_user_by_id(feedback_by_id)
                fb['feedback_by_name'] = giver['name'] if giver else 'Unknown'
            else:
                fb['feedback_by_name'] = 'Unknown'

    # Show only official feedbacks
    official_feedbacks = [f for f in feedbacks if f['feedback_type'] in (
        'CMD Feedback', 'VP Feedback', 'Manager Feedback'
    )]

    if official_feedbacks:
        st.markdown("---")
        
        for fb in official_feedbacks:
            label_to_show = fb['feedback_type']
            
            # Get the actual feedback giver's name
            feedback_by_id = fb.get('feedback_by')
            feedback_giver = db.get_user_by_id(feedback_by_id) if feedback_by_id else None
            feedback_giver_name = fb.get('feedback_by_name', 'Unknown')
            
            # Format the created_at date properly
            created_date = fb.get('created_at', '')
            if created_date:
                try:
                    if 'T' in created_date:
                        created_date = created_date.split('T')[0]
                    else:
                        created_date = created_date[:10]
                except:
                    created_date = 'Unknown date'
            
            # Main feedback container
            with st.container():
                st.markdown(f"""
                <div style="
                    background-color: #dbeafe;
                    border-left: 6px solid #3b82f6;
                    padding: 16px;
                    border-radius: 8px;
                    margin: 12px 0;
                    box-shadow: 0 2px 6px rgba(0,0,0,0.1);
                ">
                    <p style="margin:0; font-weight:600; color:#1e40af;">
                        {label_to_show}
                    </p>
                    <p style="margin:8px 0 4px 0; color:#1f2937;">
                        <strong>Rating:</strong> {'★' * fb.get('rating',0)} ({fb.get('rating',0)}/5)
                    </p>
                    <p style="margin:8px 0; color:#374151;">{fb.get('comment','')}</p>
                    <p style="margin:12px 0 0 0; font-size:13px; color:#6b7280;">
                        — {feedback_giver_name} • {created_date}
                    </p>
                </div>
                """, unsafe_allow_html=True)
                
                # ──────────────────────────────────────────────
                # SHOW REPLIES SECTION
                # ──────────────────────────────────────────────
                replies = db.get_feedback_replies(fb.get('feedback_id'))
                
                if replies:
                    st.markdown("**💬 Replies:**")
                    for reply in replies:
                        reply_by_name = reply.get('reply_by_name', 'Unknown')
                        reply_text = reply.get('reply_text', '')
                        reply_date = reply.get('created_at', '')
                        
                        # Format reply date
                        if reply_date:
                            try:
                                if 'T' in reply_date:
                                    reply_date = reply_date.split('T')[0]
                                else:
                                    reply_date = reply_date[:10]
                            except:
                                reply_date = ''
                        
                        st.markdown(f"""
                        <div style="
                            margin-left: 30px;
                            background-color: #f0f9ff;
                            border-left: 3px solid #60a5fa;
                            padding: 12px;
                            border-radius: 5px;
                            margin-top: 8px;
                        ">
                            <p style="margin:0; color:#1f2937; font-size: 14px;">↳ {reply_text}</p>
                            <p style="margin:8px 0 0 0; font-size:12px; color:#6b7280;">
                                — {reply_by_name} • {reply_date}
                            </p>
                        </div>
                        """, unsafe_allow_html=True)
                
                # ──────────────────────────────────────────────
                # REPLY BUTTON & FORM
                # ──────────────────────────────────────────────
                # Anyone can reply (employee can reply to manager's feedback, manager can reply back)
                reply_key = f"reply_btn_{fb.get('feedback_id')}_{level}"
                
                if st.button("💬 Reply to this feedback", key=reply_key):
                    st.session_state[f"replying_to_{fb.get('feedback_id')}_{level}"] = True
                
                # Reply Form
                if st.session_state.get(f"replying_to_{fb.get('feedback_id')}_{level}"):
                    with st.form(f"reply_form_{fb.get('feedback_id')}_{level}"):
                        reply_text = st.text_area(
                            "Your Reply*", 
                            key=f"reply_text_{fb.get('feedback_id')}_{level}",
                            height=100
                        )
                        
                        col_reply1, col_reply2 = st.columns(2)
                        with col_reply1:
                            if st.form_submit_button("📤 Send Reply", use_container_width=True):
                                if reply_text.strip():
                                    reply_data = {
                                        'feedback_id': fb.get('feedback_id'),
                                        'reply_by': user['id'],
                                        'reply_text': reply_text.strip()
                                    }
                                    if db.create_feedback_reply(reply_data):
                                        # Notify the original feedback giver
                                        feedback_giver = db.get_user_by_id(fb.get('feedback_by'))
                                        if feedback_giver:
                                            notify_feedback_reply(fb, user, feedback_giver)
                                        
                                        st.success("✅ Reply added!")
                                        del st.session_state[f"replying_to_{fb.get('feedback_id')}_{level}"]
                                        st.rerun()
                                else:
                                    st.error("❌ Please enter a reply")
                        
                        with col_reply2:
                            if st.form_submit_button("❌ Cancel", use_container_width=True):
                                del st.session_state[f"replying_to_{fb.get('feedback_id')}_{level}"]
                                st.rerun()
                
                st.markdown("---")
    else:
        st.info("💡 No feedback yet for this goal")

    # ──────────────────────────────────────────────
    # 3. Add new feedback form (only if allowed)
    # ──────────────────────────────────────────────
    if feedback_label:   # ← this means the logged-in user IS allowed
        st.markdown("---")
        with st.expander("➕ Add New Feedback", expanded=False):
            with st.form(key=f"fb_form_{selected_goal['goal_id']}_{level}"):
                rating = st.slider("Rating", 1, 5, 3, key=f"rating_{selected_goal['goal_id']}_{level}")
                comment = st.text_area("Comment *", height=120, key=f"comment_{selected_goal['goal_id']}_{level}")

                if st.form_submit_button("📤 Submit Feedback", use_container_width=True):
                    if not comment.strip():
                        st.error("❌ Comment is required")
                    else:
                        feedback_data = {
                            'goal_id': selected_goal['goal_id'],
                            'user_id': target_user_id,
                            'feedback_by': user['id'],
                            'feedback_type': feedback_label,          # ← exactly "CMD Feedback" / "VP Feedback" / "Manager Feedback"
                            'rating': rating,
                            'comment': comment.strip(),
                            'level': level
                        }
                        if db.create_feedback(feedback_data):
                            # Notify goal owner
                            goal_owner = db.get_user_by_id(target_user_id)
                            if goal_owner:
                                notify_feedback_given(selected_goal, user, goal_owner)
                            
                            st.success("✅ Feedback submitted successfully!")
                            st.rerun()
                        else:
                            st.error("❌ Failed to save feedback")
    else:
        if user['id'] != target_user_id:
            st.caption("ℹ️ You do not have permission to give feedback to this user.")

def display_profile():
    """Display and edit user profile"""
    user = st.session_state.user
    
    if not user:
        st.warning("⚠️ Session expired. Please login again.")
        st.rerun()

    st.title("👤 My Profile")
    
    tab1, tab2 = st.tabs(["📝 Edit Profile", "🔒 Change Password"])
    
    with tab1:
        with st.form("edit_profile"):
            st.subheader("Personal Information")
            
            col1, col2 = st.columns(2)
            with col1:
                new_name = st.text_input("Full Name*", value=user['name'])
                new_email = st.text_input("Email*", value=user['email'])
            
            with col2:
                new_designation = st.text_input("Designation", value=user.get('designation', ''))
                new_department = st.text_input("Department", value=user.get('department', ''))
            
            if st.form_submit_button("💾 Save Changes", use_container_width=True):
                if new_name and new_email:
                    updates = {
                        'name': new_name,
                        'email': new_email.lower().strip(),
                        'designation': new_designation,
                        'department': new_department
                    }
                    
                    if db.update_user(user['id'], updates):
                        st.session_state.user.update(updates)
                        st.success("✅ Profile updated successfully!")
                        st.rerun()
                    else:
                        st.error("❌ Failed to update profile")
                else:
                    st.error("❌ Name and Email are required")
    
    with tab2:
        # Initialize session state for forgot password in profile
        if 'profile_forgot_password' not in st.session_state:
            st.session_state.profile_forgot_password = False
        if 'profile_reset_token_sent' not in st.session_state:
            st.session_state.profile_reset_token_sent = False
        
        # Track password input for strength meter (OUTSIDE FORM)
        if 'temp_new_password' not in st.session_state:
            st.session_state.temp_new_password = ""
        
        # ===== NORMAL CHANGE PASSWORD =====
        if not st.session_state.profile_forgot_password and not st.session_state.profile_reset_token_sent:
            st.subheader("Change Password")
            
            old_password = st.text_input("Current Password*", type="password", key="old_pass_input")
            new_password = st.text_input("New Password*", type="password", key="new_pass_input")
            
            # Update temp password for strength meter
            if new_password != st.session_state.temp_new_password:
                st.session_state.temp_new_password = new_password
            
            # Password strength meter (OUTSIDE FORM)
            if new_password:
                render_password_strength_meter(new_password, "profile_change")
            
            confirm_password = st.text_input("Confirm New Password*", type="password", key="confirm_pass_input")
            
            col_submit, col_forgot = st.columns(2)
            
            with col_submit:
                if st.button("🔒 Change Password", use_container_width=True, key="change_pass_btn"):
                    if old_password == user['password']:
                        if new_password == confirm_password:
                            if len(new_password) >= 6:
                                # Check password strength
                                score, _, strength, _ = check_password_strength(new_password)
                                
                                if db.update_user(user['id'], {'password': new_password}):
                                    st.session_state.user['password'] = new_password
                                    st.session_state.temp_new_password = ""
                                    st.success("✅ Password changed successfully!")
                                    if score >= 70:
                                        st.info("🔒 Great! Your password is strong.")
                                    elif score < 30:
                                        st.warning(f"⚠️ Your password is {strength}. Consider making it stronger.")
                                    st.rerun()
                                else:
                                    st.error("❌ Failed to change password")
                            else:
                                st.error("❌ Password must be at least 6 characters")
                        else:
                            st.error("❌ Passwords don't match")
                    else:
                        st.error("❌ Current password is incorrect")
            
            with col_forgot:
                if st.button("🔑 Forgot Password?", use_container_width=True, key="forgot_pass_btn"):
                    st.session_state.profile_forgot_password = True
                    st.rerun()
        
        # ===== FORGOT PASSWORD - REQUEST TOKEN =====
        elif st.session_state.profile_forgot_password and not st.session_state.profile_reset_token_sent:
            st.subheader("🔑 Reset Password via Email")
            st.info("We'll send a reset token to your registered email address.")
            
            st.markdown(f"**Your Email:** {user['email']}")
            st.caption("A reset token will be sent to this email address")
            
            col_send, col_cancel = st.columns(2)
            
            with col_send:
                if st.button("📧 Send Reset Token", use_container_width=True, key="send_token_btn"):
                    # Generate token
                    token = db.create_password_reset_token(user['email'])
                    
                    if token:
                        # Try to send email
                        email_sent = send_password_reset_email(user['email'], token)
                        
                        if email_sent:
                            st.success("✅ Reset token sent to your email!")
                            st.info(f"**Backup Token:** `{token}`\n\n(In case you don't receive the email)")
                        else:
                            st.warning("⚠️ Could not send email. Please use this token:")
                            st.code(token, language=None)
                        
                        st.session_state.profile_reset_token_sent = True
                        st.rerun()
                    else:
                        st.error("❌ Failed to generate reset token")
            
            with col_cancel:
                if st.button("❌ Cancel", use_container_width=True, key="cancel_forgot_btn"):
                    st.session_state.profile_forgot_password = False
                    st.rerun()
        
        # ===== RESET PASSWORD WITH TOKEN =====
        elif st.session_state.profile_reset_token_sent:
            st.subheader("🔒 Enter Reset Token")
            st.info("Check your email for the reset token and enter it below with your new password.")
            
            reset_token = st.text_input("🎫 Reset Token*", placeholder="Enter 8-character token", key="reset_token_input")
            new_password_reset = st.text_input("🔒 New Password*", type="password", placeholder="••••••••", key="new_pass_reset_input")
            
            # Password strength meter (OUTSIDE FORM)
            if new_password_reset:
                render_password_strength_meter(new_password_reset, "profile_reset")
            
            confirm_password_reset = st.text_input("🔒 Confirm New Password*", type="password", placeholder="••••••••", key="confirm_pass_reset_input")
            
            col_reset, col_cancel = st.columns(2)
            
            with col_reset:
                if st.button("✅ Reset Password", use_container_width=True, key="reset_pass_btn"):
                    if reset_token and new_password_reset and confirm_password_reset:
                        if new_password_reset == confirm_password_reset:
                            if len(new_password_reset) >= 6:
                                # Check password strength
                                score, _, strength, _ = check_password_strength(new_password_reset)
                                
                                # Reset password using token
                                if db.reset_password_with_token(reset_token, new_password_reset):
                                    st.success("✅ Password reset successful!")
                                    if score >= 70:
                                        st.info("🔒 Great! Your password is strong.")
                                    elif score < 30:
                                        st.warning(f"⚠️ Your password is {strength}. Consider making it stronger.")
                                    st.balloons()
                                    
                                    # Update session
                                    st.session_state.user['password'] = new_password_reset
                                    st.session_state.profile_forgot_password = False
                                    st.session_state.profile_reset_token_sent = False
                                    st.rerun()
                                else:
                                    st.error("❌ Invalid or expired token. Please request a new one.")
                            else:
                                st.error("❌ Password must be at least 6 characters")
                        else:
                            st.error("❌ Passwords don't match")
                    else:
                        st.warning("⚠️ Please fill all fields")
            
            with col_cancel:
                if st.button("❌ Cancel", use_container_width=True, key="cancel_reset_btn"):
                    st.session_state.profile_forgot_password = False
                    st.session_state.profile_reset_token_sent = False
                    st.rerun()
            
            # Option to resend token
            st.markdown("---")
            if st.button("📧 Resend Reset Token", use_container_width=True, key="resend_token_btn"):
                token = db.create_password_reset_token(user['email'])
                if token:
                    email_sent = send_password_reset_email(user['email'], token)
                    if email_sent:
                        st.success("✅ New token sent to your email!")
                    st.info(f"**Backup Token:** `{token}`")
                else:
                    st.error("❌ Failed to generate new token")
                    
def display_permissions():
    """Manage user permissions (HR only)"""
    user = st.session_state.user
    
    if not user:
        st.warning("⚠️ Session expired. Please login again.")
        st.rerun()

    if user['role'] != 'HR':
        st.warning("⚠️ Only HR can access this page")
        return
    
    st.title("🔐 Permissions Management")
    
    st.info("**Note:** This section allows HR to manage granular permissions for users")
    
    all_users = db.get_all_users()
    
    # Permission categories
    permissions = {
        'view_all_goals': 'View All Goals',
        'edit_all_goals': 'Edit All Goals',
        'delete_all_goals': 'Delete All Goals',
        'view_all_feedback': 'View All Feedback',
        'edit_all_feedback': 'Edit All Feedback',
        'delete_all_feedback': 'Delete All Feedback',
        'create_users': 'Create Users',
        'edit_users': 'Edit Users',
        'delete_users': 'Delete Users',
        'manage_teams': 'Manage Teams',
        'view_analytics': 'View Analytics',
        'export_data': 'Export Data'
    }
    
    selected_user = st.selectbox(
        "Select User",
        [f"{u['name']} ({u['role']}) - {u['email']}" for u in all_users if u['id'] != user['id']]
    )
    
    if selected_user:
        user_email = selected_user.split(' - ')[1]
        selected_user_obj = next(u for u in all_users if u['email'] == user_email)
        
        st.subheader(f"Permissions for {selected_user_obj['name']}")
        
        # Get current permissions
        current_perms = db.get_user_permissions(selected_user_obj['id'])
        
        with st.form("update_permissions"):
            st.markdown("**Grant Permissions:**")
            
            selected_perms = []
            cols = st.columns(2)
            for idx, (perm_key, perm_label) in enumerate(permissions.items()):
                with cols[idx % 2]:
                    if st.checkbox(perm_label, value=perm_key in current_perms, key=perm_key):
                        selected_perms.append(perm_key)
            
            if st.form_submit_button("💾 Update Permissions", use_container_width=True):
                if db.update_user_permissions(selected_user_obj['id'], selected_perms):
                    st.success(f"✅ Permissions updated for {selected_user_obj['name']}")
                    st.rerun()
                else:
                    st.error("❌ Failed to update permissions")
        
        # Show current permissions
        st.markdown("---")
        st.subheader("Current Permissions")
        if current_perms:
            for perm in current_perms:
                st.success(f"✓ {permissions.get(perm, perm)}")
        else:
            st.info("No special permissions granted")


def display_employee_management():
    """Employee management - create users, teams"""
    user = st.session_state.user
    
    if not user:
        st.warning("⚠️ Session expired. Please login again.")
        st.rerun()

    # Check if user has permission to create employees
    if user['role'] not in ['CMD', 'VP', 'HR']:
        st.warning("⚠️ You don't have permission to access this page")
        return
    
    st.title(" Employee Management")
    
    tab1, tab2, tab3, tab4 = st.tabs(["👤 Create Employee", "👥 All Employees", "👥 Manage Teams", "📋 View All Teams"])
    
    # ===== CREATE EMPLOYEE TAB =====
    # ===== CREATE EMPLOYEE TAB =====
    with tab1:
        st.subheader("Create New Employee")
        
        with st.form("create_employee"):
            col1, col2 = st.columns(2)
            
            with col1:
                new_name = st.text_input("Full Name*")
                new_email = st.text_input("Email*")
                new_password = st.text_input("Password*", type="password")
                new_designation = st.text_input("Designation")
            
            with col2:
                # Show only roles that current user can create
                modifiable_roles = get_modifiable_roles(user['role'])
                new_role = st.selectbox("Role*", modifiable_roles)
                new_department = st.text_input("Department")
                new_joining_date = st.date_input("Joining Date*", value=date.today())
                
                managers = [u for u in db.get_all_users() if u['role'] == 'Manager']
                manager_options = ["None"] + [f"{m['name']} ({m['email']})" for m in managers]
                selected_manager = st.selectbox("Assign to Manager", manager_options)
            
            if st.form_submit_button("Create Employee", use_container_width=True):
                if new_name and new_email and new_password and new_role:
                    manager_id = None
                    if selected_manager != "None":
                        manager_email = selected_manager.split('(')[1].strip(')')
                        manager_id = next((m['id'] for m in managers if m['email'] == manager_email), None)
                    
                    employee_data = {
                        'name': new_name,
                        'email': new_email.lower().strip(),
                        'password': new_password,
                        'designation': new_designation,
                        'role': new_role,
                        'department': new_department,
                        'manager_id': manager_id,
                        'joining_date': str(new_joining_date)
                    }
                    
                    if db.create_user(employee_data):
                        st.success(f"✅ Employee {new_name} created successfully!")
                        # No notification needed for employee creation per new requirements
                    else:
                        st.error("❌ Failed to create employee")
                else:
                    st.error("❌ Please fill all required fields")
    
    # ===== ALL EMPLOYEES TAB (NEW) =====
    with tab2:
        st.subheader("All Employees in Organization")
        
        # Search and filter
        col_search1, col_search2, col_search3 = st.columns(3)
        with col_search1:
            search_emp = st.text_input("🔍 Search by name or email", key="all_emp_search")
        with col_search2:
            filter_role = st.selectbox("Filter by Role", ["All"] + ["CMD", "VP", "HR", "Manager", "Employee"], key="all_emp_role")
        with col_search3:
            all_depts = list(set([normalize_department(u.get('department')) for u in db.get_all_users()]))
            filter_dept = st.selectbox("Filter by Department", 
                ["All"] + sorted([d for d in all_depts if d != 'UNASSIGNED']))
        
        all_users_list = db.get_all_users()
        
        # Apply filters
        filtered_users = all_users_list
        if search_emp:
            filtered_users = [u for u in filtered_users if search_emp.lower() in u['name'].lower() or search_emp.lower() in u['email'].lower()]
        if filter_role != "All":
            filtered_users = [u for u in filtered_users if u['role'] == filter_role]
        if filter_dept != "All":
            filtered_users = [u for u in filtered_users if normalize_department(u.get('department')) == filter_dept]
        
        st.markdown(f"**Showing {len(filtered_users)} of {len(all_users_list)} employees**")
        st.markdown("---")
        
        # ============================================
        # ✅ NEW: SIMPLE LIST VIEW INSTEAD OF CARDS
        # ============================================
        
        # Create employee list with action buttons
        for idx, emp in enumerate(filtered_users):
            # Role badge color
            role_colors = {
                'CMD': '#8B0000',
                'VP': '#FF4500',
                'HR': '#4facfe',
                'Manager': '#f093fb',
                'Employee': '#dbeafe'
            }
            role_color = role_colors.get(emp['role'], '#dbeafe')
            
            # Create horizontal layout
            col1, col2, col3, col4, col5 = st.columns([3, 2, 2, 1, 1])
            
            with col1:
                st.markdown(f"**{emp['name']}**")
            
            with col2:
                st.markdown(f"{emp.get('designation', 'N/A')}")
            
            with col3:
                st.markdown(f"*{emp.get('department', 'N/A')}*")
            
            with col4:
                st.markdown(f"""<span style='background: {role_color}; color: white; padding: 4px 12px; 
                            border-radius: 12px; font-size: 11px; font-weight: bold; display: inline-block;'>
                    {emp['role']}</span>""", unsafe_allow_html=True)
            
            with col5:
                if st.button("👁️", key=f"view_all_emp_{emp['id']}_{idx}", help=f"View {emp['name']}'s details"):
                    st.session_state.viewing_employee_details = emp
                    st.rerun()
            
            st.markdown("---")
        
        # Employee Details Modal (keep existing code)
        if 'viewing_employee_details' in st.session_state:
            st.markdown("---")
            
            view_emp = st.session_state.viewing_employee_details
            
            st.markdown(f"""
            <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                        padding: 20px; border-radius: 10px; margin-bottom: 20px;">
                <h2 style="color: white; margin: 0;">👤 {view_emp['name']}'s Details</h2>
            </div>
            """, unsafe_allow_html=True)
            
            # Display current details
            col_detail1, col_detail2 = st.columns(2)
            
            with col_detail1:
                st.markdown("**Personal Information**")
                st.info(f"**Name:** {view_emp['name']}")
                st.info(f"**Email:** {view_emp['email']}")
                st.info(f"**Role:** {view_emp['role']}")
                st.info(f"**Designation:** {view_emp.get('designation', 'N/A')}")
            
            with col_detail2:
                st.markdown("**Organization Details**")
                st.info(f"**Department:** {view_emp.get('department', 'N/A')}")
                st.info(f"**Joining Date:** {view_emp.get('joining_date', 'N/A')}")
                st.info(f"**Exit Date:** {view_emp.get('exit_date', 'Not Set')}")
                
                # Manager info
                if view_emp.get('manager_id'):
                    manager = db.get_user_by_id(view_emp['manager_id'])
                    st.info(f"**Manager:** {manager['name'] if manager else 'N/A'}")
                else:
                    st.info("**Manager:** None")
            
            st.markdown("---")
            
            # Edit Employee Details Form
            with st.expander("✏️ Edit Employee Details", expanded=False):
                with st.form(f"edit_emp_details_{view_emp['id']}"):
                    col_edit1, col_edit2 = st.columns(2)
                    
                    with col_edit1:
                        edit_name = st.text_input("Full Name*", value=view_emp['name'])
                        edit_email = st.text_input("Email*", value=view_emp['email'])
                        edit_designation = st.text_input("Designation", value=view_emp.get('designation', ''))
                        edit_department = st.text_input("Department", value=view_emp.get('department', ''))
                    
                    with col_edit2:
                        edit_role = st.selectbox("Role*", ["CMD", "VP", "HR", "Manager", "Employee"], 
                                                index=["CMD", "VP", "HR", "Manager", "Employee"].index(view_emp['role']))
                        
                        # Joining date
                        joining_date_val = view_emp.get('joining_date')
                        if joining_date_val:
                            if isinstance(joining_date_val, str):
                                joining_date_val = datetime.strptime(joining_date_val, '%Y-%m-%d').date()
                        else:
                            joining_date_val = date.today()
                        edit_joining_date = st.date_input("Joining Date", value=joining_date_val)

                        # Exit date - Allow removal
                        exit_date_val = view_emp.get('exit_date')
                        remove_exit = st.checkbox("Remove Exit Date", value=(exit_date_val is None))

                        if not remove_exit:
                            if exit_date_val:
                                if isinstance(exit_date_val, str):
                                    exit_date_val = datetime.strptime(exit_date_val, '%Y-%m-%d').date()
                                edit_exit_date = st.date_input("Exit Date", value=exit_date_val)
                            else:
                                edit_exit_date = st.date_input("Exit Date", value=date.today())
                        else:
                            edit_exit_date = None
                        
                        # Manager assignment
                        managers = [u for u in db.get_all_users() if u['role'] == 'Manager' and u['id'] != view_emp['id']]
                        manager_options = ["None"] + [f"{m['name']} ({m['email']})" for m in managers]
                        
                        current_manager_idx = 0
                        if view_emp.get('manager_id'):
                            current_manager = db.get_user_by_id(view_emp['manager_id'])
                            if current_manager:
                                current_manager_str = f"{current_manager['name']} ({current_manager['email']})"
                                if current_manager_str in manager_options:
                                    current_manager_idx = manager_options.index(current_manager_str)
                        
                        edit_manager = st.selectbox("Assign to Manager", manager_options, index=current_manager_idx)
                    
                    col_save, col_cancel = st.columns(2)
                    
                    with col_save:
                        if st.form_submit_button("💾 Save Changes", use_container_width=True):
                            if edit_name and edit_email and edit_role:
                                manager_id = None
                                if edit_manager != "None":
                                    manager_email = edit_manager.split('(')[1].strip(')')
                                    manager_id = next((m['id'] for m in managers if m['email'] == manager_email), None)
                                
                                updates = {
                                    'name': edit_name,
                                    'email': edit_email.lower().strip(),
                                    'designation': edit_designation,
                                    'role': edit_role,
                                    'department': edit_department,
                                    'manager_id': manager_id,
                                    'joining_date': str(edit_joining_date),
                                    'exit_date': str(edit_exit_date) if edit_exit_date and not remove_exit else None
                                }
                                
                                if db.update_user(view_emp['id'], updates):
                                    st.success(f"✅ Employee {edit_name} updated successfully!")
                                    del st.session_state.viewing_employee_details
                                    st.rerun()
                                else:
                                    st.error("❌ Failed to update employee")
                            else:
                                st.error("❌ Please fill all required fields")
                    
                    with col_cancel:
                        if st.form_submit_button("❌ Cancel", use_container_width=True):
                            # Just close the form, keep modal open
                            pass
            
            # Close Details Button
            if st.button("← Back to All Employees", use_container_width=True):
                del st.session_state.viewing_employee_details
                st.rerun()
    
    
    
    # ===== MANAGE TEAMS TAB =====
    with tab3:
        st.subheader("Assign Employees to Managers")
        
        managers = [u for u in db.get_all_users() if u['role'] == 'Manager']
        employees = [u for u in db.get_all_users() if u['role'] == 'Employee']
        
        if not managers:
            st.info("No managers found. Create a manager first.")
        elif not employees:
            st.info("No employees found. Create employees first.")
        else:
            selected_manager_name = st.selectbox(
                "Select Manager",
                [f"{m['name']} ({m['email']})" for m in managers]
            )
            
            manager_email = selected_manager_name.split('(')[1].strip(')')
            selected_manager = next(m for m in managers if m['email'] == manager_email)
            
            current_team = db.get_team_members(selected_manager['id'])
            st.write(f"**Current Team Size:** {len(current_team)}")
            
            unassigned_employees = [e for e in employees if not e.get('manager_id')]
            
            if unassigned_employees:
                selected_employees = st.multiselect(
                    "Select Employees to Add",
                    [f"{e['name']} ({e['email']})" for e in unassigned_employees]
                )
                
                if st.button("Assign to Team", use_container_width=True):
                    for emp_str in selected_employees:
                        emp_email = emp_str.split('(')[1].strip(')')
                        emp_id = next(e['id'] for e in unassigned_employees if e['email'] == emp_email)
                        db.update_user(emp_id, {'manager_id': selected_manager['id']})
                    
                    st.success(f"✅ Assigned {len(selected_employees)} employees to {selected_manager['name']}")
                    st.rerun()
            else:
                st.info("No unassigned employees available")
    
    # ===== VIEW ALL TEAMS TAB =====
    with tab4:
        st.subheader("All Teams Overview")
        
        managers = [u for u in db.get_all_users() if u['role'] == 'Manager']
        for manager in managers:
            with st.expander(f"👔 {manager['name']}'s Team ({manager.get('department', 'N/A')})"):
                team = db.get_team_members(manager['id'])
                
                if team:
                    team_data = []
                    for member in team:
                        team_data.append({
                            'Name': member['name'],
                            'Email': member['email'],
                            'Designation': member.get('designation', 'N/A'),
                            'Department': member.get('department', 'N/A')
                        })
                    
                    df = pd.DataFrame(team_data)
                    st.dataframe(df, use_container_width=True)
                    
                    # Option to remove members
                    remove_member = st.selectbox(
                        "Remove Member",
                        ["None"] + [m['name'] for m in team],
                        key=f"remove_{manager['id']}"
                    )
                    
                    if remove_member != "None" and st.button(f"Remove {remove_member}", key=f"btn_remove_{manager['id']}"):
                        member_id = next(m['id'] for m in team if m['name'] == remove_member)
                        db.update_user(member_id, {'manager_id': None})
                        st.success(f"✅ Removed {remove_member} from team")
                        st.rerun()
                else:
                    st.info("No team members yet")

    

def display_approval_page():
    """Display pending goal and achievement approvals for managers"""
    user = st.session_state.user
    
    if user['role'] not in ['Manager', 'HR', 'VP']:
        st.warning("⚠️ Only Managers can access this page")
        return
    
    st.title("✅ Approval Dashboard")
    
    # Create tabs for different approval types
    tab1, tab2 = st.tabs(["📋 Goal Approvals", "📊 Achievement Approvals"])
    
    # ===== TAB 1: GOAL APPROVALS =====
    with tab1:
        st.subheader("Pending Goal Approvals")
        
        # Get pending goal approvals
        pending_goals = db.get_pending_approvals(user['id'])
        
        if not pending_goals:
            st.success("🎉 No pending goal approvals!")
        else:
            st.markdown(f"**You have {len(pending_goals)} goal(s) awaiting approval**")
            st.markdown("---")
            
            for goal in pending_goals:
                # Get employee info
                employee = db.get_user_by_id(goal['user_id'])
                employee_name = employee['name'] if employee else 'Unknown'
                
                with st.expander(f"📋 {goal['goal_title']} - {employee_name}"):
                    col1, col2 = st.columns([2, 1])
                    
                    with col1:
                        st.markdown(f"**Employee:** {employee_name}")
                        st.markdown(f"**Department:** {goal.get('department', 'N/A')}")
                        st.markdown(f"**KPI:** {goal.get('kpi', 'N/A')}")
                        st.markdown(f"**Monthly Target:** {goal.get('monthly_target', 0)}")
                        st.markdown(f"**Period:** {goal.get('year')}-Q{goal.get('quarter')}-M{goal.get('month')}")
                        
                        # Weekly Targets Section
                        st.markdown("**Weekly Breakdown:**")
                        col_w1, col_w2, col_w3, col_w4 = st.columns(4)
                        
                        with col_w1:
                            st.metric("Week 1", goal.get('week1_target', 0))
                        with col_w2:
                            st.metric("Week 2", goal.get('week2_target', 0))
                        with col_w3:
                            st.metric("Week 3", goal.get('week3_target', 0))
                        with col_w4:
                            st.metric("Week 4", goal.get('week4_target', 0))
                        
                        st.markdown(f"**Description:** {goal.get('goal_description', 'No description')}")
                    
                    with col2:
                        st.markdown("**Actions:**")
                        
                        if st.button("✅ Approve", key=f"approve_goal_{goal['goal_id']}", use_container_width=True):
                            if db.update_goal_approval(goal['goal_id'], 'approved', user['id']):
                                st.success("✅ Goal approved!")
                                
                                # Notify employee and HR
                                employee = db.get_user_by_id(goal['user_id'])
                                if employee:
                                    notify_goal_approved(goal, user, employee)
                                
                                st.rerun()
                            else:
                                st.error("❌ Failed to approve goal")
                        
                        if st.button("❌ Reject", key=f"reject_goal_{goal['goal_id']}", use_container_width=True):
                            st.session_state[f'rejecting_goal_{goal["goal_id"]}'] = True
                        
                        # Rejection reason form
                        if st.session_state.get(f'rejecting_goal_{goal["goal_id"]}'):
                            with st.form(f"reject_goal_form_{goal['goal_id']}"):
                                reason = st.text_area("Rejection Reason*", key=f"reason_goal_{goal['goal_id']}")
                                
                                col_r1, col_r2 = st.columns(2)
                                with col_r1:
                                    if st.form_submit_button("Confirm Reject"):
                                        if reason.strip():
                                            if db.update_goal_approval(
                                                goal['goal_id'], 
                                                'rejected', 
                                                user['id'],
                                                reason
                                            ):
                                                st.success("Goal rejected")
                                                del st.session_state[f'rejecting_goal_{goal["goal_id"]}']
                                                st.rerun()
                                        else:
                                            st.error("Please provide a reason")
                                
                                with col_r2:
                                    if st.form_submit_button("Cancel"):
                                        del st.session_state[f'rejecting_goal_{goal["goal_id"]}']
                                        st.rerun()
                    
                    st.markdown("---")
    
    # ===== TAB 2: ACHIEVEMENT APPROVALS =====
    with tab2:
        st.subheader("Pending Achievement Approvals")
        
        # Get team members
        team_members = db.get_team_members(user['id'])
        
        pending_achievements = []
        for member in team_members:
            goals = db.get_user_all_goals(member['id'])
            for goal in goals:
                if goal.get('achievement_approval_status') == 'pending':
                    goal['employee'] = member
                    pending_achievements.append(goal)
        
        if not pending_achievements:
            st.success("🎉 No pending achievement approvals!")
        else:
            st.markdown(f"**You have {len(pending_achievements)} achievement update(s) awaiting approval**")
            st.markdown("---")
            
            for goal in pending_achievements:
                employee = goal['employee']
                
                with st.expander(f" {goal['goal_title']} - {employee['name']}"):
                    col1, col2 = st.columns([2, 1])
                    
                    with col1:
                        st.markdown(f"**Employee:** {employee['name']}")
                        st.markdown(f"**Goal:** {goal['goal_title']}")
                        st.markdown(f"**Department:** {goal.get('department', 'N/A')}")
                        st.markdown(f"**Monthly Target:** {goal.get('monthly_target', 0)}")
                        
                        # Show pending achievements
                        st.markdown("**Reported Achievements:**")
                        
                        col_w1, col_w2, col_w3, col_w4 = st.columns(4)
                        with col_w1:
                            w1_val = goal.get('week1_achievement_pending')
                            st.metric("Week 1", w1_val if w1_val is not None else '-')
                        with col_w2:
                            w2_val = goal.get('week2_achievement_pending')
                            st.metric("Week 2", w2_val if w2_val is not None else '-')
                        with col_w3:
                            w3_val = goal.get('week3_achievement_pending')
                            st.metric("Week 3", w3_val if w3_val is not None else '-')
                        with col_w4:
                            w4_val = goal.get('week4_achievement_pending')
                            st.metric("Week 4", w4_val if w4_val is not None else '-')
                        
                        monthly_pending = goal.get('monthly_achievement_pending', 0)
                        st.markdown(f"**Monthly Total:** {monthly_pending if monthly_pending is not None else 0}")
                        
                        # Show submission time
                        updated_at = goal.get('achievement_updated_at')
                        if updated_at:
                            st.caption(f"Submitted: {updated_at[:16]}")
                    
                    with col2:
                        st.markdown("**Actions:**")
                        
                        if st.button("✅ Approve", key=f"approve_ach_{goal['goal_id']}", use_container_width=True):
                            # Move pending to actual
                            updates = {
                                'week1_achievement': goal.get('week1_achievement_pending'),
                                'week2_achievement': goal.get('week2_achievement_pending'),
                                'week3_achievement': goal.get('week3_achievement_pending'),
                                'week4_achievement': goal.get('week4_achievement_pending'),
                                'monthly_achievement': goal.get('monthly_achievement_pending'),
                                'week1_rating': goal.get('week1_rating_pending'),
                                'week2_rating': goal.get('week2_rating_pending'),
                                'week3_rating': goal.get('week3_rating_pending'),
                                'week4_rating': goal.get('week4_rating_pending'),
                                'achievement_approval_status': 'approved',
                                'achievement_approved_by': user['id'],
                                'achievement_approved_at': datetime.now(IST).isoformat()
                            }
                            
                            if db.update_goal(goal['goal_id'], updates):
                                st.success("✅ Achievements approved!")
                                
                                # Notify employee
                                create_notification({
                                    'user_id': employee['id'],
                                    'action_by': user['id'],
                                    'action_by_name': user['name'],
                                    'action_type': 'achievement_approved',
                                    'details': f"Your achievements for '{goal['goal_title']}' have been approved",
                                    'is_read': False,
                                    'created_at': datetime.now(IST).isoformat()
                                })
                                
                                st.rerun()
                            else:
                                st.error("❌ Failed to approve achievements")
                        
                        if st.button("❌ Reject", key=f"reject_ach_{goal['goal_id']}", use_container_width=True):
                            st.session_state[f'rejecting_ach_{goal["goal_id"]}'] = True
                        
                        # Rejection form
                        if st.session_state.get(f'rejecting_ach_{goal["goal_id"]}'):
                            with st.form(f"reject_ach_form_{goal['goal_id']}"):
                                reason = st.text_area("Rejection Reason*", key=f"reason_ach_{goal['goal_id']}")
                                
                                col_r1, col_r2 = st.columns(2)
                                with col_r1:
                                    if st.form_submit_button("Confirm Reject"):
                                        if reason.strip():
                                            updates = {
                                                'achievement_approval_status': 'rejected',
                                                'achievement_rejection_reason': reason,
                                                'achievement_rejected_by': user['id'],
                                                'achievement_rejected_at': datetime.now(IST).isoformat()
                                            }
                                            
                                            if db.update_goal(goal['goal_id'], updates):
                                                st.success("Achievements rejected")
                                                
                                                # Notify employee
                                                create_notification({
                                                    'user_id': employee['id'],
                                                    'action_by': user['id'],
                                                    'action_by_name': user['name'],
                                                    'action_type': 'achievement_rejected',
                                                    'details': f"Your achievements for '{goal['goal_title']}' were rejected. Reason: {reason}",
                                                    'is_read': False,
                                                    'created_at': datetime.now(IST).isoformat()
                                                })
                                                
                                                del st.session_state[f'rejecting_ach_{goal["goal_id"]}']
                                                st.rerun()
                                        else:
                                            st.error("Please provide a reason")
                                
                                with col_r2:
                                    if st.form_submit_button("Cancel"):
                                        del st.session_state[f'rejecting_ach_{goal["goal_id"]}']
                                        st.rerun()
                    
                    st.markdown("---")
# ============================================
# SIDEBAR
# ============================================

import base64

def img_to_base64(path):
    with open(path, "rb") as f:
        return base64.b64encode(f.read()).decode()
    

def render_sidebar():
    """Render sidebar with navigation"""
    user = st.session_state.user
    role = user['role']
    
    with st.sidebar:
        # User profile with role indicator
        
        logo = img_to_base64("infopaceee.jpg")

        st.markdown(
        f"""
        <div style="
            width: 100%;
            padding: 6px 12px 16px 12px; 
            box-sizing: border-box;
            text-align: center;
            border-bottom: 1px solid rgba(255,255,255,0.08);">
            <img src="data:image/png;base64,{logo}"
                style="
                    height: 64px;          /* ⬅ BIG but controlled */
                    width: auto;
                    max-width: 90%;
                    object-fit: contain;
                    display: block;
                    margin: 0 auto;        /* ⬅ perfect centering */
                    pointer-events: none;
                    filter: brightness(0.92) contrast(0.95);
                    background: transparent;
                    ">
        </div>
        """,
        unsafe_allow_html=True
    )
        render_user_avatar(user)

        st.markdown("---")
        
        # Navigation Menu
        
        if st.button("Dashboard", use_container_width=True, key="nav_dashboard"):
            st.session_state.page = 'dashboard'
            st.session_state.pop('viewing_employee', None)
            st.session_state.pop('viewing_employee_year', None)
            save_session_to_storage()
            st.rerun()
        
        if st.button("My Goals", use_container_width=True, key="nav_my_goals"):
            st.session_state.page = 'my_goals'
            st.session_state.pop('viewing_employee', None)
            st.session_state.pop('viewing_employee_year', None)
            save_session_to_storage()
            st.rerun()
        
        
        # Role-specific navigation
        # Role-specific navigation - UPDATED LABELS
        if role in ['CMD', 'VP', 'HR', 'Manager']:
            # Determine label based on role
            if role == 'CMD':
                team_label = "Organization Overview"
            elif role == 'VP':
                team_label = "HR & Managers"
            elif role == 'HR':
                team_label = "Managers"
            else:  # Manager
                team_label = "My Team"
            
            if st.button(team_label, use_container_width=True, key="nav_employees"):
                st.session_state.page = 'employees'
                st.session_state.pop('viewing_employee', None)
                st.session_state.pop('viewing_employee_year', None)
                save_session_to_storage()
                st.rerun()
        
        
        if role in ['Manager']:
            # Get pending count
            pending_count = len(db.get_pending_approvals(user['id']))
            
            approval_label = f" Approvals ({pending_count})" if pending_count > 0 else " Approvals"
            
            if st.button(approval_label, use_container_width=True, key="nav_approvals"):
                st.session_state.page = 'approvals'
                save_session_to_storage()
                st.rerun()

        if role in 'HR':
            if st.button("Employee Management", use_container_width=True, key="nav_emp_mgmt"):
                st.session_state.page = 'employee_management'
                st.session_state.pop('viewing_employee', None)
                st.session_state.pop('viewing_employee_year', None)
                save_session_to_storage()
                st.rerun()
            
            if st.button("Organization Info", use_container_width=True, key="nav_hr_info"):
                st.session_state.page = 'hr_info'
                st.session_state.pop('viewing_employee', None)
                st.session_state.pop('viewing_employee_year', None)
                save_session_to_storage()
                st.rerun()
        
        if st.button("Analytics", use_container_width=True, key="nav_analytics"):
            st.session_state.page = 'analytics'
            save_session_to_storage()
            st.rerun()
        
        if st.button("Feedback History", use_container_width=True, key="nav_feedback"):
            st.session_state.page = 'feedback_history'
            st.session_state.pop('viewing_employee', None)
            st.session_state.pop('viewing_employee_year', None)
            save_session_to_storage()
            st.rerun()
        
        
        # Settings
        st.markdown("---")
        st.markdown("""
        <span style="color:white; font-weight:700;">Settings</span>
        """, unsafe_allow_html=True)
        
        if st.button("Profile", use_container_width=True, key="nav_profile"):
            st.session_state.page = 'profile'
            save_session_to_storage()
            st.rerun()
        # Settings

        


        # Logout
        st.markdown("---")
        if st.button("Logout", use_container_width=True):
            st.query_params.clear()
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.rerun()
# ============================================
# MAIN APPLICATION
# ============================================
def main():

    # ✅ NEW: Handle approval/rejection from email links
    query_params = st.query_params
    
    if 'action' in query_params and 'token' in query_params:
        action = query_params['action']
        token = query_params['token']
        
        # Get goal by token
        token_data = db.get_goal_by_token(token)
        
        if token_data:
            goal = token_data.get('goals')
            
            if action == 'approve':
                # Auto-login as manager or show login
                st.success(f"✅ Approving goal: {goal.get('goal_title')}")
                
                # You need to handle authentication here
                # Option 1: Auto-login with token
                # Option 2: Redirect to login page with action stored
                
                if db.update_goal_approval(goal['goal_id'], 'approved', token_data['approved_by']):
                    db.mark_token_used(token)
                    st.success("✅ Goal approved successfully!")
                    st.balloons()
                    
            elif action == 'reject':
                st.warning(f"❌ Rejecting goal: {goal.get('goal_title')}")
                
                reason = st.text_area("Please provide a rejection reason:")
                if st.button("Confirm Rejection"):
                    if reason.strip():
                        if db.update_goal_approval(goal['goal_id'], 'rejected', token_data['approved_by'], reason):
                            db.mark_token_used(token)
                            st.success("Goal rejected")
        else:
            st.error("❌ Invalid or expired approval link")
        
        # Clear query params
        st.query_params.clear()
        return
    
    def test_notifications():
    
        st.title("🧪 Notification Testing Dashboard")
        
        user = st.session_state.user
        if not user:
            st.warning("Please login first")
            return
        
        st.markdown(f"**Testing notifications for:** {user['name']} ({user['role']})")

        with st.expander("📋 Notification Routing Rules", expanded=False):
            st.markdown("""
            **Based on your role, notifications will be sent to:**
            
            - **If you are Employee:**
            - Goal created → Manager + HR
            - Weekly achievement → Manager + HR
            - Goal completed → Manager + HR
            
            - **If you are Manager:**
            - Goal created → VP
            - Weekly achievement → VP
            - Goal completed → VP
            
            - **If you are HR:**
            - Goal created → VP
            - Weekly achievement → VP
            - Goal completed → VP
            
            - **If you are VP:**
            - Goal created → CMD
            - Weekly achievement → CMD
            - Goal completed → CMD
            """)
        
        st.markdown("---")
        
        
        # Test 1: Goal Created
        st.subheader("1️⃣ Test Goal Creation Notification")
        if st.button("Create Test Goal Notification"):
            test_goal = {
                'goal_title': 'Test Goal - Notification Check',
                'user_id': user['id']
            }
            notify_goal_created(test_goal, user)
            st.success("✅ Goal creation notification sent!")
            st.info("Check Dashboard to see notification")
        
        st.markdown("---")
    
        # Test 2: Goal Approved
        st.subheader("2️⃣ Test Goal Approval Notification")
        if st.button("Create Test Approval Notification"):
            test_goal = {
                'goal_title': 'Test Goal - Approved',
                'goal_id': 'test_123'
            }
            test_employee = user
            notify_goal_approved(test_goal, user, test_employee)
            st.success("✅ Goal approval notification sent!")
        
        st.markdown("---")
        
        # Test 3: Weekly Achievement Updated
        st.subheader("3️⃣ Test Weekly Achievement Notification")
        if st.button("Create Test Weekly Update Notification"):
            test_goal = {
                'goal_title': 'Test Goal - Weekly Update',
                'user_id': user['id']
            }
            # ✅ Test as if HR is updating (should notify VP)
            notify_weekly_achievement_updated(test_goal, user, 1)
            st.success("✅ Weekly achievement notification sent!")
            st.info(f"Notification should go to: {'VP' if user['role'] == 'HR' else 'HR/Manager'}")

        st.markdown("---")

        # Test 4: Goal Completed
        st.subheader("4️⃣ Test Goal Completion Notification")
        if st.button("Create Test Completion Notification"):
            test_goal = {
                'goal_title': 'Test Goal - Completed',
                'user_id': user['id']
            }
            # ✅ Test as if HR is completing (should notify VP)
            notify_goal_completed(test_goal, user)
            st.success("✅ Goal completion notification sent!")
            st.info(f"Notification should go to: {'VP' if user['role'] == 'HR' else 'HR/Manager'}")
        st.markdown("---")

        # Test 5: Goal Not Completed
        st.subheader("5️⃣ Test Goal Deadline Missed Notification")
        if st.button("Create Test Deadline Missed Notification"):
            test_goal = {
                'goal_title': 'Test Goal - Deadline Missed',
                'user_id': user['id']
            }
            notify_goal_not_completed(test_goal, user)
            st.success("✅ Goal deadline missed notification sent!")
        
        st.markdown("---")
        
        # Test 6: Feedback Given
        st.subheader("6️⃣ Test Feedback Notification")
        if st.button("Create Test Feedback Notification"):
            test_goal = {
                'goal_title': 'Test Goal - Feedback',
                'user_id': user['id']
            }
            notify_feedback_given(test_goal, user, user)
            st.success("✅ Feedback notification sent!")
        
        st.markdown("---")
        
        # Test 7: Goal Due Soon
        st.subheader("7️⃣ Test Goal Due Soon Notification")
        if st.button("Create Test Due Soon Notification"):
            test_goal = {
                'goal_title': 'Test Goal - Due Soon',
                'user_id': user['id']
            }
            notify_goal_due_soon(test_goal, user, 3)
            st.success("✅ Goal due soon notification sent!")
        
        st.markdown("---")
        
        # View all notifications
        st.subheader("📋 Your Recent Notifications")
        notifications = get_user_notifications(user['id'], limit=20)
        
        if notifications:
            for notif in notifications:
                status = "✅ Read" if notif.get('is_read') else "🔴 Unread"
                st.markdown(f"""
                **{status}** | **{notif.get('action_type')}**  
                {notif.get('details')}  
                *{notif.get('created_at', 'Unknown time')}*
                """)
                st.markdown("---")
        else:
            st.info("No notifications found")
        
        # Clear test notifications
        st.markdown("---")
        if st.button("🗑️ Clear All Test Notifications", type="primary"):
            try:
                # Delete notifications containing "Test Goal"
                result = supabase.table('notifications').delete().eq(
                    'user_id', user['id']
                ).like('details', '%Test Goal%').execute()
                st.success("✅ Test notifications cleared!")
                st.rerun()
            except Exception as e:
                st.error(f"Error clearing notifications: {str(e)}")
        

    if not st.session_state.user:
        login_page()
        return
    
    # Render sidebar
    render_sidebar()
    
    # Route to appropriate page
    page = st.session_state.get('page', 'dashboard')
    
    if page == 'dashboard':
        display_dashboard()
    elif page == 'my_goals':
        display_my_goals()
    elif page == 'view_all_goals':
        display_view_all_goals()
    elif page == 'quarters':
        display_quarter_selection()
    elif page == 'months':
        display_month_selection()
    elif page == 'month_goals':
        display_month_goals()
    elif page == 'employees':
        display_employees_page()
    elif page == 'vp_team_view':  # ADD THIS
        display_vp_team_view()
    elif page == 'hr_team_view':  # ADD THIS
        display_hr_team_view()
    elif page == 'manager_team_view':  # ADD THIS
        display_manager_team_view()
    elif page == 'employee_goals':
        display_employee_goals()
    elif page == 'employee_quarters':
        display_quarter_selection()
    elif page == 'employee_months':
        display_month_selection()
    elif page == 'employee_month_goals':
        display_month_goals()
    elif page == 'employee_management':
        display_employee_management()
    elif page == 'hr_info':
        display_hr_info()
    elif page == 'approvals':
        display_approval_page()
    elif page == 'analytics':  # ADD THIS
        display_analytics_page()
    elif page == 'feedback_history':
        display_feedback_history()
    elif page == 'profile':
        display_profile()
    elif page == 'permissions':
        display_permissions()


    else:
        # Default to dashboard
        st.session_state.page = 'dashboard'
        st.rerun()

st.markdown(
    """
    <style>
        .pms-footer {
            position: fixed;
            left: 0;
            bottom: 0;
            width: 100%;
            background: #F8FAFC;
            border-top: 1px solid #E2E8F0;
            text-align: center;
            padding: 8px 0;
            font-size: 12px;
            color: #475569;
            z-index: 999;
        }
    </style>

    <div class="pms-footer">
        © 2025 Infopace Management Pvt. Ltd. • All rights reserved
    </div>
    """,
    unsafe_allow_html=True
)

 

if __name__ == "__main__":
    main()
